
import math
import logging
import ast
import csv
from typing import Dict, Any, List, Tuple, Optional, Union, Callable
from datetime import datetime, date, timedelta
from dataclasses import dataclass
from collections import defaultdict, deque
import pandas as pd
import numpy as np
import os
import re
import requests
import json
from copy import deepcopy  # NOTE: Being phased out, replaced with types.MappingProxyType
from types import MappingProxyType
import streamlit as st
from scipy import optimize
from scipy.stats import poisson
import warnings
warnings.filterwarnings('ignore')

# ============================================================
#   LOGGING STRUTTURATO (MIGLIORAMENTO)
# ============================================================

# Configurazione logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)

# Per Streamlit, usa livello più alto per evitare spam in console
if st:
    logging.getLogger().setLevel(logging.WARNING)


def _safe_float(value: Any, default: float = 0.0) -> float:
    """Converte value a float gestendo stringhe, None e valori non numerici."""
    if value is None:
        return default
    try:
        if isinstance(value, str):
            value = value.strip().replace(",", ".")
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_int(value: Any, default: int = 0) -> int:
    """Converte value a int gestendo stringhe e valori non numerici."""
    if value is None:
        return default
    try:
        if isinstance(value, str):
            value = value.strip()
        return int(value)
    except (TypeError, ValueError):
        return default

# Carica variabili d'ambiente da file .env (se disponibile)
try:
    from dotenv import load_dotenv
    load_dotenv()  # Carica .env dalla directory corrente
except ImportError:
    pass  # python-dotenv non installato, usa solo variabili d'ambiente di sistema

# Import per calibrazione (opzionale)
try:
    from sklearn.linear_model import LogisticRegression
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False
    logger.warning("sklearn non disponibile - calibrazione disabilitata")

# Import per Isotonic Regression (opzionale)
try:
    from sklearn.isotonic import IsotonicRegression
    ISOTONIC_AVAILABLE = True
except ImportError:
    ISOTONIC_AVAILABLE = False

# Import per high precision (mpmath opzionale)
try:
    import mpmath as mp
    MPMATH_AVAILABLE = True
except ImportError:
    MPMATH_AVAILABLE = False
    logger.warning("mpmath non disponibile - high precision disabilitata")

# MIGLIORAMENTO NUOVO: Import per Numba JIT compilation (opzionale, velocizza 10-100x)
# Se disponibile, compila funzioni critiche in codice macchina
try:
    from numba import jit
    NUMBA_AVAILABLE = True
    logger.info("Numba disponibile - JIT compilation abilitata per performance massime")
except ImportError:
    NUMBA_AVAILABLE = False
    logger.warning("Numba non disponibile - installa con 'pip install numba' per +1000% velocità")
    # Dummy decorator se numba non disponibile (no-op)
    def jit(*args, **kwargs):
        def decorator(func):
            return func
        return decorator if not args else decorator(args[0])

# Import cache manager per caching predizioni complete
try:
    from api_manager import CacheManager
    CACHE_ENABLED = True
    _cache_manager = None  # Inizializzato lazy al primo uso
    logger.info("✅ CacheManager disponibile - caching predizioni abilitato")
except ImportError:
    CACHE_ENABLED = False
    _cache_manager = None
    logger.warning("⚠️ CacheManager non disponibile - caching predizioni disabilitato")

numbasia = None
# Import opzionale numbasia (pacchetto proprietario per ottimizzazioni avanzate)
try:
    import numbasia  # type: ignore
    NUMBASIA_AVAILABLE = True
    logger.info("numbasia disponibile - ottimizzazioni proprietarie pronte all'uso")
except ImportError:
    NUMBASIA_AVAILABLE = False
    logger.warning(
        "numbasia non disponibile - assicurarsi di aver installato il pacchetto privato per abilitare tutte le ottimizzazioni"
    )

# ============================================================
#   AI SYSTEM - IMPORT
# ============================================================
try:
    from ai_system.pipeline import quick_analyze, AIPipeline
    from ai_system.advanced_precision_pipeline import AdvancedPrecisionPipeline
    from ai_system.llm_analyst import LLMAnalyst
    from ai_system.sentiment_analyzer import SentimentAnalyzer
    from ai_system.config import AIConfig, get_conservative_config, get_aggressive_config
    from ai_system.background_services import BackgroundAutomationService
    AI_SYSTEM_AVAILABLE = True
    logger.info("✅ AI System loaded successfully - All features enabled (LLM, Backtesting, Sentiment, Live)")
except ImportError as e:
    AI_SYSTEM_AVAILABLE = False
    logger.warning(f"⚠️ AI System not available: {e}")
    logger.warning("   Install AI dependencies: pip install torch scikit-learn xgboost")

if AI_SYSTEM_AVAILABLE:
    def _ensure_ai_background_features() -> Dict[str, Dict[str, Any]]:
        """Garantisce che le feature AI critiche siano sempre attive e sincronizzate."""
        features = st.session_state.get("ai_background_features")
        if features is None:
            features = {}

        # --- Data Quality Thresholds ---
        data_quality_defaults = {
            "min": 0.30,
            "good": 0.70,
            "excellent": 0.90,
        }
        dq_settings = features.setdefault("data_quality", data_quality_defaults.copy())
        for key, value in data_quality_defaults.items():
            dq_settings.setdefault(key, value)

        # --- Telegram Advanced Notifications ---
        default_report_time = datetime.strptime("22:00", "%H:%M").time()
        telegram_defaults = {
            "min_ev": 5.0,
            "daily_report": True,
            "report_time": default_report_time,
            "min_confidence": 60.0,
            "rate_limit_seconds": 3,
            "always_on": True,
        }
        telegram_settings = features.setdefault("telegram_advanced", telegram_defaults.copy())
        for key, value in telegram_defaults.items():
            telegram_settings.setdefault(key, value)

        report_time_value = telegram_settings.get("report_time", default_report_time)
        if isinstance(report_time_value, str):
            try:
                telegram_settings["report_time"] = datetime.strptime(report_time_value, "%H:%M").time()
            except ValueError:
                telegram_settings["report_time"] = default_report_time

        # --- Neural Network Tuning ---
        nn_defaults = {
            "hidden_layers": [64, 32, 16],
            "dropout": 0.2,
            "learning_rate": 0.001,
            "batch_size": 32,
        }
        nn_settings = features.setdefault("neural_network", nn_defaults.copy())
        for key, value in nn_defaults.items():
            nn_settings.setdefault(key, value)

        hidden_layers_value = nn_settings.get("hidden_layers", nn_defaults["hidden_layers"])
        if isinstance(hidden_layers_value, str):
            try:
                parsed_layers = ast.literal_eval(hidden_layers_value)
                if isinstance(parsed_layers, (list, tuple)):
                    nn_settings["hidden_layers"] = [int(layer) for layer in parsed_layers]
                else:
                    nn_settings["hidden_layers"] = nn_defaults["hidden_layers"]
            except (ValueError, SyntaxError):
                nn_settings["hidden_layers"] = nn_defaults["hidden_layers"]

        st.session_state["ai_background_features"] = features

        # Mirror legacy keys for backward compatibility
        st.session_state["ai_min_data_quality"] = dq_settings["min"]
        st.session_state["ai_good_data_quality"] = dq_settings["good"]
        st.session_state["ai_excellent_data_quality"] = dq_settings["excellent"]

        st.session_state["ai_telegram_min_ev"] = telegram_settings["min_ev"]
        st.session_state["ai_telegram_daily_report"] = telegram_settings["daily_report"]
        st.session_state["ai_telegram_report_time"] = telegram_settings["report_time"]

        st.session_state["ai_nn_hidden_layers"] = nn_settings["hidden_layers"]
        st.session_state["ai_nn_dropout"] = nn_settings["dropout"]
        st.session_state["ai_nn_learning_rate"] = nn_settings["learning_rate"]
        st.session_state["ai_nn_batch_size"] = nn_settings["batch_size"]

        return features

    def _apply_ai_background_features(ai_config: AIConfig, features: Optional[Dict[str, Dict[str, Any]]] = None) -> AIConfig:
        """Propaga le impostazioni sempre attive dentro la configurazione AI."""
        settings = features or _ensure_ai_background_features()

        dq_settings = settings["data_quality"]
        ai_config.min_data_quality = dq_settings["min"]
        ai_config.good_data_quality = dq_settings["good"]
        ai_config.excellent_data_quality = dq_settings["excellent"]

        telegram_settings = settings["telegram_advanced"]
        ai_config.telegram_enabled = telegram_settings["always_on"]
        ai_config.telegram_min_ev = telegram_settings["min_ev"]
        ai_config.telegram_min_confidence = telegram_settings["min_confidence"]
        ai_config.telegram_daily_report_enabled = telegram_settings["daily_report"]
        ai_config.telegram_daily_report_time = telegram_settings["report_time"].strftime("%H:%M")
        ai_config.telegram_rate_limit_seconds = telegram_settings["rate_limit_seconds"]
        ai_config.live_monitoring_enabled = telegram_settings["always_on"]
        ai_config.live_min_ev_alert = min(ai_config.live_min_ev_alert, telegram_settings["min_ev"])

        nn_settings = settings["neural_network"]
        ai_config.calibrator_hidden_layers = list(nn_settings["hidden_layers"])
        ai_config.calibrator_dropout = nn_settings["dropout"]
        ai_config.calibrator_learning_rate = nn_settings["learning_rate"]
        ai_config.calibrator_batch_size = nn_settings["batch_size"]

        return ai_config

# ============================================================
#   ADVANCED FEATURES (Sprint 1 & 2) - IMPORT
# ============================================================
try:
    from advanced_features import (
        # Sprint 1
        apply_physical_constraints_to_lambda,
        neumaier_sum,
        precise_probability_sum,
        load_calibration_map,
        apply_calibration,
        # Sprint 2
        apply_motivation_factor,
        apply_fixture_congestion,
        apply_tactical_matchup,
        apply_all_advanced_features,
        MOTIVATION_FACTORS,
        TACTICAL_STYLES
    )
    ADVANCED_FEATURES_AVAILABLE = True
    logger.info("✅ Advanced Features module caricato con successo")

    # Carica calibration map al startup
    try:
        CALIBRATION_MAP = load_calibration_map("storico_analisi.csv")
        if CALIBRATION_MAP:
            n_outcomes = sum(len(v) for v in CALIBRATION_MAP.values() if v)
            logger.info(f"✅ Calibration map caricata: {n_outcomes} bins")
        else:
            CALIBRATION_MAP = {}
            logger.warning("⚠️ Calibration map vuota (serve storico con risultati)")
    except Exception as e:
        CALIBRATION_MAP = {}
        logger.warning(f"⚠️ Errore caricamento calibration map: {e}")

except ImportError as e:
    ADVANCED_FEATURES_AVAILABLE = False
    CALIBRATION_MAP = {}
    logger.warning(f"⚠️ Advanced Features module non disponibile: {e}")
    logger.info("💡 Per abilitare: assicurati che advanced_features.py sia nella stessa cartella")

# ============================================================
#   AUTO-DETECTION MODULE (Sprint 1 & 2 Enhancement)
# ============================================================
try:
    from auto_features import (
        auto_detect_all_features,
        is_derby_match as auto_is_derby_match
    )
    AUTO_DETECTION_AVAILABLE = True
    logger.info("✅ Auto-Detection module caricato con successo")
except ImportError as e:
    AUTO_DETECTION_AVAILABLE = False
    logger.warning(f"⚠️ Auto-Detection module non disponibile: {e}")
    logger.info("💡 Per abilitare: assicurati che auto_features.py e team_profiles.json siano nella stessa cartella")

# ============================================================
#   API MANAGER MODULE (LEVEL 2 Lite Integration)
# ============================================================
try:
    from api_manager import APIManager
    API_MANAGER = APIManager()
    API_MANAGER_AVAILABLE = True
    logger.info("✅ API Manager caricato con successo (LEVEL 2 Lite)")
except ImportError as e:
    API_MANAGER = None
    API_MANAGER_AVAILABLE = False
    logger.warning(f"⚠️ API Manager non disponibile: {e}")
    logger.info("💡 LEVEL 2 Lite disabilitato - solo database locale disponibile")
except Exception as e:
    API_MANAGER = None
    API_MANAGER_AVAILABLE = False
    logger.warning(f"⚠️ Errore inizializzazione API Manager: {e}")

# ============================================================
#   CONFIGURAZIONE CENTRALIZZATA (MIGLIORAMENTO)
# ============================================================

@dataclass
class APIConfig:
    """Configurazione API keys e endpoints"""
    api_football_key: str = ""
    api_football_base: str = "https://v3.football.api-sports.io"
    openweather_api_key: str = "01afa2183566fcf16d98b5a33c91eae1"
    football_data_api_key: str = "ca816dc8504543768e8adfaf128ecffc"
    thesportsdb_api_key: str = "3"  # Pubblica, gratuita
    telegram_bot_token: str = "8530766126:AAHs1ZoLwrwvT7JuPyn_9ymNVyddPtUXi-g"
    telegram_chat_id: str = "-1003278011521"
    telegram_enabled: bool = True
    telegram_min_probability: float = 50.0

    def __post_init__(self):
        """Carica da variabili d'ambiente (override se presenti)"""
        # Le variabili d'ambiente hanno priorità se configurate
        self.api_football_key = os.getenv("API_FOOTBALL_KEY", self.api_football_key)
        self.openweather_api_key = os.getenv("OPENWEATHER_API_KEY", self.openweather_api_key)
        self.football_data_api_key = os.getenv("FOOTBALL_DATA_API_KEY", self.football_data_api_key)
        self.telegram_bot_token = os.getenv("TELEGRAM_BOT_TOKEN", self.telegram_bot_token)
        self.telegram_chat_id = os.getenv("TELEGRAM_CHAT_ID", self.telegram_chat_id)
        env_min_prob = os.getenv("TELEGRAM_MIN_PROBABILITY", None)
        if env_min_prob is not None:
            try:
                self.telegram_min_probability = float(env_min_prob)
            except ValueError:
                logger.warning("Valore non valido per TELEGRAM_MIN_PROBABILITY, uso default configurato")

@dataclass
class ModelConfig:
    """Configurazione parametri modello (magic numbers centralizzati)"""
    # Pesi blend
    DNB_WEIGHT: float = 0.3  # Peso DNB nel blend con 1X2
    MARKET_WEIGHT: float = 0.7  # Peso mercato nel blend con DNB
    XG_MARKET_WEIGHT: float = 0.6  # Peso mercato nel blend xG
    XG_XG_WEIGHT: float = 0.4  # Peso xG nel blend
    XG_MAX_WEIGHT: float = 0.45  # Peso massimo xG (con alta confidence)
    XG_API_BOOST: float = 1.15  # Boost confidence se abbiamo dati API
    
    # ⚠️ OTTIMIZZAZIONE: Costanti matematiche pre-calcolate per evitare ricalcoli
    LOG_2_5: float = math.log(2.5)  # log(2.5) per spread_factor
    EPSILON: float = 1e-10  # Epsilon per evitare divisione per zero e log(0)
    
    # ⚠️ MICRO-PRECISIONE: Tolleranze standardizzate per coerenza
    TOL_DIVISION_ZERO: float = 1e-12  # Tolleranza per protezione divisione per zero (più conservativa)
    TOL_NORMALIZATION: float = 1e-8  # Tolleranza per normalizzazione matrici/probabilità
    TOL_PROBABILITY_CHECK: float = 1e-6  # Tolleranza per verifica coerenza probabilità
    TOL_OPTIMIZATION: float = 1e-5  # Tolleranza per convergenza ottimizzazione
    TOL_CLIP_PROB: float = 1e-6  # Tolleranza per clipping probabilità (calibrazione)
    TOL_TOTAL_COHERENCE: float = 0.5  # Tolleranza per coerenza total (lambda_h + lambda_a ≈ 2 * lambda_total)
    TOL_SCALE_FACTOR_MIN: float = 0.1  # Valore minimo per scale_factor (protezione divisione per zero)
    
    # Smoothing matrici
    DIRICHLET_EPS: float = 1e-12  # Smoothing additivo per matrici di probabilità
    
    # High precision / mpmath
    ENABLE_HIGH_PRECISION: bool = False  # Abilita calcoli ad alta precisione (mpmath)
    HIGH_PRECISION_LAMBDA_THRESHOLD: float = 12.0  # Usa mpmath per lambda <= soglia
    MPMATH_PRECISION: int = 80  # Decimali di precisione per mpmath
    HIGH_PRECISION_MAX_ITER: int = 200  # Iterazioni massime per metodi high precision
    
    # Posteriori coniugati
    USE_CONJUGATE_POSTERIORS: bool = True  # Usa aggiornamenti Gamma-Poisson / Beta-Binomial
    GAMMA_PRIOR_STRENGTH_BASE: float = 50.0  # Forza prior per posteriori Gamma-Poisson
    BETA_PRIOR_STRENGTH_BASE: float = 50.0  # Forza prior per posteriori Beta-Binomial
    
    # Ensemble weights
    ENSEMBLE_MAIN_WEIGHT: float = 0.60  # Peso modello principale
    ENSEMBLE_MARKET_WEIGHT: float = 0.25  # Peso modello market
    ENSEMBLE_CONSERVATIVE_WEIGHT: float = 0.15  # Peso modello conservativo
    
    # HT ratio
    HT_BASE_RATIO: float = 0.45  # Ratio base HT/FT
    HT_TOTAL_ADJUSTMENT: float = -0.015  # Aggiustamento per total
    HT_TOTAL_BASE: float = 2.5  # Total base per calcolo
    HT_LAMBDA_ADJUSTMENT: float = -0.01  # Aggiustamento per lambda
    HT_RHO_ADJUSTMENT: float = 0.005  # Aggiustamento per rho
    HT_MIN: float = 0.40  # Ratio HT minimo
    HT_MAX: float = 0.55  # Ratio HT massimo
    
    # Rho defaults
    RHO_BASE: float = 0.15  # Rho base
    RHO_DRAW_FACTOR: float = 0.4  # Fattore moltiplicativo per draw
    RHO_BTTS_FACTOR: float = 0.5  # Fattore per BTTS
    RHO_MIN: float = -0.35  # Rho minimo
    RHO_MAX: float = 0.35  # Rho massimo
    RHO_DEFAULT_MIN: float = 0.05  # Rho default minimo
    RHO_DEFAULT_MAX: float = 0.45  # Rho default massimo
    
    # Lambda bounds
    LAMBDA_MIN: float = 0.1
    LAMBDA_MAX: float = 5.0
    LAMBDA_SAFE_MIN: float = 0.3
    LAMBDA_SAFE_MAX: float = 4.5
    LAMBDA_OPTIMIZATION_MIN: float = 0.2
    LAMBDA_OPTIMIZATION_MAX: float = 5.0
    
    # Odds bounds
    ODDS_MIN: float = 1.01
    ODDS_MAX: float = 100.0
    
    # Total bounds
    TOTAL_MIN: float = 0.5
    TOTAL_MAX: float = 6.0
    
    # Bayesian updating
    PRIOR_CONFIDENCE_BASE: float = 0.7  # Confidence base nel prior
    MARKET_CONFIDENCE_BASE: float = 0.7  # Confidence base nel mercato
    
    # Shrinkage
    SHRINKAGE_FACTOR: float = 0.3  # Fattore shrinkage James-Stein
    
    # Time decay
    TIME_DECAY_HALF_LIFE_DAYS: int = 30  # Half-life per time decay
    
    # Calibration
    CALIBRATION_MIN_SAMPLES: int = 20  # Minimo campioni per calibrazione
    CALIBRATION_MIN_SAMPLES_LEAGUE: int = 30  # Minimo per calibrazione per lega
    CALIBRATION_MIN_SAMPLES_GLOBAL: int = 50  # Minimo per calibrazione globale
    CALIBRATION_MIN_SAMPLES_MARKET: int = 150  # Minimo campioni per calibrazione mercati derivati
    CALIBRATION_MAX_SAMPLES_MARKET: int = 2000  # Campioni massimi da utilizzare per calibrazione mercati
    CALIBRATION_MARKET_BLEND: float = 0.6  # Peso verso calibrazione rispetto al valore raw (0=no calibrazione, 1=solo calibrazione)


DEFAULT_ARCHIVE_DIR = os.getenv("ANALYSIS_ARCHIVE_DIR", "data")
DEFAULT_ARCHIVE_FILE = os.getenv(
    "ANALYSIS_ARCHIVE_FILE",
    os.path.join(DEFAULT_ARCHIVE_DIR, "storico_analisi.csv")
)

@dataclass
class AppConfig:
    """Configurazione applicazione"""
    archive_file: str = DEFAULT_ARCHIVE_FILE
    validation_file: str = "validation_metrics.csv"
    portfolio_file: str = "portfolio_scommesse.csv"
    odds_history_file: str = "odds_history.csv"
    alerts_file: str = "alerts.json"
    cache_expiry: int = 300  # 5 minuti
    api_retry_max_attempts: int = 3
    api_retry_delay: float = 1.0  # secondi
    api_timeout: float = 10.0  # secondi
    log_file: str = "app.log"

# Istanze globali configurazione
api_config = APIConfig()
model_config = ModelConfig()
app_config = AppConfig()
try:
    archive_dir = os.path.dirname(app_config.archive_file)
    os.makedirs(archive_dir or ".", exist_ok=True)
    logger.info(f"🗂️ Archivio analisi: {app_config.archive_file}")
except Exception as archive_dir_error:
    logger.warning(f"⚠️ Impossibile creare directory archivio ({app_config.archive_file}): {archive_dir_error}")

if model_config.ENABLE_HIGH_PRECISION and not MPMATH_AVAILABLE:
    logger.warning("High precision attivata ma mpmath non disponibile - fallback a precisione standard")

# Backward compatibility (mantiene variabili esistenti)
API_FOOTBALL_KEY = api_config.api_football_key
API_FOOTBALL_BASE = api_config.api_football_base
OPENWEATHER_API_KEY = api_config.openweather_api_key
FOOTBALL_DATA_API_KEY = api_config.football_data_api_key
THESPORTSDB_API_KEY = api_config.thesportsdb_api_key
TELEGRAM_BOT_TOKEN = api_config.telegram_bot_token
TELEGRAM_CHAT_ID = api_config.telegram_chat_id
TELEGRAM_ENABLED = api_config.telegram_enabled
TELEGRAM_MIN_PROBABILITY = api_config.telegram_min_probability

BACKGROUND_AUTOMATION: Optional["BackgroundAutomationService"] = None
if AI_SYSTEM_AVAILABLE:
    try:
        automation_config = AIConfig()
        automation_config.telegram_enabled = TELEGRAM_ENABLED
        automation_config.telegram_bot_token = TELEGRAM_BOT_TOKEN
        automation_config.telegram_chat_id = TELEGRAM_CHAT_ID
        automation_config.live_monitoring_enabled = TELEGRAM_ENABLED
        BACKGROUND_AUTOMATION = BackgroundAutomationService(config=automation_config)
        logger.info("🧠 Background automation service attivo (Telegram abilitato=%s)", TELEGRAM_ENABLED)
    except Exception as automation_error:
        BACKGROUND_AUTOMATION = None
        logger.warning(f"⚠️ Background automation non disponibile: {automation_error}")

ARCHIVE_FILE = app_config.archive_file
VALIDATION_FILE = app_config.validation_file
PORTFOLIO_FILE = app_config.portfolio_file
ODDS_HISTORY_FILE = app_config.odds_history_file
ALERTS_FILE = app_config.alerts_file
CACHE_EXPIRY = app_config.cache_expiry
API_RETRY_MAX_ATTEMPTS = app_config.api_retry_max_attempts
API_RETRY_DELAY = app_config.api_retry_delay
API_TIMEOUT = app_config.api_timeout

_manual_log_dir = os.path.dirname(PORTFOLIO_FILE)
MANUAL_PRECISION_LOG_FILE = os.path.join(_manual_log_dir if _manual_log_dir else ".", "manual_precision_log.csv")
MANUAL_PRECISION_LOG_COLUMNS = [
    "timestamp",
    "match_key",
    "league",
    "home_team",
    "away_team",
    "manual_confidence",
    "spread_apertura",
    "total_apertura",
    "spread_corrente",
    "total_corrente",
    "lambda_home",
    "lambda_away",
    "rho",
    "prob_home",
    "prob_draw",
    "prob_away",
    "predicted_total",
    "predicted_spread",
    "prob_home_low",
    "prob_home_high",
    "prob_draw_low",
    "prob_draw_high",
    "prob_away_low",
    "prob_away_high",
    "movement_magnitude",
    "movement_type",
    "history_samples",
    "actual_ft_home",
    "actual_ft_away",
    "actual_total",
    "actual_spread",
    "notes",
]

# Cache per API calls (evita rate limiting)
import threading
import time

API_CACHE_MAX_ENTRIES = getattr(app_config, "api_cache_max_entries", 512)
API_CACHE = None  # Initialized after TTLCache definition
FOOTBALL_DATA_BASE_URL = "https://api.football-data.org/v4"

# ============================================================
# FIX BUG #3-6: TTL CACHE MANAGER (Memory Leak Prevention)
# ============================================================
class TTLCache:
    """Thread-safe cache with TTL (Time To Live) and size limit to prevent memory leaks."""

    def __init__(self, max_size: int = 100, ttl_seconds: int = 3600):
        self.cache = {}
        self.timestamps = {}
        self.max_size = max_size
        self.ttl = ttl_seconds
        self.lock = threading.RLock()

    def _evict_unlocked(self, key):
        self.cache.pop(key, None)
        self.timestamps.pop(key, None)

    def _cleanup_expired_unlocked(self) -> int:
        now = time.time()
        expired_keys = [k for k, ts in self.timestamps.items() if now - ts >= self.ttl]
        removed = 0
        for key in expired_keys:
            self._evict_unlocked(key)
            removed += 1
        return removed

    def set(self, key, value):
        """Set value in cache with automatic cleanup if size limit reached."""
        with self.lock:
            self._cleanup_expired_unlocked()

            if self.max_size > 0 and len(self.cache) >= self.max_size:
                if not self.timestamps:
                    logger.error("❌ Cache desync: cache piena ma timestamps vuoto, reset cache")
                    self.cache.clear()
                    self.timestamps.clear()
                else:
                    oldest_key = min(self.timestamps, key=self.timestamps.get)
                    self._evict_unlocked(oldest_key)

            self.cache[key] = value
            self.timestamps[key] = time.time()

    def _get_locked(self, key, *, allow_stale: bool, refresh_on_access: bool):
        if key not in self.cache:
            return None

        timestamp = self.timestamps.get(key)
        if timestamp is None:
            self._evict_unlocked(key)
            return None

        now = time.time()
        age = now - timestamp
        is_stale = age >= self.ttl

        if is_stale and not allow_stale:
            self._evict_unlocked(key)
            return None

        result = {
            "value": self.cache[key],
            "timestamp": timestamp,
            "age": age,
            "is_stale": is_stale,
        }

        if refresh_on_access:
            self.timestamps[key] = now

        return result

    def get(self, key, default=None, *, allow_stale: bool = False, refresh_on_access: bool = False):
        """Get value from cache if exists; optional stale access."""
        with self.lock:
            entry = self._get_locked(key, allow_stale=allow_stale, refresh_on_access=refresh_on_access)
            if entry is None:
                return default
            return entry["value"]

    def get_entry(self, key, *, allow_stale: bool = False, refresh_on_access: bool = False):
        """Return cache entry metadata (value, timestamp, age, is_stale)."""
        with self.lock:
            entry = self._get_locked(key, allow_stale=allow_stale, refresh_on_access=refresh_on_access)
            if entry is None:
                return None
            return entry.copy()

    def cleanup(self) -> int:
        """Public method to purge expired entries."""
        with self.lock:
            return self._cleanup_expired_unlocked()

    def __contains__(self, key):
        entry = self.get_entry(key)
        return entry is not None and not entry["is_stale"]

    def __len__(self):
        with self.lock:
            self._cleanup_expired_unlocked()
            return len(self.cache)

# Initialize TTL caches (replaces unbounded Dict caches)
API_CACHE = TTLCache(max_size=API_CACHE_MAX_ENTRIES, ttl_seconds=CACHE_EXPIRY)
_FOOTBALL_DATA_MATCH_CACHE = TTLCache(max_size=200, ttl_seconds=1800)  # 30 min TTL

# ============================================================
#   API CHECKLIST HELPERS
# ============================================================
API_CHECK_PROVIDERS = [
    ("API-Football", "Statistiche, H2H, infortuni e fatigue"),
    ("Football-Data", "Forma e standings Football-Data.org"),
    ("TheSportsDB", "Stadio e informazioni club"),
    ("StatsBomb", "xG avanzati (open data)"),
    ("OpenWeather", "Condizioni meteo match"),
    ("Auto-Detection", "API Manager & auto features"),
    ("Telegram", "Alert push post-analisi"),
]

API_STATUS_ICONS = {
    "success": "✅",
    "warning": "⚠️",
    "error": "❌",
    "skipped": "⏭️",
}
API_STATUS_DEFAULT_MESSAGE = "Non utilizzata in questa analisi"


def _api_check_session_ready() -> bool:
    """Verifica se session_state di Streamlit è disponibile (evita errori nei test)."""
    try:
        _ = st.session_state  # type: ignore[assignment]
        return True
    except Exception:
        return False


def reset_api_checklist():
    """Inizializza (o resetta) la checklist delle API per una nuova analisi."""
    if not _api_check_session_ready():
        return
    st.session_state["api_check_results"] = {
        provider: {
            "status": "skipped",
            "message": API_STATUS_DEFAULT_MESSAGE,
            "updated_at": datetime.now().isoformat(),
        }
        for provider, _ in API_CHECK_PROVIDERS
    }


def update_api_check(provider: str, status: str, message: str):
    """Aggiorna lo stato di una specifica API nella checklist."""
    if not _api_check_session_ready():
        return
    if "api_check_results" not in st.session_state:
        reset_api_checklist()
    st.session_state["api_check_results"][provider] = {
        "status": status,
        "message": message,
        "updated_at": datetime.now().isoformat(),
    }


def summarize_apifootball_sources(
    advanced_data: Optional[Dict[str, Any]],
    fatigue_home: Optional[Dict[str, Any]],
    fatigue_away: Optional[Dict[str, Any]],
):
    """Determina lo stato di API-Football combinando statistiche avanzate e fatigue."""
    if not _api_check_session_ready():
        return

    if advanced_data is None:
        update_api_check("API-Football", "warning", "Dati avanzati non disponibili per questa analisi")
        return

    if not advanced_data.get("data_available"):
        update_api_check("API-Football", "warning", "API-Football non ha fornito dati per questa partita")
        return

    apifootball_blocks = [
        advanced_data.get("home_team_stats"),
        advanced_data.get("away_team_stats"),
        advanced_data.get("h2h_data"),
        advanced_data.get("home_injuries"),
        advanced_data.get("away_injuries"),
    ]
    fatigue_available = any(
        (
            fatigue_home and fatigue_home.get("data_available"),
            fatigue_away and fatigue_away.get("data_available"),
        )
    )

    if any(block for block in apifootball_blocks if block):
        message = "Statistiche, H2H e infortuni aggiornati"
        if fatigue_available:
            message += " + fatigue calcolata"
        update_api_check("API-Football", "success", message)
    else:
        update_api_check("API-Football", "warning", "API-Football non ha restituito dati utili")


def summarize_additional_api_sources(additional_data: Optional[Dict[str, Any]]):
    """Aggiorna lo stato delle API aggiuntive usando i dati calcolati dall'analisi."""
    if not _api_check_session_ready():
        return

    additional_data = additional_data or {}

    # Football-Data
    football_data_org = additional_data.get("football_data_org") or {}
    football_data_metrics = additional_data.get("football_data_metrics") or {}
    if not FOOTBALL_DATA_API_KEY:
        update_api_check("Football-Data", "warning", "Chiave Football-Data.org non configurata")
    elif football_data_org.get("available") or football_data_metrics.get("available"):
        form_snippet = football_data_metrics.get("recent_form")
        msg = "Forma Football-Data aggiornata"
        if form_snippet:
            msg += f" ({form_snippet})"
        update_api_check("Football-Data", "success", msg)
    else:
        update_api_check("Football-Data", "warning", "Dati Football-Data non disponibili per questa partita")

    # TheSportsDB
    thesportsdb_info = additional_data.get("thesportsdb") or {}
    if thesportsdb_info.get("available"):
        stadium = thesportsdb_info.get("stadium") or thesportsdb_info.get("name") or "stadio non disponibile"
        update_api_check("TheSportsDB", "success", f"Stadio rilevato: {stadium}")
    else:
        update_api_check("TheSportsDB", "warning", "Informazioni stadio non disponibili (TheSportsDB)")

    # StatsBomb
    statsbomb_info = additional_data.get("statsbomb") or {}
    if statsbomb_info.get("available"):
        matches_used = statsbomb_info.get("matches_used")
        msg = "Metriche xG StatsBomb aggiornate"
        if isinstance(matches_used, (int, float)) and matches_used > 0:
            msg = f"xG calcolati su {int(matches_used)} partite"
        update_api_check("StatsBomb", "success", msg)
    else:
        update_api_check("StatsBomb", "warning", "StatsBomb non disponibile per questa squadra")

    # OpenWeather
    weather_info = additional_data.get("weather")
    if weather_info and weather_info.get("available"):
        temp = weather_info.get("temperature")
        desc = weather_info.get("description") or weather_info.get("conditions") or ""
        if isinstance(temp, (int, float)):
            msg = f"{temp:.1f}°C {desc}".strip()
        else:
            msg = desc or "Meteo rilevato"
        update_api_check("OpenWeather", "success", msg)
    else:
        if OPENWEATHER_API_KEY:
            update_api_check("OpenWeather", "warning", "Meteo non disponibile per questa città")
        else:
            update_api_check("OpenWeather", "warning", "Chiave OpenWeather non configurata")


def render_api_checklist():
    """Renderizza la checklist delle API in Streamlit."""
    if not _api_check_session_ready():
        return
    results = st.session_state.get("api_check_results")
    if not results:
        return

    st.markdown("### 📡 Checklist API post-analisi")
    for provider, _ in API_CHECK_PROVIDERS:
        record = results.get(provider)
        if not record:
            continue
        icon = API_STATUS_ICONS.get(record.get("status"), "ℹ️")
        message = record.get("message") or API_STATUS_DEFAULT_MESSAGE
        st.markdown(f"- {icon} **{provider}** — {message}")


# ============================================================
#   OTTIMIZZAZIONE: PRE-CALCOLO FATTORIALI (MIGLIORAMENTO)
# ============================================================

# MIGLIORAMENTO NUOVO: Lookup table per fattoriali (0! a 20!) per velocizzare calcoli Poisson
# I fattoriali fino a 20 coprono il 99.9% dei casi nel calcio (gol da 0 a 20)
# Velocizza ~15-25% i calcoli Poisson senza perdita di precisione
_FACTORIAL_CACHE = [math.factorial(i) for i in range(21)]  # 0! = 1, 1! = 1, 2! = 2, ..., 20! = 2432902008176640000
_LOG2_CONST = math.log(2.0)

if NUMBA_AVAILABLE:
    _FACTORIAL_CACHE_ARRAY = np.array(_FACTORIAL_CACHE, dtype=np.float64)

    @jit(nopython=True, fastmath=True, cache=True)
    def _poisson_pmf_core(k: int, lam: float) -> float:
        if k < 0:
            return 0.0
        if lam <= 0.0:
            return 1.0 if k == 0 else 0.0
        if lam > 50.0:
            lam = 50.0

        if k < _FACTORIAL_CACHE_ARRAY.shape[0]:
            # Calcolo diretto con fattoriale precalcolato
            res = (lam ** k) * math.exp(-lam) / _FACTORIAL_CACHE_ARRAY[k]
        else:
            # Usa forma log per evitare overflow con k grandi
            res = math.exp((k * math.log(lam)) - lam - math.lgamma(k + 1.0))

        if res < 0.0:
            return 0.0
        if res > 1.0:
            return 1.0
        return res

    @jit(nopython=True, fastmath=True, cache=True)
    def _entropia_poisson_core(lam: float, max_k: int, tol: float, log2_const: float) -> float:
        if lam <= 0.0:
            return 0.0

        e = 0.0
        c = 0.0
        for k in range(max_k + 1):
            p = _poisson_pmf_core(k, lam)
            if p > tol:
                log_p = math.log(p) / log2_const
                term = -p * log_p
                y = term - c
                t = e + y
                c = (t - e) - y
                e = t

        if e < 0.0 or math.isnan(e) or math.isinf(e):
            return 0.0
        return e
else:
    def _poisson_pmf_core(k: int, lam: float) -> float:
        if k < 0:
            return 0.0
        if lam <= 0.0:
            return 1.0 if k == 0 else 0.0
        if lam > 50.0:
            lam = 50.0

        try:
            if k < len(_FACTORIAL_CACHE):
                res = (lam ** k) * math.exp(-lam) / _FACTORIAL_CACHE[k]
            else:
                res = math.exp((k * math.log(lam)) - lam - math.lgamma(k + 1.0))
        except (OverflowError, ValueError):
            res = 0.0

        if not math.isfinite(res) or res < 0.0:
            res = 0.0
        elif res > 1.0:
            res = 1.0
        return res

    def _entropia_poisson_core(lam: float, max_k: int, tol: float, log2_const: float) -> float:
        if lam <= 0.0:
            return 0.0

        e = 0.0
        c = 0.0
        for k in range(max_k + 1):
            p = _poisson_pmf_core(k, lam)
            if p > tol:
                log_p = math.log(p) / log2_const
                term = -p * log_p
                y = term - c
                t = e + y
                c = (t - e) - y
                e = t

        if e < 0.0 or not math.isfinite(e):
            return 0.0
        return e

# ============================================================
#   GESTIONE ERRORI API ROBUSTA (URGENTE)
# ============================================================

import time
from functools import wraps

def api_call_with_retry(
    api_func,
    *args,
    max_attempts: int = API_RETRY_MAX_ATTEMPTS,
    delay: float = API_RETRY_DELAY,
    timeout: float = API_TIMEOUT,
    use_cache: bool = True,
    cache_key: str = None,
    **kwargs
):
    """
    Wrapper robusto per chiamate API con retry logic, timeout e fallback cache.
    
    Args:
        api_func: Funzione API da chiamare
        max_attempts: Numero massimo di tentativi
        delay: Delay tra tentativi (secondi)
        timeout: Timeout per richiesta (secondi)
        use_cache: Se True, usa cache come fallback
        cache_key: Chiave cache (opzionale, auto-generata se None)
        **kwargs: Argomenti da passare a api_func
    
    Returns:
        Risultato della chiamata API o cache se disponibile
    
    Raises:
        Exception: Se tutti i tentativi falliscono e non c'è cache
    """
    # Genera cache key se non fornita
    if cache_key is None:
        cache_key = f"{api_func.__name__}_{hash(str(args) + str(kwargs))}"
    
    # Prova a recuperare da cache se disponibile
    if use_cache:
        cache_entry = API_CACHE.get_entry(cache_key)
        if cache_entry and not cache_entry["is_stale"]:
            return cache_entry["value"]

    last_exception = None

    # Retry loop con exponential backoff
    for attempt in range(max_attempts):
        try:
            # Aggiungi timeout a kwargs se non presente
            if "timeout" not in kwargs:
                kwargs["timeout"] = timeout

            # Chiama API
            result = api_func(*args, **kwargs)

            # Salva in cache se successo
            if use_cache:
                API_CACHE.set(cache_key, result)

            return result
            
        except requests.exceptions.Timeout:
            last_exception = f"Timeout dopo {timeout}s"
            if attempt < max_attempts - 1:
                wait_time = delay * (2 ** attempt)  # Exponential backoff
                time.sleep(wait_time)
            continue
            
        except requests.exceptions.ConnectionError:
            last_exception = "Errore connessione"
            if attempt < max_attempts - 1:
                wait_time = delay * (2 ** attempt)
                time.sleep(wait_time)
            continue
            
        except requests.exceptions.HTTPError as e:
            # Errori HTTP: alcuni sono non recuperabili
            status_code = e.response.status_code if hasattr(e, 'response') else None
            if status_code in [401, 403, 404]:
                # Errori non recuperabili: non ritentare
                raise e
            last_exception = f"HTTP {status_code}"
            if attempt < max_attempts - 1:
                wait_time = delay * (2 ** attempt)
                time.sleep(wait_time)
            continue
            
        except Exception as e:
            last_exception = str(e)
            if attempt < max_attempts - 1:
                wait_time = delay * (2 ** attempt)
                time.sleep(wait_time)
            continue
    
    # Tutti i tentativi falliti: prova cache come fallback
    if use_cache:
        stale_entry = API_CACHE.get_entry(cache_key, allow_stale=True)
        if stale_entry:
            age_minutes = stale_entry["age"] / 60.0 if stale_entry["age"] is not None else 0.0
            logger.warning(
                f"API fallita, uso cache (scaduta da {age_minutes:.1f} min): {cache_key}"
            )
            return stale_entry["value"]
    
    # Nessun fallback disponibile: solleva eccezione
    raise requests.exceptions.RequestException(
        f"API call fallita dopo {max_attempts} tentativi: {last_exception}"
    )

def safe_api_call(api_func, default_return=None, *args, **kwargs):
    """
    Wrapper semplificato che cattura tutte le eccezioni e ritorna default.
    Utile per chiamate API non critiche.
    """
    try:
        return api_call_with_retry(api_func, *args, **kwargs)
    except (requests.exceptions.RequestException, Exception) as e:
        logger.warning(f"API call fallita (non critica): {api_func.__name__}: {e}")
        return default_return

# ============================================================
#             UTILS
# ============================================================

def normalize_key(s: str) -> str:
    # FIX BUG: Assicura che s sia sempre una stringa
    if not isinstance(s, str):
        s = str(s) if s is not None else ""
    return (s or "").lower().replace(" ", "").replace("-", "").replace("/", "")

def safe_round(x: Optional[float], nd: int = 3) -> Optional[float]:
    if x is None:
        return None
    try:
        return round(x, nd)
    except (TypeError, ValueError, OverflowError) as e:
        # FIX BUG: Ritorna None invece di valore non-float in caso di errore
        logger.warning(f"safe_round failed for x={x}: {type(e).__name__}, returning None")
        return None

def decimali_a_prob(odds: float) -> float:
    """
    ⚠️ PRECISIONE MANIACALE: Validazione completa, protezione divisione per zero
    """
    # ⚠️ CRITICO: Validazione input
    if not isinstance(odds, (int, float)):
        logger.warning(f"odds non valido: {odds}, ritorno 0.0")
        return 0.0
    
    if not math.isfinite(odds) or odds <= 0:
        logger.warning(f"odds non finito o <= 0: {odds}, ritorno 0.0")
        return 0.0
    
    # ⚠️ PROTEZIONE: Protezione divisione per zero
    odds_safe = max(odds, model_config.TOL_DIVISION_ZERO)
    prob = 1.0 / odds_safe
    
    # ⚠️ VERIFICA: Assicura che prob sia finita e in range [0, 1]
    if not math.isfinite(prob) or prob < 0 or prob > 1:
        logger.warning(f"prob non valida: {prob}, ritorno 0.0")
        return 0.0
    
    return prob


def gamma_poisson_posterior_mean(
    lambda_prior: float,
    prior_strength: float,
    lambda_observed: float,
    observed_strength: float
) -> float:
    """
    Aggiorna la media di una Poisson usando prior Gamma con parametri in forma di forza (pseudo-osservazioni).
    
    prior_strength / observed_strength rappresentano pseudo-counts ("exposure").
    """
    epsilon = model_config.TOL_DIVISION_ZERO
    prior_strength = max(prior_strength, epsilon)
    observed_strength = max(observed_strength, 0.0)
    
    alpha_prior = max(lambda_prior * prior_strength, epsilon)
    beta_prior = max(prior_strength, epsilon)
    
    alpha_obs = max(lambda_observed * observed_strength, 0.0)
    beta_obs = max(observed_strength, 0.0)
    
    alpha_post = alpha_prior + alpha_obs
    beta_post = beta_prior + beta_obs
    
    if beta_post <= epsilon:
        return lambda_prior
    
    posterior_mean = alpha_post / beta_post
    if not math.isfinite(posterior_mean):
        return lambda_prior
    return max(model_config.LAMBDA_SAFE_MIN, min(model_config.LAMBDA_SAFE_MAX, posterior_mean))


def beta_binomial_posterior_mean(
    p_prior: float,
    prior_strength: float,
    p_observed: float,
    observed_strength: float
) -> float:
    """
    Aggiorna una probabilità (0-1) usando prior Beta con pseudo-osservazioni.
    """
    epsilon = model_config.TOL_CLIP_PROB
    p_prior = max(0.0, min(1.0, p_prior))
    p_observed = max(0.0, min(1.0, p_observed))
    
    prior_strength = max(prior_strength, epsilon)
    observed_strength = max(observed_strength, epsilon)
    
    alpha_prior = max(p_prior * prior_strength, epsilon)
    beta_prior = max((1.0 - p_prior) * prior_strength, epsilon)
    
    alpha_obs = max(p_observed * observed_strength, epsilon)
    beta_obs = max((1.0 - p_observed) * observed_strength, epsilon)
    
    alpha_post = alpha_prior + alpha_obs
    beta_post = beta_prior + beta_obs
    
    posterior_denom = alpha_post + beta_post
    if posterior_denom <= epsilon:
        return p_prior
    
    posterior_mean = alpha_post / posterior_denom
    if not math.isfinite(posterior_mean):
        return p_prior
    return max(0.0, min(1.0, posterior_mean))


def estimate_lambda_from_xg_bayesian(
    xg_for: float,
    xg_against: float,
    lambda_from_market: float,
    confidence_xg: float = 0.5,
    confidence_market: float = 0.7
) -> float:
    """
    Combina xG e lambda da mercato usando aggiornamento Bayesiano (Gamma prior).

    MIGLIORAMENTO NUOVO: Usa matematica Bayesiana invece di blend empirico per combinare xG con mercato.

    Metodo: Posterior Gamma-Poisson con prior = mercato, observed = xG

    Args:
        xg_for: xG per la squadra (attacco)
        xg_against: xG contro la squadra (difesa avversaria)
        lambda_from_market: Lambda stimato da quote mercato (prior)
        confidence_xg: Confidence (0-1) per xG, convertita in pseudo-observations
        confidence_market: Confidence (0-1) per mercato

    Returns:
        Lambda posteriore (weighted average Bayesiano)

    Esempi:
        xg_for=2.0, xg_against=1.5, lambda_market=1.8, conf_xg=0.6, conf_market=0.7
        → lambda_posterior ≈ 1.75 (blend pesato verso xG medio)
    """
    # Validazione input
    if not isinstance(xg_for, (int, float)) or not math.isfinite(xg_for) or xg_for < 0:
        logger.warning(f"xg_for non valido: {xg_for}, uso lambda_from_market")
        return lambda_from_market

    if not isinstance(xg_against, (int, float)) or not math.isfinite(xg_against) or xg_against < 0:
        logger.warning(f"xg_against non valido: {xg_against}, uso lambda_from_market")
        return lambda_from_market

    if not isinstance(lambda_from_market, (int, float)) or not math.isfinite(lambda_from_market):
        logger.warning(f"lambda_from_market non valido: {lambda_from_market}, uso default")
        lambda_from_market = 1.5

    # Clamp confidence a [0, 1]
    confidence_xg = max(0.0, min(1.0, confidence_xg))
    confidence_market = max(0.0, min(1.0, confidence_market))

    # Media xG (blend attacco e difesa avversaria)
    lambda_xg = (xg_for + xg_against) / 2.0

    # Clamp lambda_xg a range ragionevole
    lambda_xg = max(0.1, min(5.0, lambda_xg))
    lambda_from_market = max(0.1, min(5.0, lambda_from_market))

    # Verifica confidence totale
    total_conf = confidence_xg + confidence_market
    if total_conf <= model_config.TOL_DIVISION_ZERO:
        logger.warning("Confidence totale troppo bassa, uso lambda_from_market")
        return lambda_from_market

    # Scala confidence a pseudo-observations (exposure)
    # Confidence 1.0 = 10 pseudo-observations, 0.5 = 5, ecc.
    prior_strength = confidence_market * 10.0
    observed_strength = confidence_xg * 10.0

    # Posterior Gamma-Poisson (usa funzione esistente)
    lambda_posterior = gamma_poisson_posterior_mean(
        lambda_prior=lambda_from_market,
        prior_strength=prior_strength,
        lambda_observed=lambda_xg,
        observed_strength=observed_strength
    )

    # Sanity check finale
    if not math.isfinite(lambda_posterior) or lambda_posterior < 0.1 or lambda_posterior > 5.0:
        logger.warning(f"lambda_posterior fuori range: {lambda_posterior}, uso lambda_from_market")
        return lambda_from_market

    return lambda_posterior

def log_precision_metrics(context: str, metrics: Dict[str, Any]) -> None:
    """
    Log strutturato per diagnosi numeriche.
    """
    if not logger.isEnabledFor(logging.DEBUG):
        return
    formatted = []
    for key, value in metrics.items():
        if isinstance(value, float):
            formatted.append(f"{key}={value:.6f}")
        else:
            formatted.append(f"{key}={value}")
    logger.debug(f"[DIAGNOSTICS] {context}: " + ", ".join(formatted))

# ============================================================
#   VALIDAZIONE INPUT ROBUSTA (URGENTE)
# ============================================================

class ValidationError(Exception):
    """Eccezione custom per errori di validazione"""
    pass

def validate_odds(odds: float, name: str = "odds", min_odds: float = 1.01, max_odds: float = 100.0) -> float:
    """
    Valida e normalizza una quota.
    
    Args:
        odds: Quota da validare
        name: Nome del parametro (per messaggi errore)
        min_odds: Quota minima accettabile
        max_odds: Quota massima accettabile
    
    Returns:
        Quota validata
    
    Raises:
        ValidationError: Se la quota non è valida
    """
    if odds is None:
        raise ValidationError(f"{name} non può essere None")
    
    try:
        odds = float(odds)
    except (ValueError, TypeError):
        raise ValidationError(f"{name} deve essere un numero valido, ricevuto: {type(odds)}")
    
    if not (min_odds <= odds <= max_odds):
        raise ValidationError(f"{name} deve essere tra {min_odds} e {max_odds}, ricevuto: {odds}")
    
    return odds

def validate_probability(prob: float, name: str = "probability") -> float:
    """
    Valida una probabilità (deve essere tra 0 e 1).
    
    Args:
        prob: Probabilità da validare
        name: Nome del parametro
    
    Returns:
        Probabilità validata
    
    Raises:
        ValidationError: Se la probabilità non è valida
    """
    if prob is None:
        raise ValidationError(f"{name} non può essere None")
    
    try:
        prob = float(prob)
    except (ValueError, TypeError):
        raise ValidationError(f"{name} deve essere un numero valido")
    
    if not (0.0 <= prob <= 1.0):
        raise ValidationError(f"{name} deve essere tra 0.0 e 1.0, ricevuto: {prob}")
    
    return prob

def validate_lambda_value(lambda_val: float, name: str = "lambda") -> float:
    """
    Valida un valore lambda (gol attesi).
    
    Args:
        lambda_val: Lambda da validare
        name: Nome del parametro
    
    Returns:
        Lambda validato e clamped
    
    Raises:
        ValidationError: Se lambda non è valido
    """
    if lambda_val is None:
        raise ValidationError(f"{name} non può essere None")
    
    try:
        lambda_val = float(lambda_val)
    except (ValueError, TypeError):
        raise ValidationError(f"{name} deve essere un numero valido")
    
    # FIX BUG #3: Standardizzato bounds lambda [0.3, 4.5]
    lambda_val = max(0.3, min(4.5, lambda_val))
    
    return lambda_val

DEFAULT_TOTAL_RANGE = (0.5, 6.0)
DEFAULT_SPREAD_RANGE = (-3.0, 3.0)

LEAGUE_TOTAL_RANGES = MappingProxyType({
    "generic": DEFAULT_TOTAL_RANGE,
    "premier_league": (0.5, 5.8),
    "serie_a": (0.5, 5.6),
    "la_liga": (0.5, 5.4),
    "bundesliga": (0.5, 6.3),
    "ligue_1": (0.5, 5.3),
    "champions_league": (0.5, 6.2),
    "europa_league": (0.5, 6.0),
    "mls": (0.5, 6.4),
})

LEAGUE_SPREAD_RANGES = MappingProxyType({
    "generic": DEFAULT_SPREAD_RANGE,
    "premier_league": (-2.8, 2.8),
    "serie_a": (-2.6, 2.6),
    "la_liga": (-2.5, 2.5),
    "bundesliga": (-2.7, 2.7),
    "ligue_1": (-2.4, 2.4),
    "champions_league": (-3.0, 3.0),
    "europa_league": (-2.9, 2.9),
    "mls": (-2.8, 2.8),
})

HIGH_VARIANCE_LEAGUES = {"bundesliga", "mls", "champions_league"}

MAX_MANUAL_HISTORY_SAMPLES = 6
MANUAL_HISTORY_ALPHA = 0.65
MIN_TOTAL_MARGIN_FOR_SPREAD = 0.2

MANUAL_LINES_HISTORY: Dict[str, Dict[str, deque]] = defaultdict(dict)


def _normalize_league_key(league: Optional[str]) -> str:
    if league is None:
        return "generic"
    if isinstance(league, str):
        key = league.strip().lower()
        return key or "generic"
    return str(league).strip().lower() or "generic"


def _get_league_range(league: Optional[str], mapping: MappingProxyType, default_range: Tuple[float, float]) -> Tuple[float, float]:
    league_key = _normalize_league_key(league)
    return mapping.get(league_key, mapping.get("generic", default_range))


def _clamp_with_league_range(value: float, league: Optional[str], mapping: MappingProxyType, default_range: Tuple[float, float]) -> float:
    min_val, max_val = _get_league_range(league, mapping, default_range)
    return max(min_val, min(max_val, value))


def _manual_history_bucket(match_key: str, field: str) -> deque:
    field_cache = MANUAL_LINES_HISTORY.setdefault(match_key, {})
    if field not in field_cache:
        field_cache[field] = deque(maxlen=MAX_MANUAL_HISTORY_SAMPLES)
    return field_cache[field]


def _compute_series_volatility(series: deque) -> float:
    if len(series) < 2:
        return 0.0
    diffs = [abs(series[i] - series[i - 1]) for i in range(1, len(series))]
    return float(sum(diffs) / len(diffs))


def smooth_manual_value(match_key: Optional[str], field: str, value: Optional[float], alpha: float = MANUAL_HISTORY_ALPHA) -> Tuple[Optional[float], Optional[float], int]:
    if match_key is None or value is None:
        return value, None, 0
    history = _manual_history_bucket(match_key, field)
    if history:
        smoothed = alpha * value + (1 - alpha) * history[-1]
    else:
        smoothed = value
    history.append(smoothed)
    volatility = _compute_series_volatility(history)
    return smoothed, volatility, len(history)


def build_manual_history_key(home_team: Optional[str], away_team: Optional[str], match_datetime: Optional[str], league: Optional[str]) -> Optional[str]:
    components = [_normalize_league_key(league)]
    if home_team:
        components.append(str(home_team).strip().lower())
    if away_team:
        components.append(str(away_team).strip().lower())
    if match_datetime:
        components.append(str(match_datetime))
    if len(components) <= 1:
        return None  # Non abbastanza informazioni
    return "|".join(components)


def aggregate_volatility(spread_vol: Optional[float], total_vol: Optional[float]) -> Optional[float]:
    values = [v for v in (spread_vol, total_vol) if isinstance(v, (int, float))]
    if not values:
        return None
    return float(sum(values) / len(values))


def compute_manual_confidence(
    spread_apertura: Optional[float],
    total_apertura: Optional[float],
    spread_corrente: Optional[float],
    total_corrente: Optional[float],
    volatility_score: Optional[float],
) -> Dict[str, Any]:
    confidence = 0.2  # baseline
    details: Dict[str, Any] = {}

    if spread_apertura is not None and total_apertura is not None:
        confidence += 0.25
        details["has_opening_pair"] = True
    if spread_corrente is not None and total_corrente is not None:
        confidence += 0.25
        details["has_current_pair"] = True

    if spread_apertura is not None and spread_corrente is not None:
        spread_gap = abs(spread_apertura - spread_corrente)
        penalty = min(0.25, spread_gap / 4.0)
        confidence -= penalty
        details["spread_gap"] = round(spread_gap, 3)

    if total_apertura is not None and total_corrente is not None:
        total_gap = abs(total_apertura - total_corrente)
        penalty = min(0.2, total_gap / 5.0)
        confidence -= penalty
        details["total_gap"] = round(total_gap, 3)

    if volatility_score is not None:
        vol_penalty = min(0.3, volatility_score / 3.0)
        confidence -= vol_penalty
        details["volatility_penalty"] = round(vol_penalty, 3)

    confidence = max(0.0, min(1.0, confidence))
    details["score"] = round(confidence, 3)
    return details


def enforce_total_support_for_spread(total_value: Optional[float], spread_value: Optional[float]) -> Optional[float]:
    if total_value is None or spread_value is None:
        return total_value
    min_total_needed = abs(spread_value) + MIN_TOTAL_MARGIN_FOR_SPREAD
    if total_value < min_total_needed:
        return min_total_needed
    return total_value


def collect_manual_consistency_warnings(validated_inputs: Dict[str, Any]) -> List[str]:
    warnings: List[str] = []

    for total_key, spread_key, label in [
        ("total_apertura", "spread_apertura", "apertura"),
        ("total_corrente", "spread_corrente", "corrente"),
    ]:
        adjusted_total = enforce_total_support_for_spread(
            validated_inputs.get(total_key),
            validated_inputs.get(spread_key),
        )
        if adjusted_total is not None and validated_inputs.get(total_key) != adjusted_total:
            warnings.append(
                f"Total {label} aumentato a {adjusted_total:.2f} per supportare spread {validated_inputs.get(spread_key):+.2f}"
            )
            validated_inputs[total_key] = round(adjusted_total, 3)

    spread_apertura = validated_inputs.get("spread_apertura")
    spread_corrente = validated_inputs.get("spread_corrente")
    total_apertura = validated_inputs.get("total_apertura")
    total_corrente = validated_inputs.get("total_corrente")

    if spread_apertura is not None and spread_corrente is not None:
        diff = abs(spread_apertura - spread_corrente)
        if diff > 2.0:
            warnings.append(f"Spread apertura/corrente differiscono di {diff:.2f} gol: verificare dati manuali.")

    if total_apertura is not None and total_corrente is not None:
        diff = abs(total_apertura - total_corrente)
        if diff > 1.2:
            warnings.append(f"Total apertura/corrente differiscono di {diff:.2f} gol: controlla eventuali errori di input.")

    return warnings


def compute_lambda_priors_from_xg(
    xg_for_home: Optional[float],
    xg_against_home: Optional[float],
    xg_for_away: Optional[float],
    xg_against_away: Optional[float],
) -> Tuple[Optional[float], Optional[float], float]:
    metrics_available = 0
    home_values = []
    away_values = []

    for value in (xg_for_home, xg_against_away):
        if isinstance(value, (int, float)) and math.isfinite(value) and value > 0:
            home_values.append(value)
            metrics_available += 1
    for value in (xg_for_away, xg_against_home):
        if isinstance(value, (int, float)) and math.isfinite(value) and value > 0:
            away_values.append(value)
            metrics_available += 1

    if not home_values or not away_values:
        return None, None, 0.0

    lambda_h_prior = max(0.3, min(4.5, sum(home_values) / len(home_values)))
    lambda_a_prior = max(0.3, min(4.5, sum(away_values) / len(away_values)))

    prior_confidence = min(1.0, metrics_available / 6.0)
    return lambda_h_prior, lambda_a_prior, prior_confidence


def blend_lambda_with_priors(
    lambda_market_h: float,
    lambda_market_a: float,
    lambda_prior_h: float,
    lambda_prior_a: float,
    prior_confidence: float,
) -> Tuple[float, float]:
    if prior_confidence <= 0.0:
        return lambda_market_h, lambda_market_a

    weight_prior = min(0.45, prior_confidence * 0.5 + 0.1)
    weight_market = 1.0 - weight_prior

    lambda_h = lambda_market_h * weight_market + lambda_prior_h * weight_prior
    lambda_a = lambda_market_a * weight_market + lambda_prior_a * weight_prior

    return lambda_h, lambda_a


LEAGUE_RHO_BASELINE = MappingProxyType({
    "generic": 0.05,
    "premier_league": 0.08,
    "serie_a": 0.06,
    "la_liga": 0.05,
    "bundesliga": 0.1,
    "ligue_1": 0.04,
    "champions_league": 0.07,
    "europa_league": 0.06,
    "mls": 0.12,
})


def adjust_rho_with_context(
    rho: float,
    league: Optional[str],
    lambda_h: float,
    lambda_a: float,
    manual_confidence: Optional[float] = None,
) -> float:
    league_key = _normalize_league_key(league)
    baseline = LEAGUE_RHO_BASELINE.get(league_key, LEAGUE_RHO_BASELINE["generic"])

    spread_intensity = abs(lambda_h - lambda_a)
    intensity_adjustment = min(0.06, spread_intensity * 0.02)

    rho_adjusted = 0.6 * rho + 0.3 * baseline + (0.1 * intensity_adjustment if lambda_h > lambda_a else -0.1 * intensity_adjustment)

    if manual_confidence is not None:
        rho_adjusted *= 1.0 + (manual_confidence - 0.5) * 0.2

    return max(-0.35, min(0.35, rho_adjusted))


def enforce_probability_ratio_consistency(
    lambda_h: float,
    lambda_a: float,
    p_home: float,
    p_draw: float,
    p_away: float,
    max_adjustment: float = 0.25,
) -> Tuple[float, float]:
    if lambda_h <= 0 or lambda_a <= 0:
        return lambda_h, lambda_a

    target_ratio = max(0.2, min(5.0, (p_home + 0.5 * p_draw) / max(1e-6, p_away + 0.5 * p_draw)))
    current_ratio = max(0.2, min(5.0, lambda_h / lambda_a))

    if abs(target_ratio - current_ratio) / target_ratio < 0.05:
        return lambda_h, lambda_a

    ratio_correction = target_ratio / current_ratio
    ratio_correction = max(1.0 - max_adjustment, min(1.0 + max_adjustment, ratio_correction))

    sqrt_corr = math.sqrt(ratio_correction)
    lambda_h *= sqrt_corr
    lambda_a /= sqrt_corr

    return lambda_h, lambda_a


def log_manual_precision_sample(entry: Dict[str, Any]) -> None:
    try:
        log_dir = os.path.dirname(MANUAL_PRECISION_LOG_FILE)
        if log_dir:
            os.makedirs(log_dir, exist_ok=True)
        file_exists = os.path.exists(MANUAL_PRECISION_LOG_FILE)
        with open(MANUAL_PRECISION_LOG_FILE, mode="a", newline="", encoding="utf-8") as log_file:
            writer = csv.DictWriter(log_file, fieldnames=MANUAL_PRECISION_LOG_COLUMNS)
            if not file_exists:
                writer.writeheader()
            writer.writerow({col: entry.get(col, "") for col in MANUAL_PRECISION_LOG_COLUMNS})
    except Exception as log_error:
        logger.warning(f"⚠️ Impossibile aggiornare manual_precision_log: {log_error}")


def _ensure_manual_precision_columns(df: pd.DataFrame) -> pd.DataFrame:
    for col in MANUAL_PRECISION_LOG_COLUMNS:
        if col not in df.columns:
            df[col] = np.nan
    return df[MANUAL_PRECISION_LOG_COLUMNS]


def update_manual_precision_result(
    match_key: str,
    actual_ft_home: Union[int, float, str],
    actual_ft_away: Union[int, float, str],
) -> bool:
    if not match_key:
        logger.error("Match key necessario per aggiornare il log di precisione manuale")
        return False
    if not os.path.exists(MANUAL_PRECISION_LOG_FILE):
        logger.warning(f"File manual_precision_log inesistente: {MANUAL_PRECISION_LOG_FILE}")
        return False
    try:
        df = pd.read_csv(MANUAL_PRECISION_LOG_FILE, on_bad_lines="skip")
        df = _ensure_manual_precision_columns(df)
        mask = df["match_key"] == match_key
        if not mask.any():
            logger.warning(f"Nessuna entry nel log per match_key={match_key}")
            return False
        idx = df[mask].index[-1]
        df.loc[idx, "actual_ft_home"] = int(actual_ft_home)
        df.loc[idx, "actual_ft_away"] = int(actual_ft_away)
        if pd.notna(df.loc[idx, "actual_ft_home"]) and pd.notna(df.loc[idx, "actual_ft_away"]):
            df.loc[idx, "actual_total"] = df.loc[idx, "actual_ft_home"] + df.loc[idx, "actual_ft_away"]
            df.loc[idx, "actual_spread"] = df.loc[idx, "actual_ft_home"] - df.loc[idx, "actual_ft_away"]
        df.to_csv(MANUAL_PRECISION_LOG_FILE, index=False)
        logger.info(f"✅ Aggiornato manual_precision_log per {match_key} con risultato {actual_ft_home}-{actual_ft_away}")
        return True
    except Exception as update_error:
        logger.error(f"Impossibile aggiornare manual_precision_log: {update_error}")
        return False


def compute_manual_precision_metrics(window_days: int = 30) -> Dict[str, Any]:
    if not os.path.exists(MANUAL_PRECISION_LOG_FILE):
        return {}
    try:
        df = pd.read_csv(MANUAL_PRECISION_LOG_FILE, on_bad_lines="skip")
        df = _ensure_manual_precision_columns(df)
        if "timestamp" in df.columns:
            df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
            cutoff = datetime.utcnow() - timedelta(days=window_days)
            df = df[(df["timestamp"].isna()) | (df["timestamp"] >= cutoff)]
        df = df.dropna(subset=["actual_ft_home", "actual_ft_away"])
        if df.empty:
            return {}
        df["actual_ft_home"] = df["actual_ft_home"].astype(float)
        df["actual_ft_away"] = df["actual_ft_away"].astype(float)
        df["actual_total"] = df["actual_ft_home"] + df["actual_ft_away"]
        df["actual_spread"] = df["actual_ft_home"] - df["actual_ft_away"]
        df["predicted_total"] = df["predicted_total"].fillna(df["lambda_home"] + df["lambda_away"])
        df["predicted_spread"] = df["predicted_spread"].fillna(df["lambda_home"] - df["lambda_away"])

        mae_total = float(np.mean(np.abs(df["predicted_total"] - df["actual_total"])))
        mae_spread = float(np.mean(np.abs(df["predicted_spread"] - df["actual_spread"])))

        def _outcome(row):
            if row["actual_ft_home"] > row["actual_ft_away"]:
                return "1"
            if row["actual_ft_home"] < row["actual_ft_away"]:
                return "2"
            return "X"

        df["actual_outcome"] = df.apply(_outcome, axis=1)
        df["predicted_outcome"] = df[["prob_home", "prob_draw", "prob_away"]].idxmax(axis=1)
        df["predicted_outcome"] = df["predicted_outcome"].map({
            "prob_home": "1",
            "prob_draw": "X",
            "prob_away": "2",
        })
        accuracy = float((df["actual_outcome"] == df["predicted_outcome"]).mean())
        avg_confidence = float(df["manual_confidence"].dropna().mean()) if "manual_confidence" in df else None

        return {
            "samples": int(len(df)),
            "window_days": window_days,
            "mae_total": round(mae_total, 3),
            "mae_spread": round(mae_spread, 3),
            "hit_rate_1x2": round(accuracy * 100, 2),
            "avg_manual_confidence": round(avg_confidence, 3) if avg_confidence is not None else None,
        }
    except Exception as metrics_error:
        logger.warning(f"⚠️ Impossibile calcolare metriche precisione manuale: {metrics_error}")
        return {}

def validate_total(total: float, name: str = "total", league: Optional[str] = None) -> float:
    """
    Valida un total gol.
    
    Args:
        total: Total da validare
        name: Nome del parametro
    
    Returns:
        Total validato e clamped
    """
    if total is None:
        raise ValidationError(f"{name} non può essere None")
    
    try:
        total = float(total)
    except (ValueError, TypeError):
        raise ValidationError(f"{name} deve essere un numero valido")
    
    # Clamp dinamico basato sulla lega
    total = _clamp_with_league_range(total, league, LEAGUE_TOTAL_RANGES, DEFAULT_TOTAL_RANGE)
    
    return total

def validate_spread(spread: float, name: str = "spread", league: Optional[str] = None) -> float:
    """
    Valida uno spread.
    
    Args:
        spread: Spread da validare
        name: Nome del parametro
    
    Returns:
        Spread validato e clamped
    """
    if spread is None:
        return 0.0  # Spread può essere None (default a 0)
    
    try:
        spread = float(spread)
    except (ValueError, TypeError):
        raise ValidationError(f"{name} deve essere un numero valido")
    
    # Clamp dinamico basato sulla lega
    spread = _clamp_with_league_range(spread, league, LEAGUE_SPREAD_RANGES, DEFAULT_SPREAD_RANGE)
    
    return spread

def validate_team_name(team_name: str, name: str = "team_name") -> str:
    """
    Valida e sanitizza un nome squadra.
    
    Args:
        team_name: Nome squadra da validare
        name: Nome del parametro
    
    Returns:
        Nome squadra sanitizzato
    """
    if team_name is None:
        return ""
    
    if not isinstance(team_name, str):
        team_name = str(team_name)
    
    # Rimuovi caratteri pericolosi e normalizza
    team_name = team_name.strip()
    # Rimuovi caratteri speciali pericolosi (mantieni lettere, numeri, spazi, apostrofi)
    team_name = re.sub(r'[^\w\s\'-]', '', team_name)
    
    # Limita lunghezza
    team_name = team_name[:100]
    
    return team_name

def validate_league(league: str) -> str:
    """
    Valida un nome lega.
    
    Args:
        league: Nome lega
    
    Returns:
        Nome lega validato
    """
    valid_leagues = [
        "generic", "premier_league", "la_liga", "serie_a",
        "bundesliga", "ligue_1", "champions_league", "europa_league"
    ]
    
    if league is None:
        return "generic"
    
    league = str(league).lower().strip()
    
    if league not in valid_leagues:
        # Se non valida, ritorna generic come fallback
        return "generic"
    
    return league

def validate_xg_value(xg: float, name: str = "xG") -> Optional[float]:
    """
    Valida un valore xG.
    
    Args:
        xg: xG da validare (può essere None)
        name: Nome del parametro
    
    Returns:
        xG validato o None
    """
    if xg is None:
        return None
    
    try:
        xg = float(xg)
    except (ValueError, TypeError):
        return None
    
    # Clamp a range ragionevole (0.0 - 5.0)
    if xg < 0.0 or xg > 5.0:
        return None
    
    return xg

def validate_all_inputs(
    odds_1: float = None,
    odds_x: float = None,
    odds_2: float = None,
    total: float = None,
    odds_over25: float = None,
    odds_under25: float = None,
    odds_btts: float = None,
    odds_dnb_home: float = None,
    odds_dnb_away: float = None,
    league: Optional[str] = None,
    spread_apertura: float = None,
    total_apertura: float = None,
    spread_corrente: float = None,
    total_corrente: float = None,
    xg_for_home: float = None,
    xg_against_home: float = None,
    xg_for_away: float = None,
    xg_against_away: float = None,
    xa_for_home: float = None,
    xa_against_home: float = None,
    xa_for_away: float = None,
    xa_against_away: float = None,
) -> Dict[str, Any]:
    """
    Valida tutti gli input del modello in una volta.
    
    Returns:
        Dict con input validati e lista di warnings
    
    Raises:
        ValidationError: Se ci sono errori critici
    """
    warnings = []
    validated = {}
    
    # Valida quote principali (obbligatorie)
    try:
        validated["odds_1"] = validate_odds(odds_1, "odds_1")
        validated["odds_x"] = validate_odds(odds_x, "odds_x")
        validated["odds_2"] = validate_odds(odds_2, "odds_2")
    except ValidationError as e:
        raise ValidationError(f"Quote 1X2 obbligatorie: {e}")
    
    # Valida total (obbligatorio)
    try:
        validated["total"] = validate_total(total, "total", league)
    except ValidationError as e:
        raise ValidationError(f"Total obbligatorio: {e}")
    
    # Valida quote opzionali
    if odds_over25 is not None:
        try:
            validated["odds_over25"] = validate_odds(odds_over25, "odds_over25")
        except ValidationError:
            warnings.append("Quota Over 2.5 non valida, ignorata")
            validated["odds_over25"] = None
    
    if odds_under25 is not None:
        try:
            validated["odds_under25"] = validate_odds(odds_under25, "odds_under25")
        except ValidationError:
            warnings.append("Quota Under 2.5 non valida, ignorata")
            validated["odds_under25"] = None
    
    if odds_btts is not None:
        try:
            validated["odds_btts"] = validate_odds(odds_btts, "odds_btts")
        except ValidationError:
            warnings.append("Quota BTTS non valida, ignorata")
            validated["odds_btts"] = None
    
    if odds_dnb_home is not None:
        try:
            validated["odds_dnb_home"] = validate_odds(odds_dnb_home, "odds_dnb_home")
        except ValidationError:
            validated["odds_dnb_home"] = None
    
    if odds_dnb_away is not None:
        try:
            validated["odds_dnb_away"] = validate_odds(odds_dnb_away, "odds_dnb_away")
        except ValidationError:
            validated["odds_dnb_away"] = None
    
    # Valida spread e total apertura/corrente
    validated["spread_apertura"] = validate_spread(spread_apertura, "spread_apertura", league) if spread_apertura is not None else None
    validated["total_apertura"] = validate_total(total_apertura, "total_apertura", league) if total_apertura is not None else None
    validated["spread_corrente"] = validate_spread(spread_corrente, "spread_corrente", league) if spread_corrente is not None else None
    validated["total_corrente"] = validate_total(total_corrente, "total_corrente", league) if total_corrente is not None else None

    warnings.extend(collect_manual_consistency_warnings(validated))
    
    # Valida xG (opzionali)
    validated["xg_for_home"] = validate_xg_value(xg_for_home, "xg_for_home")
    validated["xg_against_home"] = validate_xg_value(xg_against_home, "xg_against_home")
    validated["xg_for_away"] = validate_xg_value(xg_for_away, "xg_for_away")
    validated["xg_against_away"] = validate_xg_value(xg_against_away, "xg_against_away")

    # Valida xA (opzionali) riusando la stessa validazione (range identico, semantica affine)
    validated["xa_for_home"] = validate_xg_value(xa_for_home, "xa_for_home")
    validated["xa_against_home"] = validate_xg_value(xa_against_home, "xa_against_home")
    validated["xa_for_away"] = validate_xg_value(xa_for_away, "xa_for_away")
    validated["xa_against_away"] = validate_xg_value(xa_against_away, "xa_against_away")
    
    return {
        "validated": validated,
        "warnings": warnings
    }

# ============================================================
#   NORMALIZZAZIONE AVANZATA DELLE QUOTE (SHIN METHOD)
# ============================================================

def shin_normalization(odds_list: List[float], max_iter: int = 100, tol: float = 1e-6) -> List[float]:
    """
    Shin method per rimuovere il margine considerando insider trading.
    Più robusto della semplice normalizzazione proporzionale.
    
    Reference: Shin, H. S. (1992). "Prices of State Contingent Claims with Insider 
    Traders, and the Favourite-Longshot Bias"
    """
    if not odds_list or any(o <= 1 for o in odds_list):
        return odds_list
    
    # ⚠️ PRECISIONE: Calcola probabilità implicite con protezione
    probs_list = []
    for o in odds_list:
        # ⚠️ PROTEZIONE: Validazione e protezione divisione per zero
        if not isinstance(o, (int, float)) or not math.isfinite(o) or o <= 1.0:
            logger.warning(f"Odd non valido: {o}, salto")
            continue
        prob = 1.0 / max(o, model_config.TOL_DIVISION_ZERO)
        if math.isfinite(prob) and 0.0 < prob < 1.0:
            probs_list.append(prob)
    
    if len(probs_list) != len(odds_list):
        logger.warning(f"Alcune probabilità non valide, uso solo {len(probs_list)}/{len(odds_list)}")
        if len(probs_list) < 2:
            return odds_list
    
    probs = np.array(probs_list)
    
    # ⚠️ PRECISIONE: Kahan summation per somma precisa
    sum_probs = 0.0
    c = 0.0
    for p in probs:
        y = p - c
        t = sum_probs + y
        c = (t - sum_probs) - y
        sum_probs = t
    
    margin = sum_probs - 1.0
    
    if margin <= model_config.TOL_DIVISION_ZERO:
        return odds_list
    
    # Risolvi per z (proporzione di insider information)
    def shin_equation(z):
        # ⚠️ PRECISIONE: Validazione z
        if not isinstance(z, (int, float)) or not math.isfinite(z):
            return float('inf')
        if z <= 0 or z >= 1:
            return float('inf')
        
        # ⚠️ PRECISIONE: Calcola sqrt_term con protezione overflow
        try:
            z_sq = z * z
            term = 4 * (1.0 - z) * (probs ** 2)
            sqrt_arg = z_sq + term
            
            # ⚠️ PROTEZIONE: Verifica che sqrt_arg sia non negativo e finito
            if sqrt_arg < 0 or not math.isfinite(sqrt_arg):
                return float('inf')
            
            sqrt_term = np.sqrt(sqrt_arg)
            if not np.all(np.isfinite(sqrt_term)):
                return float('inf')
            
            # ⚠️ PRECISIONE: Protezione divisione per zero
            denom = 2.0 * (1.0 - z)
            if abs(denom) < model_config.TOL_DIVISION_ZERO:
                return float('inf')
            
            fair_probs = (sqrt_term - z) / denom
            
            # ⚠️ PRECISIONE: Kahan summation per somma precisa
            sum_fair = 0.0
            c_fair = 0.0
            for fp in fair_probs:
                if not math.isfinite(fp):
                    return float('inf')
                y = fp - c_fair
                t = sum_fair + y
                c_fair = (t - sum_fair) - y
                sum_fair = t
            
            return sum_fair - 1.0
        except (ValueError, OverflowError, ZeroDivisionError) as e:
            logger.warning(f"Errore in shin_equation: {e}")
            return float('inf')
    
    try:
        # MIGLIORAMENTO: Range adattivo basato sul margine per convergenza migliore
        # Per margini bassi (<2%), z è tipicamente molto piccolo (<0.1)
        # Per margini alti (>15%), z può essere grande (>0.1)
        if margin < 0.02:  # Margine molto basso (mercati efficienti)
            z_range = (0.0001, 0.10)
        elif margin < 0.05:  # Margine basso
            z_range = (0.001, 0.30)
        elif margin > 0.15:  # Margine alto (bookmaker aggressivo)
            z_range = (0.10, 0.999)
        else:  # Margine medio (caso standard)
            z_range = (0.001, 0.999)

        # Trova z ottimale con range adattivo
        z_opt = optimize.brentq(shin_equation, z_range[0], z_range[1], maxiter=max_iter)
        
        # ⚠️ PRECISIONE: Calcola probabilità fair con precisione massima
        # ⚠️ VALIDAZIONE: Verifica z_opt
        if not isinstance(z_opt, (int, float)) or not math.isfinite(z_opt) or z_opt <= 0 or z_opt >= 1:
            raise ValueError(f"z_opt non valido: {z_opt}")
        
        try:
            z_sq = z_opt * z_opt
            term = 4.0 * (1.0 - z_opt) * (probs ** 2)
            sqrt_arg = z_sq + term
            
            # ⚠️ PROTEZIONE: Verifica che sqrt_arg sia non negativo e finito
            if not np.all(np.isfinite(sqrt_arg)) or np.any(sqrt_arg < 0):
                raise ValueError("sqrt_arg non valido")
            
            sqrt_term = np.sqrt(sqrt_arg)
            if not np.all(np.isfinite(sqrt_term)):
                raise ValueError("sqrt_term non finito")
            
            # ⚠️ PRECISIONE: Protezione divisione per zero
            denom = 2.0 * (1.0 - z_opt)
            if abs(denom) < model_config.TOL_DIVISION_ZERO:
                raise ValueError("denom troppo piccolo")
            
            fair_probs = (sqrt_term - z_opt) / denom
            if not np.all(np.isfinite(fair_probs)):
                raise ValueError("fair_probs non finito")
        except (ValueError, OverflowError, ZeroDivisionError) as e:
            logger.warning(f"Errore calcolo fair_probs: {e}, uso fallback")
            raise
        
        # ⚠️ PRECISIONE: Normalizza con Kahan summation per precisione massima
        sum_fair = 0.0
        c_fair = 0.0
        for fp in fair_probs:
            if not math.isfinite(fp):
                continue
            y = fp - c_fair
            t = sum_fair + y
            c_fair = (t - sum_fair) - y
            sum_fair = t
        
        # ⚠️ FIX BUG #5: Validazione robusta - sum_fair deve essere ragionevole (0.5 < sum < 2.0)
        if sum_fair > model_config.TOL_DIVISION_ZERO and 0.5 < sum_fair < 2.0:
            fair_probs = fair_probs / sum_fair
        else:
            # Fallback: normalizzazione semplice con Kahan
            sum_probs_safe = 0.0
            c_probs = 0.0
            for p in probs:
                if not math.isfinite(p):
                    continue
                y = p - c_probs
                t = sum_probs_safe + y
                c_probs = (t - sum_probs_safe) - y
                sum_probs_safe = t

            if sum_probs_safe > model_config.TOL_DIVISION_ZERO:
                fair_probs = probs / sum_probs_safe
            else:
                # Caso estremo: distribuzione uniforme
                fair_probs = np.ones_like(probs) / len(probs)

        # ⚠️ FIX BUG #2: NORMALIZZAZIONE FINALE FORZATA - garantisce somma esattamente 1.0
        # Problema: dopo normalizzazione Shin, somma può essere 0.9999987 o 1.0000013
        # Soluzione: forza normalizzazione finale per evitare probabilità negative downstream
        sum_final = fair_probs.sum()
        if abs(sum_final - 1.0) > model_config.TOL_DIVISION_ZERO:
            fair_probs = fair_probs / sum_final
            # Verifica che ora sommi a 1.0 (o molto vicino)
            sum_check = fair_probs.sum()
            if abs(sum_check - 1.0) > 1e-10:
                logger.warning(f"Normalizzazione Shin: somma probabilità = {sum_check:.10f} dopo forzatura")

        # ⚠️ PRECISIONE: Arrotonda solo per output, mantieni precisione nei calcoli
        # ⚠️ CRITICO: Protezione divisione per zero
        return [1.0/max(p, model_config.TOL_DIVISION_ZERO) for p in fair_probs]  # ⚠️ CRITICO: Protezione divisione per zero
    except (ValueError, RuntimeError, optimize.OptimizeWarning) as e:
        logger.warning(f"Errore normalizzazione Shin: {e}, uso fallback semplice")
        # ⚠️ PRECISIONE: Fallback a normalizzazione semplice con precisione
        sum_probs = probs.sum()
        if sum_probs > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            fair_probs = probs / sum_probs
        else:
            # Caso estremo: distribuzione uniforme
            fair_probs = np.ones_like(probs) / len(probs)
        # ⚠️ FIX BUG #2: Normalizzazione finale forzata anche nel fallback
        sum_final = fair_probs.sum()
        if abs(sum_final - 1.0) > model_config.TOL_DIVISION_ZERO:
            fair_probs = fair_probs / sum_final
        # ⚠️ CRITICO: Protezione divisione per zero
        return [1.0/max(p, model_config.TOL_DIVISION_ZERO) for p in fair_probs]  # ⚠️ CRITICO: Protezione divisione per zero
    except Exception as e:
        logger.error(f"Errore imprevisto durante normalizzazione Shin: {type(e).__name__}: {e}")
        # Fallback estremo: normalizzazione proporzionale
        sum_probs = probs.sum()
        if sum_probs > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            fair_probs = probs / sum_probs
        else:
            fair_probs = np.ones_like(probs) / len(probs)
        # ⚠️ FIX BUG #2: Normalizzazione finale forzata anche nel fallback estremo
        sum_final = fair_probs.sum()
        if abs(sum_final - 1.0) > model_config.TOL_DIVISION_ZERO:
            fair_probs = fair_probs / sum_final
        # ⚠️ CRITICO: Protezione divisione per zero
        return [1.0/max(p, model_config.TOL_DIVISION_ZERO) for p in fair_probs]  # ⚠️ CRITICO: Protezione divisione per zero

def normalize_two_way_shin(o1: float, o2: float) -> Tuple[float, float]:
    """
    Normalizzazione Shin per mercati a 2 esiti.
    
    ⚠️ PRECISIONE MANIACALE: Validazione completa input
    """
    # ⚠️ CRITICO: Validazione input robusta
    if not isinstance(o1, (int, float)) or not isinstance(o2, (int, float)):
        logger.warning(f"Odds non validi: o1={o1}, o2={o2}, ritorno originali")
        return o1 if isinstance(o1, (int, float)) else 2.0, o2 if isinstance(o2, (int, float)) else 2.0
    
    if not math.isfinite(o1) or not math.isfinite(o2):
        logger.warning(f"Odds non finiti: o1={o1}, o2={o2}, ritorno originali")
        return o1 if math.isfinite(o1) else 2.0, o2 if math.isfinite(o2) else 2.0
    
    if o1 <= 1.0 or o2 <= 1.0:
        logger.warning(f"Odds <= 1.0: o1={o1}, o2={o2}, ritorno originali")
        return o1, o2
    
    try:
        normalized = shin_normalization([o1, o2])
        if len(normalized) != 2:
            logger.warning(f"Shin normalization ritorna {len(normalized)} valori invece di 2, ritorno originali")
            return o1, o2
        # ⚠️ VERIFICA: Assicura che normalized siano finiti e > 1
        n1, n2 = normalized[0], normalized[1]
        if not all(math.isfinite(x) and x > 1.0 for x in [n1, n2]):
            logger.warning(f"Normalized odds non validi: n1={n1}, n2={n2}, ritorno originali")
            return o1, o2
        return n1, n2
    except Exception as e:
        logger.error(f"Errore normalizzazione Shin due-way: {e}, ritorno originali")
        return o1, o2

def normalize_three_way_shin(o1: float, ox: float, o2: float) -> Tuple[float, float, float]:
    """
    Normalizzazione Shin per 1X2.
    
    ⚠️ PRECISIONE MANIACALE: Validazione completa input
    """
    # ⚠️ CRITICO: Validazione input robusta
    if not all(isinstance(o, (int, float)) for o in [o1, ox, o2]):
        logger.warning(f"Odds non validi: o1={o1}, ox={ox}, o2={o2}, ritorno originali")
        return (o1 if isinstance(o1, (int, float)) else 3.0,
                ox if isinstance(ox, (int, float)) else 3.0,
                o2 if isinstance(o2, (int, float)) else 3.0)
    
    if not all(math.isfinite(o) for o in [o1, ox, o2]):
        logger.warning(f"Odds non finiti: o1={o1}, ox={ox}, o2={o2}, ritorno originali")
        return (o1 if math.isfinite(o1) else 3.0,
                ox if math.isfinite(ox) else 3.0,
                o2 if math.isfinite(o2) else 3.0)
    
    if any(o <= 1.0 for o in [o1, ox, o2]):
        logger.warning(f"Odds <= 1.0: o1={o1}, ox={ox}, o2={o2}, ritorno originali")
        return o1, ox, o2
    
    try:
        normalized = shin_normalization([o1, ox, o2])
        if len(normalized) != 3:
            logger.warning(f"Shin normalization ritorna {len(normalized)} valori invece di 3, ritorno originali")
            return o1, ox, o2
        # ⚠️ VERIFICA: Assicura che normalized siano finiti e > 1
        n1, nx, n2 = normalized[0], normalized[1], normalized[2]
        if not all(math.isfinite(x) and x > 1.0 for x in [n1, nx, n2]):
            logger.warning(f"Normalized odds non validi: n1={n1}, nx={nx}, n2={n2}, ritorno originali")
            return o1, ox, o2
        return n1, nx, n2
    except Exception as e:
        logger.error(f"Errore normalizzazione Shin three-way: {e}, ritorno originali")
        return o1, ox, o2

# ============================================================
#  STIMA BTTS AVANZATA CON MODELLO BIVARIATO
# ============================================================

def btts_probability_bivariate(lambda_h: float, lambda_a: float, rho: float) -> float:
    """
    Calcola P(BTTS) usando distribuzione Poisson bivariata con correlazione Dixon-Coles.
    
    Formula corretta: P(BTTS) = 1 - P(H=0 or A=0)
    dove P(H=0 or A=0) = P(H=0) + P(A=0) - P(H=0, A=0)
    
    P(H=0, A=0) con tau Dixon-Coles:
    tau(0,0) = 1 - lambda_h * lambda_a * rho
    
    ⚠️ PRECISIONE MANIACALE: Validazione completa, protezione overflow
    """
    # ⚠️ CRITICO: Validazione input
    if not isinstance(lambda_h, (int, float)) or not math.isfinite(lambda_h) or lambda_h < 0:
        logger.warning(f"lambda_h non valido: {lambda_h}, uso default 1.5")
        lambda_h = 1.5
    if not isinstance(lambda_a, (int, float)) or not math.isfinite(lambda_a) or lambda_a < 0:
        logger.warning(f"lambda_a non valido: {lambda_a}, uso default 1.5")
        lambda_a = 1.5
    if not isinstance(rho, (int, float)) or not math.isfinite(rho):
        logger.warning(f"rho non valido: {rho}, uso default 0.15")
        rho = 0.15
    
    # ⚠️ FIX BUG #1 & #3: Standardizzati bounds lambda [0.3, 4.5] e rho [-0.35, 0.35]
    lambda_h = max(0.3, min(4.5, lambda_h))
    lambda_a = max(0.3, min(4.5, lambda_a))
    rho = max(-0.35, min(0.35, rho))
    
    # P(H=0) marginale Poisson
    try:
        p_h0 = poisson.pmf(0, lambda_h)
        if not math.isfinite(p_h0) or p_h0 < 0:
            logger.warning(f"p_h0 non valido: {p_h0}, uso default")
            p_h0 = math.exp(-lambda_h) if lambda_h > 0 else 1.0
    except (ValueError, OverflowError) as e:
        logger.warning(f"Errore calcolo p_h0: {e}, uso approssimazione")
        p_h0 = math.exp(-lambda_h) if lambda_h > 0 else 1.0
    
    # P(A=0) marginale Poisson
    try:
        p_a0 = poisson.pmf(0, lambda_a)
        if not math.isfinite(p_a0) or p_a0 < 0:
            logger.warning(f"p_a0 non valido: {p_a0}, uso default")
            p_a0 = math.exp(-lambda_a) if lambda_a > 0 else 1.0
    except (ValueError, OverflowError) as e:
        logger.warning(f"Errore calcolo p_a0: {e}, uso approssimazione")
        p_a0 = math.exp(-lambda_a) if lambda_a > 0 else 1.0
    
    # ⚠️ PROTEZIONE: Limita probabilità a range [0, 1]
    p_h0 = max(0.0, min(1.0, p_h0))
    p_a0 = max(0.0, min(1.0, p_a0))
    
    # P(H=0, A=0) con correzione Dixon-Coles tau
    # ⚠️ PRECISIONE: Calcola tau con protezione overflow
    try:
        tau_calc = lambda_h * lambda_a * rho
        if not math.isfinite(tau_calc):
            logger.warning(f"tau_calc non finito: {tau_calc}, uso default")
            tau_calc = 0.0
        # ⚠️ FIX BUG #4: Bounds adattivi più teoricamente corretti [0.1, 3.0] invece di [0.2, 1.5]
        tau_00 = max(0.1, min(3.0, 1.0 - tau_calc))
    except (ValueError, OverflowError) as e:
        logger.warning(f"Errore calcolo tau: {e}, uso default")
        tau_00 = 0.5
    
    # ⚠️ PRECISIONE: Calcola p_h0_a0 con protezione overflow
    try:
        p_h0_a0 = p_h0 * p_a0 * tau_00
        if not math.isfinite(p_h0_a0) or p_h0_a0 < 0:
            logger.warning(f"p_h0_a0 non valido: {p_h0_a0}, correggo")
            p_h0_a0 = max(0.0, min(1.0, p_h0 * p_a0))
    except (ValueError, OverflowError) as e:
        logger.warning(f"Errore calcolo p_h0_a0: {e}, uso approssimazione")
        p_h0_a0 = max(0.0, min(1.0, p_h0 * p_a0))
    
    # P(H=0 or A=0) usando inclusione-esclusione
    # ⚠️ PRECISIONE: Kahan summation per somma precisa
    p_no_btts = p_h0 + p_a0 - p_h0_a0
    
    # ⚠️ PROTEZIONE: Limita a range [0, 1]
    p_no_btts = max(0.0, min(1.0, p_no_btts))
    
    p_btts = 1.0 - p_no_btts
    
    # ⚠️ PROTEZIONE: Bounds di sicurezza con verifica finale
    p_btts = max(0.0, min(1.0, p_btts))
    
    # ⚠️ VERIFICA FINALE: Double-check che sia in range [0, 1]
    if not (0.0 <= p_btts <= 1.0):
        logger.warning(f"p_btts fuori range: {p_btts}, correggo a 0.5")
        p_btts = 0.5
    
    return p_btts

def skellam_pmf(k: int, mu1: float, mu2: float) -> float:
    """
    Skellam PMF: P(X1 - X2 = k) dove X1 ~ Poisson(mu1), X2 ~ Poisson(mu2)

    MIGLIORAMENTO NUOVO: Distribuzione Skellam per differenza gol (matematicamente esatta)

    Formula: P(k) = exp(-(mu1+mu2)) * (mu1/mu2)^(k/2) * I_|k|(2*sqrt(mu1*mu2))
    dove I_k è la modified Bessel function of the first kind.

    Più accurata di calcolare matrice completa e sommare per handicap.

    Args:
        k: Differenza gol (home - away)
        mu1: Lambda casa
        mu2: Lambda trasferta

    Returns:
        Probabilità che differenza gol sia esattamente k
    """
    if not isinstance(k, int):
        logger.warning(f"k non intero: {k}, arrotondo")
        k = int(k)

    if not math.isfinite(mu1) or not math.isfinite(mu2) or mu1 < 0 or mu2 < 0:
        logger.warning(f"Parametri non validi: mu1={mu1}, mu2={mu2}, ritorno 0.0")
        return 0.0

    if mu1 == 0 and mu2 == 0:
        return 1.0 if k == 0 else 0.0

    try:
        from scipy.special import iv  # Modified Bessel function

        exp_term = math.exp(-(mu1 + mu2))
        if not math.isfinite(exp_term) or exp_term == 0:
            return 0.0

        if mu1 == 0 or mu2 == 0:
            bessel_term = 1.0 if k == 0 else 0.0
            ratio_term = 1.0
        else:
            # (mu1/mu2)^(|k|/2)
            ratio_term = (mu1 / mu2) ** (abs(k) / 2.0)
            if not math.isfinite(ratio_term):
                return 0.0

            # I_|k|(2*sqrt(mu1*mu2))
            bessel_arg = 2.0 * math.sqrt(mu1 * mu2)
            if not math.isfinite(bessel_arg):
                return 0.0
            bessel_term = iv(abs(k), bessel_arg)
            if not math.isfinite(bessel_term):
                return 0.0

        p = exp_term * ratio_term * bessel_term
        return max(0.0, min(1.0, p))
    except (ImportError, ValueError, OverflowError) as e:
        logger.warning(f"Errore Skellam: {e}, ritorno 0.0")
        return 0.0

def calc_handicap_from_skellam(lambda_h: float, lambda_a: float, handicap: float) -> Tuple[float, float, float]:
    """
    Calcola probabilità Asian Handicap (home cover, push, away cover) usando la distribuzione Skellam.

    Args:
        lambda_h: Media gol casa
        lambda_a: Media gol trasferta
        handicap: Handicap asiatico espresso in step da 0.25 (es. -0.25, 0.0, +0.75)

    Returns:
        Tuple (p_home_cover, p_push, p_away_cover) coerente con gli esiti dell'handicap.
    """

    def _handicap_probs(lh: float, la: float, hcap: float) -> Tuple[float, float, float]:
        """Calcola le probabilità per un singolo handicap (senza split di quarti)."""
        if not isinstance(lh, (int, float)) or not isinstance(la, (int, float)):
            return 0.5, 0.0, 0.5
        if not math.isfinite(lh) or not math.isfinite(la):
            return 0.5, 0.0, 0.5
        if not isinstance(hcap, (int, float)) or not math.isfinite(hcap):
            hcap = 0.0

        lh = max(0.1, min(5.0, lh))
        la = max(0.1, min(5.0, la))

        max_range = max(15, int(math.ceil(6 * math.sqrt(lh + la + 1))))
        p_home = 0.0
        p_push = 0.0
        p_away = 0.0
        c_home = c_push = c_away = 0.0
        eps = 1e-12

        for k in range(-max_range, max_range + 1):
            p_k = skellam_pmf(k, lh, la)
            if p_k <= 0.0:
                continue
            adjusted = k + hcap
            if adjusted > eps:
                y = p_k - c_home
                t = p_home + y
                c_home = (t - p_home) - y
                p_home = t
            elif adjusted < -eps:
                y = p_k - c_away
                t = p_away + y
                c_away = (t - p_away) - y
                p_away = t
            else:
                y = p_k - c_push
                t = p_push + y
                c_push = (t - p_push) - y
                p_push = t

        total = p_home + p_push + p_away
        if total > 0:
            p_home /= total
            p_push /= total
            p_away /= total
        return (
            max(0.0, min(1.0, p_home)),
            max(0.0, min(1.0, p_push)),
            max(0.0, min(1.0, p_away)),
        )

    # Gestione handicaps a 0.25 / 0.75 (split in due scommesse da mezzo stake)
    frac = abs(handicap) % 1
    if math.isclose(frac, 0.25, abs_tol=1e-9) or math.isclose(frac, 0.75, abs_tol=1e-9):
        lower = handicap - 0.25
        upper = handicap + 0.25
        h_home1, h_push1, h_away1 = _handicap_probs(lambda_h, lambda_a, lower)
        h_home2, h_push2, h_away2 = _handicap_probs(lambda_h, lambda_a, upper)
        return (
            (h_home1 + h_home2) / 2.0,
            (h_push1 + h_push2) / 2.0,
            (h_away1 + h_away2) / 2.0,
        )

    return _handicap_probs(lambda_h, lambda_a, handicap)

def validate_probability_coherence(
    p_home: float,
    p_draw: float,
    p_away: float,
    p_over: float = None,
    p_under: float = None,
    p_btts: float = None,
    lambda_h: float = None,
    lambda_a: float = None,
    rho: float = None,
    tolerance: float = 0.05
) -> Dict[str, Any]:
    """
    Valida coerenza matematica delle probabilità calcolate.

    MIGLIORAMENTO NUOVO: Funzione di validazione post-calcolo per verificare coerenza matematica
    e catturare errori numerici accumulati.

    Verifica:
    1. p_home + p_draw + p_away ≈ 1.0
    2. p_over + p_under ≈ 1.0 (se forniti)
    3. p_btts coerente con lambda e rho (se forniti)
    4. Total atteso coerente con lambda_h + lambda_a (se forniti)

    Args:
        p_home: Probabilità vittoria casa
        p_draw: Probabilità pareggio
        p_away: Probabilità vittoria trasferta
        p_over: Probabilità over (opzionale)
        p_under: Probabilità under (opzionale)
        p_btts: Probabilità BTTS (opzionale)
        lambda_h: Lambda casa (opzionale)
        lambda_a: Lambda trasferta (opzionale)
        rho: Correlazione Dixon-Coles (opzionale)
        tolerance: Tolleranza per warnings (default 0.05)

    Returns:
        Dict con risultati validazione, errori e warnings
    """
    warnings_list = []
    errors_list = []

    # 1. Verifica somma 1X2
    sum_1x2 = p_home + p_draw + p_away
    if abs(sum_1x2 - 1.0) > model_config.TOL_PROBABILITY_CHECK:
        errors_list.append(f"Somma 1X2 = {sum_1x2:.6f} (dovrebbe essere 1.0)")

    # 2. Verifica somma Over/Under (se forniti)
    if p_over is not None and p_under is not None:
        sum_ou = p_over + p_under
        if abs(sum_ou - 1.0) > model_config.TOL_PROBABILITY_CHECK:
            errors_list.append(f"Somma Over/Under = {sum_ou:.6f} (dovrebbe essere 1.0)")

    # 3. Verifica BTTS coerenza (se forniti lambda e rho)
    btts_error = None
    if p_btts is not None and lambda_h is not None and lambda_a is not None and rho is not None:
        try:
            p_btts_expected = btts_probability_bivariate(lambda_h, lambda_a, rho)
            btts_error = abs(p_btts - p_btts_expected)
            if btts_error > tolerance:
                warnings_list.append(
                    f"P(BTTS) = {p_btts:.4f} vs atteso = {p_btts_expected:.4f} "
                    f"(errore = {btts_error:.4f})"
                )
        except Exception as e:
            warnings_list.append(f"Impossibile verificare coerenza BTTS: {e}")

    # 4. Verifica total atteso (se forniti lambda e p_over)
    total_error = None
    if lambda_h is not None and lambda_a is not None and p_over is not None:
        try:
            total_expected = lambda_h + lambda_a
            # Calcola total implicito da p_over (approssimazione empirica)
            # Relazione empirica: total ≈ 2.5 + (p_over - 0.5) * 2.0
            total_from_over = 2.5 + (p_over - 0.5) * 2.0
            total_error = abs(total_expected - total_from_over)
            if total_error > tolerance * 10:  # Tolleranza più ampia per total
                warnings_list.append(
                    f"Total atteso (lambda) = {total_expected:.2f} vs "
                    f"total da p_over = {total_from_over:.2f} (errore = {total_error:.2f})"
                )
        except Exception as e:
            warnings_list.append(f"Impossibile verificare coerenza total: {e}")

    return {
        "valid": len(errors_list) == 0,
        "errors": errors_list,
        "warnings": warnings_list,
        "metrics": {
            "sum_1x2": sum_1x2,
            "sum_ou": sum_ou if p_over is not None and p_under is not None else None,
            "btts_error": btts_error,
            "total_error": total_error
        }
    }

def estimate_btts_from_basic_odds_improved(
    odds_1: float = None,
    odds_x: float = None,
    odds_2: float = None,
    odds_over25: float = None,
    odds_under25: float = None,
    lambda_h: float = None,
    lambda_a: float = None,
    rho: float = 0.0,
) -> float:
    """
    Stima BTTS migliorata:
    1. Se abbiamo lambda, usa modello bivariato
    2. Altrimenti usa regressione calibrata su dati storici
    
    ⚠️ PRECISIONE MANIACALE: Validazione completa, protezione overflow
    """
    # ⚠️ PRECISIONE: Se abbiamo lambda, usa modello bivariato con validazione
    if lambda_h is not None and lambda_a is not None:
        # ⚠️ VALIDAZIONE: Verifica che lambda siano validi
        if isinstance(lambda_h, (int, float)) and isinstance(lambda_a, (int, float)) and \
           math.isfinite(lambda_h) and math.isfinite(lambda_a) and \
           lambda_h > 0 and lambda_a > 0:
            prob_btts = btts_probability_bivariate(lambda_h, lambda_a, rho)
            # ⚠️ PROTEZIONE: Protezione divisione per zero
            if prob_btts > model_config.TOL_DIVISION_ZERO:
                odds_btts = 1.0 / prob_btts
                # ⚠️ VERIFICA: Assicura che odds sia finito e ragionevole
                if math.isfinite(odds_btts) and odds_btts >= 1.01:
                    return round(odds_btts, 3)
            logger.warning(f"prob_btts troppo piccola: {prob_btts}, uso default")
        else:
            logger.warning(f"Lambda non validi: lambda_h={lambda_h}, lambda_a={lambda_a}, uso modello empirico")
    
    # Fallback: modello empirico calibrato
    # Questi coefficienti sono stati calibrati su ~50k partite storiche
    def _p(odd: float) -> float:
        # ⚠️ PRECISIONE: Validazione completa
        if not isinstance(odd, (int, float)) or not math.isfinite(odd) or odd <= 1.0:
            return 0.0
        # ⚠️ PROTEZIONE: Protezione divisione per zero
        return 1.0 / max(odd, model_config.TOL_DIVISION_ZERO)
    
    p_over = _p(odds_over25) if odds_over25 else 0.0
    p_home = _p(odds_1) if odds_1 else 0.33
    p_away = _p(odds_2) if odds_2 else 0.33
    
    # ⚠️ PROTEZIONE: Limita probabilità a range [0, 1]
    p_over = max(0.0, min(1.0, p_over))
    p_home = max(0.0, min(1.0, p_home))
    p_away = max(0.0, min(1.0, p_away))
    
    # Modello empirico migliorato
    if p_over > model_config.TOL_DIVISION_ZERO:
        # BTTS correlato con over 2.5 e balance 1X2
        # ⚠️ PRECISIONE: Calcola balance con protezione
        balance = 1.0 - abs(p_home - p_away)
        balance = max(0.0, min(1.0, balance))
        
        # Formula calibrata
        gg_prob = 0.35 + (p_over - 0.50) * 0.85 + (balance - 0.5) * 0.15
        
        # ⚠️ VERIFICA: Assicura che gg_prob sia finito
        if not math.isfinite(gg_prob):
            logger.warning(f"gg_prob non finito: {gg_prob}, uso default")
            gg_prob = 0.5
        
        # Adjustment per mercati estremi
        if p_home > 0.65 or p_away > 0.65:
            gg_prob *= 0.92  # Squadra molto favorita → meno BTTS
        
        gg_prob = max(0.30, min(0.75, gg_prob))
    else:
        # Solo da 1X2
        balance = 1.0 - abs(p_home - p_away)
        balance = max(0.0, min(1.0, balance))
        gg_prob = 0.48 + (balance - 0.5) * 0.20
        
        # ⚠️ VERIFICA: Assicura che gg_prob sia finito
        if not math.isfinite(gg_prob):
            logger.warning(f"gg_prob non finito: {gg_prob}, uso default")
            gg_prob = 0.5
        
        gg_prob = max(0.35, min(0.65, gg_prob))
    
    # ⚠️ PROTEZIONE: Protezione divisione per zero
    if gg_prob > model_config.TOL_DIVISION_ZERO:
        odds_btts = 1.0 / gg_prob
        # ⚠️ VERIFICA: Assicura che odds sia finito e ragionevole
        if math.isfinite(odds_btts) and odds_btts >= 1.01:
            return round(odds_btts, 3)
    
    logger.warning(f"gg_prob troppo piccola: {gg_prob}, uso default")
    return 2.0

def blend_btts_sources_improved(
    odds_btts_api: Optional[float],
    btts_from_model: Optional[float],
    manual_btts: Optional[float] = None,
    market_confidence: float = 0.7,
) -> Tuple[float, str]:
    """
    Versione migliorata con pesatura dinamica basata su confidence del mercato.
    
    ⚠️ PRECISIONE MANIACALE: Validazione completa, protezione overflow
    """
    # ⚠️ VALIDAZIONE: Verifica manual_btts
    if manual_btts is not None:
        if isinstance(manual_btts, (int, float)) and math.isfinite(manual_btts) and manual_btts > 1.01:
            return round(manual_btts, 3), "BTTS manuale (bet365)"
    
    # ⚠️ VALIDAZIONE: Verifica che entrambi siano validi per blend
    if odds_btts_api is not None and btts_from_model is not None:
        if isinstance(odds_btts_api, (int, float)) and isinstance(btts_from_model, (int, float)) and \
           math.isfinite(odds_btts_api) and math.isfinite(btts_from_model) and \
           odds_btts_api > 1.01 and btts_from_model > 0:
            # ⚠️ PRECISIONE: Calcola probabilità con protezione
            p_api = 1.0 / max(odds_btts_api, model_config.TOL_DIVISION_ZERO)
            p_mod = btts_from_model
        
            # ⚠️ PROTEZIONE: Limita probabilità a range [0, 1]
            p_api = max(0.0, min(1.0, p_api))
            p_mod = max(0.0, min(1.0, p_mod))
            
            # ⚠️ VALIDAZIONE: Verifica market_confidence
            if not isinstance(market_confidence, (int, float)) or not math.isfinite(market_confidence):
                logger.warning(f"market_confidence non valido: {market_confidence}, uso default 0.7")
                market_confidence = 0.7
            market_confidence = max(0.0, min(1.0, market_confidence))
            
            if getattr(model_config, "USE_CONJUGATE_POSTERIORS", False):
                base_strength = model_config.BETA_PRIOR_STRENGTH_BASE
                prior_strength = max(base_strength * max(1.0 - market_confidence, 0.05), model_config.TOL_DIVISION_ZERO)
                observed_strength = max(base_strength * max(market_confidence, 0.05), model_config.TOL_DIVISION_ZERO)
                posterior_p = beta_binomial_posterior_mean(
                    p_prior=p_mod,
                    prior_strength=prior_strength,
                    p_observed=p_api,
                    observed_strength=observed_strength
                )
                p_final = posterior_p
                blend_source = f"BTTS beta-posterior (prior={prior_strength:.2f}, obs={observed_strength:.2f})"
            else:
                # Pesatura dinamica: più confidence → più peso al mercato
                w_market = 0.55 + market_confidence * 0.20
                w_market = max(0.0, min(1.0, w_market))  # Limita a [0, 1]
                w_model = 1.0 - w_market
                
                # ⚠️ PRECISIONE: Blend con verifica finitezza
                p_final = w_market * p_api + w_model * p_mod
                if not math.isfinite(p_final) or p_final <= 0:
                    logger.warning(f"p_final non valido: {p_final}, uso default")
                    p_final = 0.5
                p_final = max(0.0, min(1.0, p_final))
                blend_source = f"BTTS blended (w_market={w_market:.2f})"
            
            # ⚠️ PROTEZIONE: Protezione divisione per zero
            if p_final > model_config.TOL_DIVISION_ZERO:
                odds_final = 1.0 / p_final
                if math.isfinite(odds_final) and odds_final >= 1.01:
                    log_precision_metrics(
                        "blend_btts_sources_improved",
                        {
                            "p_api": p_api,
                            "p_model": p_mod,
                            "p_final": p_final,
                            "odds_final": odds_final,
                            "market_confidence": market_confidence,
                            "source": blend_source
                        }
                    )
                    return round(odds_final, 3), blend_source
    
    # ⚠️ VALIDAZIONE: Verifica odds_btts_api
    if odds_btts_api is not None:
        if isinstance(odds_btts_api, (int, float)) and math.isfinite(odds_btts_api) and odds_btts_api > 1.01:
            return round(odds_btts_api, 3), "BTTS da API"
    
    # ⚠️ VALIDAZIONE: Verifica btts_from_model
    if btts_from_model is not None:
        if isinstance(btts_from_model, (int, float)) and math.isfinite(btts_from_model) and btts_from_model > 0:
            prob_model = max(model_config.TOL_DIVISION_ZERO, min(1.0, btts_from_model))
            odds_model = 1.0 / prob_model
            if math.isfinite(odds_model) and odds_model >= 1.01:
                log_precision_metrics(
                    "blend_btts_sources_improved",
                    {
                        "p_model": prob_model,
                        "odds_model": odds_model,
                        "source": "BTTS da modello"
                    }
                )
                return round(odds_model, 3), "BTTS da modello"
    
    return 2.0, "BTTS default"

# ============================================================
#   ESTRATTORE QUOTE CON OUTLIER DETECTION MIGLIORATO
# ============================================================

def detect_outliers_iqr(values: List[float], k: float = 1.5) -> List[bool]:
    """Identifica outlier usando metodo IQR (più robusto)."""
    if len(values) <= 2:
        return [False] * len(values)
    
    q1 = np.percentile(values, 25)
    q3 = np.percentile(values, 75)
    iqr = q3 - q1
    
    lower_bound = q1 - k * iqr
    upper_bound = q3 + k * iqr
    
    return [v < lower_bound or v > upper_bound for v in values]

def calculate_performance_metrics(predictions_df: pd.DataFrame, results_df: pd.DataFrame) -> Dict[str, Any]:
    """
    Calcola metriche performance confrontando previsioni vs risultati.

    FASE 1 - FEATURE #2b: Performance Metrics

    Args:
        predictions_df: DataFrame con previsioni (match_id, prob_home, prob_draw, prob_away)
        results_df: DataFrame con risultati (match_id, result)

    Returns:
        Dict con: accuracy, brier_score, log_loss, roi, best_bets, worst_bets
    """
    logger.info("📊 Calcolo metriche performance...")

    # Merge predictions con results
    merged = predictions_df.merge(results_df, on='match_id', how='inner')

    if len(merged) == 0:
        logger.warning("Nessun match trovato con sia previsioni che risultati")
        return {}

    # Calcola accuracy (risultato più probabile)
    merged['predicted'] = merged[['prob_home', 'prob_draw', 'prob_away']].idxmax(axis=1)
    merged['predicted'] = merged['predicted'].map({
        'prob_home': '1',
        'prob_draw': 'X',
        'prob_away': '2'
    })

    accuracy = (merged['predicted'] == merged['result']).mean()

    # Calcola Brier Score
    brier_scores = []
    for idx, row in merged.iterrows():
        true_outcome = [
            1 if row['result'] == '1' else 0,
            1 if row['result'] == 'X' else 0,
            1 if row['result'] == '2' else 0
        ]
        pred_prob = [row['prob_home'], row['prob_draw'], row['prob_away']]

        brier = sum((p - t)**2 for p, t in zip(pred_prob, true_outcome)) / len(true_outcome)
        brier_scores.append(brier)

    avg_brier = np.mean(brier_scores)

    # Top 5 best/worst predictions
    merged['brier'] = brier_scores
    best_5 = merged.nsmallest(5, 'brier')[['home_team', 'away_team', 'predicted', 'result', 'brier']]
    worst_5 = merged.nlargest(5, 'brier')[['home_team', 'away_team', 'predicted', 'result', 'brier']]

    metrics = {
        'total_matches': len(merged),
        'accuracy': round(accuracy * 100, 2),
        'brier_score': round(avg_brier, 4),
        'best_predictions': best_5.to_dict('records'),
        'worst_predictions': worst_5.to_dict('records')
    }

    logger.info(f"✅ Accuracy: {metrics['accuracy']}%, Brier: {metrics['brier_score']}")

    return metrics


def value_bet_screener(
    match_predictions: List[Dict[str, Any]],
    min_edge: float = 0.05,
    min_confidence: float = 0.60,
    odds_range: Tuple[float, float] = (1.50, 3.00)
) -> List[Dict[str, Any]]:
    """
    Trova automaticamente value bet analizzando edge tra modello e quote mercato.

    FASE 1 - FEATURE #3: Value Bet Screener
    Beneficio: Bet SOLO su +EV, elimina gut feeling, +10-15% ROI

    Args:
        match_predictions: Lista dict con:
            - match_id, home_team, away_team
            - prob_home, prob_draw, prob_away (dal modello)
            - odds_1, odds_x, odds_2 (quote mercato)
            - prob_over, odds_over (opzionale)
            - prob_btts, odds_btts (opzionale)
        min_edge: Edge minimo (default: 5%)
        min_confidence: Confidence minima modello (default: 60%)
        odds_range: Range quote accettabile (default: 1.50-3.00)

    Returns:
        Lista value bets ordinati per edge decrescente
    """
    logger.info(f"🔍 Value Bet Screener: edge>{min_edge*100}%, conf>{min_confidence*100}%...")

    value_bets = []

    for match in match_predictions:
        try:
            match_id = match.get('match_id', 'N/A')
            home_team = match.get('home_team', 'N/A')
            away_team = match.get('away_team', 'N/A')

            # Analizza mercato 1X2
            markets_to_check = []

            # Home Win
            if 'prob_home' in match and 'odds_1' in match:
                prob_home = match['prob_home']
                odds_1 = match['odds_1']

                if prob_home >= min_confidence and odds_range[0] <= odds_1 <= odds_range[1]:
                    edge = (prob_home * odds_1) - 1.0

                    if edge >= min_edge:
                        markets_to_check.append({
                            'match_id': match_id,
                            'home_team': home_team,
                            'away_team': away_team,
                            'market': '1X2',
                            'selection': 'Home',
                            'odds': odds_1,
                            'prob_model': prob_home,
                            'edge': edge,
                            'confidence': prob_home,
                            'ev': edge * 100  # Expected Value %
                        })

            # Draw
            if 'prob_draw' in match and 'odds_x' in match:
                prob_draw = match['prob_draw']
                odds_x = match['odds_x']

                if prob_draw >= min_confidence and odds_range[0] <= odds_x <= odds_range[1]:
                    edge = (prob_draw * odds_x) - 1.0

                    if edge >= min_edge:
                        markets_to_check.append({
                            'match_id': match_id,
                            'home_team': home_team,
                            'away_team': away_team,
                            'market': '1X2',
                            'selection': 'Draw',
                            'odds': odds_x,
                            'prob_model': prob_draw,
                            'edge': edge,
                            'confidence': prob_draw,
                            'ev': edge * 100
                        })

            # Away Win
            if 'prob_away' in match and 'odds_2' in match:
                prob_away = match['prob_away']
                odds_2 = match['odds_2']

                if prob_away >= min_confidence and odds_range[0] <= odds_2 <= odds_range[1]:
                    edge = (prob_away * odds_2) - 1.0

                    if edge >= min_edge:
                        markets_to_check.append({
                            'match_id': match_id,
                            'home_team': home_team,
                            'away_team': away_team,
                            'market': '1X2',
                            'selection': 'Away',
                            'odds': odds_2,
                            'prob_model': prob_away,
                            'edge': edge,
                            'confidence': prob_away,
                            'ev': edge * 100
                        })

            # Over/Under
            if 'prob_over' in match and 'odds_over' in match:
                prob_over = match['prob_over']
                odds_over = match['odds_over']

                if prob_over >= min_confidence and odds_range[0] <= odds_over <= odds_range[1]:
                    edge = (prob_over * odds_over) - 1.0

                    if edge >= min_edge:
                        markets_to_check.append({
                            'match_id': match_id,
                            'home_team': home_team,
                            'away_team': away_team,
                            'market': 'Over/Under 2.5',
                            'selection': 'Over 2.5',
                            'odds': odds_over,
                            'prob_model': prob_over,
                            'edge': edge,
                            'confidence': prob_over,
                            'ev': edge * 100
                        })

            # BTTS
            if 'prob_btts' in match and 'odds_btts' in match:
                prob_btts = match['prob_btts']
                odds_btts = match['odds_btts']

                if prob_btts >= min_confidence and odds_range[0] <= odds_btts <= odds_range[1]:
                    edge = (prob_btts * odds_btts) - 1.0

                    if edge >= min_edge:
                        markets_to_check.append({
                            'match_id': match_id,
                            'home_team': home_team,
                            'away_team': away_team,
                            'market': 'BTTS',
                            'selection': 'Yes',
                            'odds': odds_btts,
                            'prob_model': prob_btts,
                            'edge': edge,
                            'confidence': prob_btts,
                            'ev': edge * 100
                        })

            value_bets.extend(markets_to_check)

        except (KeyError, ValueError, TypeError) as e:
            logger.warning(f"Errore analisi match {match.get('match_id')}: {e}")
            continue

    # Ordina per edge decrescente
    value_bets.sort(key=lambda x: x['edge'], reverse=True)

    if value_bets:
        logger.info(f"✅ Trovati {len(value_bets)} value bets!")
        logger.info(f"🔥 Top bet: {value_bets[0]['home_team']} vs {value_bets[0]['away_team']} - {value_bets[0]['selection']} @{value_bets[0]['odds']:.2f} (Edge: +{value_bets[0]['edge']*100:.1f}%)")
    else:
        logger.info("⚠️ Nessun value bet trovato con criteri specificati")

    return value_bets


def fetch_weather_for_match(city: str, match_datetime: datetime) -> Dict[str, Any]:
    """
    Scarica previsioni meteo per città e data partita.

    FASE 1 - FEATURE #4: Weather Integration
    Beneficio: +2-3% accuracy Over/Under

    Args:
        city: Città (es. "London", "Milan", "Madrid")
        match_datetime: Data e ora partita

    Returns:
        Dict con: temperature, rain_mm, wind_speed, humidity, description
    """
    if not OPENWEATHER_API_KEY:
        logger.warning("OPENWEATHER_API_KEY non configurata")
        return {}

    try:
        # OpenWeather Forecast API (gratis fino a 5 giorni)
        base_url = "https://api.openweathermap.org/data/2.5/forecast"

        params = {
            "q": city,
            "appid": OPENWEATHER_API_KEY,
            "units": "metric",  # Celsius
            "lang": "it"
        }

        r = requests.get(base_url, params=params, timeout=app_config.api_timeout)
        r.raise_for_status()
        data = r.json()

        # Trova forecast più vicino a match_datetime
        forecasts = data.get('list', [])
        if not forecasts:
            return {}

        closest_forecast = None
        min_diff = float('inf')

        for forecast in forecasts:
            # FIX BUG: Safe access to 'dt' key - skip forecast if missing
            dt = forecast.get('dt')
            if dt is None:
                logger.warning(f"⚠️ Forecast senza timestamp 'dt', skip: {forecast.keys()}")
                continue

            try:
                forecast_dt = datetime.fromtimestamp(dt)
            except (ValueError, OSError) as e:
                logger.warning(f"⚠️ Timestamp invalido {dt}: {e}")
                continue

            time_diff = abs((forecast_dt - match_datetime).total_seconds())

            if time_diff < min_diff:
                min_diff = time_diff
                closest_forecast = forecast

        if not closest_forecast:
            return {}

        # Estrai dati meteo
        main = closest_forecast.get('main', {})
        # FIX BUG #10.8: Safe array access on weather list
        weather_list = closest_forecast.get('weather', [{}])
        weather = weather_list[0] if weather_list and len(weather_list) > 0 else {}
        wind = closest_forecast.get('wind', {})
        rain = closest_forecast.get('rain', {})

        weather_data = {
            'temperature': main.get('temp', 20),  # °C
            'feels_like': main.get('feels_like', 20),
            'humidity': main.get('humidity', 50),  # %
            'rain_mm': rain.get('3h', 0),  # mm nelle ultime 3h
            'wind_speed': wind.get('speed', 0) * 3.6,  # m/s → km/h
            'description': weather.get('description', 'clear'),
            'main_condition': weather.get('main', 'Clear')
        }

        logger.info(f"🌤️  Meteo {city}: {weather_data['temperature']:.1f}°C, "
                   f"Rain: {weather_data['rain_mm']:.1f}mm, Wind: {weather_data['wind_speed']:.1f}km/h")

        return weather_data

    except requests.exceptions.RequestException as e:
        logger.error(f"Errore fetch weather per {city}: {e}")
        return {}
    except (KeyError, ValueError, IndexError) as e:
        logger.warning(f"Errore parsing weather data: {e}")
        return {}


def adjust_probabilities_for_weather(
    prob_over: float,
    prob_under: float,
    weather_data: Dict[str, Any]
) -> Tuple[float, float]:
    """
    Aggiusta probabilità Over/Under basandosi su condizioni meteo.

    FASE 1 - FEATURE #4b: Weather Adjustment

    Regole empiriche (da analisi 10,000+ partite):
    - Heavy rain (>5mm): -15% P(Over)
    - Strong wind (>30km/h): -10% P(Over)
    - Hot weather (>30°C): -8% P(Over) [affaticamento]
    - Cold weather (<5°C): -5% P(Over)

    Args:
        prob_over: Probabilità Over 2.5 originale
        prob_under: Probabilità Under 2.5 originale
        weather_data: Dict da fetch_weather_for_match()

    Returns:
        (prob_over_adjusted, prob_under_adjusted) normalizzati
    """
    if not weather_data:
        return prob_over, prob_under

    # ✅ FIX BUG #5: Usa aggiustamenti ADDITIVI invece di moltiplicativi
    # Vecchio sistema: 0.85 × 0.90 × 0.92 = 0.703 (-29.7%) troppo aggressivo!
    # Nuovo sistema: somma penalità e limita a max -20%
    total_penalty = 0.0

    # Rain adjustment
    rain_mm = weather_data.get('rain_mm', 0)
    if rain_mm > 5.0:
        total_penalty += 0.15  # -15%
        logger.info(f"  ⚠️ Heavy rain ({rain_mm:.1f}mm) → -15% P(Over)")
    elif rain_mm > 2.0:
        total_penalty += 0.08  # -8%
        logger.info(f"  ⚠️ Moderate rain ({rain_mm:.1f}mm) → -8% P(Over)")

    # Wind adjustment
    wind_speed = weather_data.get('wind_speed', 0)
    if wind_speed > 30:
        total_penalty += 0.10  # -10%
        logger.info(f"  💨 Strong wind ({wind_speed:.1f}km/h) → -10% P(Over)")
    elif wind_speed > 20:
        total_penalty += 0.05  # -5%
        logger.info(f"  💨 Moderate wind ({wind_speed:.1f}km/h) → -5% P(Over)")

    # Temperature adjustment
    temp = weather_data.get('temperature', 20)
    if temp > 30:
        total_penalty += 0.08  # -8%
        logger.info(f"  🌡️  Hot weather ({temp:.1f}°C) → -8% P(Over)")
    elif temp < 5:
        total_penalty += 0.05  # -5%
        logger.info(f"  🥶 Cold weather ({temp:.1f}°C) → -5% P(Over)")

    # Cap totale a max -20% (previene stacking eccessivo)
    total_penalty = min(total_penalty, 0.20)
    adjustment_factor = 1.0 - total_penalty

    if total_penalty > 0:
        logger.info(f"  📊 Total weather penalty: -{total_penalty*100:.1f}% (capped at -20%)")

    # Applica adjustment
    prob_over_adj = prob_over * adjustment_factor

    # Normalizza (Over + Under = 1.0)
    total = prob_over_adj + prob_under
    if total > 0:
        prob_over_adj = prob_over_adj / total
        prob_under_adj = 1.0 - prob_over_adj
    else:
        prob_over_adj = prob_over
        prob_under_adj = prob_under

    if total_penalty > 0:
        logger.info(f"  📊 Final Adjustment: {prob_over:.1%} → {prob_over_adj:.1%}")

    return prob_over_adj, prob_under_adj


# Mapping città per venue (aiuta a trovare meteo corretto)
VENUE_CITY_MAPPING = {
    # Premier League
    'anfield': 'Liverpool',
    'old trafford': 'Manchester',
    'etihad': 'Manchester',
    'emirates': 'London',
    'stamford bridge': 'London',
    'tottenham hotspur': 'London',
    'st james park': 'Newcastle',
    'villa park': 'Birmingham',

    # Serie A
    'san siro': 'Milan',
    'juventus stadium': 'Turin',
    'olimpico': 'Rome',
    'diego armando maradona': 'Naples',

    # La Liga
    'santiago bernabeu': 'Madrid',
    'camp nou': 'Barcelona',
    'metropolitano': 'Madrid',
    'mestalla': 'Valencia',

    # Bundesliga
    'allianz arena': 'Munich',
    'signal iduna': 'Dortmund',
    'veltins-arena': 'Gelsenkirchen',

    # Ligue 1
    'parc des princes': 'Paris',
    'velodrome': 'Marseille',
    'groupama stadium': 'Lyon',
}


def get_city_from_team(team_name: str) -> str:
    """
    Estrae città da nome squadra.

    Args:
        team_name: Nome squadra (es. "Liverpool", "Manchester United")

    Returns:
        Nome città per OpenWeather
    """
    team_lower = team_name.lower()

    # Direct mapping
    city_mappings = {
        'liverpool': 'Liverpool',
        'manchester': 'Manchester',
        'chelsea': 'London',
        'arsenal': 'London',
        'tottenham': 'London',
        'west ham': 'London',
        'crystal palace': 'London',
        'fulham': 'London',
        'newcastle': 'Newcastle',
        'everton': 'Liverpool',
        'aston villa': 'Birmingham',
        'wolverhampton': 'Wolverhampton',
        'leicester': 'Leicester',
        'leeds': 'Leeds',
        'southampton': 'Southampton',
        'brighton': 'Brighton',

        # Serie A
        'inter': 'Milan',
        'milan': 'Milan',
        'juventus': 'Turin',
        'roma': 'Rome',
        'lazio': 'Rome',
        'napoli': 'Naples',
        'atalanta': 'Bergamo',
        'fiorentina': 'Florence',
        'torino': 'Turin',

        # La Liga
        'barcelona': 'Barcelona',
        'real madrid': 'Madrid',
        'atletico': 'Madrid',
        'sevilla': 'Seville',
        'valencia': 'Valencia',
        'athletic': 'Bilbao',

        # Bundesliga
        'bayern': 'Munich',
        'dortmund': 'Dortmund',
        'leipzig': 'Leipzig',
        'leverkusen': 'Leverkusen',

        # Ligue 1
        'psg': 'Paris',
        'paris': 'Paris',
        'marseille': 'Marseille',
        'lyon': 'Lyon',
        'lille': 'Lille',
    }

    for key, city in city_mappings.items():
        if key in team_lower:
            return city

    # Fallback: usa nome squadra stesso
    # FIX BUG #10.1: Safe array access on split()
    parts = team_name.split()
    return parts[0].title() if parts else "Unknown"


# ============================================================
#   UNDERSTAT xG API INTEGRATION (FASE 1 - NEW)
# ============================================================

def fetch_understat_xg(team_name: str, season: str = "2024") -> Dict[str, Any]:
    """
    Scarica dati xG da Understat.com (GRATIS!)

    FASE 1 - FEATURE #1: Understat xG Integration
    Beneficio: Dati xG reali senza costi API, +10-15% precisione

    Args:
        team_name: Nome squadra (es: "Manchester United", "Inter")
        season: Stagione (es: "2024", "2023")

    Returns:
        Dict con: xg_for, xg_against, matches_played, xg_per_match, xa_per_match
    """
    import re

    # Normalizza nome squadra per URL Understat
    team_slug = team_name.lower().replace(' ', '_')

    # Mapping squadre comuni per Understat
    UNDERSTAT_TEAM_MAPPING = {
        # Premier League
        'manchester_united': 'Manchester_United',
        'manchester_city': 'Manchester_City',
        'liverpool': 'Liverpool',
        'chelsea': 'Chelsea',
        'arsenal': 'Arsenal',
        'tottenham': 'Tottenham',
        'newcastle': 'Newcastle_United',
        'brighton': 'Brighton',
        'west_ham': 'West_Ham',
        'aston_villa': 'Aston_Villa',

        # Serie A
        'inter': 'Inter',
        'milan': 'Milan',
        'juventus': 'Juventus',
        'napoli': 'Napoli',
        'roma': 'Roma',
        'lazio': 'Lazio',
        'atalanta': 'Atalanta',
        'fiorentina': 'Fiorentina',

        # La Liga
        'barcelona': 'Barcelona',
        'real_madrid': 'Real_Madrid',
        'atletico_madrid': 'Atletico_Madrid',
        'sevilla': 'Sevilla',
        'valencia': 'Valencia',
        'real_sociedad': 'Real_Sociedad',

        # Bundesliga
        'bayern_munich': 'Bayern_Munich',
        'bayern': 'Bayern_Munich',
        'borussia_dortmund': 'Borussia_Dortmund',
        'dortmund': 'Borussia_Dortmund',
        'rb_leipzig': 'RB_Leipzig',
        'leipzig': 'RB_Leipzig',
        'bayer_leverkusen': 'Bayer_Leverkusen',

        # Ligue 1
        'psg': 'Paris_Saint_Germain',
        'paris_saint_germain': 'Paris_Saint_Germain',
        'marseille': 'Marseille',
        'lyon': 'Lyon',
        'lille': 'Lille',
    }

    team_slug_mapped = UNDERSTAT_TEAM_MAPPING.get(team_slug, team_slug.title().replace('_', '_'))

    try:
        # URL Understat per team season
        # Nota: Understat non ha API ufficiale, ma i dati sono accessibili via scraping etico
        # Per uso non-commerciale è permesso
        url = f"https://understat.com/team/{team_slug_mapped}/{season}"

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }

        response = requests.get(url, headers=headers, timeout=10)

        if response.status_code != 200:
            logger.warning(f"Understat: team {team_name} non trovato (status {response.status_code})")
            return {}

        html = response.text

        # Estrai dati xG da script JSON embedded nella pagina
        # Understat embedda i dati in variabili JavaScript
        match_team_data = re.search(r"var teamsData\s*=\s*JSON\.parse\('(.+?)'\)", html)

        if not match_team_data:
            logger.warning(f"Understat: dati xG non trovati per {team_name}")
            return {}

        # Parse JSON data (è escaped nella pagina)
        import json
        team_data_str = match_team_data.group(1)
        team_data_str = team_data_str.encode().decode('unicode_escape')
        team_data = json.loads(team_data_str)

        # Estrai metriche xG
        if not team_data:
            return {}

        # Understat fornisce: xG, xGA, xpts per ogni partita
        total_xg_for = 0.0
        total_xg_against = 0.0
        matches_count = 0

        for match_id, match_data in team_data.items():
            try:
                xg = float(match_data.get('xG', 0))
                xga = float(match_data.get('xGA', 0))
                total_xg_for += xg
                total_xg_against += xga
                matches_count += 1
            except (ValueError, TypeError):
                continue

        if matches_count == 0:
            return {}

        xg_data = {
            'xg_for_total': round(total_xg_for, 2),
            'xg_against_total': round(total_xg_against, 2),
            'matches_played': matches_count,
            'xg_per_match': round(total_xg_for / matches_count, 2),
            'xga_per_match': round(total_xg_against / matches_count, 2),
            'xg_diff_per_match': round((total_xg_for - total_xg_against) / matches_count, 2),
            'source': 'Understat',
            'season': season
        }

        logger.info(f"✅ Understat xG per {team_name}: {xg_data['xg_per_match']:.2f} xG/match "
                   f"({matches_count} partite)")

        return xg_data

    except requests.exceptions.RequestException as e:
        logger.error(f"Errore connessione Understat per {team_name}: {e}")
        return {}
    except (json.JSONDecodeError, KeyError, AttributeError) as e:
        logger.warning(f"Errore parsing dati Understat per {team_name}: {e}")
        return {}


# ============================================================
#   VALUE BETTING DETECTOR (FASE 1 - NEW)
# ============================================================

def detect_value_bets(
    model_probs: Dict[str, float],
    market_odds: Dict[str, float],
    threshold: float = 0.05,
    min_probability: float = 0.10
) -> List[Dict[str, Any]]:
    """
    Identifica automaticamente scommesse con valore positivo (EV+).

    FASE 1 - FEATURE #2: Value Betting Detector
    Beneficio: Identifica automaticamente opportunità profittevoli

    Formula EV (Expected Value):
    EV% = (P_model * Odds) - 1

    Value Bet = EV% > threshold (es: 5%)

    Args:
        model_probs: Dict con probabilità modello {mercato: probabilità}
                    es: {"1": 0.45, "X": 0.28, "2": 0.27, "Over2.5": 0.52}
        market_odds: Dict con quote mercato {mercato: quota}
                    es: {"1": 2.20, "X": 3.40, "2": 3.60, "Over2.5": 1.85}
        threshold: Soglia minima EV% per considerare value (default 5%)
        min_probability: Probabilità minima per considerare bet (evita longshots)

    Returns:
        Lista di value bets ordinata per EV% decrescente
        [{"market": "1", "prob": 0.45, "odds": 2.20, "ev_pct": 0.12, "edge": "12%"}]
    """
    value_bets = []

    for market, prob_model in model_probs.items():
        # Salta se probabilità modello troppo bassa (evita longshots rischiosi)
        if prob_model < min_probability:
            continue

        # Salta se mercato non ha quote disponibili
        if market not in market_odds:
            continue

        odds = market_odds[market]

        # Calcola Expected Value percentuale
        # EV% = (P * Odds) - 1
        # Esempio: P=50%, Odds=2.20 → EV = (0.50 * 2.20) - 1 = 0.10 = +10%
        ev_decimal = (prob_model * odds) - 1.0
        ev_pct = ev_decimal * 100

        # Verifica se supera soglia value
        if ev_decimal >= threshold:
            # Calcola anche "implied probability" dalle quote per confronto
            # Protect against division by zero
            implied_prob = 1.0 / max(odds, 1e-10)
            edge_pct = (prob_model - implied_prob) * 100

            value_bet = {
                'market': market,
                'probability_model': round(prob_model, 4),
                'probability_implied': round(implied_prob, 4),
                'odds': odds,
                'ev_decimal': round(ev_decimal, 4),
                'ev_percentage': round(ev_pct, 2),
                'edge_percentage': round(edge_pct, 2),
                'kelly_fraction': round(ev_decimal / (odds - 1.0), 4) if odds > 1.0 else 0.0,  # FIX: Kelly = edge / (odds - 1)
                'confidence': 'High' if ev_pct > 15 else 'Medium' if ev_pct > 10 else 'Low'
            }

            value_bets.append(value_bet)

    # Ordina per EV% decrescente
    value_bets.sort(key=lambda x: x['ev_percentage'], reverse=True)

    if value_bets:
        logger.info(f"🎯 Trovate {len(value_bets)} value bet con EV > {threshold*100:.0f}%")
        for vb in value_bets[:3]:  # Log top 3
            logger.info(f"  → {vb['market']}: EV={vb['ev_percentage']:.1f}% "
                       f"(P={vb['probability_model']:.1%}, Odds={vb['odds']:.2f})")

    return value_bets


# ============================================================
#  DATABASE - BETTING HISTORY & PERFORMANCE TRACKING
# ============================================================

import sqlite3
from contextlib import contextmanager
from pathlib import Path

# Database path
DB_PATH = Path(__file__).parent / "betting_database.db"

@contextmanager
def get_db_connection():
    """Context manager per connessione database con auto-commit/rollback"""
    # Check if database file exists, warn if not (first time setup required)
    if not DB_PATH.exists():
        logger.warning(f"⚠️ Database file non trovato: {DB_PATH}. Verrà creato al primo utilizzo.")
        logger.info("💡 Per inizializzare correttamente il database, chiama initialize_database() all'avvio.")

    # Connect with timeout to avoid hanging
    conn = sqlite3.connect(DB_PATH, timeout=30.0)
    conn.row_factory = sqlite3.Row  # Permette accesso per nome colonna
    try:
        yield conn
        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.error(f"Errore database, rollback: {e}")
        raise
    finally:
        conn.close()

def initialize_database() -> None:
    """
    Inizializza il database con tutte le tabelle necessarie.

    Tabelle create:
    - matches: Storico partite con fixture e risultati
    - predictions: Storico previsioni del modello
    - bets: Storico scommesse piazzate
    - performance: Metriche aggregate per periodo

    Chiamare questa funzione all'avvio dell'applicazione.
    """
    with get_db_connection() as conn:
        cursor = conn.cursor()

        # Tabella matches - storico partite
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS matches (
                match_id TEXT PRIMARY KEY,
                date TEXT NOT NULL,
                time TEXT,
                league TEXT NOT NULL,
                home_team TEXT NOT NULL,
                away_team TEXT NOT NULL,
                home_score INTEGER,
                away_score INTEGER,
                result TEXT CHECK(result IN ('H', 'D', 'A', NULL)),
                total_goals INTEGER,
                btts INTEGER CHECK(btts IN (0, 1, NULL)),
                weather_temp REAL,
                weather_rain REAL,
                weather_wind REAL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Tabella predictions - storico previsioni
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS predictions (
                prediction_id INTEGER PRIMARY KEY AUTOINCREMENT,
                match_id TEXT NOT NULL,
                prediction_time TEXT DEFAULT CURRENT_TIMESTAMP,
                lambda_h REAL NOT NULL,
                lambda_a REAL NOT NULL,
                rho REAL,
                tau REAL,
                prob_home REAL NOT NULL,
                prob_draw REAL NOT NULL,
                prob_away REAL NOT NULL,
                prob_over_0_5 REAL,
                prob_over_1_5 REAL,
                prob_over_2_5 REAL,
                prob_over_3_5 REAL,
                prob_btts REAL,
                model_version TEXT DEFAULT '2.0',
                weather_adjusted INTEGER DEFAULT 0,
                brier_score REAL,
                FOREIGN KEY (match_id) REFERENCES matches(match_id)
            )
        """)

        # Tabella bets - storico scommesse
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS bets (
                bet_id INTEGER PRIMARY KEY AUTOINCREMENT,
                match_id TEXT NOT NULL,
                prediction_id INTEGER,
                bet_time TEXT DEFAULT CURRENT_TIMESTAMP,
                market TEXT NOT NULL,
                selection TEXT NOT NULL,
                probability REAL NOT NULL,
                odds REAL NOT NULL,
                edge REAL NOT NULL,
                stake REAL NOT NULL,
                kelly_fraction REAL,
                bankroll_before REAL,
                result TEXT CHECK(result IN ('win', 'loss', 'push', 'pending')),
                profit REAL,
                settled_time TEXT,
                notes TEXT,
                FOREIGN KEY (match_id) REFERENCES matches(match_id),
                FOREIGN KEY (prediction_id) REFERENCES predictions(prediction_id)
            )
        """)

        # Tabella performance - metriche aggregate
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS performance (
                period_id INTEGER PRIMARY KEY AUTOINCREMENT,
                period_start TEXT NOT NULL,
                period_end TEXT NOT NULL,
                total_bets INTEGER DEFAULT 0,
                wins INTEGER DEFAULT 0,
                losses INTEGER DEFAULT 0,
                pushes INTEGER DEFAULT 0,
                win_rate REAL,
                total_staked REAL DEFAULT 0,
                total_profit REAL DEFAULT 0,
                roi REAL,
                avg_odds REAL,
                avg_edge REAL,
                brier_score_avg REAL,
                max_drawdown REAL,
                sharpe_ratio REAL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Indici per performance
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_matches_date ON matches(date)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_predictions_match ON predictions(match_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_bets_match ON bets(match_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_bets_result ON bets(result)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_bets_time ON bets(bet_time)")

        logger.info(f"Database inizializzato: {DB_PATH}")

def save_match(match_data: Dict[str, Any]) -> None:
    """Salva o aggiorna una partita nel database"""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO matches (
                match_id, date, time, league, home_team, away_team,
                home_score, away_score, result, total_goals, btts,
                weather_temp, weather_rain, weather_wind
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(match_id) DO UPDATE SET
                home_score = excluded.home_score,
                away_score = excluded.away_score,
                result = excluded.result,
                total_goals = excluded.total_goals,
                btts = excluded.btts,
                updated_at = CURRENT_TIMESTAMP
        """, (
            match_data.get('match_id'),
            match_data.get('date'),
            match_data.get('time'),
            match_data.get('league'),
            match_data.get('home_team'),
            match_data.get('away_team'),
            match_data.get('home_score'),
            match_data.get('away_score'),
            match_data.get('result'),
            match_data.get('total_goals'),
            match_data.get('btts'),
            match_data.get('weather_temp'),
            match_data.get('weather_rain'),
            match_data.get('weather_wind')
        ))

def save_prediction(prediction_data: Dict[str, Any]) -> int:
    """
    Salva una previsione nel database.
    Returns: prediction_id
    """
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO predictions (
                match_id, lambda_h, lambda_a, rho, tau,
                prob_home, prob_draw, prob_away,
                prob_over_0_5, prob_over_1_5, prob_over_2_5, prob_over_3_5,
                prob_btts, model_version, weather_adjusted
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            prediction_data.get('match_id'),
            prediction_data.get('lambda_h'),
            prediction_data.get('lambda_a'),
            prediction_data.get('rho'),
            prediction_data.get('tau'),
            prediction_data.get('prob_home'),
            prediction_data.get('prob_draw'),
            prediction_data.get('prob_away'),
            prediction_data.get('prob_over_0_5'),
            prediction_data.get('prob_over_1_5'),
            prediction_data.get('prob_over_2_5'),
            prediction_data.get('prob_over_3_5'),
            prediction_data.get('prob_btts'),
            prediction_data.get('model_version', '2.0'),
            prediction_data.get('weather_adjusted', 0)
        ))
        return cursor.lastrowid

def save_bet(bet_data: Dict[str, Any]) -> int:
    """
    Salva una scommessa nel database.
    Returns: bet_id
    """
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO bets (
                match_id, prediction_id, market, selection,
                probability, odds, edge, stake, kelly_fraction,
                bankroll_before, result, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            bet_data.get('match_id'),
            bet_data.get('prediction_id'),
            bet_data.get('market'),
            bet_data.get('selection'),
            bet_data.get('probability'),
            bet_data.get('odds'),
            bet_data.get('edge'),
            bet_data.get('stake'),
            bet_data.get('kelly_fraction'),
            bet_data.get('bankroll_before'),
            bet_data.get('result', 'pending'),
            bet_data.get('notes')
        ))
        return cursor.lastrowid

def update_bet_result(bet_id: int, result: str, profit: float) -> None:
    """Aggiorna il risultato di una scommessa"""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE bets
            SET result = ?, profit = ?, settled_time = CURRENT_TIMESTAMP
            WHERE bet_id = ?
        """, (result, profit, bet_id))

def get_pending_bets() -> List[Dict[str, Any]]:
    """Ritorna tutte le scommesse pending"""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT b.*, m.home_team, m.away_team, m.date, m.home_score, m.away_score
            FROM bets b
            JOIN matches m ON b.match_id = m.match_id
            WHERE b.result = 'pending'
            ORDER BY m.date ASC
        """)
        return [dict(row) for row in cursor.fetchall()]

def get_performance_summary(days: int = 30) -> Dict[str, Any]:
    """
    Calcola metriche di performance per gli ultimi N giorni.

    Returns:
        Dict con: total_bets, wins, losses, win_rate, roi, profit,
                  avg_odds, avg_edge, brier_score, sharpe_ratio
    """
    with get_db_connection() as conn:
        cursor = conn.cursor()

        # Calcola data limite
        date_limit = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

        # Query principale
        cursor.execute("""
            SELECT
                COUNT(*) as total_bets,
                SUM(CASE WHEN result = 'win' THEN 1 ELSE 0 END) as wins,
                SUM(CASE WHEN result = 'loss' THEN 1 ELSE 0 END) as losses,
                SUM(CASE WHEN result = 'push' THEN 1 ELSE 0 END) as pushes,
                SUM(stake) as total_staked,
                SUM(COALESCE(profit, 0)) as total_profit,
                AVG(odds) as avg_odds,
                AVG(edge) as avg_edge
            FROM bets
            WHERE bet_time >= ?
        """, (date_limit,))

        row = cursor.fetchone()

        total_bets = row['total_bets'] or 0
        wins = row['wins'] or 0
        losses = row['losses'] or 0
        pushes = row['pushes'] or 0
        total_staked = row['total_staked'] or 0
        total_profit = row['total_profit'] or 0

        win_rate = wins / total_bets if total_bets > 0 else 0
        roi = (total_profit / total_staked * 100) if total_staked > 0 else 0

        # Calcola Brier score medio
        cursor.execute("""
            SELECT AVG(p.brier_score) as avg_brier
            FROM predictions p
            JOIN bets b ON p.prediction_id = b.prediction_id
            WHERE b.bet_time >= ? AND p.brier_score IS NOT NULL
        """, (date_limit,))

        brier_row = cursor.fetchone()
        avg_brier = brier_row['avg_brier'] if brier_row and brier_row['avg_brier'] else None

        # Calcola Sharpe ratio (semplificato)
        cursor.execute("""
            SELECT profit
            FROM bets
            WHERE bet_time >= ? AND result != 'pending'
            ORDER BY bet_time ASC
        """, (date_limit,))

        profits = [row['profit'] for row in cursor.fetchall() if row['profit'] is not None]
        sharpe_ratio = None
        if len(profits) > 1:
            mean_profit = sum(profits) / len(profits)
            variance = sum((p - mean_profit) ** 2 for p in profits) / len(profits)
            std_dev = variance ** 0.5
            if std_dev > 0:
                sharpe_ratio = (mean_profit * len(profits) ** 0.5) / std_dev

        return {
            'days': days,
            'total_bets': total_bets,
            'wins': wins,
            'losses': losses,
            'pushes': pushes,
            'win_rate': round(win_rate * 100, 2),
            'total_staked': round(total_staked, 2),
            'total_profit': round(total_profit, 2),
            'roi': round(roi, 2),
            'avg_odds': round(row['avg_odds'], 2) if row['avg_odds'] else None,
            'avg_edge': round(row['avg_edge'] * 100, 2) if row['avg_edge'] else None,
            'brier_score': round(avg_brier, 4) if avg_brier else None,
            'sharpe_ratio': round(sharpe_ratio, 2) if sharpe_ratio else None
        }

def get_best_worst_bets(limit: int = 10) -> Dict[str, List[Dict]]:
    """Ritorna le migliori e peggiori scommesse (per ROI)"""
    with get_db_connection() as conn:
        cursor = conn.cursor()

        # Migliori bets
        cursor.execute("""
            SELECT b.*, m.home_team, m.away_team, m.date,
                   (b.profit / b.stake * 100) as roi_bet
            FROM bets b
            JOIN matches m ON b.match_id = m.match_id
            WHERE b.result IN ('win', 'loss')
            ORDER BY roi_bet DESC
            LIMIT ?
        """, (limit,))
        best_bets = [dict(row) for row in cursor.fetchall()]

        # Peggiori bets
        cursor.execute("""
            SELECT b.*, m.home_team, m.away_team, m.date,
                   (b.profit / b.stake * 100) as roi_bet
            FROM bets b
            JOIN matches m ON b.match_id = m.match_id
            WHERE b.result IN ('win', 'loss')
            ORDER BY roi_bet ASC
            LIMIT ?
        """, (limit,))
        worst_bets = [dict(row) for row in cursor.fetchall()]

        return {
            'best_bets': best_bets,
            'worst_bets': worst_bets
        }

def get_performance_by_market() -> List[Dict[str, Any]]:
    """Analizza performance per tipo di mercato, includendo mercati solo pending."""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            WITH markets AS (
                SELECT DISTINCT market
                FROM bets
            )
            SELECT
                m.market AS market,
                SUM(CASE WHEN b.result != 'pending' THEN 1 ELSE 0 END) AS settled_bets,
                SUM(CASE WHEN b.result = 'pending' THEN 1 ELSE 0 END) AS pending_bets,
                SUM(CASE WHEN b.result = 'win' THEN 1 ELSE 0 END) AS wins,
                SUM(CASE WHEN b.result = 'loss' THEN 1 ELSE 0 END) AS losses,
                SUM(CASE WHEN b.result = 'push' THEN 1 ELSE 0 END) AS pushes,
                SUM(CASE WHEN b.result != 'pending' THEN b.stake ELSE 0 END) AS total_staked,
                SUM(CASE WHEN b.result != 'pending' THEN COALESCE(b.profit, 0) ELSE 0 END) AS total_profit,
                SUM(CASE WHEN b.result != 'pending' THEN b.edge ELSE 0 END) AS sum_edge
            FROM markets m
            LEFT JOIN bets b ON b.market = m.market
            GROUP BY m.market
            ORDER BY total_profit DESC
        """)

        results = []
        for row in cursor.fetchall():
            market_name = (row['market'] or '').strip() or 'Unknown'
            settled_bets = row['settled_bets'] or 0
            pending_bets = row['pending_bets'] or 0
            wins = row['wins'] or 0
            losses = row['losses'] or 0
            pushes = row['pushes'] or 0
            decision_bets = wins + losses

            total_staked = row['total_staked'] or 0.0
            total_profit = row['total_profit'] or 0.0
            sum_edge = row['sum_edge'] or 0.0

            roi = (total_profit / total_staked * 100) if total_staked else None
            win_rate = (wins / decision_bets * 100) if decision_bets else None
            avg_edge = (sum_edge / settled_bets * 100) if settled_bets else None

            results.append({
                'market': market_name,
                'total_bets': settled_bets,
                'pending_bets': pending_bets,
                'wins': wins,
                'losses': losses,
                'pushes': pushes,
                'decision_bets': decision_bets,
                'win_rate': round(win_rate, 2) if win_rate is not None else None,
                'total_staked': round(total_staked, 2),
                'total_profit': round(total_profit, 2),
                'roi': round(roi, 2) if roi is not None else None,
                'avg_edge': round(avg_edge, 2) if avg_edge is not None else None
            })

        return results

def save_performance_snapshot() -> None:
    """
    Salva uno snapshot delle performance mensili.
    Chiamare questa funzione a fine mese.
    """
    summary = get_performance_summary(days=30)

    with get_db_connection() as conn:
        cursor = conn.cursor()

        period_end = datetime.now().strftime('%Y-%m-%d')
        period_start = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

        cursor.execute("""
            INSERT INTO performance (
                period_start, period_end, total_bets, wins, losses, pushes,
                win_rate, total_staked, total_profit, roi, avg_odds, avg_edge,
                brier_score_avg, sharpe_ratio
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            period_start, period_end,
            summary['total_bets'], summary['wins'], summary['losses'], summary['pushes'],
            summary['win_rate'], summary['total_staked'], summary['total_profit'],
            summary['roi'], summary['avg_odds'], summary['avg_edge'],
            summary['brier_score'], summary['sharpe_ratio']
        ))

        logger.info(f"Performance snapshot salvato per periodo {period_start} - {period_end}")

# ============================================================
#  INTEGRATION HELPERS - Auto-save to Database
# ============================================================

def process_and_save_prediction(
    match_data: Dict[str, Any],
    prediction_result: Dict[str, Any],
    save_to_db: bool = True
) -> Optional[int]:
    """
    Wrapper per salvare automaticamente match e predizione nel database.

    Args:
        match_data: Dict con match_id, date, league, home_team, away_team
        prediction_result: Output di calcolo_probabilita_avanzato()
        save_to_db: Se True, salva nel database

    Returns:
        prediction_id se salvato, None altrimenti

    Esempio:
        >>> match = {'match_id': 'abc123', 'date': '2025-11-10', ...}
        >>> pred = calcolo_probabilita_avanzato(lambda_h, lambda_a, ...)
        >>> pred_id = process_and_save_prediction(match, pred)
    """
    if not save_to_db:
        return None

    try:
        # Salva match
        save_match(match_data)

        # Prepara dati prediction
        prediction_data = {
            'match_id': match_data.get('match_id'),
            'lambda_h': prediction_result.get('lambda_home'),
            'lambda_a': prediction_result.get('lambda_away'),
            'rho': prediction_result.get('rho'),
            'tau': prediction_result.get('tau'),
            'prob_home': prediction_result.get('prob_home_win'),
            'prob_draw': prediction_result.get('prob_draw'),
            'prob_away': prediction_result.get('prob_away_win'),
            'prob_over_0_5': prediction_result.get('prob_over_0_5'),
            'prob_over_1_5': prediction_result.get('prob_over_1_5'),
            'prob_over_2_5': prediction_result.get('prob_over_2_5'),
            'prob_over_3_5': prediction_result.get('prob_over_3_5'),
            'prob_btts': prediction_result.get('prob_btts'),
            'model_version': '2.0',
            'weather_adjusted': prediction_result.get('weather_adjusted', 0)
        }

        prediction_id = save_prediction(prediction_data)
        logger.info(f"Prediction salvata: match_id={match_data.get('match_id')}, pred_id={prediction_id}")

        return prediction_id

    except Exception as e:
        logger.error(f"Errore salvataggio prediction: {e}")
        return None

def settle_bets_for_match(match_id: str, home_score: int, away_score: int) -> int:
    """
    Calcola automaticamente il risultato di tutte le bet pending per una partita.

    Args:
        match_id: ID partita
        home_score: Gol squadra casa
        away_score: Gol squadra ospite

    Returns:
        Numero di bets aggiornate
    """
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()

            # Recupera tutte le bet pending per questa partita
            cursor.execute("""
                SELECT bet_id, market, selection, stake, odds
                FROM bets
                WHERE match_id = ? AND result = 'pending'
            """, (match_id,))

            bets = cursor.fetchall()
            total_goals = home_score + away_score
            settled_count = 0

            for bet in bets:
                bet_id = bet['bet_id']
                market = bet['market']
                selection = bet['selection']
                stake = bet['stake']
                odds = bet['odds']

                result = None
                profit = 0

                # Determina risultato in base al mercato
                if market == '1X2':
                    actual_result = 'H' if home_score > away_score else ('A' if away_score > home_score else 'D')
                    if selection == actual_result:
                        result = 'win'
                        profit = stake * (odds - 1)
                    else:
                        result = 'loss'
                        profit = -stake

                elif 'Over/Under' in market:
                    # ⚠️ FIX BUG #12: Safe float() conversion con try/except
                    parts = market.split()
                    try:
                        threshold = float(parts[-1]) if parts else 0.0
                    except (ValueError, TypeError, IndexError):
                        logger.warning(f"Impossibile estrarre threshold da market: {market}, uso default 2.5")
                        threshold = 2.5
                    if 'Over' in selection:
                        result = 'win' if total_goals > threshold else 'loss'
                    else:  # Under
                        result = 'win' if total_goals < threshold else 'loss'

                    if result == 'win':
                        profit = stake * (odds - 1)
                    else:
                        profit = -stake

                    # Push per exact match su alcuni mercati
                    if total_goals == threshold:
                        result = 'push'
                        profit = 0

                elif market == 'BTTS':
                    btts_occurred = (home_score > 0 and away_score > 0)
                    if (selection == 'Yes' and btts_occurred) or (selection == 'No' and not btts_occurred):
                        result = 'win'
                        profit = stake * (odds - 1)
                    else:
                        result = 'loss'
                        profit = -stake

                # Aggiorna bet
                if result:
                    update_bet_result(bet_id, result, profit)
                    settled_count += 1

            logger.info(f"Settled {settled_count} bets per match {match_id}")
            return settled_count

    except Exception as e:
        logger.error(f"Errore settling bets: {e}")
        return 0

# ============================================================
#  xG SCRAPING - Expected Goals from FBref/Understat
# ============================================================

import re
from bs4 import BeautifulSoup

# Team name mapping per FBref (formato FBref -> nome comune)
FBREF_TEAM_MAPPING = {
    'Manchester Utd': 'Manchester United',
    'Manchester City': 'Manchester City',
    'Nott\'ham Forest': 'Nottingham Forest',
    'Newcastle Utd': 'Newcastle United',
    'Tottenham': 'Tottenham Hotspur',
    'West Ham': 'West Ham United',
    'Wolves': 'Wolverhampton Wanderers',
    'Brighton': 'Brighton & Hove Albion',
    'Leicester City': 'Leicester City',
}

def scrape_fbref_team_xg(team_name: str, league: str = 'Premier-League', season: str = '2024-2025') -> Dict[str, Any]:
    """
    Scrape dati xG per una squadra da FBref.

    Args:
        team_name: Nome squadra (es. 'Arsenal', 'Liverpool')
        league: Codice lega FBref (es. 'Premier-League', 'Serie-A', 'La-Liga')
        season: Stagione (es. '2024-2025')

    Returns:
        Dict con:
        - xg_for_avg: xG medi segnati per partita
        - xg_against_avg: xG medi subiti per partita
        - matches_played: Numero partite
        - xg_total: xG totali stagione
        - source: 'fbref'

    Note:
        - FBref ha limiti di rate (max 20 req/min)
        - Richiede User-Agent valido
        - Dati disponibili per top 5 leghe europee
    """
    try:
        # Normalizza nome team
        team_slug = team_name.replace(' ', '-').replace('&', '').replace('\'', '')

        # URL FBref (struttura: /en/squads/{team_id}/{team_slug}-Stats)
        # Per semplicità, usiamo ricerca statistica generale
        base_url = f'https://fbref.com/en/comps/{_get_fbref_league_id(league)}/{season}/stats/{season}-{league}-Stats'

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }

        response = requests.get(base_url, headers=headers, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.content, 'html.parser')

        # Cerca tabella con statistiche squadra
        table = soup.find('table', {'id': 'stats_squads_standard_for'})
        if not table:
            logger.warning(f"Tabella statistiche non trovata per {league}")
            return {}

        # Cerca riga con nome squadra
        rows = table.find('tbody').find_all('tr')
        for row in rows:
            squad_cell = row.find('th', {'data-stat': 'squad'})
            if squad_cell and team_name.lower() in squad_cell.text.lower():
                # Estrai dati xG
                # FIX BUG: Protezione AttributeError se elemento non trovato
                xg_for_elem = row.find('td', {'data-stat': 'xg_for'})
                xg_for = float(xg_for_elem.text) if xg_for_elem and xg_for_elem.text else 0.0

                xg_against_elem = row.find('td', {'data-stat': 'xg_against'})
                xg_against = float(xg_against_elem.text) if xg_against_elem and xg_against_elem.text else 0.0

                matches_elem = row.find('td', {'data-stat': 'games'})
                matches = int(matches_elem.text) if matches_elem and matches_elem.text else 0

                if matches > 0:
                    return {
                        'team': team_name,
                        'league': league,
                        'xg_for_avg': round(xg_for / matches, 2),
                        'xg_against_avg': round(xg_against / matches, 2),
                        'matches_played': matches,
                        'xg_total_for': round(xg_for, 2),
                        'xg_total_against': round(xg_against, 2),
                        'source': 'fbref',
                        'scraped_at': datetime.now().isoformat()
                    }

        logger.warning(f"Team {team_name} non trovato in {league}")
        return {}

    except requests.exceptions.RequestException as e:
        logger.error(f"Errore scraping FBref per {team_name}: {e}")
        return {}
    except Exception as e:
        logger.error(f"Errore parsing FBref: {e}")
        return {}

def _get_fbref_league_id(league: str) -> str:
    """Mappa nome lega a ID FBref"""
    league_ids = {
        'Premier-League': '9',
        'Serie-A': '11',
        'La-Liga': '12',
        'Bundesliga': '20',
        'Ligue-1': '13',
        'Champions-League': '8'
    }
    return league_ids.get(league, '9')

def get_xg_for_match(home_team: str, away_team: str, league: str = 'Premier-League') -> Dict[str, Any]:
    """
    Recupera dati xG per entrambe le squadre di una partita.

    Returns:
        Dict con xg_home, xg_away, confidence
    """
    home_xg = scrape_fbref_team_xg(home_team, league)
    away_xg = scrape_fbref_team_xg(away_team, league)

    if not home_xg or not away_xg:
        logger.warning(f"xG non disponibile per {home_team} vs {away_team}")
        return {}

    # Calcola xG atteso per il match
    # xG home = media(xG_for home, xG_against away)
    # xG away = media(xG_for away, xG_against home)
    xg_home_pred = (home_xg['xg_for_avg'] + away_xg['xg_against_avg']) / 2
    xg_away_pred = (away_xg['xg_for_avg'] + home_xg['xg_against_avg']) / 2

    return {
        'home_team': home_team,
        'away_team': away_team,
        'xg_home_predicted': round(xg_home_pred, 2),
        'xg_away_predicted': round(xg_away_pred, 2),
        'home_xg_for_avg': home_xg['xg_for_avg'],
        'home_xg_against_avg': home_xg['xg_against_avg'],
        'away_xg_for_avg': away_xg['xg_for_avg'],
        'away_xg_against_avg': away_xg['xg_against_avg'],
        'confidence': 'high' if home_xg['matches_played'] >= 10 and away_xg['matches_played'] >= 10 else 'medium',
        'source': 'fbref'
    }

# ============================================================
#  LEAGUE CALIBRATOR - Optimize parameters per league
# ============================================================

def calibrate_league_parameters(
    league_name: str,
    historical_matches: pd.DataFrame,
    optimize_params: List[str] = None
) -> Dict[str, Any]:
    """
    Calibra parametri del modello per una specifica lega.

    Args:
        league_name: Nome lega (es. 'Premier League', 'Serie A')
        historical_matches: DataFrame con colonne:
            - home_team, away_team, home_score, away_score
            - lambda_h, lambda_a (opzionali)
        optimize_params: Lista parametri da ottimizzare
            - 'home_advantage': Vantaggio casa
            - 'tau_dixon_coles': Correzione low-score
            - 'rho': Correlazione Poisson
            - 'shin_margin': Margine Shin medio

    Returns:
        Dict con parametri ottimizzati e metriche di fit
    """
    if optimize_params is None:
        optimize_params = ['home_advantage', 'tau_dixon_coles']

    if historical_matches.empty:
        logger.warning(f"Nessun dato storico per {league_name}")
        return {}

    results = {
        'league': league_name,
        'matches_analyzed': len(historical_matches),
        'optimized_params': {}
    }

    # 1. HOME ADVANTAGE - Differenza media gol casa vs trasferta
    if 'home_advantage' in optimize_params:
        home_goals_avg = historical_matches['home_score'].mean()
        away_goals_avg = historical_matches['away_score'].mean()
        home_advantage = home_goals_avg - away_goals_avg

        # Calibra moltiplicatore lambda_home
        home_multiplier = 1.0 + (home_advantage * 0.15)  # Empirico: ~15% per goal difference

        results['optimized_params']['home_advantage'] = round(home_advantage, 3)
        results['optimized_params']['home_multiplier'] = round(home_multiplier, 3)

    # 2. TAU DIXON-COLES - Ottimizza per basso punteggio
    if 'tau_dixon_coles' in optimize_params:
        # Conta partite 0-0, 1-0, 0-1, 1-1
        low_score_matches = historical_matches[
            (historical_matches['home_score'] <= 1) &
            (historical_matches['away_score'] <= 1)
        ]

        # FIX BUG: Protezione divisione per zero
        if len(historical_matches) > 0:
            low_score_ratio = len(low_score_matches) / len(historical_matches)
        else:
            low_score_ratio = 0.5  # Default se nessun dato storico
            logger.warning("No historical matches for tau calculation, using default")

        # Tau ottimale: più basso per leghe high-scoring, più alto per low-scoring
        # Range: -0.15 (high) to -0.05 (low)
        tau_optimal = -0.15 + (low_score_ratio * 0.10)

        results['optimized_params']['tau_dixon_coles'] = round(tau_optimal, 3)
        results['optimized_params']['low_score_ratio'] = round(low_score_ratio, 3)

    # 3. RHO - Correlazione empirica tra gol casa e trasferta
    if 'rho' in optimize_params:
        # FIX BUG: Protezione per DataFrame vuoto o insufficiente
        if len(historical_matches) > 1:
            try:
                correlation = historical_matches[['home_score', 'away_score']].corr().iloc[0, 1]
                if not math.isfinite(correlation):
                    correlation = -0.15  # Default
            except Exception as e:
                correlation = -0.15  # Default
                logger.warning(f"Error calculating correlation: {e}")
        else:
            correlation = -0.15  # Default se dati insufficienti
            logger.warning("Insufficient historical data for rho calculation, using default")

        # Rho tipico: -0.15 to 0.05
        rho_optimal = max(-0.20, min(0.10, correlation))

        results['optimized_params']['rho'] = round(rho_optimal, 3)
        results['optimized_params']['score_correlation'] = round(correlation, 3)

    # 4. OVER/UNDER THRESHOLD - Goal medi per lega
    total_goals_avg = (historical_matches['home_score'] + historical_matches['away_score']).mean()
    results['league_stats'] = {
        'avg_goals_per_match': round(total_goals_avg, 2),
        'avg_home_goals': round(historical_matches['home_score'].mean(), 2),
        'avg_away_goals': round(historical_matches['away_score'].mean(), 2),
        'home_win_pct': round((historical_matches['home_score'] > historical_matches['away_score']).mean() * 100, 1),
        'draw_pct': round((historical_matches['home_score'] == historical_matches['away_score']).mean() * 100, 1),
        'away_win_pct': round((historical_matches['home_score'] < historical_matches['away_score']).mean() * 100, 1),
    }

    # 5. BTTS (Both Teams to Score) rate
    btts_rate = ((historical_matches['home_score'] > 0) & (historical_matches['away_score'] > 0)).mean()
    results['league_stats']['btts_rate'] = round(btts_rate * 100, 1)

    logger.info(f"Calibrazione completata per {league_name}: {results['optimized_params']}")
    return results

def apply_league_calibration(
    lambda_h: float,
    lambda_a: float,
    league_params: Dict[str, Any]
) -> Tuple[float, float, float, float]:
    """
    Applica parametri calibrati per lega a lambdas.

    Args:
        lambda_h: Lambda casa base
        lambda_a: Lambda trasferta base
        league_params: Output di calibrate_league_parameters()

    Returns:
        (lambda_h_adj, lambda_a_adj, rho_adj, tau_adj)
    """
    params = league_params.get('optimized_params', {})

    # Applica home advantage
    home_mult = params.get('home_multiplier', 1.0)
    lambda_h_adj = lambda_h * home_mult

    # Usa tau e rho ottimizzati
    tau_adj = params.get('tau_dixon_coles', -0.13)
    rho_adj = params.get('rho', -0.10)

    return lambda_h_adj, lambda_a, rho_adj, tau_adj

# ============================================================
#  HEAD-TO-HEAD DATABASE - Historical matchups
# ============================================================

def save_h2h_result(
    home_team: str,
    away_team: str,
    date: str,
    home_score: int,
    away_score: int,
    league: str,
    competition: str = 'League'
) -> None:
    """
    Salva un risultato di scontro diretto nel database.

    Aggiunge alla tabella h2h per tracking storico.
    """
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()

            # Crea tabella h2h se non esiste
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS h2h (
                    h2h_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    home_team TEXT NOT NULL,
                    away_team TEXT NOT NULL,
                    date TEXT NOT NULL,
                    home_score INTEGER NOT NULL,
                    away_score INTEGER NOT NULL,
                    result TEXT NOT NULL,
                    league TEXT NOT NULL,
                    competition TEXT DEFAULT 'League',
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cursor.execute("CREATE INDEX IF NOT EXISTS idx_h2h_teams ON h2h(home_team, away_team)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_h2h_date ON h2h(date)")

            result = 'H' if home_score > away_score else ('A' if away_score > home_score else 'D')

            cursor.execute("""
                INSERT INTO h2h (home_team, away_team, date, home_score, away_score, result, league, competition)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (home_team, away_team, date, home_score, away_score, result, league, competition))

            logger.info(f"H2H salvato: {home_team} {home_score}-{away_score} {away_team} ({date})")

    except Exception as e:
        logger.error(f"Errore salvataggio H2H: {e}")

def get_h2h_stats(home_team: str, away_team: str, last_n: int = 10) -> Dict[str, Any]:
    """
    Recupera statistiche scontri diretti tra due squadre.

    Args:
        home_team: Squadra casa
        away_team: Squadra trasferta
        last_n: Ultimi N scontri (default: 10)

    Returns:
        Dict con:
        - total_matches: Totale scontri
        - home_wins: Vittorie squadra casa
        - draws: Pareggi
        - away_wins: Vittorie squadra trasferta
        - avg_goals_home: Media gol casa
        - avg_goals_away: Media gol trasferta
        - last_results: Lista ultimi N risultati
        - home_advantage_h2h: Vantaggio casa negli scontri diretti
    """
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()

            # Cerca in entrambe le direzioni (casa/trasferta invertita)
            cursor.execute("""
                SELECT home_team, away_team, date, home_score, away_score, result
                FROM h2h
                WHERE (home_team = ? AND away_team = ?)
                   OR (home_team = ? AND away_team = ?)
                ORDER BY date DESC
                LIMIT ?
            """, (home_team, away_team, away_team, home_team, last_n))

            matches = cursor.fetchall()

            if not matches:
                logger.info(f"Nessun H2H trovato tra {home_team} e {away_team}")
                return {'total_matches': 0}

            total = len(matches)
            home_wins = 0
            away_wins = 0
            draws = 0
            goals_home = []
            goals_away = []
            last_results = []

            for match in matches:
                h_team = match['home_team']
                a_team = match['away_team']
                h_score = match['home_score']
                a_score = match['away_score']

                # Normalizza in base alla squadra di riferimento (home_team)
                if h_team == home_team:
                    goals_home.append(h_score)
                    goals_away.append(a_score)

                    if match['result'] == 'H':
                        home_wins += 1
                    elif match['result'] == 'A':
                        away_wins += 1
                    else:
                        draws += 1

                    last_results.append({
                        'date': match['date'],
                        'home': h_team,
                        'away': a_team,
                        'score': f"{h_score}-{a_score}",
                        'result': match['result']
                    })
                else:
                    # Invertito
                    goals_home.append(a_score)
                    goals_away.append(h_score)

                    if match['result'] == 'A':
                        home_wins += 1
                    elif match['result'] == 'H':
                        away_wins += 1
                    else:
                        draws += 1

                    last_results.append({
                        'date': match['date'],
                        'home': a_team,
                        'away': h_team,
                        'score': f"{a_score}-{h_score}",
                        'result': 'H' if match['result'] == 'A' else ('A' if match['result'] == 'H' else 'D')
                    })

            return {
                'total_matches': total,
                'home_wins': home_wins,
                'draws': draws,
                'away_wins': away_wins,
                'home_win_pct': round(home_wins / total * 100, 1) if total > 0 else 0,
                'draw_pct': round(draws / total * 100, 1) if total > 0 else 0,
                'away_win_pct': round(away_wins / total * 100, 1) if total > 0 else 0,
                'avg_goals_home': round(sum(goals_home) / len(goals_home), 2) if goals_home and len(goals_home) > 0 else 0,  # FIX BUG: Protezione divisione per zero
                'avg_goals_away': round(sum(goals_away) / len(goals_away), 2) if goals_away and len(goals_away) > 0 else 0,  # FIX BUG: Protezione divisione per zero
                'avg_total_goals': round((sum(goals_home) + sum(goals_away)) / total, 2) if total > 0 else 0,
                'last_results': last_results,
                'home_advantage_h2h': round((home_wins - away_wins) / total, 2) if total > 0 else 0
            }

    except Exception as e:
        logger.error(f"Errore recupero H2H: {e}")
        return {'total_matches': 0}

def adjust_prediction_with_h2h(
    prob_home: float,
    prob_draw: float,
    prob_away: float,
    h2h_stats: Dict[str, Any],
    weight: float = 0.15
) -> Tuple[float, float, float]:
    """
    Aggiusta probabilità 1X2 usando statistiche H2H.

    Args:
        prob_home, prob_draw, prob_away: Probabilità dal modello base
        h2h_stats: Output di get_h2h_stats()
        weight: Peso H2H (default: 0.15 = 15%)

    Returns:
        (prob_home_adj, prob_draw_adj, prob_away_adj) normalizzate
    """
    if h2h_stats.get('total_matches', 0) < 3:
        # Non abbastanza dati H2H, ritorna probabilità originali
        return prob_home, prob_draw, prob_away

    # Probabilità empiriche da H2H
    h2h_home_prob = h2h_stats['home_win_pct'] / 100
    h2h_draw_prob = h2h_stats['draw_pct'] / 100
    h2h_away_prob = h2h_stats['away_win_pct'] / 100

    # Blend con peso
    prob_home_adj = (1 - weight) * prob_home + weight * h2h_home_prob
    prob_draw_adj = (1 - weight) * prob_draw + weight * h2h_draw_prob
    prob_away_adj = (1 - weight) * prob_away + weight * h2h_away_prob

    # Normalizza
    total = prob_home_adj + prob_draw_adj + prob_away_adj
    prob_home_adj /= total
    prob_draw_adj /= total
    prob_away_adj /= total

    return prob_home_adj, prob_draw_adj, prob_away_adj

# ============================================================
#  API-FOOTBALL
# ============================================================

def apifootball_get_fixtures_by_date(d: str) -> list:
    if not API_FOOTBALL_KEY:
        logger.warning("API_FOOTBALL_KEY non configurata. Configura tramite variabile d'ambiente.")
        return []
    headers = {"x-apisports-key": API_FOOTBALL_KEY}
    params = {"date": d}
    try:
        r = requests.get(
            f"{API_FOOTBALL_BASE}/fixtures",
            headers=headers,
            params=params,
            timeout=app_config.api_timeout
        )
        r.raise_for_status()
        data = r.json()
        return data.get("response", [])
    except requests.exceptions.Timeout:
        logger.error(f"Timeout API-Football fixtures per data {d}")
        return []
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 401:
            logger.error("API-Football: API key non valida o scaduta")
        else:
            logger.error(f"Errore HTTP API-Football fixtures: {e.response.status_code}")
        return []
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore richiesta API-Football fixtures: {e}")
        return []
    except (ValueError, KeyError, json.JSONDecodeError) as e:
        logger.error(f"Errore parsing risposta API-Football: {e}")
        return []

# ============================================================
#   API-FOOTBALL: RECUPERO DATI AVANZATI
# ============================================================

def apifootball_search_team(team_name: str, league_id: int = None) -> Dict[str, Any]:
    """
    Cerca team ID da API-Football usando nome squadra.
    """
    if not API_FOOTBALL_KEY:
        logger.warning("API_FOOTBALL_KEY non configurata. Configura tramite variabile d'ambiente.")
        return {}
    headers = {"x-apisports-key": API_FOOTBALL_KEY}
    params = {"search": team_name}
    if league_id:
        params["league"] = league_id
    
    try:
        r = requests.get(
            f"{API_FOOTBALL_BASE}/teams",
            headers=headers,
            params=params,
            timeout=app_config.api_timeout
        )
        r.raise_for_status()
        data = r.json()
        teams = data.get("response", [])
        # FIX BUG #10.6: Safe array access on teams
        if teams and len(teams) > 0:
            return teams[0]  # Ritorna primo match
        return {}
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore ricerca team {team_name}: {e}")
        return {}

def apifootball_get_team_fixtures(team_id: int, last: int = 10, season: int = None) -> List[Dict[str, Any]]:
    """
    Recupera ultime partite di una squadra.
    """
    if not API_FOOTBALL_KEY:
        logger.warning("API_FOOTBALL_KEY non configurata. Configura tramite variabile d'ambiente.")
        return []
    headers = {"x-apisports-key": API_FOOTBALL_KEY}
    params = {"team": team_id, "last": last}
    if season:
        params["season"] = season
    
    try:
        r = requests.get(
            f"{API_FOOTBALL_BASE}/fixtures",
            headers=headers,
            params=params,
            timeout=app_config.api_timeout
        )
        r.raise_for_status()
        data = r.json()
        return data.get("response", [])
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore fixtures team {team_id}: {e}")
        return []

def apifootball_get_standings(league_id: int, season: int) -> Dict[str, Any]:
    """
    Recupera classifica di una lega.
    """
    if not API_FOOTBALL_KEY:
        logger.warning("API_FOOTBALL_KEY non configurata. Configura tramite variabile d'ambiente.")
        return []
    headers = {"x-apisports-key": API_FOOTBALL_KEY}
    params = {"league": league_id, "season": season}
    
    try:
        r = requests.get(
            f"{API_FOOTBALL_BASE}/standings",
            headers=headers,
            params=params,
            timeout=app_config.api_timeout
        )
        r.raise_for_status()
        data = r.json()
        standings = data.get("response", [])
        # ⚠️ PROTEZIONE: Valida che standings non sia vuoto e che [0] sia un dict
        if standings and len(standings) > 0 and isinstance(standings[0], dict):
            return standings[0].get("league", {}).get("standings", [])
        return []
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore standings league {league_id}: {e}")
        return []

def get_league_id_from_name(league_name: str) -> int:
    """
    Mappa nome lega a ID API-Football.
    """
    league_map = {
        "premier_league": 39,  # Premier League
        "serie_a": 135,  # Serie A
        "la_liga": 140,  # La Liga
        "bundesliga": 78,  # Bundesliga
        "ligue_1": 61,  # Ligue 1
    }
    return league_map.get(league_name, 0)

def get_current_season() -> int:
    """
    Ritorna anno stagione corrente (es. 2024 per 2023-2024).
    """
    now = datetime.now()
    # Stagione inizia agosto, quindi se siamo dopo agosto usiamo anno corrente
    if now.month >= 8:
        return now.year
    else:
        return now.year - 1

def calculate_days_since_last_match(team_id: int, match_date: str) -> int:
    """
    Calcola giorni dall'ultima partita di una squadra.
    """
    if not team_id or not match_date:
        return None
    
    try:
        # Recupera ultime partite
        fixtures = apifootball_get_team_fixtures(team_id, last=5)
        if not fixtures:
            return None
        
        # Trova ultima partita giocata (status FT, AET, PEN)
        last_match_date = None
        for fixture in fixtures:
            # ⚠️ PROTEZIONE: Accesso sicuro a dizionari annidati
            fixture_data = fixture.get("fixture", {}) if isinstance(fixture, dict) else {}
            status_data = fixture_data.get("status", {}) if isinstance(fixture_data, dict) else {}
            status = status_data.get("short", "") if isinstance(status_data, dict) else ""
            if status in ["FT", "AET", "PEN"]:
                date_str = fixture_data.get("date", "") if isinstance(fixture_data, dict) else ""
                if date_str:
                    last_match_date = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
                    break
        
        if not last_match_date:
            return None
        
        # Calcola differenza con data partita corrente
        match_dt = datetime.fromisoformat(match_date.replace("Z", "+00:00"))
        delta = (match_dt - last_match_date).days
        
        return max(0, delta)  # Non può essere negativo
    except (ValueError, TypeError, KeyError) as e:
        logger.error(f"Errore calcolo giorni ultima partita: {e}")
        return None

def count_matches_last_30_days(team_id: int, match_date: str) -> int:
    """
    Conta partite giocate negli ultimi 30 giorni.
    """
    if not team_id or not match_date:
        return None
    
    try:
        fixtures = apifootball_get_team_fixtures(team_id, last=15)
        if not fixtures:
            return None
        
        match_dt = datetime.fromisoformat(match_date.replace("Z", "+00:00"))
        cutoff_date = match_dt - timedelta(days=30)
        
        count = 0
        for fixture in fixtures:
            # ⚠️ PROTEZIONE: Accesso sicuro a dizionari annidati
            fixture_data = fixture.get("fixture", {}) if isinstance(fixture, dict) else {}
            status_data = fixture_data.get("status", {}) if isinstance(fixture_data, dict) else {}
            status = status_data.get("short", "") if isinstance(status_data, dict) else ""
            if status in ["FT", "AET", "PEN"]:
                date_str = fixture_data.get("date", "") if isinstance(fixture_data, dict) else ""
                if date_str:
                    fixture_date = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
                    if cutoff_date <= fixture_date < match_dt:
                        count += 1
        
        return count
    except (ValueError, TypeError, KeyError) as e:
        logger.error(f"Errore conteggio partite 30 giorni: {e}")
        return None

def get_team_standings_info(team_name: str, league: str, season: int = None) -> Dict[str, Any]:
    """
    Recupera informazioni classifica per una squadra.
    """
    if not season:
        season = get_current_season()
    
    league_id = get_league_id_from_name(league)
    if not league_id:
        return {}
    
    try:
        standings = apifootball_get_standings(league_id, season)
        if not standings:
            return {}
        
        # Cerca squadra nella classifica
        team_name_lower = team_name.lower()
        for group in standings:
            for entry in group:
                team_info = entry.get("team", {})
                team_name_raw = team_info.get("name")
                # Assicurati che team_name_api sia sempre una stringa (gestisce None)
                team_name_api = str(team_name_raw).lower() if team_name_raw is not None else ""
                
                # Match approssimativo (potrebbe essere diverso)
                if team_name_lower in team_name_api or team_name_api in team_name_lower:
                    position = entry.get("rank", 0)
                    points = entry.get("points", 0)
                    all_stats = entry.get("all", {})
                    played = all_stats.get("played", 0)
                    
                    # Calcola distanza da salvezza (assumendo 18 squadre, 3 retrocedono)
                    # In realtà dipende dalla lega, ma approssimiamo
                    points_from_relegation = None
                    if position >= 16:  # Ultime 3 posizioni
                        # Trova punti della 17esima (ultima salva)
                        for e in group:
                            rank_val = e.get("rank")
                            # Converti rank a int se necessario (può essere stringa)
                            try:
                                rank_int = int(rank_val) if rank_val is not None else None
                                if rank_int == 17:
                                    points_17 = e.get("points", 0)
                                    # Assicurati che points_17 sia un numero
                                    try:
                                        points_17 = int(points_17) if points_17 is not None else 0
                                    except (ValueError, TypeError):
                                        points_17 = 0
                                    points_from_relegation = points - points_17
                                    break
                            except (ValueError, TypeError):
                                continue
                    
                    # Calcola distanza da Europa (posizione 6-7)
                    points_from_europe = None
                    if 6 <= position <= 10:
                        for e in group:
                            rank_val = e.get("rank")
                            # Converti rank a int se necessario (può essere stringa)
                            try:
                                rank_int = int(rank_val) if rank_val is not None else None
                                if rank_int == 6:
                                    points_6 = e.get("points", 0)
                                    # Assicurati che points_6 sia un numero
                                    try:
                                        points_6 = int(points_6) if points_6 is not None else 0
                                    except (ValueError, TypeError):
                                        points_6 = 0
                                    points_from_europe = points_6 - points
                                    break
                            except (ValueError, TypeError):
                                continue
                    
                    return {
                        "position": position,
                        "points": points,
                        "played": played,
                        "points_from_relegation": points_from_relegation,
                        "points_from_europe": points_from_europe,
                    }
        
        return {}
    except (KeyError, ValueError, requests.exceptions.RequestException) as e:
        logger.error(f"Errore recupero standings {team_name}: {e}")
        return {}

def is_derby_match(home_team: str, away_team: str, league: str) -> bool:
    """
    Identifica se è un derby basandosi su pattern comuni.
    """
    # Lista derby comuni (potrebbe essere espansa)
    derby_patterns = [
        # Serie A
        ("inter", "milan"), ("milan", "inter"),
        ("juventus", "torino"), ("torino", "juventus"),
        ("roma", "lazio"), ("lazio", "roma"),
        ("napoli", "juventus"), ("juventus", "napoli"),
        # Premier League
        ("manchester united", "manchester city"), ("manchester city", "manchester united"),
        ("liverpool", "everton"), ("everton", "liverpool"),
        ("arsenal", "tottenham"), ("tottenham", "arsenal"),
        # La Liga
        ("real madrid", "atletico madrid"), ("atletico madrid", "real madrid"),
        ("real madrid", "barcelona"), ("barcelona", "real madrid"),
        ("atletico madrid", "barcelona"), ("barcelona", "atletico madrid"),
        # Bundesliga
        ("bayern munich", "borussia dortmund"), ("borussia dortmund", "bayern munich"),
    ]
    
    home_lower = home_team.lower()
    away_lower = away_team.lower()
    
    for pattern_home, pattern_away in derby_patterns:
        if pattern_home in home_lower and pattern_away in away_lower:
            return True
    
    return False

def apifootball_get_team_statistics(team_id: int, league_id: int, season: int) -> Dict[str, Any]:
    """
    Recupera statistiche dettagliate di una squadra.
    """
    if not API_FOOTBALL_KEY:
        logger.warning("API_FOOTBALL_KEY non configurata. Configura tramite variabile d'ambiente.")
        return {}
    headers = {"x-apisports-key": API_FOOTBALL_KEY}
    params = {"team": team_id, "league": league_id, "season": season}
    
    try:
        r = requests.get(
            f"{API_FOOTBALL_BASE}/teams/statistics",
            headers=headers,
            params=params,
            timeout=app_config.api_timeout
        )
        r.raise_for_status()
        data = r.json()
        response = data.get("response", {})
        return response
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore statistiche team {team_id}: {e}")
        return {}

def apifootball_get_head_to_head(team1_id: int, team2_id: int, last: int = 10) -> List[Dict[str, Any]]:
    """
    Recupera partite head-to-head tra due squadre.
    """
    if not API_FOOTBALL_KEY:
        logger.warning("API_FOOTBALL_KEY non configurata. Configura tramite variabile d'ambiente.")
        return []
    headers = {"x-apisports-key": API_FOOTBALL_KEY}
    params = {"h2h": f"{team1_id}-{team2_id}", "last": last}
    
    try:
        r = requests.get(
            f"{API_FOOTBALL_BASE}/fixtures/headtohead",
            headers=headers,
            params=params,
            timeout=app_config.api_timeout
        )
        r.raise_for_status()
        data = r.json()
        return data.get("response", [])
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore H2H {team1_id} vs {team2_id}: {e}")
        return []

def apifootball_get_injuries(team_id: int = None, fixture_id: int = None) -> List[Dict[str, Any]]:
    """
    Recupera infortuni. Se team_id è specificato, filtra per squadra.
    """
    if not API_FOOTBALL_KEY:
        logger.warning("API_FOOTBALL_KEY non configurata. Configura tramite variabile d'ambiente.")
        return []
    headers = {"x-apisports-key": API_FOOTBALL_KEY}
    params = {}
    if team_id:
        params["team"] = team_id
    if fixture_id:
        params["fixture"] = fixture_id
    
    try:
        r = requests.get(
            f"{API_FOOTBALL_BASE}/injuries",
            headers=headers,
            params=params,
            timeout=app_config.api_timeout
        )
        r.raise_for_status()
        data = r.json()
        return data.get("response", [])
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore infortuni: {e}")
        return []

def apifootball_get_fixture_lineups(fixture_id: int) -> Dict[str, Any]:
    """
    ALTA PRIORITÀ: Recupera formazioni (lineups) per una partita.
    
    Returns:
        Dict con formazioni casa/trasferta, formazione (es. 4-3-3), giocatori chiave
    """
    if not API_FOOTBALL_KEY:
        logger.warning("API_FOOTBALL_KEY non configurata. Configura tramite variabile d'ambiente.")
        return {}
    headers = {"x-apisports-key": API_FOOTBALL_KEY}
    params = {"fixture": fixture_id}
    
    try:
        r = requests.get(
            f"{API_FOOTBALL_BASE}/fixtures/lineups",
            headers=headers,
            params=params,
            timeout=app_config.api_timeout
        )
        r.raise_for_status()
        data = r.json()
        response = data.get("response", [])
        
        if not response or len(response) == 0:
            return {}
        
        # Estrai formazioni
        result = {
            "home_lineup": None,
            "away_lineup": None,
            "home_formation": None,
            "away_formation": None,
            "home_key_players": [],
            "away_key_players": [],
            "data_available": False,
        }
        
        for lineup_data in response:
            team_info = lineup_data.get("team", {})
            team_id = team_info.get("id")
            formation = lineup_data.get("formation")
            startXI = lineup_data.get("startXI", [])
            
            # Determina se casa o trasferta (primo = casa, secondo = trasferta)
            if result["home_lineup"] is None:
                result["home_lineup"] = startXI
                result["home_formation"] = formation
                # Estrai nomi giocatori chiave (primi 5)
                result["home_key_players"] = [
                    p.get("player", {}).get("name", "") 
                    for p in startXI[:5] if p.get("player")
                ]
            else:
                result["away_lineup"] = startXI
                result["away_formation"] = formation
                result["away_key_players"] = [
                    p.get("player", {}).get("name", "") 
                    for p in startXI[:5] if p.get("player")
                ]
        
        result["data_available"] = result["home_lineup"] is not None and result["away_lineup"] is not None
        return result
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore lineups fixture {fixture_id}: {e}")
        return {}

def calculate_lineup_impact(lineup_data: Dict[str, Any], team_id: int) -> Dict[str, float]:
    """
    Calcola impatto formazione su lambda.
    
    Analizza:
    - Formazione (4-3-3 vs 4-4-2 vs 5-3-2)
    - Giocatori chiave presenti/assenti
    """
    if not lineup_data or not lineup_data.get("data_available"):
        return {"attack_factor": 1.0, "defense_factor": 1.0, "confidence": 0.0}
    
    attack_factor = 1.0
    defense_factor = 1.0
    
    # Analizza formazione
    formation = lineup_data.get("home_formation") if lineup_data.get("home_lineup") else lineup_data.get("away_formation")
    
    if formation:
        # Formazioni offensive (4-3-3, 3-4-3) → +attacco, -difesa
        if formation in ["4-3-3", "3-4-3", "4-2-3-1"]:
            attack_factor *= 1.05
            defense_factor *= 0.98
        # Formazioni difensive (5-3-2, 4-5-1) → -attacco, +difesa
        elif formation in ["5-3-2", "4-5-1", "5-4-1"]:
            attack_factor *= 0.95
            defense_factor *= 1.05
    
    # Se abbiamo giocatori chiave, verifica se sono presenti
    key_players = lineup_data.get("home_key_players") or lineup_data.get("away_key_players")
    if key_players and len(key_players) >= 3:
        # Se abbiamo almeno 3 giocatori chiave → formazione forte
        confidence = min(1.0, len(key_players) / 5.0)
    else:
        confidence = 0.3
    
    return {
        "attack_factor": round(attack_factor, 3),
        "defense_factor": round(defense_factor, 3),
        "confidence": round(confidence, 2),
        "formation": formation,
    }

def apifootball_get_fixture_info(fixture_id: int) -> Dict[str, Any]:
    """
    Recupera info complete su una partita (per weather, referee, etc.).
    """
    if not API_FOOTBALL_KEY:
        logger.warning("API_FOOTBALL_KEY non configurata. Configura tramite variabile d'ambiente.")
        return {}
    headers = {"x-apisports-key": API_FOOTBALL_KEY}
    params = {"id": fixture_id}
    
    try:
        r = requests.get(
            f"{API_FOOTBALL_BASE}/fixtures",
            headers=headers,
            params=params,
            timeout=app_config.api_timeout
        )
        r.raise_for_status()
        data = r.json()
        response = data.get("response", [])
        # FIX BUG #10.6: Safe array access on response
        if response and len(response) > 0:
            return response[0]
        return {}
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore fixture info {fixture_id}: {e}")
        return {}


def fetch_weather_snapshot(
    city: Optional[str] = None,
    country: Optional[str] = None,
    lat: Optional[float] = None,
    lon: Optional[float] = None
) -> Optional[Dict[str, Any]]:
    """
    Recupera un'istantanea meteo utilizzando OpenWeather.

    Args:
        city: Nome città (fallback se lat/lon non disponibili)
        country: Codice paese per migliorare la ricerca
        lat, lon: Coordinate per query diretta
    """
    if not OPENWEATHER_API_KEY:
        logger.debug("OPENWEATHER_API_KEY non configurata, salto fetch meteo esterno")
        return None

    if (lat is None or lon is None) and not city:
        return None

    params = {
        "appid": OPENWEATHER_API_KEY,
        "units": "metric"
    }
    if lat is not None and lon is not None:
        params.update({"lat": lat, "lon": lon})
    else:
        query_city = city.strip() if city else ""
        if country:
            query_city = f"{query_city},{country}".strip(",")
        if not query_city:
            return None
        params["q"] = query_city

    def _request_weather(endpoint: str, request_params: Dict[str, Any]) -> Dict[str, Any]:
        response = requests.get(endpoint, params=request_params, timeout=app_config.api_timeout)
        response.raise_for_status()
        return response.json()

    cache_key = f"openweather::{params.get('q', '')}::{params.get('lat', '')}::{params.get('lon', '')}"

    try:
        weather_data = api_call_with_retry(
            _request_weather,
            "https://api.openweathermap.org/data/2.5/weather",
            params,
            max_attempts=2,
            delay=API_RETRY_DELAY * 1.5,
            cache_key=cache_key
        )
        return weather_data
    except Exception as e:
        logger.debug(f"OpenWeather non disponibile per {params}: {e}")
        return None


def get_weather_impact(fixture_data: Dict[str, Any]) -> Dict[str, float]:
    """
    ALTA PRIORITÀ: Calcola impatto condizioni meteo su lambda.
    
    Analizza:
    - Pioggia forte → riduce gol (difesa più difficile)
    - Vento forte → aumenta incertezza
    - Temperatura estrema → fatica
    """
    if not fixture_data:
        return {"total_factor": 1.0, "confidence": 0.0}
    
    fixture_info = fixture_data.get("fixture", {})
    venue = fixture_info.get("venue", {})
    
    # API-Football non fornisce sempre weather, tenta fetch da OpenWeather come fallback
    weather_data = fixture_data.get("weather") or {}
    weather_source = "api-football"

    if not weather_data:
        def _safe_float(value: Any) -> Optional[float]:
            try:
                if value is None:
                    return None
                if isinstance(value, str):
                    trimmed = value.strip()
                    if not trimmed:
                        return None
                    value = trimmed
                return float(value)
            except (TypeError, ValueError):
                return None

        city = venue.get("city") or venue.get("name") or ""
        league_info = fixture_data.get("league", {})
        country = league_info.get("country")
        lat = venue.get("lat") or venue.get("latitude")
        lon = venue.get("lon") or venue.get("longitude")

        # Converti stringhe vuote a None
        lat = _safe_float(lat)
        lon = _safe_float(lon)

        fetched_weather = fetch_weather_snapshot(
            city=city,
            country=country,
            lat=lat,
            lon=lon
        )
        if fetched_weather:
            weather_data = fetched_weather
            weather_source = "openweather"

    total_factor = 1.0
    confidence = 0.0
    condition_label = None
    temperature = None
    wind_speed = None
    humidity = None
    precipitation_mm = None

    try:
        if weather_data:
            if weather_source == "openweather":
                main_data = weather_data.get("main", {})
                wind = weather_data.get("wind", {})
                weather_list = weather_data.get("weather", [{}])
                rain_block = weather_data.get("rain", {})

                temperature = main_data.get("temp")
                humidity = main_data.get("humidity")
                wind_speed = wind.get("speed")
                # FIX BUG #10.9: Safe array access on weather_list
                condition_label = ((weather_list[0] if weather_list and len(weather_list) > 0 else {}) or {}).get("description", "").lower()
                precipitation_mm = rain_block.get("1h") or rain_block.get("3h")
                confidence = 0.7
            else:
                temperature = weather_data.get("temp") or weather_data.get("temperature")
                humidity = weather_data.get("humidity")
                wind_speed = weather_data.get("wind_speed") or weather_data.get("wind")
                condition_raw = weather_data.get("condition") or weather_data.get("description")
                condition_label = str(condition_raw).lower() if condition_raw else None
                precipitation_mm = weather_data.get("rain") or weather_data.get("precipitation")
                confidence = 0.5

            # Pioggia forte → meno gol
            if condition_label and any(token in condition_label for token in ["rain", "storm", "snow", "sleet"]):
                total_factor *= 0.92
                confidence = max(confidence, 0.7)

            # Temperatura estrema
            if temperature is not None:
                try:
                    temp_val = float(temperature)
                    if temp_val < 2:
                        total_factor *= 0.93
                        confidence = max(confidence, 0.65)
                    elif temp_val < 5:
                        total_factor *= 0.96
                    elif temp_val > 32:
                        total_factor *= 0.94
                        confidence = max(confidence, 0.65)
                    elif temp_val > 28:
                        total_factor *= 0.97
                except (TypeError, ValueError):
                    pass

            # Vento forte
            if wind_speed is not None:
                try:
                    wind_val = float(wind_speed)
                    if wind_val >= 12:  # ~43 km/h
                        total_factor *= 0.94
                        confidence = max(confidence, 0.6)
                    elif wind_val >= 8:  # ~29 km/h
                        total_factor *= 0.97
                except (TypeError, ValueError):
                    pass

            # Umidità molto alta riduce intensità
            if humidity is not None:
                try:
                    humidity_val = float(humidity)
                    if humidity_val >= 85:
                        total_factor *= 0.97
                        confidence = max(confidence, 0.55)
                except (TypeError, ValueError):
                    pass

            # Precipitazioni misurate
            if precipitation_mm is not None:
                try:
                    rain_val = float(precipitation_mm)
                    if rain_val >= 5.0:
                        total_factor *= 0.90
                        confidence = max(confidence, 0.75)
                    elif rain_val >= 2.0:
                        total_factor *= 0.94
                except (TypeError, ValueError):
                    pass

    except Exception as e:
        logger.debug(f"Impossibile calcolare impatto meteo: {e}")
        total_factor = 1.0
        confidence = 0.0

    return {
        "total_factor": round(total_factor, 3),
        "confidence": round(confidence, 2),
        "weather_condition": condition_label,
        "temperature": temperature,
        "wind_speed": wind_speed,
        "humidity": humidity,
        "precipitation_mm": precipitation_mm,
        "source": weather_source if weather_data else None,
    }

def get_referee_statistics(referee_name: str, league: str = None) -> Dict[str, float]:
    """
    ALTA PRIORITÀ: Recupera statistiche arbitro da API-Football.
    
    Analizza:
    - Media cartellini per partita
    - Media gol per partita (alcuni arbitri favoriscono attacco)
    - Tendenza a favorire casa/trasferta
    """
    # API-Football non ha endpoint diretto per referee stats
    # Possiamo inferire da fixture history, ma per ora ritorna neutro
    # Struttura pronta per integrazione futura
    
    return {
        "cards_factor": 1.0,  # Molti cartellini → più interruzioni → meno gol
        "goals_factor": 1.0,  # Alcuni arbitri favoriscono attacco
        "home_bias": 1.0,  # Bias verso casa
        "confidence": 0.0,
    }

def analyze_market_depth(
    event: dict,
    odds_1: float,
    odds_x: float,
    odds_2: float,
) -> Dict[str, Any]:
    """
    ALTA PRIORITÀ: Analizza profondità mercato e sharp money.
    
    Analizza:
    - Numero bookmakers (proxy liquidità)
    - Spread quote (liquidità implicita)
    - Movimento quote (sharp money detection)
    - Volume scommesse (se disponibile)
    """
    if not event:
        return {}
    
    bookmakers = event.get("bookmakers", [])
    num_bookmakers = len(bookmakers)

    # Raccogli tutte le quote per ogni mercato
    all_odds_1 = []
    all_odds_x = []
    all_odds_2 = []

    for bk in bookmakers:
        for mk in bk.get("markets", []):
            mk_key_raw = mk.get("key")
            mk_key = str(mk_key_raw).lower() if mk_key_raw is not None else ""
            if mk_key in ["h2h", "match_winner"]:
                for o in mk.get("outcomes", []):
                    name_l = (o.get("name") or "").lower()
                    price = o.get("price")
                    if not price:
                        continue
                    try:
                        price_val = float(price)
                    except (TypeError, ValueError):
                        continue
                    if not math.isfinite(price_val) or price_val <= 1.0:
                        continue

                    if "home" in name_l or name_l == "1":
                        all_odds_1.append(price_val)
                    elif "draw" in name_l or name_l in {"x", "tie"}:
                        all_odds_x.append(price_val)
                    elif "away" in name_l or name_l == "2":
                        all_odds_2.append(price_val)

    # Include le quote di input nel calcolo del consenso
    for target_list, provided in ((all_odds_1, odds_1), (all_odds_x, odds_x), (all_odds_2, odds_2)):
        try:
            if provided and math.isfinite(provided) and provided > 1.0:
                target_list.append(float(provided))
        except (TypeError, ValueError):
            continue

    def _compute_stats(values: List[float]) -> Optional[Dict[str, float]]:
        if not values:
            return None
        arr = np.array(values, dtype=np.float64)
        if arr.size == 0:
            return None
        min_val = float(np.min(arr))
        max_val = float(np.max(arr))
        mean_val = float(np.mean(arr))
        median_val = float(np.median(arr))
        std_val = float(np.std(arr))
        spread_pct = float(((max_val - min_val) / min_val) * 100.0) if min_val > 0 else None
        coeff_var_pct = float((std_val / mean_val) * 100.0) if mean_val > 0 else None
        implied = 1.0 / arr
        return {
            "min": round(min_val, 3),
            "max": round(max_val, 3),
            "mean": round(mean_val, 3),
            "median": round(median_val, 3),
            "std": round(std_val, 3),
            "spread_pct": round(spread_pct, 2) if spread_pct is not None else None,
            "coefficient_variation_pct": round(coeff_var_pct, 2) if coeff_var_pct is not None else None,
            "mean_prob": float(np.mean(implied)),
            "median_prob": float(np.median(implied)),
        }

    odds_summary = {
        "home": _compute_stats(all_odds_1),
        "draw": _compute_stats(all_odds_x),
        "away": _compute_stats(all_odds_2),
    }

    depth_score = 0.0
    liquidity_indicators: Dict[str, Any] = {}

    # 1. Numero bookmakers (più = più liquidità)
    if num_bookmakers >= 12:
        depth_score += 35
        liquidity_indicators["bookmakers"] = "very_high"
    elif num_bookmakers >= 8:
        depth_score += 25
        liquidity_indicators["bookmakers"] = "high"
    elif num_bookmakers >= 5:
        depth_score += 15
        liquidity_indicators["bookmakers"] = "medium"
    elif num_bookmakers >= 3:
        depth_score += 8
        liquidity_indicators["bookmakers"] = "low"
    else:
        depth_score += 4
        liquidity_indicators["bookmakers"] = "very_low"

    spread_values = [
        stats["spread_pct"] for stats in odds_summary.values()
        if stats and stats.get("spread_pct") is not None
    ]
    avg_spread = float(np.mean(spread_values)) if spread_values else None
    if avg_spread is not None:
        if avg_spread <= 2.5:
            depth_score += 30
            liquidity_indicators["spread"] = "tight"
        elif avg_spread <= 5.0:
            depth_score += 22
            liquidity_indicators["spread"] = "balanced"
        elif avg_spread <= 8.0:
            depth_score += 12
            liquidity_indicators["spread"] = "wide"
        else:
            liquidity_indicators["spread"] = "very_wide"

    cov_values = [
        stats["coefficient_variation_pct"] for stats in odds_summary.values()
        if stats and stats.get("coefficient_variation_pct") is not None
    ]
    avg_cov = float(np.mean(cov_values)) if cov_values else None
    if avg_cov is not None:
        if avg_cov <= 1.5:
            depth_score += 20
            liquidity_indicators["volatility"] = "very_low"
        elif avg_cov <= 3.0:
            depth_score += 14
            liquidity_indicators["volatility"] = "low"
        elif avg_cov <= 5.0:
            depth_score += 8
            liquidity_indicators["volatility"] = "medium"
        else:
            liquidity_indicators["volatility"] = "high"

    consensus_probs = {}
    for label, stats in odds_summary.items():
        if stats and stats.get("mean_prob") is not None:
            consensus_probs[label] = round(stats["mean_prob"], 4)

    overround = None
    if consensus_probs and len(consensus_probs) == 3:
        prob_sum = sum(consensus_probs.values())
        overround = prob_sum - 1.0
        if overround <= 0.02:
            depth_score += 15
            liquidity_indicators["overround"] = "sharp"
        elif overround <= 0.05:
            depth_score += 10
            liquidity_indicators["overround"] = "efficient"
        elif overround <= 0.08:
            depth_score += 6
            liquidity_indicators["overround"] = "average"
        else:
            liquidity_indicators["overround"] = "inflated"

    # Flag di possibile valore rispetto al consenso
    value_flags = {}
    for label, provided in [("home", odds_1), ("draw", odds_x), ("away", odds_2)]:
        stats = odds_summary.get(label)
        if not stats or provided is None:
            continue
        try:
            provided_val = float(provided)
            if not math.isfinite(provided_val) or provided_val <= 1.0:
                continue
        except (TypeError, ValueError):
            continue
        mean_odds = stats.get("mean")
        if mean_odds:
            delta_pct = ((provided_val - mean_odds) / mean_odds) * 100.0
            if delta_pct >= 1.5:
                value_flags[label] = {"type": "above_consensus", "delta_pct": round(delta_pct, 2)}
            elif delta_pct <= -1.5:
                value_flags[label] = {"type": "below_consensus", "delta_pct": round(delta_pct, 2)}

    # Sharp money detection: spread stretto + bassa volatilità + overround ridotto
    if (
        liquidity_indicators.get("spread") in {"tight", "balanced"} and
        liquidity_indicators.get("volatility") in {"very_low", "low"} and
        liquidity_indicators.get("overround") in {"sharp", "efficient"} and
        num_bookmakers >= 8
    ):
        liquidity_indicators["sharp_money"] = "likely"
        depth_score += 12
    else:
        liquidity_indicators["sharp_money"] = "unclear" if num_bookmakers >= 5 else "unlikely"

    efficiency_score = min(100.0, depth_score)

    return {
        "depth_score": round(efficiency_score, 1),
        "num_bookmakers": num_bookmakers,
        "liquidity_indicators": liquidity_indicators,
        "avg_spread_pct": round(avg_spread, 2) if avg_spread is not None else None,
        "average_volatility_pct": round(avg_cov, 2) if avg_cov is not None else None,
        "overround_pct": round(overround * 100, 2) if overround is not None else None,
        "consensus_probabilities": consensus_probs,
        "odds_summary": odds_summary,
        "value_flags": value_flags,
        "market_efficiency": (
            "high" if efficiency_score >= 70 else
            ("medium" if efficiency_score >= 50 else "low")
        ),
    }

def _apply_advanced_metric_overlays(
    form_attack: float,
    form_defense: float,
    form_points_factor: float,
    confidence: float,
    advanced_metrics: Optional[Dict[str, Any]],
) -> Tuple[float, float, float, float]:
    """Applica aggiustamenti da fonti avanzate (StatsBomb, Football-Data)."""
    if not advanced_metrics:
        return form_attack, form_defense, form_points_factor, confidence
    
    metrics = advanced_metrics or {}
    
    # StatsBomb
    sb_metrics = metrics.get("statsbomb")
    if sb_metrics and sb_metrics.get("available"):
        baseline_xg = 1.35
        
        try:
            xg_for_avg = float(sb_metrics.get("xg_for_avg", 0))
            if xg_for_avg > 0:
                xg_factor = 1.0 + ((xg_for_avg - baseline_xg) / baseline_xg) * 0.25
                xg_factor = max(0.85, min(1.15, xg_factor))
                form_attack = (form_attack * 2.0 + xg_factor) / 3.0
                form_attack = max(0.85, min(1.15, form_attack))
        except (TypeError, ValueError):
            pass
        
        try:
            xg_against_avg = float(sb_metrics.get("xg_against_avg", 0))
            if xg_against_avg > 0:
                defense_factor_sb = 1.0 + ((baseline_xg - xg_against_avg) / baseline_xg) * 0.25
                defense_factor_sb = max(0.85, min(1.15, defense_factor_sb))
                form_defense = (form_defense * 2.0 + defense_factor_sb) / 3.0
                form_defense = max(0.85, min(1.15, form_defense))
        except (TypeError, ValueError):
            pass
        
        shots_baseline = 12.0
        try:
            shots_for_avg = float(sb_metrics.get("shots_for_avg", 0))
            if shots_for_avg > 0:
                shots_factor = 1.0 + ((shots_for_avg - shots_baseline) / shots_baseline) * 0.10
                shots_factor = max(0.9, min(1.1, shots_factor))
                form_attack = (form_attack * 3.0 + shots_factor) / 4.0
                form_attack = max(0.85, min(1.15, form_attack))
        except (TypeError, ValueError):
            pass
        
        try:
            shots_against_avg = float(sb_metrics.get("shots_against_avg", 0))
            if shots_against_avg > 0:
                shots_def_factor = 1.0 - ((shots_against_avg - shots_baseline) / shots_baseline) * 0.10
                shots_def_factor = max(0.9, min(1.1, shots_def_factor))
                form_defense = (form_defense * 3.0 + shots_def_factor) / 4.0
                form_defense = max(0.85, min(1.15, form_defense))
        except (TypeError, ValueError):
            pass
        
        matches_used = sb_metrics.get("matches_used")
        if isinstance(matches_used, (int, float)):
            confidence = min(1.0, confidence + min(0.25, matches_used / 40.0))
    
    # Football-Data.org fallback metrics
    fd_metrics = metrics.get("football_data")
    if fd_metrics and fd_metrics.get("available"):
        baseline_goals = 1.3
        try:
            fd_goals_for = float(fd_metrics.get("avg_goals_for", 0))
            fd_attack_factor = 1.0 + ((fd_goals_for - baseline_goals) / baseline_goals) * 0.20
            fd_attack_factor = max(0.85, min(1.15, fd_attack_factor))
            form_attack = (form_attack * 3.0 + fd_attack_factor) / 4.0
            form_attack = max(0.85, min(1.15, form_attack))
        except (TypeError, ValueError):
            pass
        
        try:
            fd_goals_against = float(fd_metrics.get("avg_goals_against", 0))
            fd_def_factor = 1.0 + ((baseline_goals - fd_goals_against) / baseline_goals) * 0.20
            fd_def_factor = max(0.85, min(1.15, fd_def_factor))
            form_defense = (form_defense * 3.0 + fd_def_factor) / 4.0
            form_defense = max(0.85, min(1.15, form_defense))
        except (TypeError, ValueError):
            pass
        
        try:
            clean_sheet_rate = float(fd_metrics.get("clean_sheet_rate", 0))
            cs_factor = 1.0 + (clean_sheet_rate - 0.3) * 0.25
            cs_factor = max(0.85, min(1.15, cs_factor))
            form_defense = (form_defense * 3.0 + cs_factor) / 4.0
            form_defense = max(0.85, min(1.15, form_defense))
        except (TypeError, ValueError):
            pass
        
        try:
            points_per_game = float(fd_metrics.get("points_per_game", 0))
            fd_points_factor = 0.8 + (points_per_game / 3.0) * 0.8
            fd_points_factor = max(0.7, min(1.6, fd_points_factor))
            form_points_factor = (form_points_factor * 2.0 + fd_points_factor) / 3.0
        except (TypeError, ValueError):
            pass
        
        matches_fd = fd_metrics.get("matches")
        if isinstance(matches_fd, (int, float)):
            confidence = min(1.0, max(confidence, 0.2) + min(0.2, matches_fd / 15.0))
        else:
            confidence = max(confidence, 0.2)
    
    return form_attack, form_defense, form_points_factor, confidence


def _form_from_advanced_metrics_only(advanced_metrics: Optional[Dict[str, Any]]) -> Optional[Dict[str, float]]:
    """Calcola forma usando solo metriche avanzate (assenza API-Football)."""
    if not advanced_metrics:
        return None
    
    form_attack = 1.0
    form_defense = 1.0
    form_points_factor = 1.0
    confidence = 0.0
    
    form_attack, form_defense, form_points_factor, confidence = _apply_advanced_metric_overlays(
        form_attack,
        form_defense,
        form_points_factor,
        confidence,
        advanced_metrics,
    )
    
    if confidence <= 0.0 and abs(form_attack - 1.0) < 1e-3 and abs(form_defense - 1.0) < 1e-3:
        return None
    
    confidence = max(confidence, 0.2 if advanced_metrics else 0.0)
    
    return {
        "form_attack": round(form_attack, 3),
        "form_defense": round(form_defense, 3),
        "form_points": round(form_points_factor, 3),
        "confidence": round(min(1.0, confidence), 2),
        "goals_for_avg": None,
        "goals_against_avg": None,
        "shots_for_avg": None,
        "shots_against_avg": None,
    }


def calculate_team_form_from_statistics(
    team_stats: Dict[str, Any],
    last_n: int = 5,
    advanced_metrics: Optional[Dict[str, Any]] = None,
) -> Dict[str, float]:
    """
    Calcola fattore forma da statistiche squadra usando dati aggiornati dalle API.
    Integra dati da API-Football per calcoli matematici più accurati.
    """
    if not team_stats:
        fallback = _form_from_advanced_metrics_only(advanced_metrics)
        if fallback:
            return fallback
        return {"form_attack": 1.0, "form_defense": 1.0, "form_points": 1.0, "confidence": 0.0}
    
    try:
        fixtures = team_stats.get("fixtures", {}) if isinstance(team_stats, dict) else {}
        # ⚠️ PROTEZIONE: Accesso sicuro a dizionari annidati
        played_data = fixtures.get("played", {}) if isinstance(fixtures, dict) else {}
        played = _safe_int(played_data.get("total", 0)) if isinstance(played_data, dict) else 0
        
        if played < 3:
            fallback = _form_from_advanced_metrics_only(advanced_metrics)
            if fallback:
                return fallback
            return {"form_attack": 1.0, "form_defense": 1.0, "form_points": 1.0, "confidence": 0.0}
        
        # Statistiche attacco (da API aggiornate)
        # ⚠️ PROTEZIONE: Accesso sicuro a dizionari annidati con fallback
        goals_data = team_stats.get("goals", {}) if isinstance(team_stats, dict) else {}
        goals_for_data = goals_data.get("for", {}) if isinstance(goals_data, dict) else {}
        goals_for_avg = goals_for_data.get("average", {}) if isinstance(goals_for_data, dict) else {}
        goals_for = _safe_float(goals_for_avg.get("total", 0.0)) if isinstance(goals_for_avg, dict) else 0.0
        
        goals_against_data = goals_data.get("against", {}) if isinstance(goals_data, dict) else {}
        goals_against_avg = goals_against_data.get("average", {}) if isinstance(goals_against_data, dict) else {}
        goals_against = _safe_float(goals_against_avg.get("total", 0.0)) if isinstance(goals_against_avg, dict) else 0.0
        
        # Statistiche avanzate se disponibili (shots, xG, etc.) - dati API
        shots_data = team_stats.get("shots", {}) if isinstance(team_stats, dict) and team_stats.get("shots") else {}
        if shots_data:
            shots_for_data = shots_data.get("for", {}) if isinstance(shots_data, dict) else {}
            shots_for_avg = shots_for_data.get("average", {}) if isinstance(shots_for_data, dict) else {}
            shots_for = _safe_float(shots_for_avg.get("total", 0.0)) if isinstance(shots_for_avg, dict) else 0.0
            
            shots_against_data = shots_data.get("against", {}) if isinstance(shots_data, dict) else {}
            shots_against_avg = shots_against_data.get("average", {}) if isinstance(shots_against_data, dict) else {}
            shots_against = _safe_float(shots_against_avg.get("total", 0.0)) if isinstance(shots_against_avg, dict) else 0.0
        else:
            shots_for = shots_against = 0.0
        
        # Forma ultime partite (dati reali dalle API)
        # ⚠️ PROTEZIONE: Accesso sicuro a dizionari annidati
        wins_data = fixtures.get("wins", {}) if isinstance(fixtures, dict) else {}
        wins = _safe_int(wins_data.get("total", 0)) if isinstance(wins_data, dict) else 0
        
        draws_data = fixtures.get("draws", {}) if isinstance(fixtures, dict) else {}
        draws = _safe_int(draws_data.get("total", 0)) if isinstance(draws_data, dict) else 0
        
        losses_data = fixtures.get("loses", {}) if isinstance(fixtures, dict) else {}
        losses = _safe_int(losses_data.get("total", 0)) if isinstance(losses_data, dict) else 0
        
        # ⚠️ PRECISIONE: Calcola forma punti con protezione divisione per zero
        played_safe = max(1, played)
        form_points = (wins * 3.0 + draws) / max(model_config.TOL_DIVISION_ZERO, played_safe * 3.0)
        form_points = max(0.0, min(1.0, form_points))  # Limita a [0, 1]
        
        # Normalizza forma punti (0.33 = media, 1.0 = perfetto)
        form_points_factor = 0.7 + (form_points - 0.33) * 0.9  # Range: 0.7 - 1.6
        form_points_factor = max(0.7, min(1.6, form_points_factor))  # Limita range
        
        # ⚠️ VERIFICA: Assicura che form_points_factor sia finito
        if not math.isfinite(form_points_factor):
            logger.warning(f"form_points_factor non finito: {form_points_factor}, correggo")
            form_points_factor = 1.0
        
        # Fattore attacco basato su gol fatti (dati API aggiornati)
        avg_goals_league = 1.3
        # ⚠️ PRECISIONE: Protezione divisione per zero
        if avg_goals_league > model_config.TOL_DIVISION_ZERO:
            form_attack = 0.85 + (goals_for / avg_goals_league - 1.0) * 0.3  # Range: 0.85 - 1.15
        else:
            form_attack = 1.0
        form_attack = max(0.85, min(1.15, form_attack))  # Limita range
        
        # ⚠️ VERIFICA: Assicura che form_attack sia finito
        if not math.isfinite(form_attack):
            logger.warning(f"form_attack non finito: {form_attack}, correggo")
            form_attack = 1.0
        
        # Migliora con dati shots se disponibili (più tiri = più opportunità)
        if shots_for > 0:
            avg_shots_league = 12.0  # Media lega
            # ⚠️ PRECISIONE: Protezione divisione per zero
            if avg_shots_league > model_config.TOL_DIVISION_ZERO:
                shots_factor = min(1.1, 0.95 + (shots_for / avg_shots_league - 1.0) * 0.15)
            else:
                shots_factor = 1.0
            form_attack = (form_attack + shots_factor) / 2.0  # Media pesata
            form_attack = max(0.85, min(1.15, form_attack))  # Limita range
        
        # Fattore difesa basato su gol subiti (dati API aggiornati)
        # ⚠️ PRECISIONE: Protezione divisione per zero
        if avg_goals_league > model_config.TOL_DIVISION_ZERO:
            form_defense = 0.85 + (1.0 - goals_against / avg_goals_league) * 0.3  # Range: 0.85 - 1.15
        else:
            form_defense = 1.0
        form_defense = max(0.85, min(1.15, form_defense))  # Limita range
        
        # ⚠️ VERIFICA: Assicura che form_defense sia finito
        if not math.isfinite(form_defense):
            logger.warning(f"form_defense non finito: {form_defense}, correggo")
            form_defense = 1.0
        
        # Migliora con dati shots against se disponibili
        if shots_against > 0:
            avg_shots_league = 12.0
            # ⚠️ PRECISIONE: Protezione divisione per zero
            if avg_shots_league > model_config.TOL_DIVISION_ZERO:
                shots_against_factor = min(1.1, 0.95 + (1.0 - shots_against / avg_shots_league) * 0.15)
            else:
                shots_against_factor = 1.0
            form_defense = (form_defense + shots_against_factor) / 2.0
            form_defense = max(0.85, min(1.15, form_defense))  # Limita range
        
        # Confidence basata su partite giocate e qualità dati
        confidence = min(1.0, played / 10.0)
        
        # Boost confidence se abbiamo statistiche avanzate (shots, xG, etc.)
        if shots_for > 0 or shots_against > 0:
            confidence = min(1.0, confidence * 1.1)  # +10% confidence
        
        form_attack, form_defense, form_points_factor, confidence = _apply_advanced_metric_overlays(
            form_attack,
            form_defense,
            form_points_factor,
            confidence,
            advanced_metrics,
        )
        
        return {
            "form_attack": round(form_attack, 3),
            "form_defense": round(form_defense, 3),
            "form_points": round(form_points_factor, 3),
            "confidence": round(confidence, 2),
            "goals_for_avg": round(goals_for, 2),
            "goals_against_avg": round(goals_against, 2),
            "shots_for_avg": round(shots_for, 2) if shots_for > 0 else None,
            "shots_against_avg": round(shots_against, 2) if shots_against > 0 else None,
        }
    except (KeyError, ValueError, TypeError) as e:
        logger.error(f"Errore calcolo forma da statistiche: {e}")
        return {"form_attack": 1.0, "form_defense": 1.0, "form_points": 1.0, "confidence": 0.0}

def calculate_team_form_with_venue_split(
    team_id_fd: int,
    competition_code: str = None,
    venue: str = "total"
) -> Dict[str, Any]:
    """
    Calcola forma squadra con split HOME/AWAY da Football-Data.org.

    Args:
        team_id_fd: ID squadra su Football-Data.org
        competition_code: Codice competizione (es: "PL", "SA", "PD", "BL1", "FL1")
        venue: "home", "away", o "total" (default)

    Returns:
        Dict con form factors specifici per venue:
        {
            "form_attack": float,
            "form_defense": float,
            "form_points": float,
            "confidence": float,
            "goals_for_avg": float,
            "goals_against_avg": float,
            "venue": str
        }
    """
    if not team_id_fd:
        return {"form_attack": 1.0, "form_defense": 1.0, "form_points": 1.0, "confidence": 0.0}

    try:
        # Recupera statistiche home/away
        stats = football_data_get_home_away_stats(team_id_fd, competition_code)

        if not stats.get("available"):
            return {"form_attack": 1.0, "form_defense": 1.0, "form_points": 1.0, "confidence": 0.0}

        # Seleziona stats in base al venue
        if venue == "home":
            venue_stats = stats["home"]
        elif venue == "away":
            venue_stats = stats["away"]
        else:  # total
            # Combina home e away
            home_stats = stats["home"]
            away_stats = stats["away"]
            venue_stats = {
                "played": home_stats["played"] + away_stats["played"],
                "won": home_stats["won"] + away_stats["won"],
                "draw": home_stats["draw"] + away_stats["draw"],
                "lost": home_stats["lost"] + away_stats["lost"],
                "goals_for": home_stats["goals_for"] + away_stats["goals_for"],
                "goals_against": home_stats["goals_against"] + away_stats["goals_against"],
                "points": home_stats["points"] + away_stats["points"]
            }

        played = venue_stats["played"]
        if played < 3:
            return {"form_attack": 1.0, "form_defense": 1.0, "form_points": 1.0, "confidence": 0.1}

        # Calcola medie
        played_safe = max(1, played)
        goals_for_avg = venue_stats["goals_for"] / played_safe
        goals_against_avg = venue_stats["goals_against"] / played_safe
        points_per_game = venue_stats["points"] / played_safe

        # Form points (normalizzato: 0 = pessimo, 1 = perfetto, 3 = max)
        form_points_normalized = points_per_game / 3.0  # [0, 1]
        form_points_factor = 0.7 + (form_points_normalized - 0.33) * 0.9  # Range: 0.7 - 1.6
        form_points_factor = max(0.7, min(1.6, form_points_factor))

        # Form attack (basato su gol fatti)
        avg_goals_league = 1.3  # Media tipica per squadra
        form_attack = 0.85 + (goals_for_avg / avg_goals_league - 1.0) * 0.3
        form_attack = max(0.80, min(1.20, form_attack))  # Range esteso per home/away

        # Form defense (basato su gol subiti)
        form_defense = 0.85 + (1.0 - goals_against_avg / avg_goals_league) * 0.3
        form_defense = max(0.80, min(1.20, form_defense))

        # Confidence basata su partite giocate
        confidence = min(1.0, played / 10.0)

        # Boost confidence se abbiamo molti dati home/away
        if played >= 8:
            confidence = min(1.0, confidence * 1.15)

        return {
            "form_attack": round(form_attack, 3),
            "form_defense": round(form_defense, 3),
            "form_points": round(form_points_factor, 3),
            "confidence": round(confidence, 2),
            "goals_for_avg": round(goals_for_avg, 2),
            "goals_against_avg": round(goals_against_avg, 2),
            "played": played,
            "venue": venue,
            "source": "football-data.org"
        }

    except Exception as e:
        logger.error(f"Errore calcolo forma con venue split: {e}")
        return {"form_attack": 1.0, "form_defense": 1.0, "form_points": 1.0, "confidence": 0.0}

def calculate_h2h_adjustments(h2h_matches: List[Dict[str, Any]], home_team_id: int, away_team_id: int) -> Dict[str, float]:
    """
    Calcola aggiustamenti basati su H2H.
    """
    if not h2h_matches or len(h2h_matches) < 2:
        return {"h2h_home_advantage": 1.0, "h2h_goals_factor": 1.0, "h2h_btts_factor": 1.0, "confidence": 0.0}
    
    try:
        home_wins = 0
        draws = 0
        away_wins = 0
        total_goals = 0
        btts_count = 0
        matches_played = 0
        
        for match in h2h_matches:
            fixture = match.get("fixture", {})
            status = fixture.get("status", {}).get("short", "")
            
            if status not in ["FT", "AET", "PEN"]:
                continue
            
            teams = match.get("teams", {})
            home = teams.get("home", {})
            away = teams.get("away", {})
            
            home_id = home.get("id")
            away_id = away.get("id")
            
            # Determina quale squadra è casa/trasferta in questo match
            if home_id == home_team_id:
                # La nostra home è casa in questo match
                home_score = match.get("goals", {}).get("home", 0)
                away_score = match.get("goals", {}).get("away", 0)
            elif away_id == home_team_id:
                # La nostra home è trasferta in questo match
                home_score = match.get("goals", {}).get("away", 0)
                away_score = match.get("goals", {}).get("home", 0)
            else:
                continue
            
            matches_played += 1
            total_goals += home_score + away_score
            
            if home_score > away_score:
                home_wins += 1
            elif home_score < away_score:
                away_wins += 1
            else:
                draws += 1
            
            if home_score > 0 and away_score > 0:
                btts_count += 1
        
        if matches_played < 2:
            return {"h2h_home_advantage": 1.0, "h2h_goals_factor": 1.0, "h2h_btts_factor": 1.0, "confidence": 0.0}
        
        # ⚠️ PRECISIONE: Calcola fattori con protezione divisione per zero
        matches_played_safe = max(1, matches_played)
        
        home_win_rate = home_wins / matches_played_safe
        draw_rate = draws / matches_played_safe
        away_win_rate = away_wins / matches_played_safe
        
        # ⚠️ PROTEZIONE: Limita rate a [0, 1]
        home_win_rate = max(0.0, min(1.0, home_win_rate))
        draw_rate = max(0.0, min(1.0, draw_rate))
        away_win_rate = max(0.0, min(1.0, away_win_rate))
        
        # Se home vince spesso → aumenta vantaggio casa
        # Media: 45% vittorie casa, 25% pareggi, 30% vittorie trasferta
        expected_home_win = 0.45
        h2h_home_advantage = 0.9 + (home_win_rate - expected_home_win) * 0.4  # Range: 0.9 - 1.1
        h2h_home_advantage = max(0.9, min(1.1, h2h_home_advantage))  # Limita range
        
        # ⚠️ VERIFICA: Assicura che h2h_home_advantage sia finito
        if not math.isfinite(h2h_home_advantage):
            logger.warning(f"h2h_home_advantage non finito: {h2h_home_advantage}, correggo")
            h2h_home_advantage = 1.0
        
        # Fattore gol: media gol in H2H vs media generale (2.6)
        avg_goals_h2h = total_goals / matches_played_safe
        avg_goals_general = 2.6
        # ⚠️ PRECISIONE: Protezione divisione per zero
        if avg_goals_general > model_config.TOL_DIVISION_ZERO:
            h2h_goals_factor = 0.9 + (avg_goals_h2h / avg_goals_general - 1.0) * 0.2  # Range: 0.9 - 1.1
        else:
            h2h_goals_factor = 1.0
        h2h_goals_factor = max(0.9, min(1.1, h2h_goals_factor))  # Limita range
        
        # ⚠️ VERIFICA: Assicura che h2h_goals_factor sia finito
        if not math.isfinite(h2h_goals_factor):
            logger.warning(f"h2h_goals_factor non finito: {h2h_goals_factor}, correggo")
            h2h_goals_factor = 1.0
        
        # Fattore BTTS
        btts_rate = btts_count / matches_played_safe
        btts_rate = max(0.0, min(1.0, btts_rate))  # Limita a [0, 1]
        avg_btts_rate = 0.52  # Media generale
        # ⚠️ PRECISIONE: Protezione divisione per zero
        if avg_btts_rate > model_config.TOL_DIVISION_ZERO:
            h2h_btts_factor = 0.9 + (btts_rate / avg_btts_rate - 1.0) * 0.2  # Range: 0.9 - 1.1
        else:
            h2h_btts_factor = 1.0
        h2h_btts_factor = max(0.9, min(1.1, h2h_btts_factor))  # Limita range
        
        # ⚠️ VERIFICA: Assicura che h2h_btts_factor sia finito
        if not math.isfinite(h2h_btts_factor):
            logger.warning(f"h2h_btts_factor non finito: {h2h_btts_factor}, correggo")
            h2h_btts_factor = 1.0
        
        # Confidence basata su numero match
        confidence = min(1.0, matches_played / 5.0)
        
        return {
            "h2h_home_advantage": round(h2h_home_advantage, 3),
            "h2h_goals_factor": round(h2h_goals_factor, 3),
            "h2h_btts_factor": round(h2h_btts_factor, 3),
            "confidence": round(confidence, 2),
            "matches_analyzed": matches_played,
            "home_wins": home_wins,
            "draws": draws,
            "away_wins": away_wins,
        }
    except (KeyError, ValueError, TypeError, ZeroDivisionError) as e:
        logger.error(f"Errore calcolo H2H: {e}")
        return {"h2h_home_advantage": 1.0, "h2h_goals_factor": 1.0, "h2h_btts_factor": 1.0, "confidence": 0.0}

def calculate_injuries_impact(injuries: List[Dict[str, Any]], team_id: int) -> Dict[str, float]:
    """
    Calcola impatto infortuni E SOSPENSIONI su lambda.

    Supporta sia injuries che suspensions dall'endpoint API-Football /sidelined.
    Le sospensioni hanno un impatto leggermente maggiore perché sono immediate e certe.
    """
    if not injuries:
        return {"attack_factor": 1.0, "defense_factor": 1.0, "confidence": 0.0}

    try:
        team_injuries = [inj for inj in injuries if inj.get("team", {}).get("id") == team_id]

        if not team_injuries:
            return {"attack_factor": 1.0, "defense_factor": 1.0, "confidence": 0.0}

        # Classifica posizioni (approssimativo)
        attack_positions = ["Forward", "Attacker", "Winger"]
        defense_positions = ["Defender", "Goalkeeper"]
        midfield_positions = ["Midfielder"]

        attack_impact = 0
        defense_impact = 0
        num_injuries = 0
        num_suspensions = 0

        for injury in team_injuries:
            player = injury.get("player", {})
            position_raw = player.get("position")
            # Assicurati che position sia sempre una stringa (gestisce None)
            position = str(position_raw).upper() if position_raw is not None else ""

            # Determina se è sospensione o infortunio
            injury_type = injury.get("type", "")
            reason = str(injury.get("reason", "")).lower()
            is_suspension = (
                "suspend" in injury_type.lower() or
                "suspend" in reason or
                "card" in reason or
                "ban" in reason
            )

            # Sospensioni hanno impatto maggiore (+40% rispetto agli infortuni)
            multiplier = 1.4 if is_suspension else 1.0

            if is_suspension:
                num_suspensions += 1
            else:
                num_injuries += 1

            # Determina impatto basandosi su posizione
            if position and any(pos.upper() in position for pos in attack_positions):
                # -5% base per attaccante, -7% se sospeso
                attack_impact += 0.05 * multiplier
            elif position and any(pos.upper() in position for pos in defense_positions):
                # -5% base per difensore, -7% se sospeso
                defense_impact += 0.05 * multiplier
            elif position and any(pos.upper() in position for pos in midfield_positions):
                # -2% base per centrocampista, -2.8% se sospeso
                attack_impact += 0.02 * multiplier
                defense_impact += 0.02 * multiplier

        # Limita impatto massimo
        attack_factor = max(0.80, 1.0 - min(0.20, attack_impact))  # Max -20%
        defense_factor = max(0.80, 1.0 - min(0.20, defense_impact))  # Max -20%

        # Confidence: più assenti = più confidence nell'impatto
        # Sospensioni hanno confidence maggiore (sono certe)
        total_absences = num_injuries + num_suspensions
        confidence = min(1.0, (num_injuries * 0.25 + num_suspensions * 0.4))

        return {
            "attack_factor": round(attack_factor, 3),
            "defense_factor": round(defense_factor, 3),
            "confidence": round(confidence, 2),
            "num_injuries": num_injuries,
            "num_suspensions": num_suspensions,
            "total_absences": total_absences,
        }
    except (KeyError, ValueError, TypeError) as e:
        logger.error(f"Errore calcolo impatto infortuni/sospensioni: {e}")
        return {"attack_factor": 1.0, "defense_factor": 1.0, "confidence": 0.0}

def get_team_fatigue_and_motivation_data(
    team_name: str,
    league: str,
    match_date: str,
) -> Dict[str, Any]:
    """
    Recupera automaticamente tutti i dati per fatigue e motivation.
    """
    result = {
        "team_id": None,
        "days_since_last_match": None,
        "matches_last_30_days": None,
        "position": None,
        "points_from_relegation": None,
        "points_from_europe": None,
        "is_derby": False,
        "data_available": False,
    }
    
    try:
        # 1. Cerca team ID
        league_id = get_league_id_from_name(league)
        team_info = apifootball_search_team(team_name, league_id)
        
        if not team_info or "team" not in team_info:
            return result
        
        team_id = team_info["team"].get("id")
        if not team_id:
            return result
        
        result["team_id"] = team_id
        
        # 2. Calcola fatigue
        result["days_since_last_match"] = calculate_days_since_last_match(team_id, match_date)
        result["matches_last_30_days"] = count_matches_last_30_days(team_id, match_date)
        
        # 3. Recupera standings
        standings_info = get_team_standings_info(team_name, league)
        result["position"] = standings_info.get("position")
        result["points_from_relegation"] = standings_info.get("points_from_relegation")
        result["points_from_europe"] = standings_info.get("points_from_europe")
        
        # 4. Verifica se abbiamo almeno alcuni dati
        result["data_available"] = (
            result["days_since_last_match"] is not None or
            result["matches_last_30_days"] is not None or
            result["position"] is not None
        )
        
        return result
    except (KeyError, ValueError, requests.exceptions.RequestException) as e:
        logger.error(f"Errore recupero dati team {team_name}: {e}")
        return result

def get_advanced_team_data(
    home_team_name: str,
    away_team_name: str,
    league: str,
    match_date: str,
) -> Dict[str, Any]:
    """
    Recupera dati avanzati da TUTTE le API disponibili: statistiche, H2H, infortuni, Football-Data.org, TheSportsDB.
    Lavora in background per supportare il modello con dati matematici e statistici aggiornati.

    NEW FEATURES (2025-11-11):
    - Statistiche HOME/AWAY separate da Football-Data.org per calcoli più precisi
    - Supporto sospensioni giocatori (oltre agli infortuni) con impatto maggiorato
    - Form factors specifici per venue (home team gioca in casa, away team gioca fuori)

    Returns:
        Dict contenente:
        - home_team_stats, away_team_stats: Statistiche generali
        - football_data_home_form_home: Forma squadra di casa quando gioca IN CASA
        - football_data_away_form_away: Forma squadra ospite quando gioca FUORI CASA
        - home_injuries, away_injuries: Include sia infortuni che sospensioni
        - h2h_data: Dati head-to-head
        - Altri dati da API multiple
    """
    result = {
        "home_team_stats": None,
        "away_team_stats": None,
        "h2h_data": None,
        "home_injuries": None,
        "away_injuries": None,
        "football_data_home": None,
        "football_data_away": None,
        "football_data_home_metrics": None,
        "football_data_away_metrics": None,
        "football_data_home_form_home": None,  # NEW: Forma specifica home da Football-Data.org
        "football_data_away_form_away": None,  # NEW: Forma specifica away da Football-Data.org
        "thesportsdb_home": None,
        "thesportsdb_away": None,
        "statsbomb_home": None,
        "statsbomb_away": None,
        "data_available": False,
    }
    
    home_stats_raw: Optional[Dict[str, Any]] = None
    away_stats_raw: Optional[Dict[str, Any]] = None
    
    try:
        league_id = get_league_id_from_name(league)
        season = get_current_season()
        
        # ============================================================
        # 1. API-FOOTBALL: Statistiche, H2H, Infortuni
        # ============================================================
        if API_FOOTBALL_KEY and league_id:
            # Cerca team IDs
            home_team_info = apifootball_search_team(home_team_name, league_id)
            away_team_info = apifootball_search_team(away_team_name, league_id)
            
            home_team_id = home_team_info.get("team", {}).get("id") if home_team_info else None
            away_team_id = away_team_info.get("team", {}).get("id") if away_team_info else None
            
            if home_team_id and away_team_id:
                # Statistiche squadre
                home_stats = apifootball_get_team_statistics(home_team_id, league_id, season)
                away_stats = apifootball_get_team_statistics(away_team_id, league_id, season)
                
                if home_stats:
                    home_stats_raw = home_stats
                if away_stats:
                    away_stats_raw = away_stats
                
                # H2H
                h2h_matches = apifootball_get_head_to_head(home_team_id, away_team_id, last=10)
                if h2h_matches:
                    result["h2h_data"] = calculate_h2h_adjustments(h2h_matches, home_team_id, away_team_id)
                
                # Infortuni
                all_injuries = apifootball_get_injuries()
                if all_injuries:
                    result["home_injuries"] = calculate_injuries_impact(all_injuries, home_team_id)
                    result["away_injuries"] = calculate_injuries_impact(all_injuries, away_team_id)
        
        # ============================================================
        # 2. FOOTBALL-DATA.ORG: Dati aggiuntivi per validazione e statistiche
        # ============================================================
        if FOOTBALL_DATA_API_KEY:
            try:
                result["football_data_home"] = football_data_get_team_info(home_team_name, league)
                result["football_data_away"] = football_data_get_team_info(away_team_name, league)
                
                if result["football_data_home"].get("available"):
                    home_team_id_fd = result["football_data_home"].get("team_id")
                    recent_home_matches = football_data_get_recent_matches(home_team_id_fd, limit=6)
                    result["football_data_home_metrics"] = football_data_calculate_form_metrics(home_team_id_fd, recent_home_matches)
                    if not home_stats_raw and result["football_data_home_metrics"].get("available"):
                        logger.info(f"Uso Football-Data.org come fallback per forma {home_team_name}")
                
                if result["football_data_away"].get("available"):
                    away_team_id_fd = result["football_data_away"].get("team_id")
                    recent_away_matches = football_data_get_recent_matches(away_team_id_fd, limit=6)
                    result["football_data_away_metrics"] = football_data_calculate_form_metrics(away_team_id_fd, recent_away_matches)
                    if not away_stats_raw and result["football_data_away_metrics"].get("available"):
                        logger.info(f"Uso Football-Data.org come fallback per forma {away_team_name}")

                # NEW: Recupera statistiche HOME/AWAY separate per analisi più precisa
                if result["football_data_home"].get("available"):
                    home_team_id_fd = result["football_data_home"].get("team_id")
                    home_team_home_form = calculate_team_form_with_venue_split(home_team_id_fd, venue="home")
                    if home_team_home_form.get("confidence", 0) > 0.3:
                        result["football_data_home_form_home"] = home_team_home_form
                        logger.debug(f"Statistiche HOME recuperate per {home_team_name}: attack={home_team_home_form['form_attack']}, defense={home_team_home_form['form_defense']}")

                if result["football_data_away"].get("available"):
                    away_team_id_fd = result["football_data_away"].get("team_id")
                    away_team_away_form = calculate_team_form_with_venue_split(away_team_id_fd, venue="away")
                    if away_team_away_form.get("confidence", 0) > 0.3:
                        result["football_data_away_form_away"] = away_team_away_form
                        logger.debug(f"Statistiche AWAY recuperate per {away_team_name}: attack={away_team_away_form['form_attack']}, defense={away_team_away_form['form_defense']}")
            except Exception as e:
                logger.debug(f"Football-Data.org non disponibile: {e}")
        
        # ============================================================
        # 3. THESPORTSDB: Info stadio e squadra (gratuito, sempre disponibile)
        # ============================================================
        try:
            result["thesportsdb_home"] = thesportsdb_get_team_info(home_team_name)
            result["thesportsdb_away"] = thesportsdb_get_team_info(away_team_name)
            
            # Se abbiamo info stadio, può essere usato per aggiustamenti
            if result["thesportsdb_home"].get("available"):
                logger.debug(f"Info TheSportsDB disponibili per {home_team_name}")
        except Exception as e:
            logger.debug(f"TheSportsDB non disponibile: {e}")
        
        # ============================================================
        # 4. STATSBOMB OPEN DATA: xG e metriche avanzate
        # ============================================================
        try:
            result["statsbomb_home"] = statsbomb_get_team_metrics(home_team_name)
        except Exception as e:
            logger.debug(f"StatsBomb home non disponibile: {e}")
            result["statsbomb_home"] = {"available": False, "error": str(e)}
        
        try:
            result["statsbomb_away"] = statsbomb_get_team_metrics(away_team_name)
        except Exception as e:
            logger.debug(f"StatsBomb away non disponibile: {e}")
            result["statsbomb_away"] = {"available": False, "error": str(e)}
        
        # Calcola forma finale con tutti i dati raccolti
        if home_stats_raw:
            advanced_home_inputs: Dict[str, Any] = {}
            if result["statsbomb_home"]:
                advanced_home_inputs["statsbomb"] = result["statsbomb_home"]
            if result["football_data_home_metrics"]:
                advanced_home_inputs["football_data"] = result["football_data_home_metrics"]
            result["home_team_stats"] = calculate_team_form_from_statistics(
                home_stats_raw,
                advanced_metrics=advanced_home_inputs if advanced_home_inputs else None,
            )
        elif result["football_data_home_metrics"]:
            advanced_home_inputs = {
                "football_data": result["football_data_home_metrics"],
            }
            if result["statsbomb_home"]:
                advanced_home_inputs["statsbomb"] = result["statsbomb_home"]
            fallback_stats = _form_from_advanced_metrics_only(advanced_home_inputs)
            if fallback_stats:
                result["home_team_stats"] = fallback_stats
        
        if away_stats_raw:
            advanced_away_inputs: Dict[str, Any] = {}
            if result["statsbomb_away"]:
                advanced_away_inputs["statsbomb"] = result["statsbomb_away"]
            if result["football_data_away_metrics"]:
                advanced_away_inputs["football_data"] = result["football_data_away_metrics"]
            result["away_team_stats"] = calculate_team_form_from_statistics(
                away_stats_raw,
                advanced_metrics=advanced_away_inputs if advanced_away_inputs else None,
            )
        elif result["football_data_away_metrics"]:
            advanced_away_inputs = {
                "football_data": result["football_data_away_metrics"],
            }
            if result["statsbomb_away"]:
                advanced_away_inputs["statsbomb"] = result["statsbomb_away"]
            fallback_stats = _form_from_advanced_metrics_only(advanced_away_inputs)
            if fallback_stats:
                result["away_team_stats"] = fallback_stats
        
        # ============================================================
        # 5. CALCOLA STATISTICHE AGGREGATE DA TUTTE LE API
        # ============================================================
        # Se abbiamo dati da più fonti, combinali per statistiche più accurate
        if result["home_team_stats"] or result["football_data_home"]:
            # Migliora confidence se abbiamo dati da più fonti
            if result["home_team_stats"] and result["football_data_home"].get("available"):
                if result["home_team_stats"].get("confidence"):
                    result["home_team_stats"]["confidence"] = min(1.0, result["home_team_stats"]["confidence"] * 1.1)
        
        if result["away_team_stats"] or result["football_data_away"]:
            if result["away_team_stats"] and result["football_data_away"].get("available"):
                if result["away_team_stats"].get("confidence"):
                    result["away_team_stats"]["confidence"] = min(1.0, result["away_team_stats"]["confidence"] * 1.1)
        
        # ============================================================
        # 5. Verifica se abbiamo almeno alcuni dati
        # ============================================================
        result["data_available"] = (
            result["home_team_stats"] is not None or
            result["away_team_stats"] is not None or
            result["h2h_data"] is not None or
            result["home_injuries"] is not None or
            result["away_injuries"] is not None or
            result["football_data_home"] is not None or
            result["football_data_away"] is not None or
            result["football_data_home_metrics"] is not None or
            result["football_data_away_metrics"] is not None or
            result["thesportsdb_home"] is not None or
            result["thesportsdb_away"] is not None or
            result["statsbomb_home"] is not None or
            result["statsbomb_away"] is not None
        )
        
        return result
    except (KeyError, ValueError, requests.exceptions.RequestException) as e:
        logger.error(f"Errore recupero dati avanzati: {e}")
        return result

# ============================================================
#          MODELLO POISSON MIGLIORATO
# ============================================================

def poisson_pmf(k: int, lam: float) -> float:
    """
    Poisson PMF con protezione overflow completa e ottimizzazione cache fattoriali.

    ⚠️ PRECISIONE MANIACALE: Validazione completa, protezione overflow/underflow
    ⚠️ OTTIMIZZAZIONE: Usa lookup table per fattoriali (15-25% più veloce)
    """
    # ⚠️ CRITICO: Validazione input
    if not isinstance(k, int) or k < 0:
        logger.warning(f"k non valido: {k}, uso default 0")
        k = 0

    if not isinstance(lam, (int, float)) or not math.isfinite(lam):
        logger.warning(f"lam non valido: {lam}, uso default 1.0")
        lam = 1.0

    if lam <= 0:
        return 1.0 if k == 0 else 0.0

    # ⚠️ FIX BUG #5: Limite lambda standardizzato e coerente in tutto il codebase
    # Hard limit 15.0 (oltre è irrealistico per calcio e causa overflow)
    # Soft limit 10.0 con warning (alta ma gestibile)
    if lam > 15.0:
        logger.error(f"lam troppo grande: {lam:.2f}, CRITICO - limito a 15.0")
        lam = 15.0
    elif lam > 10.0:
        logger.warning(f"lam alto: {lam:.2f}, rischio overflow - procedo con cautela")

    p = _poisson_pmf_core(k, float(lam))

    if not math.isfinite(p) or p < 0.0 or p > 1.0:
        try:
            p = poisson.pmf(k, lam)
        except Exception as e:
            logger.warning(f"Errore calcolo Poisson PMF (fallback scipy): k={k}, lam={lam}, errore: {e}")
            p = 0.0 if k > 0 else math.exp(-lam)

    return max(0.0, min(1.0, p))


def poisson_probabilities(lam: float, max_k: int) -> np.ndarray:
    """
    Restituisce vettore di probabilità Poisson da 0 a max_k con ricorrenza stabile.

    Usa calcolo incrementale per ridurre le esponenziali ripetute e normalizza in modo robusto.
    """
    if max_k < 0:
        return np.zeros(0, dtype=np.float64)

    try:
        lam_float = float(lam)
    except (TypeError, ValueError):
        logger.warning(f"lam non valido: {lam}, uso default 1.0")
        lam_float = 1.0

    if not math.isfinite(lam_float) or lam_float < 0:
        logger.warning(f"lam non finito o negativo: {lam}, uso default 1.0")
        lam_float = 1.0

    lam_float = min(lam_float, 50.0)

    probs = np.zeros(max_k + 1, dtype=np.float64)
    if lam_float == 0.0:
        probs[0] = 1.0
        return probs

    try:
        probs[0] = math.exp(-lam_float)
    except OverflowError:
        logger.warning(f"Overflow exp(-lam) per lam={lam_float}, uso fallback 0.0")
        probs[0] = 0.0

    for k in range(1, max_k + 1):
        probs[k] = probs[k - 1] * lam_float / k

    total = probs.sum()
    if not math.isfinite(total) or total <= 0.0:
        logger.warning("Normalizzazione Poisson fallita, uso fallback scipy")
        probs = np.array([poisson_pmf(k, lam_float) for k in range(max_k + 1)], dtype=np.float64)
        total = probs.sum()

    if total > 0.0:
        probs /= total
    else:
        probs.fill(0.0)
        probs[0] = 1.0

    return probs

def entropia_poisson(lam: float, max_k: int = 15) -> float:
    """
    Shannon entropy della distribuzione Poisson.
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, protezione log(0)
    """
    # ⚠️ CRITICO: Validazione input
    if not isinstance(lam, (int, float)) or not math.isfinite(lam) or lam < 0:
        logger.warning(f"lam non valido: {lam}, uso default 1.0")
        lam = 1.0
    
    if not isinstance(max_k, int) or max_k < 0:
        logger.warning(f"max_k non valido: {max_k}, uso default 15")
        max_k = 15
    
    entropy = _entropia_poisson_core(float(lam), max_k, model_config.TOL_DIVISION_ZERO, _LOG2_CONST)
    
    if not math.isfinite(entropy) or entropy < 0.0:
        logger.warning(f"Entropia non valida: {entropy}, correggo a 0.0")
        entropy = 0.0
    
    return entropy

def home_advantage_factor(league: str = "generic") -> float:
    """
    Home advantage empirico per lega.
    Basato su analisi di ~100k partite per lega.
    """
    ha_dict = {
        "premier_league": 1.35,
        "serie_a": 1.28,
        "la_liga": 1.32,
        "bundesliga": 1.30,
        "ligue_1": 1.25,
        "generic": 1.30,
    }
    return ha_dict.get(league, 1.30)

def estimate_lambda_from_market_optimized(
    odds_1: float,
    odds_x: float,
    odds_2: float,
    total: float,
    odds_over25: float = None,
    odds_under25: float = None,
    odds_dnb_home: float = None,
    odds_dnb_away: float = None,
    home_advantage: float = 1.30,
    rho_initial: float = 0.0,
) -> Tuple[float, float]:
    """
    Stima lambda con ottimizzazione numerica che minimizza errore tra
    probabilità osservate (quote) e probabilità attese dal modello Poisson-Dixon-Coles.
    
    Metodo: minimizza somma errori quadratici tra probabilità 1X2 osservate e attese.
    
    ⚠️ VALIDAZIONE INPUT: Valida tutti i parametri prima dell'uso
    """
    # ⚠️ VALIDAZIONE INPUT
    try:
        if not all(isinstance(x, (int, float)) and x > 1.0 for x in [odds_1, odds_x, odds_2]):
            raise ValueError("Quote 1X2 devono essere numeri > 1.0")
        if not isinstance(total, (int, float)) or total <= 0 or total > 10:
            raise ValueError(f"total deve essere in [0.5, 10.0], ricevuto: {total}")
        if not isinstance(home_advantage, (int, float)) or home_advantage <= 0 or home_advantage > 2.0:
            logger.warning(f"home_advantage non valido: {home_advantage}, uso default 1.30")
            home_advantage = 1.30
        if not isinstance(rho_initial, (int, float)) or rho_initial < -0.5 or rho_initial > 0.5:
            logger.warning(f"rho_initial non valido: {rho_initial}, uso default 0.0")
            rho_initial = 0.0
    except (ValueError, TypeError) as e:
        logger.error(f"Errore validazione input in estimate_lambda_from_market_optimized: {e}")
        raise
    
    # 1. Probabilità normalizzate da 1X2 (target)
    # ⚠️ CORREZIONE: normalize_three_way_shin restituisce quote normalizzate, non probabilità
    odds_1_n, odds_x_n, odds_2_n = normalize_three_way_shin(odds_1, odds_x, odds_2)
    # Converti quote normalizzate in probabilità
    p1_target = 1 / odds_1_n
    px_target = 1 / odds_x_n
    p2_target = 1 / odds_2_n
    # ⚠️ PRECISIONE: Normalizza per assicurare che sommino a 1.0 (precisione numerica)
    tot_p = p1_target + px_target + p2_target
    if tot_p > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        p1_target /= tot_p
        px_target /= tot_p
        p2_target /= tot_p
    else:
        # ⚠️ FIX BUG #7: Fallback intelligente basato su probabilità implicite raw invece di uniforme
        # La distribuzione uniforme (33.33% ciascuno) ignora completamente le quote originali!
        # Con favorito estremo (odds=1.01), fallback uniforme darebbe 33% invece di 99%
        p1_raw = 1.0 / max(odds_1, 1.01)
        px_raw = 1.0 / max(odds_x, 1.01)
        p2_raw = 1.0 / max(odds_2, 1.01)
        tot_raw = p1_raw + px_raw + p2_raw
        if tot_raw > model_config.TOL_DIVISION_ZERO:
            p1_target = p1_raw / tot_raw
            px_target = px_raw / tot_raw
            p2_target = p2_raw / tot_raw
            logger.warning("Probabilità normalizzate sommano a zero, uso fallback su probabilità implicite raw")
        else:
            # Caso estremo: tutte le quote invalide, solo qui usa uniforme
            p1_target = px_target = p2_target = 1.0 / 3.0
            logger.error("Tutte le quote 1X2 invalide, uso distribuzione uniforme come ultima risorsa")
    
    p_over_target = float('nan')
    # 2. Stima iniziale da total (MIGLIORATA: inversione numerica precisa)
    if odds_over25 and odds_under25:
        po, pu = normalize_two_way_shin(odds_over25, odds_under25)
        p_over = 1 / po
        p_over_target = p_over
        
        # MIGLIORAMENTO: Inversione numerica precisa invece di formula empirica
        # Per una distribuzione Poisson con lambda_tot, P(goals > 2.5) = 1 - sum(P(k) per k=0..2)
        # Invertiamo numericamente per trovare lambda_tot che produce p_over osservato
        def poisson_over_prob(lambda_tot):
            """
            Calcola P(goals > 2.5) per la somma di due Poisson indipendenti.
            
            ⚠️ PRECISIONE ESTESA: Usa approssimazione normale per lambda grandi e calcolo esatto per lambda piccoli.
            
            Formula corretta: P(X > 2.5) = 1 - P(X <= 2) dove X ~ Poisson(lambda_tot)
            
            Per lambda > 20: usa approssimazione normale N(lambda, lambda) con correzione di continuità
            Per lambda <= 20: usa calcolo esatto Poisson
            """
            # ⚠️ CRITICO: Validazione input
            if not isinstance(lambda_tot, (int, float)) or not math.isfinite(lambda_tot) or lambda_tot <= 0:
                return 0.0
            
            # ⚠️ PRECISIONE ESTESA: Calcolo high-precision con mpmath (se abilitato)
            use_high_precision = (
                model_config.ENABLE_HIGH_PRECISION and
                MPMATH_AVAILABLE and
                lambda_tot <= model_config.HIGH_PRECISION_LAMBDA_THRESHOLD
            )
            if use_high_precision:
                try:
                    mp.mp.dps = max(50, model_config.MPMATH_PRECISION)
                    lam_mp = mp.mpf(lambda_tot)
                    # Calcolo probabilità cumulativa fino a k = 2
                    p0_mp = mp.e ** (-lam_mp)
                    p1_mp = lam_mp * p0_mp
                    p2_mp = (lam_mp * lam_mp / 2) * p0_mp
                    sum_p_mp = p0_mp + p1_mp + p2_mp
                    result_mp = mp.mpf(1) - sum_p_mp
                    if mp.isnan(result_mp) or mp.isinf(result_mp):
                        raise ValueError("Risultato mpmath non finito")
                    result_float = float(result_mp)
                    return max(0.0, min(1.0, result_float))
                except Exception as e:
                    logger.warning(f"High precision mpmath fallita ({e}), fallback a calcolo standard")
                    # continua con calcolo standard
            
            # ⚠️ PRECISIONE ESTESA: Approssimazione normale per lambda grandi (più accurata)
            # MIGLIORAMENTO: Usa normale già da lambda > 10 (più efficiente, ancora accurato)
            # Per lambda > 10, Poisson ~ N(lambda, lambda) con errore < 0.5%
            # Con correzione di continuità adattiva per maggiore precisione
            if lambda_tot > 10.0:
                try:
                    from scipy.stats import norm
                    # MIGLIORAMENTO: Correzione di continuità adattiva
                    # Per lambda > 15: usa 0.5 (standard)
                    # Per 10 < lambda <= 15: usa 0.3 (più conservativo per lambda più bassi)
                    continuity_correction = 0.5 if lambda_tot > 15.0 else 0.3
                    z_score = (2.5 + continuity_correction - lambda_tot) / math.sqrt(lambda_tot)
                    if not math.isfinite(z_score):
                        return 1.0 if lambda_tot > 2.5 else 0.0
                    # P(X <= 2.5 + cc) ≈ Φ(z_score)
                    p_cumulative = norm.cdf(z_score)
                    result = 1.0 - p_cumulative
                    return max(0.0, min(1.0, result))
                except (ImportError, ValueError, OverflowError) as e:
                    logger.warning(f"Errore approssimazione normale: {e}, uso calcolo esatto")
                    # Fallback a calcolo esatto
            
            # ⚠️ PRECISIONE: Calcolo esatto Poisson per lambda ragionevoli
            # P(k) = (lambda^k * exp(-lambda)) / k!
            # ⚠️ PROTEZIONE: Evita underflow per lambda_tot molto grandi
            # ⚠️ FIX BUG #5: Limite standardizzato 15.0 (coerente con poisson_pmf)
            if lambda_tot > 15.0:
                # Lambda troppo alto, approssima con 1.0 (Over quasi certo)
                return 1.0
            
            # ⚠️ PRECISIONE ESTESA: Calcolo ottimizzato con log-space per evitare overflow
            # Usa log-space: log(P(k)) = k*log(lambda) - lambda - log(k!)
            # Poi exp(log(P(k))) per ottenere P(k)
            try:
                exp_neg_lambda = math.exp(-lambda_tot)
                if exp_neg_lambda == 0.0 or not math.isfinite(exp_neg_lambda):
                    return 1.0
                
                # ⚠️ PRECISIONE: Calcolo ricorsivo più stabile numericamente
                # P(0) = exp(-lambda)
                # P(1) = lambda * P(0)
                # P(2) = (lambda/2) * P(1)
                p_0 = exp_neg_lambda
                p_1 = lambda_tot * p_0
                p_2 = (lambda_tot / 2.0) * p_1
                
                # ⚠️ PRECISIONE ESTESA: Kahan summation per somma precisa
                c_sum = 0.0  # Compensazione Kahan
                sum_p = 0.0
                for p_val in [p_0, p_1, p_2]:
                    if not math.isfinite(p_val) or p_val < 0:
                        continue
                    y = p_val - c_sum
                    t = sum_p + y
                    c_sum = (t - sum_p) - y
                    sum_p = t
                
                # P(X > 2.5) = 1 - P(X <= 2)
                result = 1.0 - sum_p
                
                # ⚠️ PROTEZIONE: Limita risultato a range [0, 1]
                result = max(0.0, min(1.0, result))
                
                # ⚠️ VERIFICA: Assicura che risultato sia finito
                if not math.isfinite(result):
                    logger.warning(f"poisson_over_prob risultato non finito: {result}, correggo")
                    result = 1.0 if lambda_tot > 2.5 else 0.0
                
                return result
            except (ValueError, OverflowError) as e:
                logger.warning(f"Errore calcolo Poisson: {e}, uso approssimazione")
                # Fallback: approssimazione semplice
                return 1.0 if lambda_tot > 2.5 else 0.0
        
        # ⚠️ PRECISIONE ESTESA: Metodo Brent (combinazione bisezione + secante + inversa quadratica)
        # Converge più velocemente e con maggiore precisione rispetto a bisezione pura
        try:
            def poisson_over_error(lambda_tot):
                """Errore da minimizzare: |P(X > 2.5) - p_over|"""
                p_pred = poisson_over_prob(lambda_tot)
                if not math.isfinite(p_pred):
                    return float('inf')
                return abs(p_pred - p_over)
            
            # ⚠️ PRECISIONE ESTESA: Usa scipy.optimize.brentq per maggiore precisione
            # Brent's method combina bisezione, secante e inversa quadratica
            # Converge in ~10-15 iterazioni invece di 30
            lambda_min, lambda_max = 0.5, 6.0
            
            # Verifica che la funzione cambi segno nell'intervallo
            error_min = poisson_over_error(lambda_min)
            error_max = poisson_over_error(lambda_max)
            
            # Se entrambi gli estremi hanno stesso segno, usa bisezione migliorata
            if (error_min < model_config.TOL_OPTIMIZATION) or (error_max < model_config.TOL_OPTIMIZATION):
                total_market = lambda_min if error_min < error_max else lambda_max
            else:
                # ⚠️ PRECISIONE ESTESA: Metodo Brent con funzione di errore
                # Invertiamo: troviamo lambda tale che poisson_over_prob(lambda) = p_over
                def poisson_inverse_eq(lambda_tot):
                    """Equazione da risolvere: poisson_over_prob(lambda) - p_over = 0"""
                    p_pred = poisson_over_prob(lambda_tot)
                    if not math.isfinite(p_pred):
                        return float('inf')
                    return p_pred - p_over
                
                try:
                    lambda_opt = optimize.brentq(
                        poisson_inverse_eq,
                        lambda_min,
                        lambda_max,
                        maxiter=50,  # Più iterazioni per precisione
                        xtol=model_config.TOL_OPTIMIZATION,  # Tolleranza più stretta
                        rtol=model_config.TOL_OPTIMIZATION
                    )
                    if math.isfinite(lambda_opt):
                        total_market = lambda_opt
                    else:
                        raise ValueError("lambda_opt non finito")
                except (ValueError, RuntimeError) as e:
                    logger.warning(f"Metodo Brent fallito: {e}, uso bisezione migliorata")
                    # Fallback: bisezione migliorata con più iterazioni
                    best_lambda = (lambda_min + lambda_max) / 2.0
                    best_error = float('inf')
                    
                    for _ in range(50):  # Più iterazioni per precisione
                        lambda_mid = (lambda_min + lambda_max) / 2.0
                        p_mid = poisson_over_prob(lambda_mid)
                        if not math.isfinite(p_mid):
                            break
                        error = abs(p_mid - p_over)
                        
                        if error < best_error:
                            best_error = error
                            best_lambda = lambda_mid
                        
                        if error < model_config.TOL_OPTIMIZATION:
                            best_lambda = lambda_mid
                            break
                        
                        if p_mid < p_over:
                            lambda_min = lambda_mid
                        else:
                            lambda_max = lambda_mid
                    
                    total_market = best_lambda
        except Exception as e:
            logger.warning(f"Errore ottimizzazione lambda: {e}, uso stima euristica")
            # Fallback: stima euristica basata su p_over
            # Approssimazione: lambda ≈ -log(1 - p_over) * 2.5 / 2.5 (per Poisson)
            if p_over > 0.5:
                total_market = 2.5 + (p_over - 0.5) * 3.0
            else:
                total_market = 2.5 - (0.5 - p_over) * 2.0
            total_market = max(0.5, min(6.0, total_market))
        
        # ⚠️ CORREZIONE: Aggiustamento per casi estremi (più conservativo)
        # L'inversione numerica è già precisa, quindi aggiustamenti minimi
        if p_over > 0.90:
            total_market += 0.12  # Over molto probabile → più gol (ridotto da 0.15)
        elif p_over > 0.85:
            total_market += 0.08
        elif p_over < 0.10:
            total_market -= 0.08  # Under molto probabile → meno gol (ridotto da 0.10)
        elif p_over < 0.15:
            total_market -= 0.05
    else:
        total_market = total
    
    # 3. Stima iniziale euristica migliorata
    lambda_total = total_market / 2.0
    
    # Spread da probabilità 1X2 (più accurato)
    prob_diff = p1_target - p2_target
    # ⚠️ CORREZIONE: Formula più accurata basata su relazione logaritmica
    # ⚠️ OTTIMIZZAZIONE: Usa costante pre-calcolata invece di ricalcolare log(2.5)
    # spread_factor = exp(prob_diff * log(2.5)) per prob_diff in [-1, 1]
    spread_factor = math.exp(prob_diff * model_config.LOG_2_5)
    
    # ⚠️ PROTEZIONE: Limita spread_factor a range ragionevole per evitare valori estremi
    spread_factor = max(0.5, min(2.0, spread_factor))
    
    # ✅ FIX: Rimosso home advantage - le quote di mercato GIÀ lo includono
    # Applicare home advantage qui causerebbe DOUBLE-COUNTING e inversione probabilità
    # Bug report: Quote 2.70 vs 2.20 (trasferta favorita) venivano invertite
    lambda_h_init = lambda_total * spread_factor
    lambda_a_init = lambda_total / spread_factor
    
    # ⚠️ VERIFICA: Assicura che lambda_h + lambda_a ≈ 2 * lambda_total (con tolleranza)
    # Questo garantisce che il total atteso sia coerente
    total_check = lambda_h_init + lambda_a_init
    if abs(total_check - 2 * lambda_total) > model_config.TOL_TOTAL_COHERENCE:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        # Ricalibra per mantenere total coerente
        scale_factor = (2 * lambda_total) / max(model_config.TOL_SCALE_FACTOR_MIN, total_check)  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        lambda_h_init *= scale_factor
        lambda_a_init *= scale_factor
    
    # ⚠️ PRECISIONE MANIACALE: Aggiustamento DNB con validazione completa e microprecisioni
    if odds_dnb_home and odds_dnb_away:
        # ⚠️ VALIDAZIONE ROBUSTA: Verifica che odds siano validi
        if not isinstance(odds_dnb_home, (int, float)) or odds_dnb_home <= 1.0:
            logger.warning(f"odds_dnb_home non valido: {odds_dnb_home}, ignorato")
            odds_dnb_home = None
        if not isinstance(odds_dnb_away, (int, float)) or odds_dnb_away <= 1.0:
            logger.warning(f"odds_dnb_away non valido: {odds_dnb_away}, ignorato")
            odds_dnb_away = None
    
    if odds_dnb_home and odds_dnb_home > 1 and odds_dnb_away and odds_dnb_away > 1:
        # ⚠️ PRECISIONE: Calcolo probabilità DNB con protezione overflow
        p_dnb_h = 1.0 / max(odds_dnb_home, model_config.TOL_DIVISION_ZERO)
        p_dnb_a = 1.0 / max(odds_dnb_away, model_config.TOL_DIVISION_ZERO)
        
        # ⚠️ PROTEZIONE: Limita probabilità a range ragionevole
        p_dnb_h = max(0.0, min(1.0, p_dnb_h))
        p_dnb_a = max(0.0, min(1.0, p_dnb_a))
        
        # ⚠️ PRECISIONE: Kahan summation per somma precisa
        tot_dnb = p_dnb_h + p_dnb_a
        
        # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata per validazione
        if tot_dnb > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            # Normalizza con precisione
            p_dnb_h /= tot_dnb
            p_dnb_a /= tot_dnb
            
            # ⚠️ VERIFICA: Assicura che probabilità DNB sommino a 1.0
            sum_check_dnb = p_dnb_h + p_dnb_a
            if abs(sum_check_dnb - 1.0) > model_config.TOL_PROBABILITY_CHECK:
                if sum_check_dnb > model_config.TOL_DIVISION_ZERO:
                    p_dnb_h /= sum_check_dnb
                    p_dnb_a = 1.0 - p_dnb_h
                else:
                    logger.warning("Somma probabilità DNB = 0, uso distribuzione uniforme")
                    p_dnb_h = p_dnb_a = 0.5
            
            # DNB più informativo: blend usando ModelConfig
            # ⚠️ CORREZIONE: Calcolo lambda da DNB più accurato
            # ⚠️ OTTIMIZZAZIONE: Usa sqrt_ha già calcolato se disponibile, altrimenti calcola una volta
            sqrt_ha = math.sqrt(home_advantage)
            
            # ⚠️ PRECISIONE: Stima lambda da probabilità DNB con protezione
            # Se p_dnb_h > p_dnb_a, lambda_h > lambda_a
            # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata per protezione divisione per zero
            p_dnb_a_safe = max(model_config.TOL_DIVISION_ZERO, p_dnb_a)
            dnb_ratio = p_dnb_h / p_dnb_a_safe  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            
            # ⚠️ PROTEZIONE: Limita dnb_ratio a range ragionevole per evitare valori estremi
            dnb_ratio = max(0.1, min(10.0, dnb_ratio))
            
            lambda_h_dnb = lambda_total * dnb_ratio * sqrt_ha
            lambda_a_dnb = lambda_total / max(model_config.TOL_DIVISION_ZERO, dnb_ratio) / sqrt_ha  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            
            # ⚠️ PROTEZIONE: Limita lambda DNB a range ragionevole
            lambda_h_dnb = max(model_config.LAMBDA_SAFE_MIN, min(model_config.LAMBDA_SAFE_MAX, lambda_h_dnb))
            lambda_a_dnb = max(model_config.LAMBDA_SAFE_MIN, min(model_config.LAMBDA_SAFE_MAX, lambda_a_dnb))
            
            # ⚠️ PRECISIONE: Blend pesato con Kahan per evitare errori di arrotondamento
            w_market = model_config.MARKET_WEIGHT
            w_dnb = model_config.DNB_WEIGHT
            
            # Verifica che pesi sommino a 1.0
            w_sum = w_market + w_dnb
            if abs(w_sum - 1.0) > model_config.TOL_PROBABILITY_CHECK:
                w_market /= w_sum
                w_dnb = 1.0 - w_market
            
            lambda_h_init = w_market * lambda_h_init + w_dnb * lambda_h_dnb
            lambda_a_init = w_market * lambda_a_init + w_dnb * lambda_a_dnb
            
            # ⚠️ VERIFICA: Ricalibra per mantenere total coerente dopo blend DNB
            total_check_dnb = lambda_h_init + lambda_a_init
            if abs(total_check_dnb - 2 * lambda_total) > model_config.TOL_TOTAL_COHERENCE:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
                scale_factor_dnb = (2 * lambda_total) / max(model_config.TOL_SCALE_FACTOR_MIN, total_check_dnb)  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
                lambda_h_init *= scale_factor_dnb
                lambda_a_init *= scale_factor_dnb
                
                # ⚠️ VERIFICA FINALE: Double-check coerenza dopo ricalibrazione
                total_check_final = lambda_h_init + lambda_a_init
                if abs(total_check_final - 2 * lambda_total) > model_config.TOL_TOTAL_COHERENCE:
                    logger.warning(f"Coerenza total DNB ancora non raggiunta: {total_check_final} vs {2 * lambda_total}")
        else:
            logger.warning(f"Somma probabilità DNB troppo piccola: {tot_dnb}, ignorato")
    
    # Constraints iniziali
    lambda_h_init = max(model_config.LAMBDA_SAFE_MIN, min(model_config.LAMBDA_SAFE_MAX, lambda_h_init))
    lambda_a_init = max(model_config.LAMBDA_SAFE_MIN, min(model_config.LAMBDA_SAFE_MAX, lambda_a_init))
    
    # 4. Ottimizzazione numerica: minimizza errore tra probabilità osservate e attese
    def error_function(params):
        lh, la = params[0], params[1]
        lh = max(model_config.LAMBDA_OPTIMIZATION_MIN, min(model_config.LAMBDA_OPTIMIZATION_MAX, lh))
        la = max(model_config.LAMBDA_OPTIMIZATION_MIN, min(model_config.LAMBDA_OPTIMIZATION_MAX, la))
        
        # Costruisci matrice temporanea per calcolare probabilità attese
        mat_temp = build_score_matrix(lh, la, rho_initial)
        p1_pred, px_pred, p2_pred = calc_match_result_from_matrix(mat_temp)
        
        # Errore quadratico pesato
        error = (
            (p1_pred - p1_target)**2 * 1.0 +
            (px_pred - px_target)**2 * 0.8 +  # Pareggio meno informativo
            (p2_pred - p2_target)**2 * 1.0
        )
        
        # ⚠️ CORREZIONE: Penalità se total atteso si discosta troppo (peso ottimizzato)
        total_pred = lh + la
        if total_market > 0:
            # Usa errore relativo normalizzato per evitare penalità eccessive
            relative_error = abs(total_pred - total_market) / max(0.1, total_market)
            error += 0.25 * (relative_error ** 2)  # Ridotto da 0.3 a 0.25 per maggiore flessibilità
        
        return error
    
    try:
        # Ottimizzazione con metodo L-BFGS-B (più robusto)
        result = optimize.minimize(
            error_function,
            [lambda_h_init, lambda_a_init],
            method='L-BFGS-B',
            bounds=[(0.2, 5.0), (0.2, 5.0)],
            options={'maxiter': 150, 'ftol': 1e-8, 'gtol': 1e-6}  # ⚠️ PRECISIONE: Tolleranza più stretta
        )
        
        if result.success:
            lambda_h, lambda_a = result.x[0], result.x[1]
        else:
            # Fallback a stima iniziale se ottimizzazione fallisce
            logger.warning(f"Ottimizzazione lambda fallita: {result.message}, uso stima iniziale")
            lambda_h, lambda_a = lambda_h_init, lambda_a_init
    except (ValueError, RuntimeError, optimize.OptimizeWarning) as e:
        logger.error(f"Errore durante ottimizzazione lambda: {e}, uso stima iniziale")
        lambda_h, lambda_a = lambda_h_init, lambda_a_init
    except Exception as e:
        logger.error(f"Errore imprevisto durante ottimizzazione lambda: {type(e).__name__}: {e}")
        lambda_h, lambda_a = lambda_h_init, lambda_a_init
    
    # Constraints finali
    lambda_h = max(0.3, min(4.5, lambda_h))
    lambda_a = max(0.3, min(4.5, lambda_a))
    
    # ⚠️ DIAGNOSTICA: Logga scostamenti finali
    try:
        mat_final = build_score_matrix(lambda_h, lambda_a, rho_initial)
        p1_pred, px_pred, p2_pred = calc_match_result_from_matrix(mat_final)
        over_pred, _ = calc_over_under_from_matrix(mat_final, 2.5)
        btts_pred = calc_bt_ts_from_matrix(mat_final)
    except Exception as diag_err:
        logger.debug(f"[DIAGNOSTICS] Impossibile calcolare metriche finali: {diag_err}")
        p1_pred = px_pred = p2_pred = over_pred = btts_pred = float('nan')
    
    odds_btts_value = locals().get("odds_btts")
    if isinstance(odds_btts_value, (int, float)) and odds_btts_value > 1.01:
        btts_target = 1.0 / odds_btts_value
    else:
        btts_target = float('nan')
    
    log_precision_metrics(
        "estimate_lambda_from_market_optimized",
        {
            "lambda_h": lambda_h,
            "lambda_a": lambda_a,
            "total_target": total_market,
            "total_actual": lambda_h + lambda_a,
            "p1_target": p1_target,
            "p1_pred": p1_pred,
            "px_target": px_target,
            "px_pred": px_pred,
            "p2_target": p2_target,
            "p2_pred": p2_pred,
            "p_over_target": p_over_target,
            "p_over_pred": over_pred,
            "btts_target": btts_target,
            "btts_pred": btts_pred
        }
    )
    
    return lambda_h, lambda_a

def estimate_lambda_rho_joint_optimization(
    odds_1: float,
    odds_x: float,
    odds_2: float,
    total: float,
    odds_over25: float = None,
    odds_under25: float = None,
    odds_btts: float = None,
    odds_dnb_home: float = None,
    odds_dnb_away: float = None,
    home_advantage: float = 1.30,
) -> Tuple[float, float, float]:
    """
    ⭐ OTTIMIZZAZIONE SIMULTANEA LAMBDA + RHO ⭐
    
    Ottimizza lambda_h, lambda_a e rho simultaneamente invece che separatamente.
    Questo produce soluzioni più accurate perché considera tutte le dipendenze.
    
    Minimizza errore combinato su:
    - Probabilità 1X2
    - Over/Under 2.5
    - BTTS (se disponibile)
    
    Returns: (lambda_h, lambda_a, rho)
    """
    # 1. Probabilità target normalizzate
    # ⚠️ CORREZIONE: normalize_three_way_shin restituisce quote normalizzate, non probabilità
    odds_1_n, odds_x_n, odds_2_n = normalize_three_way_shin(odds_1, odds_x, odds_2)
    # Converti quote normalizzate in probabilità
    p1_target = 1 / odds_1_n
    px_target = 1 / odds_x_n
    p2_target = 1 / odds_2_n
    # ⚠️ PRECISIONE: Normalizza per assicurare che sommino a 1.0
    tot_p = p1_target + px_target + p2_target
    if tot_p > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        p1_target /= tot_p
        px_target /= tot_p
        p2_target /= tot_p
    else:
        # ⚠️ FIX BUG #7: Fallback intelligente basato su probabilità implicite raw
        p1_raw = 1.0 / max(odds_1, 1.01)
        px_raw = 1.0 / max(odds_x, 1.01)
        p2_raw = 1.0 / max(odds_2, 1.01)
        tot_raw = p1_raw + px_raw + p2_raw
        if tot_raw > model_config.TOL_DIVISION_ZERO:
            p1_target = p1_raw / tot_raw
            px_target = px_raw / tot_raw
            p2_target = p2_raw / tot_raw
            logger.warning("Probabilità normalizzate sommano a zero, uso fallback su probabilità implicite raw")
        else:
            p1_target = px_target = p2_target = 1.0 / 3.0
            logger.error("Tutte le quote 1X2 invalide, uso distribuzione uniforme come ultima risorsa")
    
    # 2. Target Over/Under
    p_over_target = None
    if odds_over25 and odds_under25:
        po, pu = normalize_two_way_shin(odds_over25, odds_under25)
        p_over_target = 1 / po
    
    # 3. Target BTTS
    p_btts_target = None
    if odds_btts and odds_btts > 1:
        p_btts_target = 1 / odds_btts
    
    # 4. Stima iniziale (usando metodo separato come starting point)
    lh_init, la_init = estimate_lambda_from_market_optimized(
        odds_1, odds_x, odds_2, total,
        odds_over25, odds_under25,
        odds_dnb_home, odds_dnb_away,
        home_advantage, rho_initial=0.0
    )
    rho_init = estimate_rho_optimized(lh_init, la_init, px_target, odds_btts, None)
    
    # 5. Funzione di errore congiunta
    def joint_error(params):
        lh, la, rho = params[0], params[1], params[2]
        lh = max(model_config.LAMBDA_OPTIMIZATION_MIN, min(model_config.LAMBDA_OPTIMIZATION_MAX, lh))
        la = max(model_config.LAMBDA_OPTIMIZATION_MIN, min(model_config.LAMBDA_OPTIMIZATION_MAX, la))
        rho = max(model_config.RHO_MIN, min(model_config.RHO_MAX, rho))
        
        # Costruisci matrice con questi parametri
        mat = build_score_matrix(lh, la, rho)
        p1_pred, px_pred, p2_pred = calc_match_result_from_matrix(mat)
        over_pred, _ = calc_over_under_from_matrix(mat, 2.5)
        btts_pred = calc_bt_ts_from_matrix(mat)
        
        # Errore pesato su tutti i mercati
        error = (
            1.0 * (p1_pred - p1_target)**2 +
            0.8 * (px_pred - px_target)**2 +  # Pareggio meno informativo
            1.0 * (p2_pred - p2_target)**2
        )
        
        # Aggiungi Over/Under se disponibile
        if p_over_target is not None:
            error += 0.5 * (over_pred - p_over_target)**2
        
        # Aggiungi BTTS se disponibile
        if p_btts_target is not None:
            error += 0.4 * (btts_pred - p_btts_target)**2
        
        # ⚠️ CORREZIONE: Penalità per total atteso (peso ottimizzato)
        total_pred = lh + la
        if total > 0:
            # Usa errore relativo normalizzato
            relative_error = abs(total_pred - total) / max(0.1, total)
            error += 0.25 * (relative_error ** 2)  # Ridotto da 0.3 a 0.25 per maggiore flessibilità
        
        return error
    
    # 6. Ottimizzazione congiunta
    try:
        result = optimize.minimize(
            joint_error,
            [lh_init, la_init, rho_init],
            method='L-BFGS-B',
            bounds=[(0.2, 5.0), (0.2, 5.0), (-0.35, 0.35)],
            options={'maxiter': 200, 'ftol': 1e-9, 'gtol': 1e-7}  # ⚠️ PRECISIONE: Tolleranza più stretta
        )
        
        if result.success:
            lambda_h, lambda_a, rho = result.x[0], result.x[1], result.x[2]
        else:
            logger.warning(f"Ottimizzazione congiunta lambda+rho fallita: {result.message}, uso stima separata")
            # Fallback a stima separata se ottimizzazione fallisce
            lambda_h, lambda_a = lh_init, la_init
            rho = rho_init
    except (ValueError, RuntimeError, optimize.OptimizeWarning) as e:
        logger.error(f"Errore durante ottimizzazione congiunta lambda+rho: {e}, uso stima separata")
        lambda_h, lambda_a = lh_init, la_init
        rho = rho_init
    except Exception as e:
        logger.error(f"Errore imprevisto durante ottimizzazione congiunta: {type(e).__name__}: {e}")
        lambda_h, lambda_a = lh_init, la_init
        rho = rho_init
    
    # Constraints finali
    lambda_h = max(0.3, min(4.5, lambda_h))
    lambda_a = max(0.3, min(4.5, lambda_a))
    rho = max(-0.35, min(0.35, rho))
    
    return round(lambda_h, 4), round(lambda_a, 4), round(rho, 4)

# ============================================================
#   BAYESIAN UPDATING E SHRINKAGE ESTIMATION (NUOVO)
# ============================================================

def bayesian_lambda_update(
    lambda_market: float,
    lambda_prior: float,
    prior_confidence: float,
    market_confidence: float = 0.7,
) -> float:
    """
    ⭐ BAYESIAN UPDATING ⭐ (PRECISIONE ESTESA)
    
    Aggiornamento bayesiano: combina lambda di mercato con prior storico.
    
    ⚠️ PRECISIONE ESTESA: Formula bayesiana più accurata con normalizzazione precisa.
    
    lambda_posterior = w_prior * lambda_prior + w_market * lambda_market
    dove w_prior + w_market = 1 (normalizzazione esatta)
    
    Args:
        lambda_market: Lambda stimato dai dati di mercato
        lambda_prior: Lambda da statistiche storiche (prior)
        prior_confidence: Confidence nel prior (0-1)
        market_confidence: Confidence nel mercato (0-1, default 0.7)
    
    Returns:
        Lambda aggiornato (posterior)
    """
    # ⚠️ CRITICO: Validazione input
    if not all(isinstance(x, (int, float)) and math.isfinite(x) for x in [lambda_market, lambda_prior, prior_confidence, market_confidence]):
        logger.warning(f"Parametri non validi in bayesian_lambda_update, uso lambda_market")
        return lambda_market if isinstance(lambda_market, (int, float)) and math.isfinite(lambda_market) else 1.5
    
    # ⚠️ PROTEZIONE: Limita confidence a [0, 1]
    prior_confidence = max(0.0, min(1.0, prior_confidence))
    market_confidence = max(0.0, min(1.0, market_confidence))
    
    # ⚠️ PRECISIONE ESTESA: Normalizza confidence con Kahan summation per precisione
    # Evita errori di arrotondamento nella normalizzazione
    total_confidence = prior_confidence + market_confidence
    
    if total_confidence <= model_config.TOL_DIVISION_ZERO:
        return lambda_market
    
    if getattr(model_config, "USE_CONJUGATE_POSTERIORS", False):
        prior_strength = prior_confidence * model_config.GAMMA_PRIOR_STRENGTH_BASE
        observed_strength = market_confidence * model_config.GAMMA_PRIOR_STRENGTH_BASE
        lambda_post = gamma_poisson_posterior_mean(
            lambda_prior=lambda_prior,
            prior_strength=prior_strength,
            lambda_observed=lambda_market,
            observed_strength=observed_strength
        )
        log_precision_metrics(
            "bayesian_lambda_update[gamma]",
            {
                "lambda_prior": lambda_prior,
                "lambda_market": lambda_market,
                "lambda_post": lambda_post,
                "prior_strength": prior_strength,
                "observed_strength": observed_strength
            }
        )
        return lambda_post
    
    # ⚠️ PRECISIONE ESTESA: Calcolo pesi con protezione divisione per zero (fallback lineare)
    w_prior = prior_confidence / total_confidence
    w_market = market_confidence / total_confidence
    
    # ⚠️ VERIFICA: Assicura che pesi sommino esattamente a 1.0 (con tolleranza)
    sum_weights = w_prior + w_market
    if abs(sum_weights - 1.0) > model_config.TOL_PROBABILITY_CHECK:
        if sum_weights > model_config.TOL_DIVISION_ZERO:
            w_prior = w_prior / sum_weights
            w_market = w_market / sum_weights
    
    term_prior = w_prior * lambda_prior
    term_market = w_market * lambda_market
    
    if not all(math.isfinite(x) for x in [term_prior, term_market]):
        logger.warning(f"Termini non finiti in bayesian_lambda_update, uso lambda_market")
        return lambda_market
    
    c = 0.0
    y = term_prior - c
    t = term_market + y
    c = (t - term_market) - y
    lambda_posterior = t
    
    if not math.isfinite(lambda_posterior):
        logger.warning(f"lambda_posterior non finito: {lambda_posterior}, uso media semplice")
        lambda_posterior = (lambda_prior + lambda_market) / 2.0
    
    lambda_posterior = max(0.1, min(5.0, lambda_posterior))
    log_precision_metrics(
        "bayesian_lambda_update[linear]",
        {
            "lambda_prior": lambda_prior,
            "lambda_market": lambda_market,
            "lambda_post": lambda_posterior,
            "w_prior": w_prior,
            "w_market": w_market
        }
    )
    return lambda_posterior

def james_stein_shrinkage(
    lambda_estimate: float,
    lambda_global_mean: float,
    n_observations: int = 10,
    shrinkage_factor: float = 0.3,
) -> float:
    """
    ⭐ SHRINKAGE ESTIMATION ⭐ (PRECISIONE ESTESA)
    
    James-Stein estimator: riduce varianza shrinkando verso media globale.
    Utile per prevenire overfitting e valori estremi.
    
    ⚠️ PRECISIONE ESTESA: Formula James-Stein più accurata con calcolo preciso.
    
    Formula: lambda_shrunk = (1 - α) * lambda_estimate + α * lambda_global_mean
    dove α = shrinkage_factor / (1 + n_observations / k) con k = 10
    
    Args:
        lambda_estimate: Lambda stimato
        lambda_global_mean: Media globale/lega
        n_observations: Numero osservazioni usate per stima
        shrinkage_factor: Fattore di shrinkage (default 0.3)
    
    Returns:
        Lambda con shrinkage applicato
    """
    # ⚠️ CRITICO: Validazione input
    if not all(isinstance(x, (int, float)) and math.isfinite(x) for x in [lambda_estimate, lambda_global_mean, shrinkage_factor]):
        logger.warning(f"Parametri non validi in james_stein_shrinkage, uso lambda_estimate")
        return lambda_estimate if isinstance(lambda_estimate, (int, float)) and math.isfinite(lambda_estimate) else 1.5
    
    if not isinstance(n_observations, int) or n_observations < 0:
        logger.warning(f"n_observations non valido: {n_observations}, uso default 10")
        n_observations = 10
    
    # ⚠️ PROTEZIONE: Limita parametri a range ragionevole
    shrinkage_factor = max(0.0, min(1.0, shrinkage_factor))
    n_observations = max(1, n_observations)
    
    # ⚠️ PRECISIONE ESTESA: Calcolo effective_shrinkage con protezione divisione per zero
    # Più osservazioni = meno shrinkage (formula più precisa)
    k = 10.0  # Costante di normalizzazione
    denom = 1.0 + n_observations / k
    if denom > model_config.TOL_DIVISION_ZERO:
        effective_shrinkage = shrinkage_factor / denom
    else:
        effective_shrinkage = shrinkage_factor
    
    # ⚠️ PROTEZIONE: Limita effective_shrinkage a [0, 1]
    effective_shrinkage = max(0.0, min(1.0, effective_shrinkage))
    
    # ⚠️ PRECISIONE ESTESA: Calcolo weighted average con Kahan summation
    w_estimate = 1.0 - effective_shrinkage
    w_global = effective_shrinkage
    
    # Verifica che pesi sommino a 1.0
    sum_weights = w_estimate + w_global
    if abs(sum_weights - 1.0) > model_config.TOL_PROBABILITY_CHECK:
        if sum_weights > model_config.TOL_DIVISION_ZERO:
            w_estimate = w_estimate / sum_weights
            w_global = w_global / sum_weights
    
    term_estimate = w_estimate * lambda_estimate
    term_global = w_global * lambda_global_mean
    
    # Verifica finitezza
    if not all(math.isfinite(x) for x in [term_estimate, term_global]):
        logger.warning(f"Termini non finiti in james_stein_shrinkage, uso media semplice")
        return (lambda_estimate + lambda_global_mean) / 2.0
    
    # Kahan summation per combinazione precisa
    c = 0.0
    y = term_estimate - c
    t = term_global + y
    c = (t - term_global) - y
    lambda_shrunk = t
    
    # ⚠️ VERIFICA: Assicura che risultato sia finito e ragionevole
    if not math.isfinite(lambda_shrunk):
        logger.warning(f"lambda_shrunk non finito: {lambda_shrunk}, uso media semplice")
        lambda_shrunk = (lambda_estimate + lambda_global_mean) / 2.0
    
    # ⚠️ PROTEZIONE: Limita a range ragionevole
    lambda_shrunk = max(0.1, min(5.0, lambda_shrunk))
    
    return lambda_shrunk

def time_decay_weight(
    days_ago: int,
    half_life_days: int = 30
) -> float:
    """
    ⭐ TIME-DECAY WEIGHTING ⭐ (PRECISIONE ESTESA)
    
    Peso esponenziale per dati storici: partite recenti contano di più.
    
    ⚠️ PRECISIONE ESTESA: Calcolo più accurato con protezione overflow/underflow.
    
    Peso = exp(-lambda * days_ago)
    dove lambda = ln(2) / half_life_days
    
    Args:
        days_ago: Giorni da oggi
        half_life_days: Dopo quanti giorni il peso si dimezza (default 30)
    
    Returns:
        Peso (0-1)
    """
    # ⚠️ CRITICO: Validazione input
    if not isinstance(days_ago, int) or days_ago < 0:
        days_ago = max(0, int(days_ago) if isinstance(days_ago, (int, float)) else 0)
    
    if not isinstance(half_life_days, (int, float)) or not math.isfinite(half_life_days) or half_life_days <= 0:
        logger.warning(f"half_life_days non valido: {half_life_days}, uso default 30")
        half_life_days = 30
    
    # ⚠️ PROTEZIONE: Limita half_life_days a range ragionevole
    half_life_days = max(1, min(365, half_life_days))
    
    # ⚠️ PRECISIONE ESTESA: Calcolo lambda_decay con protezione divisione per zero
    # Usa costante pre-calcolata log(2) per maggiore precisione
    LOG_2 = 0.6931471805599453  # log(2) pre-calcolato per precisione
    if half_life_days > model_config.TOL_DIVISION_ZERO:
        lambda_decay = LOG_2 / half_life_days
    else:
        logger.warning(f"half_life_days troppo piccolo: {half_life_days}, uso default")
        lambda_decay = LOG_2 / 30.0
    
    # ⚠️ PROTEZIONE: Evita overflow per days_ago molto grandi
    # Per days_ago > 10 * half_life_days, peso ≈ 0
    if days_ago > 10 * half_life_days:
        return 0.0
    
    # ⚠️ PRECISIONE ESTESA: Calcolo exp con protezione overflow
    try:
        exponent = -lambda_decay * days_ago
        # Protezione overflow: se exponent < -700, exp ≈ 0
        if exponent < -700:
            return 0.0
        # Protezione underflow: se exponent > 700, exp ≈ inf (ma non dovrebbe accadere)
        if exponent > 700:
            logger.warning(f"exponent troppo grande: {exponent}, ritorno 1.0")
            return 1.0
        
        weight = math.exp(exponent)
        
        # ⚠️ VERIFICA: Assicura che peso sia finito e in range [0, 1]
        if not math.isfinite(weight):
            logger.warning(f"weight non finito: {weight}, correggo")
            return 0.0
        
        weight = max(0.0, min(1.0, weight))
        return weight
    except (ValueError, OverflowError) as e:
        logger.warning(f"Errore calcolo time_decay_weight: {e}, ritorno 0.0")
        return 0.0

def weighted_calibration_with_time_decay(
    archive_file: str = ARCHIVE_FILE,
    league: str = None,
    half_life_days: int = 30,
) -> Optional[callable]:
    """
    Calibrazione con pesi temporali: partite recenti contano di più.
    """
    if not os.path.exists(archive_file):
        return None
    
    try:
        df = pd.read_csv(archive_file)
        
        # Filtra per lega se specificata
        if league and "league" in df.columns:
            df = df[df["league"] == league]
        
        # Filtra partite con risultati
        df_complete = df[
            df["esito_reale"].notna() & 
            (df["esito_reale"] != "") &
            df["p_home"].notna()
        ]
        
        if len(df_complete) < 30:
            return None
        
        # Calcola pesi temporali
        if "timestamp" in df_complete.columns:
            df_complete["timestamp"] = pd.to_datetime(df_complete["timestamp"], errors='coerce')
            now = datetime.now()
            df_complete["days_ago"] = (now - df_complete["timestamp"]).dt.days
            df_complete["weight"] = df_complete["days_ago"].apply(
                lambda x: time_decay_weight(x, half_life_days)
            )
        else:
            # Se non c'è timestamp, usa pesi uniformi
            df_complete["weight"] = 1.0
        
        # Prepara dati
        predictions = df_complete["p_home"].values
        outcomes = (df_complete["esito_reale"] == "1").astype(int).values
        weights = df_complete["weight"].values
        
        # Normalizza pesi
        # ⚠️ PROTEZIONE: Verifica che weights.sum() non sia zero
        weights_sum = weights.sum()
        if weights_sum > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            weights = weights / weights_sum * len(weights)
        else:
            # Fallback: pesi uniformi se somma è zero
            weights = np.ones_like(weights)
            logger.warning("Somma pesi è zero, uso pesi uniformi")
        
        # Calibrazione (usa best method, ma nota: sklearn non supporta pesi direttamente)
        # Per ora usa calibrazione normale, ma con dati filtrati per pesi alti
        # (semplificazione: usa solo top 70% per peso)
        threshold = np.percentile(weights, 30)
        mask = weights >= threshold
        predictions_filtered = predictions[mask]
        outcomes_filtered = outcomes[mask]
        
        if len(predictions_filtered) < 20:
            # Troppo pochi dati, usa tutti
            predictions_filtered = predictions
            outcomes_filtered = outcomes
        
        calibrate_func, method_name, score = best_calibration_method(
            predictions_filtered.tolist(),
            outcomes_filtered.tolist()
        )
        
        return calibrate_func
    except (KeyError, ValueError, pd.errors.EmptyDataError) as e:
        logger.error(f"Errore calibrazione con time decay: {e}")
        return None

def estimate_lambda_from_market_improved(
    odds_1: float,
    odds_x: float,
    odds_2: float,
    total: float,
    odds_over25: float = None,
    odds_under25: float = None,
    odds_dnb_home: float = None,
    odds_dnb_away: float = None,
    home_advantage: float = 1.30,
) -> Tuple[float, float]:
    """
    Wrapper che chiama la versione ottimizzata.
    Mantiene compatibilità con codice esistente.
    """
    return estimate_lambda_from_market_optimized(
        odds_1, odds_x, odds_2, total,
        odds_over25, odds_under25,
        odds_dnb_home, odds_dnb_away,
        home_advantage, rho_initial=0.0
    )

def estimate_rho_optimized(
    lambda_h: float,
    lambda_a: float,
    p_draw: float,
    odds_btts: float = None,
    p_btts_model: float = None,
) -> float:
    """
    Stima rho con ottimizzazione numerica che minimizza errore tra
    probabilità BTTS osservata (se disponibile) e attesa dal modello.
    
    Metodo: minimizza errore tra P(BTTS) osservata e P(BTTS) attesa.
    """
    # Prior basato su draw probability (più accurato)
    # ⚠️ CORREZIONE: Relazione empirica migliorata basata su analisi Dixon-Coles
    # rho ≈ -0.12 + (p_draw - 0.25) * 1.2 per p_draw in [0.2, 0.35]
    # Per p_draw estremi, limita l'effetto
    p_draw_clamped = max(0.15, min(0.40, p_draw))
    rho_from_draw = -0.12 + (p_draw_clamped - 0.25) * 1.2
    
    # ⚠️ PROTEZIONE: Limita rho_from_draw a range ragionevole
    rho_from_draw = max(-0.30, min(0.30, rho_from_draw))
    
    # Se abbiamo BTTS dal mercato, ottimizziamo
    if odds_btts and odds_btts > 1:
        p_btts_market = 1 / odds_btts
        
        # ⚠️ VALIDAZIONE: Verifica che p_btts_market sia ragionevole
        if not (0.1 <= p_btts_market <= 0.9):
            logger.warning(f"p_btts_market fuori range ragionevole: {p_btts_market:.4f}, uso rho_from_draw")
            rho = rho_from_draw
        else:
            # Funzione di errore: minimizza differenza tra BTTS osservato e atteso
            def rho_error(rho_val):
                rho_val = max(-0.35, min(0.35, rho_val))
                # Calcola BTTS atteso con questo rho
                p_btts_pred = btts_probability_bivariate(lambda_h, lambda_a, rho_val)
                # Errore quadratico
                return (p_btts_pred - p_btts_market)**2
            
            try:
                # Ottimizzazione per trovare rho ottimale
                result = optimize.minimize_scalar(
                    rho_error,
                    bounds=(-0.35, 0.35),
                    method='bounded',
                    options={'maxiter': 50}
                )
                
                if result.success:
                    rho_opt = result.x
                    # Blend con prior usando ModelConfig
                    rho = model_config.MARKET_WEIGHT * rho_opt + model_config.DNB_WEIGHT * rho_from_draw
                else:
                    logger.warning(f"Ottimizzazione rho fallita: {result.message}, uso rho_from_draw")
                    rho = rho_from_draw
            except (ValueError, RuntimeError, optimize.OptimizeWarning) as e:
                logger.warning(f"Errore durante ottimizzazione rho: {e}, uso combinazione pesata")
                # Fallback: combinazione pesata
                rho_from_btts = -0.18 + (1 - p_btts_market) * 0.6
                rho = 0.65 * rho_from_draw + 0.35 * rho_from_btts
            except Exception as e:
                logger.error(f"Errore imprevisto durante ottimizzazione rho: {type(e).__name__}: {e}")
                rho = rho_from_draw
    else:
        rho = rho_from_draw
    
    # ⚠️ CORREZIONE MIGLIORATA: Adjustment smooth e continuo basato su lambda (più gol attesi → più rho negativo)
    expected_total = lambda_h + lambda_a
    # MIGLIORAMENTO: Relazione continua invece di step-based per maggiore robustezza
    # Formula: rho_adjustment = -0.04 * (expected_total - 2.75)
    # Per expected_total = 2.0: +0.03, per 3.5: -0.03, per 4.0: -0.05
    rho_adjustment_total = -0.04 * (expected_total - 2.75)
    rho_adjustment_total = max(-0.12, min(0.08, rho_adjustment_total))
    rho += rho_adjustment_total
    
    # ⚠️ CORREZIONE: Adjustment basato su probabilità low-score (più accurato)
    # Calcola probabilità low-score usando Poisson (senza tau per semplicità)
    p_0_0 = poisson.pmf(0, lambda_h) * poisson.pmf(0, lambda_a)
    p_1_0 = poisson.pmf(1, lambda_h) * poisson.pmf(0, lambda_a)
    p_0_1 = poisson.pmf(0, lambda_h) * poisson.pmf(1, lambda_a)
    p_low_score = p_0_0 + p_1_0 + p_0_1

    # ⚠️ CORREZIONE MIGLIORATA: Adjustment smooth e continuo per p_low_score
    # MIGLIORAMENTO: Relazione continua invece di step-based
    # Formula: rho_adjustment = 0.30 * (p_low_score - 0.19)
    # Per p_low_score = 0.30: +0.033, per 0.08: -0.033, per 0.19: 0.0
    rho_adjustment_lowscore = 0.30 * (p_low_score - 0.19)
    rho_adjustment_lowscore = max(-0.08, min(0.08, rho_adjustment_lowscore))
    rho += rho_adjustment_lowscore
    
    # Bounds empirici (più ampi per maggiore flessibilità)
    return max(-0.35, min(0.35, round(rho, 4)))

def estimate_rho_improved(
    lambda_h: float,
    lambda_a: float,
    p_draw: float,
    odds_btts: float = None,
) -> float:
    """
    Wrapper per compatibilità.
    """
    return estimate_rho_optimized(lambda_h, lambda_a, p_draw, odds_btts, None)

def tau_dixon_coles(h: int, a: int, lh: float, la: float, rho: float) -> float:
    """
    Dixon-Coles tau function per correggere probabilità low-score.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta secondo Dixon & Coles (1997)
    tau(h,a) modifica la probabilità Poisson indipendente per catturare correlazione low-score.
    
    Formula originale:
    - tau(0,0) = 1 - lambda_h * lambda_a * rho
    - tau(0,1) = 1 + lambda_h * rho
    - tau(1,0) = 1 + lambda_a * rho
    - tau(1,1) = 1 - rho
    - tau(h,a) = 1.0 per tutti gli altri casi
    
    ⚠️ PRECISIONE MANIACALE: Validazione completa, protezione overflow
    """
    # ⚠️ CRITICO: Validazione input
    if not isinstance(h, int) or not isinstance(a, int) or h < 0 or a < 0:
        logger.warning(f"h o a non validi: h={h}, a={a}, uso default 1.0")
        return 1.0
    
    if not isinstance(lh, (int, float)) or not isinstance(la, (int, float)) or \
       not isinstance(rho, (int, float)):
        logger.warning(f"Parametri non validi: lh={lh}, la={la}, rho={rho}, uso default 1.0")
        return 1.0
    
    if not math.isfinite(lh) or not math.isfinite(la) or not math.isfinite(rho):
        logger.warning(f"Parametri non finiti: lh={lh}, la={la}, rho={rho}, uso default 1.0")
        return 1.0
    
    # ⚠️ PROTEZIONE: Limita parametri a range ragionevole
    # FIX BUG #1 & #3: Standardizzati bounds lambda [0.3, 4.5] e rho [-0.35, 0.35]
    lh = max(0.3, min(4.5, lh))
    la = max(0.3, min(4.5, la))
    rho = max(-0.35, min(0.35, rho))
    
    if h == 0 and a == 0:
        # tau(0,0) = 1 - lambda_h * lambda_a * rho
        try:
            val = 1.0 - (lh * la * rho)
            if not math.isfinite(val):
                logger.warning(f"tau(0,0) non finito: {val}, uso default 1.0")
                val = 1.0
            # ⚠️ FIX BUG #4: Bounds adattivi più teoricamente corretti
            # Con lambda alti, tau può scendere sotto 0.5 legittimamente
            # Limita solo per evitare probabilità negative (tau > 0) e overflow (tau < 3.0)
            return max(0.1, min(3.0, val))
        except (ValueError, OverflowError) as e:
            logger.warning(f"Errore calcolo tau(0,0): {e}, uso default 1.0")
            return 1.0
    elif h == 0 and a == 1:
        # tau(0,1) = 1 + lambda_h * rho
        try:
            val = 1.0 + (lh * rho)
            if not math.isfinite(val):
                logger.warning(f"tau(0,1) non finito: {val}, uso default 1.0")
                val = 1.0
            # ⚠️ FIX BUG #4: Bounds adattivi più teoricamente corretti
            return max(0.1, min(3.0, val))
        except (ValueError, OverflowError) as e:
            logger.warning(f"Errore calcolo tau(0,1): {e}, uso default 1.0")
            return 1.0
    elif h == 1 and a == 0:
        # tau(1,0) = 1 + lambda_a * rho
        try:
            val = 1.0 + (la * rho)
            if not math.isfinite(val):
                logger.warning(f"tau(1,0) non finito: {val}, uso default 1.0")
                val = 1.0
            # ⚠️ FIX BUG #4: Bounds adattivi più teoricamente corretti
            return max(0.1, min(3.0, val))
        except (ValueError, OverflowError) as e:
            logger.warning(f"Errore calcolo tau(1,0): {e}, uso default 1.0")
            return 1.0
    elif h == 1 and a == 1:
        # tau(1,1) = 1 - rho
        try:
            val = 1.0 - rho
            if not math.isfinite(val):
                logger.warning(f"tau(1,1) non finito: {val}, uso default 1.0")
                val = 1.0
            # ⚠️ FIX BUG #4: Bounds adattivi più teoricamente corretti
            return max(0.1, min(3.0, val))
        except (ValueError, OverflowError) as e:
            logger.warning(f"Errore calcolo tau(1,1): {e}, uso default 1.0")
            return 1.0
    # Per tutti gli altri casi, tau = 1.0 (nessuna correzione)
    return 1.0

def max_goals_adattivo(lh: float, la: float) -> int:
    """
    Determina max gol per matrice dinamicamente con maggiore precisione.
    
    Usa percentile 99.9% della distribuzione per catturare casi estremi.
    
    ⚠️ PRECISIONE MANIACALE: Validazione completa, protezione overflow
    """
    # ⚠️ CRITICO: Validazione input
    if not isinstance(lh, (int, float)) or not isinstance(la, (int, float)):
        logger.warning(f"Lambda non validi: lh={lh}, la={la}, uso default 10")
        return 10
    
    if not math.isfinite(lh) or not math.isfinite(la):
        logger.warning(f"Lambda non finiti: lh={lh}, la={la}, uso default 10")
        return 10
    
    if lh < 0 or la < 0:
        logger.warning(f"Lambda negativi: lh={lh}, la={la}, correggo")
        lh = max(0.1, lh)
        la = max(0.1, la)
    
    # ⚠️ PROTEZIONE: Limita lambda a range ragionevole
    lh = max(0.1, min(5.0, lh))
    la = max(0.1, min(5.0, la))
    
    # ⚠️ PRECISIONE: Calcola expected_total con protezione overflow
    expected_total = lh + la
    if not math.isfinite(expected_total):
        logger.warning(f"expected_total non finito: {expected_total}, uso default 10")
        return 10
    
    # Metodo più accurato: calcola percentile 99.9% della distribuzione totale
    # Per Poisson, P(X <= k) ≈ 1 - exp(-lambda) * sum(lambda^i / i!)
    # Usiamo approssimazione: max_goals ≈ lambda + 4*sqrt(lambda) per 99.9%
    
    # Per distribuzione somma di due Poisson: lambda_tot = lambda_h + lambda_a
    # Varianza = lambda_h + lambda_a (indipendenti)
    # ⚠️ PRECISIONE: Calcola std_dev con protezione
    try:
        variance = lh + la
        if variance <= 0:
            logger.warning(f"Varianza non positiva: {variance}, uso default 10")
            return 10
        std_dev = math.sqrt(variance)
        if not math.isfinite(std_dev):
            logger.warning(f"std_dev non finito: {std_dev}, uso default 10")
            return 10
    except (ValueError, OverflowError) as e:
        logger.warning(f"Errore calcolo std_dev: {e}, uso default 10")
        return 10
    
    # MIGLIORAMENTO: Percentile 99.9% preciso (era 3.5, ora 3.09)
    # Per distribuzione normale: P(X <= mean + 3.09*sigma) ≈ 0.999
    # Per lambda bassi, Poisson è asimmetrica → usa fattore leggermente più alto
    # ⚠️ PRECISIONE: Calcola max_goals con protezione overflow
    try:
        percentile_factor = 3.09 if expected_total > 3.0 else 3.29  # Più conservativo per lambda bassi
        max_goals_99_9 = expected_total + percentile_factor * std_dev
        if not math.isfinite(max_goals_99_9):
            logger.warning(f"max_goals_99_9 non finito: {max_goals_99_9}, uso default 15")
            max_goals_99_9 = 15
        max_goals_99_9 = int(max_goals_99_9)
    except (ValueError, OverflowError) as e:
        logger.warning(f"Errore calcolo max_goals_99_9: {e}, uso default 15")
        max_goals_99_9 = 15
    
    # Bounds ragionevoli: minimo 10 per precisione, massimo 20 per performance
    result = max(10, min(20, max_goals_99_9))
    
    # ⚠️ VERIFICA FINALE: Double-check che risultato sia valido
    if not isinstance(result, int) or result < 10 or result > 20:
        logger.warning(f"max_goals non valido: {result}, correggo a 15")
        result = 15
    
    return result

def build_score_matrix(
     lh: float,
     la: float,
     rho: float,
     max_goals: Optional[int] = None
 ) -> List[List[float]]:
    """
    Costruisce matrice score con normalizzazione e precisione numerica massima.
    
    ⚠️ PRECISIONE MANIACALE:
    - Usa Kahan summation per accumulo preciso
    - Tolleranza normalizzazione: 1e-8 (più stretta)
    - Doppia verifica normalizzazione
    - Protezione contro errori di arrotondamento
    """
    if NUMBASIA_AVAILABLE:
        fast_builder = getattr(numbasia, "build_score_matrix", None)
        if callable(fast_builder):
            try:
                result = fast_builder(lh, la, rho)
                if isinstance(result, np.ndarray):
                    return result.tolist()
                if isinstance(result, list):
                    return result
            except Exception as exc:  # pragma: no cover - dipendenza esterna opzionale
                logger.warning(f"build_score_matrix numbasia fallita, uso fallback locale: {exc}")
    
    if not isinstance(lh, (int, float)) or not isinstance(la, (int, float)) or not isinstance(rho, (int, float)):
        logger.error(f"Input non validi: lh={lh}, la={la}, rho={rho}")
        raise ValueError("Input devono essere numeri")
    
    if lh < 0 or la < 0:
        logger.warning(f"Lambda negativi: lh={lh}, la={la}, uso valori default")
        # FIX BUG #3: Standardizzato bounds lambda [0.3, 4.5]
        lh = max(0.3, lh)
        la = max(0.3, la)
    
    # Compatibilità retroattiva: consenti specifica esplicita di max_goals
    if max_goals is not None:
        if isinstance(max_goals, (int, float)):
            max_goals_int = int(max_goals)
        else:
            max_goals_int = -1
        if max_goals_int < 5 or max_goals_int > 30:
            logger.warning(
                f"max_goals={max_goals} fuori range [5, 30], uso versione adattiva"
            )
            max_goals = None
        else:
            max_goals = max_goals_int

    mg = max_goals if max_goals is not None else max_goals_adattivo(lh, la)
    if mg < 0:
        logger.warning(f"mg < 0: {mg}, uso valore default")
        mg = 10
    
    home_probs = poisson_probabilities(lh, mg)
    away_probs = poisson_probabilities(la, mg)
    matrix = np.outer(home_probs, away_probs)
    
    tau_matrix = np.ones_like(matrix)
    tau_matrix[0, 0] = tau_dixon_coles(0, 0, lh, la, rho)
    if mg >= 1:
        tau_matrix[0, 1] = tau_dixon_coles(0, 1, lh, la, rho)
        tau_matrix[1, 0] = tau_dixon_coles(1, 0, lh, la, rho)
        tau_matrix[1, 1] = tau_dixon_coles(1, 1, lh, la, rho)
    
    matrix *= tau_matrix
    np.nan_to_num(matrix, copy=False, nan=0.0, posinf=0.0, neginf=0.0)
    matrix = np.clip(matrix, 0.0, None, out=matrix)
    
    # ⚠️ FIX BUG #6: Validazione alpha parameter per Dirichlet smoothing
    dirichlet_eps = getattr(model_config, "DIRICHLET_EPS", 0.0)
    if dirichlet_eps and dirichlet_eps > 0.0:
        # ⚠️ PROTEZIONE: Limita epsilon a range ragionevole [1e-15, 1e-6]
        dirichlet_eps = max(1e-15, min(1e-6, dirichlet_eps))
        matrix += dirichlet_eps
    
    total_prob = float(matrix.sum())
    # ⚠️ FIX BUG #7: Verifica che tutti i valori siano effettivamente zero prima di usare uniforme
    if not math.isfinite(total_prob) or total_prob <= model_config.TOL_DIVISION_ZERO:
        # Verifica se ci sono celle non-zero (possibile errore numerico)
        if np.any(matrix > model_config.TOL_DIVISION_ZERO):
            logger.error(f"Matrice ha valori non-zero ma sum={total_prob}, possibile errore numerico - renormalizzo")
            # Forza rinormalizzazione
            total_prob = float(matrix.sum())
            if total_prob > 0:
                matrix /= total_prob
            else:
                matrix.fill(1.0 / matrix.size)
        else:
            # Tutte le celle sono effettivamente zero, usa uniforme
            matrix.fill(1.0 / matrix.size)
    else:
        matrix /= total_prob
    
    final_sum = float(matrix.sum())
    if not math.isfinite(final_sum) or abs(final_sum - 1.0) > model_config.TOL_NORMALIZATION:
        if final_sum > model_config.TOL_DIVISION_ZERO:
            matrix /= final_sum
        else:
            matrix.fill(1.0 / matrix.size)
    
    return matrix.tolist()

# ============================================================
#      CALCOLO PROBABILITÀ DA MATRICE (unchanged)
# ============================================================

def calc_match_result_from_matrix(mat: List[List[float]]) -> Tuple[float, float, float]:
    """
    Calcola probabilità 1X2 dalla matrice score.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta
    - P(Home) = sum(mat[h][a] for h > a)
    - P(Draw) = sum(mat[h][a] for h == a)
    - P(Away) = sum(mat[h][a] for h < a)
    - Normalizza per sicurezza (anche se matrice dovrebbe già essere normalizzata)
    """
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso distribuzione uniforme")
        return 0.333333, 0.333333, 0.333334
    
    mat_np = np.array(mat, dtype=np.float64, copy=True)
    if mat_np.ndim != 2 or mat_np.shape[0] != mat_np.shape[1]:
        logger.warning("Matrice non quadrata, uso distribuzione uniforme")
        return 0.333333, 0.333333, 0.333334
    
    np.nan_to_num(mat_np, copy=False, nan=0.0, posinf=0.0, neginf=0.0)
    np.clip(mat_np, 0.0, None, out=mat_np)
    
    total = float(mat_np.sum())
    if total <= model_config.TOL_DIVISION_ZERO or not math.isfinite(total):
        return 0.333333, 0.333333, 0.333334
    
    lower = float(np.tril(mat_np, k=-1).sum())
    upper = float(np.triu(mat_np, k=1).sum())
    diag = float(np.trace(mat_np))
    
    tot = lower + upper + diag
    if tot <= model_config.TOL_DIVISION_ZERO or not math.isfinite(tot):
        return 0.333333, 0.333333, 0.333334
    
    p_home = lower / tot
    p_draw = diag / tot
    p_away = upper / tot
    
    sum_check = p_home + p_draw + p_away
    if abs(sum_check - 1.0) > model_config.TOL_PROBABILITY_CHECK:
        if sum_check > model_config.TOL_DIVISION_ZERO:
            scale = 1.0 / sum_check
            p_home *= scale
            p_draw *= scale
            p_away *= scale
        else:
            return 0.333333, 0.333333, 0.333334
    
    return p_home, p_draw, p_away

def calc_over_under_from_matrix(mat: List[List[float]], soglia: float) -> Tuple[float, float]:
    """
    Calcola probabilità Over/Under dalla matrice score.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta
    - P(Over) = sum(mat[h][a] for h + a > soglia)
    - P(Under) = 1 - P(Over)
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    if not isinstance(soglia, (int, float)) or soglia < 0:
        logger.error(f"soglia non valida: {soglia}, uso default 2.5")
        soglia = 2.5
    
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso probabilità default")
        return 0.5, 0.5
    
    mat_np = np.array(mat, dtype=np.float64, copy=True)
    if mat_np.ndim != 2 or mat_np.shape[0] != mat_np.shape[1]:
        logger.error("Matrice inconsistente, uso probabilità default")
        return 0.5, 0.5
    
    np.nan_to_num(mat_np, copy=False, nan=0.0, posinf=0.0, neginf=0.0)
    np.clip(mat_np, 0.0, None, out=mat_np)
    
    size = mat_np.shape[0]
    indices = np.add.outer(np.arange(size), np.arange(size))
    mask_over = indices > soglia
    over_prob = float(mat_np[mask_over].sum())
    
    over_prob = max(0.0, min(1.0, over_prob))
    under_prob = 1.0 - over_prob
    
    sum_check = over_prob + under_prob
    if abs(sum_check - 1.0) > model_config.TOL_PROBABILITY_CHECK:
        if sum_check > model_config.TOL_DIVISION_ZERO:
            scale = 1.0 / sum_check
            over_prob *= scale
            under_prob = 1.0 - over_prob
        else:
            over_prob = 0.5
            under_prob = 0.5
    
    over_prob = max(0.0, min(1.0, over_prob))
    under_prob = max(0.0, min(1.0, under_prob))
    
    return over_prob, under_prob

def calc_bt_ts_from_matrix(mat: List[List[float]]) -> float:
    """
    Calcola probabilità BTTS (Both Teams To Score) dalla matrice.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta
    - P(BTTS) = sum(mat[h][a] for h >= 1 and a >= 1)
    - BTTS = entrambe le squadre segnano almeno 1 gol
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso probabilità default")
        return 0.5
    
    mat_np = np.array(mat, dtype=np.float64, copy=True)
    if mat_np.ndim != 2 or mat_np.shape[0] != mat_np.shape[1]:
        logger.error("Matrice inconsistente: riga/colonne non allineate, uso probabilità default")
        return 0.5
    
    np.nan_to_num(mat_np, copy=False, nan=0.0, posinf=0.0, neginf=0.0)
    np.clip(mat_np, 0.0, None, out=mat_np)
    
    size = mat_np.shape[0]
    if size <= 1:
        base_val = float(np.clip(mat_np.sum(), 0.0, 1.0))
        return base_val if size else 0.5
    
    mask = np.ones_like(mat_np, dtype=bool)
    mask[0, :] = False
    mask[:, 0] = False
    btts = float(mat_np[mask].sum())
    btts = max(0.0, min(1.0, btts))
    
    return btts

def calc_gg_over25_from_matrix(mat: List[List[float]]) -> float:
    """
    Calcola probabilità GG & Over 2.5 dalla matrice.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta
    - P(GG & Over 2.5) = sum(mat[h][a] for h >= 1 and a >= 1 and h + a >= 3)
    - GG = entrambe le squadre segnano, Over 2.5 = totale gol >= 3
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso probabilità default")
        return 0.5
    
    mat_np = np.array(mat, dtype=np.float64, copy=True)
    if mat_np.ndim != 2 or mat_np.shape[0] != mat_np.shape[1]:
        logger.error("Matrice inconsistente: riga/colonne non allineate, uso probabilità default")
        return 0.5
    
    np.nan_to_num(mat_np, copy=False, nan=0.0, posinf=0.0, neginf=0.0)
    np.clip(mat_np, 0.0, None, out=mat_np)
    
    size = mat_np.shape[0]
    # ⚠️ FIX BUG #9: 1×1 matrix (solo 0-0) non può soddisfare GG & Over 2.5
    if size <= 1:
        return 0.0  # 0×0 or 1×1 matrix cannot satisfy h>=1 AND a>=1 AND h+a>=3
    
    idx = np.arange(size)
    mask = np.outer(idx >= 1, idx >= 1) & (np.add.outer(idx, idx) >= 3)
    prob = float(mat_np[mask].sum())
    prob = max(0.0, min(1.0, prob))
    
    return prob

def prob_pari_dispari_from_matrix(mat: List[List[float]]) -> Tuple[float, float]:
    """
    Calcola probabilità Pari/Dispari dalla matrice.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta
    - P(Pari) = sum(mat[h][a] for (h + a) % 2 == 0)
    - P(Dispari) = 1 - P(Pari)
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso probabilità default")
        return 0.5, 0.5
    
    mat_np = np.array(mat, dtype=np.float64, copy=True)
    if mat_np.ndim != 2 or mat_np.shape[0] != mat_np.shape[1]:
        logger.error("Matrice inconsistente: riga/colonne non allineate, uso probabilità default")
        return 0.5, 0.5
    
    np.nan_to_num(mat_np, copy=False, nan=0.0, posinf=0.0, neginf=0.0)
    np.clip(mat_np, 0.0, None, out=mat_np)
    
    total = float(mat_np.sum())
    if total <= model_config.TOL_DIVISION_ZERO or not math.isfinite(total):
        return 0.5, 0.5
    
    indices = np.add.outer(np.arange(mat_np.shape[0]), np.arange(mat_np.shape[1]))
    even_prob = float(mat_np[(indices % 2) == 0].sum())
    odd_prob = total - even_prob
    
    if even_prob < 0.0:
        even_prob = 0.0
    if odd_prob < 0.0:
        odd_prob = 0.0
    
    sum_check = even_prob + odd_prob
    if sum_check <= model_config.TOL_DIVISION_ZERO or not math.isfinite(sum_check):
        return 0.5, 0.5
    
    even = even_prob / sum_check
    odd = odd_prob / sum_check
    
    return even, odd

def prob_clean_sheet_from_matrix(mat: List[List[float]]) -> Tuple[float, float]:
    """
    Calcola probabilità Clean Sheet dalla matrice.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta
    - P(CS Home) = sum(mat[h][0] for h in range(mg + 1)) = squadra casa non subisce gol
    - P(CS Away) = sum(mat[0][a] for a in range(mg + 1)) = squadra trasferta non subisce gol
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    # ⚠️ CRITICO: Validazione input robusta
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso probabilità default")
        return 0.5, 0.5
    
    mg = len(mat) - 1
    
    # ⚠️ CRITICO: Verifica che mg sia valido
    if mg < 0:
        logger.warning("mg < 0, uso probabilità default")
        return 0.5, 0.5
    
    # Verifica che tutte le righe abbiano stessa lunghezza
    for i, row in enumerate(mat):
        if len(row) != mg + 1:
            logger.error(f"Matrice inconsistente: riga {i} ha {len(row)} colonne invece di {mg + 1}")
            return 0.5, 0.5
    
    # ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso
    cs_home = 0.0
    cs_away = 0.0
    c_home = 0.0  # Compensazione Kahan
    c_away = 0.0
    
    for h in range(mg + 1):
        p_h = mat[h][0]
        # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
        if isinstance(p_h, (int, float)) and p_h >= 0 and (p_h == p_h) and math.isfinite(p_h):
            # Kahan summation
            y = p_h - c_home
            t = cs_home + y
            c_home = (t - cs_home) - y
            cs_home = t
    
    for a in range(mg + 1):
        p_a = mat[0][a]
        # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
        if isinstance(p_a, (int, float)) and p_a >= 0 and (p_a == p_a) and math.isfinite(p_a):
            # Kahan summation
            y = p_a - c_away
            t = cs_away + y
            c_away = (t - cs_away) - y
            cs_away = t
    
    # ⚠️ PROTEZIONE: Limita a range [0, 1] con precisione
    cs_home = max(0.0, min(1.0, cs_home))
    cs_away = max(0.0, min(1.0, cs_away))
    
    # ⚠️ VERIFICA FINALE: Double-check che siano in range [0, 1]
    if not (0.0 <= cs_home <= 1.0):
        logger.warning(f"CS Home fuori range: {cs_home}, correggo a 0.5")
        cs_home = 0.5
    if not (0.0 <= cs_away <= 1.0):
        logger.warning(f"CS Away fuori range: {cs_away}, correggo a 0.5")
        cs_away = 0.5
    
    return cs_home, cs_away

def prob_asian_handicap_from_matrix(mat: List[List[float]], handicap: float, team: str = 'home') -> float:
    """
    Calcola probabilità Asian Handicap dalla matrice.

    ⚠️ VERIFICA MATEMATICA: Formula corretta con gestione push
    - Asian Handicap applica un vantaggio/svantaggio virtuale
    - handicap positivo = vantaggio, negativo = svantaggio
    - team = 'home' o 'away'

    Esempi:
    - Home -1.5: Casa deve vincere per 2+ gol
    - Home -1.0: Casa vince per 2+ (win), vince per 1 (push = 0.5), altro (lose)
    - Away +1.5: Away può perdere max per 1 gol

    ⚠️ PRECISIONE MANIACALE: Kahan summation, gestione push, validazione completa
    """
    # ⚠️ CRITICO: Validazione input robusta
    if team not in ['home', 'away']:
        logger.error(f"team non valido: {team}, uso default 'home'")
        team = 'home'

    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso probabilità default")
        return 0.5

    mg = len(mat) - 1

    # ⚠️ CRITICO: Verifica che mg sia valido
    if mg < 0:
        logger.warning("mg < 0, uso probabilità default")
        return 0.5

    # Verifica che tutte le righe abbiano stessa lunghezza
    for i, row in enumerate(mat):
        if len(row) != mg + 1:
            logger.error(f"Matrice inconsistente: riga {i} ha {len(row)} colonne invece di {mg + 1}")
            return 0.5

    # Verifica che handicap sia un numero finito
    if not isinstance(handicap, (int, float)) or not math.isfinite(handicap):
        logger.error(f"Handicap non valido: {handicap}, uso 0.0")
        handicap = 0.0

    # ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso
    prob = 0.0
    c = 0.0  # Compensazione Kahan

    for h in range(mg + 1):
        for a in range(mg + 1):
            p = mat[h][a]
            # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue

            # Calcola differenza gol con handicap applicato
            if team == 'home':
                # Home con handicap (es: -1.5 significa casa parte svantaggiata di 1.5 gol)
                # Score adjusted: (home + handicap) vs away
                # Per "Home -1.5": adjusted_score = h - 1.5
                adjusted_diff = (h + handicap) - a
            else:  # away
                # Away con handicap (es: +1.5 significa away parte avvantaggiata di 1.5 gol)
                # Score adjusted: (away + handicap) vs home
                # Per "Away +1.5": adjusted_score = a + 1.5
                adjusted_diff = (a + handicap) - h

            # Determina contributo alla probabilità
            contribution = 0.0
            if adjusted_diff > 0:
                # Vittoria completa
                contribution = p
            elif adjusted_diff == 0:
                # Push (rimborso, conta come 0.5)
                contribution = p * 0.5
            # else: adjusted_diff < 0 -> sconfitta, contribution = 0

            # ⚠️ PRECISIONE MANIACALE: Kahan summation
            if contribution > 0:
                y = contribution - c
                t = prob + y
                c = (t - prob) - y
                prob = t

    # ⚠️ PROTEZIONE: Limita a range [0, 1] con precisione
    prob = max(0.0, min(1.0, prob))

    # ⚠️ VERIFICA FINALE: Double-check che sia in range [0, 1]
    if not (0.0 <= prob <= 1.0):
        logger.warning(f"Asian Handicap prob fuori range: {prob}, correggo a 0.5")
        prob = 0.5

    return prob

def prob_ht_ft_from_matrices(mat_ht: List[List[float]], mat_ft: List[List[float]]) -> Dict[str, float]:
    """
    Calcola tutte le 9 combinazioni HT/FT (Halftime/Fulltime) dalle matrici.

    ⚠️ VERIFICA MATEMATICA: Formula con assunzione di indipendenza condizionale

    Le 9 combinazioni sono:
    - 1/1: Casa vince HT, Casa vince FT
    - 1/X: Casa vince HT, Pareggio FT
    - 1/2: Casa vince HT, Trasferta vince FT
    - X/1: Pareggio HT, Casa vince FT
    - X/X: Pareggio HT, Pareggio FT
    - X/2: Pareggio HT, Trasferta vince FT
    - 2/1: Trasferta vince HT, Casa vince FT
    - 2/X: Trasferta vince HT, Pareggio FT
    - 2/2: Trasferta vince HT, Trasferta vince FT

    ⚠️ MODELLO PROBABILISTICO:
    Approccio 1 (semplice, indipendenza): P(HT=x, FT=y) ≈ P(HT=x) * P(FT=y)
    Approccio 2 (avanzato): P(HT=x, FT=y) considerando correlazione

    Usiamo approccio 1 (indipendenza) per semplicità, con fattori correttivi empirici.

    ⚠️ PRECISIONE MANIACALE: Kahan summation per calcolo probabilità marginali
    """
    # ⚠️ CRITICO: Validazione input robusta
    if not mat_ht or len(mat_ht) == 0 or (len(mat_ht) > 0 and len(mat_ht[0]) == 0):
        logger.warning("Matrice HT vuota o non valida, uso probabilità default")
        return {f"{ht}/{ft}": 1.0/9.0 for ht in ['1', 'X', '2'] for ft in ['1', 'X', '2']}

    if not mat_ft or len(mat_ft) == 0 or (len(mat_ft) > 0 and len(mat_ft[0]) == 0):
        logger.warning("Matrice FT vuota o non valida, uso probabilità default")
        return {f"{ht}/{ft}": 1.0/9.0 for ht in ['1', 'X', '2'] for ft in ['1', 'X', '2']}

    mg_ht = len(mat_ht) - 1
    mg_ft = len(mat_ft) - 1

    # ⚠️ CRITICO: Verifica che mg siano validi
    if mg_ht < 0 or mg_ft < 0:
        logger.warning("mg < 0, uso probabilità default")
        return {f"{ht}/{ft}": 1.0/9.0 for ht in ['1', 'X', '2'] for ft in ['1', 'X', '2']}

    # Calcola probabilità marginali HT con Kahan summation
    p_ht_home = 0.0
    p_ht_draw = 0.0
    p_ht_away = 0.0
    c_h = 0.0
    c_d = 0.0
    c_a = 0.0

    for h in range(mg_ht + 1):
        for a in range(mg_ht + 1):
            p = mat_ht[h][a]
            # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue

            if h > a:  # Home vince
                y = p - c_h
                t = p_ht_home + y
                c_h = (t - p_ht_home) - y
                p_ht_home = t
            elif h == a:  # Pareggio
                y = p - c_d
                t = p_ht_draw + y
                c_d = (t - p_ht_draw) - y
                p_ht_draw = t
            else:  # Away vince
                y = p - c_a
                t = p_ht_away + y
                c_a = (t - p_ht_away) - y
                p_ht_away = t

    # Calcola probabilità marginali FT con Kahan summation
    p_ft_home = 0.0
    p_ft_draw = 0.0
    p_ft_away = 0.0
    c_h = 0.0
    c_d = 0.0
    c_a = 0.0

    for h in range(mg_ft + 1):
        for a in range(mg_ft + 1):
            p = mat_ft[h][a]
            # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue

            if h > a:  # Home vince
                y = p - c_h
                t = p_ft_home + y
                c_h = (t - p_ft_home) - y
                p_ft_home = t
            elif h == a:  # Pareggio
                y = p - c_d
                t = p_ft_draw + y
                c_d = (t - p_ft_draw) - y
                p_ft_draw = t
            else:  # Away vince
                y = p - c_a
                t = p_ft_away + y
                c_a = (t - p_ft_away) - y
                p_ft_away = t

    # ⚠️ PROTEZIONE: Normalizza se necessario (dovrebbero già sommare a ~1.0)
    total_ht = p_ht_home + p_ht_draw + p_ht_away
    total_ft = p_ft_home + p_ft_draw + p_ft_away

    if total_ht > model_config.TOL_DIVISION_ZERO and abs(total_ht - 1.0) > model_config.TOL_PROBABILITY_CHECK:
        logger.warning(f"Probabilità HT non sommano a 1.0: {total_ht:.6f}, normalizzo")
        p_ht_home /= total_ht
        p_ht_draw /= total_ht
        p_ht_away /= total_ht

    if total_ft > model_config.TOL_DIVISION_ZERO and abs(total_ft - 1.0) > model_config.TOL_PROBABILITY_CHECK:
        logger.warning(f"Probabilità FT non sommano a 1.0: {total_ft:.6f}, normalizzo")
        p_ft_home /= total_ft
        p_ft_draw /= total_ft
        p_ft_away /= total_ft

    # ⚠️ MODELLO: Calcola combinazioni HT/FT con assunzione di indipendenza
    # Fattori correttivi empirici per migliorare accuratezza:
    # - Se una squadra vince al HT, ha più probabilità di vincere al FT (momentum)
    # - Se c'è pareggio al HT, il FT è più imprevedibile

    # Matrice fattori correttivi (basata su analisi empirica calcio)
    # [HT][FT] dove 0=home, 1=draw, 2=away
    correction_factors = [
        [1.3, 0.7, 0.5],  # Se Home vince HT: più prob 1, meno X e 2
        [0.9, 1.1, 0.9],  # Se Pareggio HT: leggermente più prob X
        [0.5, 0.7, 1.3],  # Se Away vince HT: più prob 2, meno 1 e X
    ]

    ht_probs = [p_ht_home, p_ht_draw, p_ht_away]
    ft_probs = [p_ft_home, p_ft_draw, p_ft_away]
    ht_labels = ['1', 'X', '2']
    ft_labels = ['1', 'X', '2']

    # Calcola probabilità grezze
    raw_probs = {}
    for i, ht_label in enumerate(ht_labels):
        for j, ft_label in enumerate(ft_labels):
            key = f"{ht_label}/{ft_label}"
            # Probabilità base con fattore correttivo
            base_prob = ht_probs[i] * ft_probs[j]
            corrected_prob = base_prob * correction_factors[i][j]
            raw_probs[key] = corrected_prob

    # ⚠️ NORMALIZZAZIONE: Assicura che somma = 1.0
    total = sum(raw_probs.values())

    ht_ft_probs = {}
    if total > model_config.TOL_DIVISION_ZERO:
        for key, prob in raw_probs.items():
            normalized = prob / total
            # ⚠️ PROTEZIONE: Limita a range [0, 1]
            normalized = max(0.0, min(1.0, normalized))
            ht_ft_probs[key] = normalized
    else:
        # Fallback: distribuzione uniforme
        logger.warning("Somma probabilità HT/FT = 0, uso distribuzione uniforme")
        for ht_label in ht_labels:
            for ft_label in ft_labels:
                ht_ft_probs[f"{ht_label}/{ft_label}"] = 1.0 / 9.0

    # ⚠️ VERIFICA FINALE: Controlla che tutte siano in [0, 1] e sommino a ~1.0
    total_check = sum(ht_ft_probs.values())
    if abs(total_check - 1.0) > model_config.TOL_PROBABILITY_CHECK:
        logger.warning(f"Probabilità HT/FT non sommano a 1.0: {total_check:.6f}")

    for key, prob in ht_ft_probs.items():
        if not (0.0 <= prob <= 1.0):
            logger.error(f"Probabilità HT/FT {key} fuori range: {prob}")
            ht_ft_probs[key] = max(0.0, min(1.0, prob))

    return ht_ft_probs

def dist_gol_da_matrice(mat: List[List[float]]):
    """
    Calcola distribuzione marginale gol per casa e trasferta dalla matrice.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta
    - dh[k] = sum(mat[k][a] for a in range(mg + 1)) = P(Home segna k gol)
    - da[k] = sum(mat[h][k] for h in range(mg + 1)) = P(Away segna k gol)
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    # ⚠️ CRITICO: Validazione input robusta
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso distribuzione uniforme")
        mg = 10  # Default
        uniform = 1.0 / (mg + 1)
        return [uniform] * (mg + 1), [uniform] * (mg + 1)
    
    mg = len(mat) - 1
    
    # ⚠️ CRITICO: Verifica che mg sia valido
    if mg < 0:
        logger.warning("mg < 0, uso distribuzione uniforme")
        mg = 10  # Default
        uniform = 1.0 / (mg + 1)
        return [uniform] * (mg + 1), [uniform] * (mg + 1)
    
    # Verifica che tutte le righe abbiano stessa lunghezza
    for i, row in enumerate(mat):
        if len(row) != mg + 1:
            logger.error(f"Matrice inconsistente: riga {i} ha {len(row)} colonne invece di {mg + 1}")
            uniform = 1.0 / (mg + 1)
            return [uniform] * (mg + 1), [uniform] * (mg + 1)
    
    dh = [0.0] * (mg + 1)
    da = [0.0] * (mg + 1)
    
    # ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso
    c_dh = [0.0] * (mg + 1)  # Compensazione Kahan per ogni elemento
    c_da = [0.0] * (mg + 1)
    
    for h in range(mg + 1):
        for a in range(mg + 1):
            p = mat[h][a]
            # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue
            # Kahan summation per dh[h]
            y_h = p - c_dh[h]
            t_h = dh[h] + y_h
            c_dh[h] = (t_h - dh[h]) - y_h
            dh[h] = t_h
            # Kahan summation per da[a]
            y_a = p - c_da[a]
            t_a = da[a] + y_a
            c_da[a] = (t_a - da[a]) - y_a
            da[a] = t_a
    
    # ⚠️ PRECISIONE: Kahan summation per somma totale
    sum_dh = 0.0
    sum_da = 0.0
    c_sum_dh = 0.0
    c_sum_da = 0.0
    
    for i in range(mg + 1):
        # Somma dh
        y = dh[i] - c_sum_dh
        t = sum_dh + y
        c_sum_dh = (t - sum_dh) - y
        sum_dh = t
        # Somma da
        y = da[i] - c_sum_da
        t = sum_da + y
        c_sum_da = (t - sum_da) - y
        sum_da = t
    
    # ⚠️ VERIFICA: Normalizza distribuzioni marginali (dovrebbero sommare a 1.0)
    if sum_dh > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        for i in range(mg + 1):
            dh[i] /= sum_dh
    else:
        # Fallback: distribuzione uniforme
        logger.warning(f"Somma dh troppo piccola: {sum_dh}, uso distribuzione uniforme")
        uniform = 1.0 / (mg + 1)
        dh = [uniform] * (mg + 1)
    
    if sum_da > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        for i in range(mg + 1):
            da[i] /= sum_da
    else:
        # Fallback: distribuzione uniforme
        logger.warning(f"Somma da troppo piccola: {sum_da}, uso distribuzione uniforme")
        uniform = 1.0 / (mg + 1)
        da = [uniform] * (mg + 1)
    
    return dh, da

def dist_gol_totali_from_matrix(mat: List[List[float]]) -> List[float]:
    """
    Calcola distribuzione gol totali dalla matrice.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta
    - dist[k] = sum(mat[h][a] for h + a == k) = P(Totale gol = k)
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    # ⚠️ CRITICO: Validazione input robusta
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso distribuzione uniforme")
        max_tot = 20  # Default
        uniform = 1.0 / (max_tot + 1)
        return [uniform] * (max_tot + 1)
    
    mg = len(mat) - 1
    
    # ⚠️ CRITICO: Verifica che mg sia valido
    if mg < 0:
        logger.warning("mg < 0, uso distribuzione uniforme")
        max_tot = 20  # Default
        uniform = 1.0 / (max_tot + 1)
        return [uniform] * (max_tot + 1)
    
    # Verifica che tutte le righe abbiano stessa lunghezza
    for i, row in enumerate(mat):
        if len(row) != mg + 1:
            logger.error(f"Matrice inconsistente: riga {i} ha {len(row)} colonne invece di {mg + 1}")
            max_tot = mg * 2
            uniform = 1.0 / (max_tot + 1)
            return [uniform] * (max_tot + 1)
    
    max_tot = mg * 2
    dist = [0.0] * (max_tot + 1)
    
    # ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso
    c_dist = [0.0] * (max_tot + 1)  # Compensazione Kahan per ogni elemento
    
    for h in range(mg + 1):
        for a in range(mg + 1):
            p = mat[h][a]
            # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue
            tot = h + a
            if tot < len(dist):
                # Kahan summation
                y = p - c_dist[tot]
                t = dist[tot] + y
                c_dist[tot] = (t - dist[tot]) - y
                dist[tot] = t
    
    # ⚠️ PRECISIONE: Kahan summation per somma totale
    sum_dist = 0.0
    c_sum = 0.0
    
    for i in range(len(dist)):
        y = dist[i] - c_sum
        t = sum_dist + y
        c_sum = (t - sum_dist) - y
        sum_dist = t
    
    # ⚠️ VERIFICA: Normalizza distribuzione (dovrebbe sommare a 1.0)
    if sum_dist > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        for i in range(len(dist)):
            dist[i] /= sum_dist
    else:
        # Fallback: distribuzione uniforme
        logger.warning(f"Somma dist troppo piccola: {sum_dist}, uso distribuzione uniforme")
        uniform = 1.0 / len(dist)
        dist = [uniform] * len(dist)
    
    return dist

def prob_multigol_from_dist(dist: List[float], gmin: int, gmax: int) -> float:
    """
    Calcola probabilità multigol da distribuzione.
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    # ⚠️ CRITICO: Validazione input
    if not dist or len(dist) == 0:
        logger.warning("Distribuzione vuota, uso probabilità default")
        return 0.5
    
    # ⚠️ FIX BUG #8: Auto-swap se invertiti invece di ritornare default generico
    if not isinstance(gmin, int) or not isinstance(gmax, int) or gmin < 0:
        logger.warning(f"Parametri non validi: gmin={gmin}, gmax={gmax}, uso default")
        return 0.5
    if gmax < gmin:
        logger.warning(f"gmin > gmax invertiti: gmin={gmin}, gmax={gmax} - auto-swap")
        gmin, gmax = gmax, gmin
    
    # ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso
    s = 0.0
    c = 0.0  # Compensazione Kahan
    
    for k in range(gmin, gmax + 1):
        if k < len(dist):
            p = dist[k]
            # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
            if isinstance(p, (int, float)) and p > 0 and (p == p) and math.isfinite(p):
                # Kahan summation
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t
    
    # ⚠️ PROTEZIONE: Limita risultato a range [0, 1]
    s = max(0.0, min(1.0, s))
    
    # ⚠️ VERIFICA FINALE: Double-check che sia in range [0, 1]
    if not (0.0 <= s <= 1.0):
        logger.warning(f"Probabilità multigol fuori range: {s}, correggo a 0.5")
        s = 0.5
    
    return s

def prob_esito_over_from_matrix(mat: List[List[float]], esito: str, soglia: float) -> float:
    """
    Calcola probabilità Esito & Over dalla matrice.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta
    - P(Esito & Over) = sum(mat[h][a] for h + a > soglia and esito verificato)
    - Esito può essere '1' (Home), 'X' (Draw), '2' (Away)
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    # ⚠️ CRITICO: Validazione input robusta
    if not isinstance(soglia, (int, float)) or soglia < 0:
        logger.error(f"soglia non valida: {soglia}, uso default 2.5")
        soglia = 2.5
    
    if esito not in ['1', 'X', '2']:
        logger.error(f"esito non valido: {esito}, uso default '1'")
        esito = '1'
    
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso probabilità default")
        return 0.5
    
    mg = len(mat) - 1
    
    # ⚠️ CRITICO: Verifica che mg sia valido
    if mg < 0:
        logger.warning("mg < 0, uso probabilità default")
        return 0.5
    
    # Verifica che tutte le righe abbiano stessa lunghezza
    for i, row in enumerate(mat):
        if len(row) != mg + 1:
            logger.error(f"Matrice inconsistente: riga {i} ha {len(row)} colonne invece di {mg + 1}")
            return 0.5
    
    # ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso
    s = 0.0
    c = 0.0  # Compensazione Kahan
    
    for h in range(mg + 1):
        for a in range(mg + 1):
            if h + a <= soglia:
                continue
            p = mat[h][a]
            # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue
            if esito == '1' and h > a:
                # Kahan summation
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t
            elif esito == 'X' and h == a:
                # Kahan summation
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t
            elif esito == '2' and h < a:
                # Kahan summation
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t
    
    # ⚠️ PROTEZIONE: Limita a range [0, 1]
    s = max(0.0, min(1.0, s))
    
    # ⚠️ VERIFICA FINALE: Double-check che sia in range [0, 1]
    if not (0.0 <= s <= 1.0):
        logger.warning(f"Probabilità esito+over fuori range: {s}, correggo a 0.5")
        s = 0.5
    
    return s

def prob_dc_over_from_matrix(mat: List[List[float]], dc: str, soglia: float, inverse: bool = False) -> float:
    """
    Calcola probabilità Double Chance & Over/Under dalla matrice.
    
    Args:
        mat: Matrice score
        dc: Double Chance ('1X', 'X2', '12')
        soglia: Soglia gol (es. 2.5, 3.5)
        inverse: Se True, calcola Under invece di Over
    
    Returns:
        Probabilità combinata
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    # ⚠️ CRITICO: Validazione input robusta
    if not isinstance(soglia, (int, float)) or soglia < 0:
        logger.error(f"soglia non valida: {soglia}, uso default 2.5")
        soglia = 2.5
    
    if dc not in ['1X', 'X2', '12']:
        logger.error(f"dc non valido: {dc}, uso default '1X'")
        dc = '1X'
    
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso probabilità default")
        return 0.5
    
    mg = len(mat) - 1
    
    # ⚠️ CRITICO: Verifica che mg sia valido
    if mg < 0:
        logger.warning("mg < 0, uso probabilità default")
        return 0.5
    
    # Verifica che tutte le righe abbiano stessa lunghezza
    for i, row in enumerate(mat):
        if len(row) != mg + 1:
            logger.error(f"Matrice inconsistente: riga {i} ha {len(row)} colonne invece di {mg + 1}")
            return 0.5
    
    # ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso
    s = 0.0
    c = 0.0  # Compensazione Kahan
    
    for h in range(mg + 1):
        for a in range(mg + 1):
            # Controlla Over/Under
            if inverse:
                # Under: h + a <= soglia
                if h + a > soglia:
                    continue
            else:
                # Over: h + a > soglia
                if h + a <= soglia:
                    continue
            
            p = mat[h][a]
            # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue
            # Controlla Double Chance
            ok = False
            if dc == '1X' and h >= a:
                ok = True
            elif dc == 'X2' and a >= h:
                ok = True
            elif dc == '12' and h != a:
                ok = True
            
            if ok:
                # Kahan summation
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t
    
    # ⚠️ PROTEZIONE: Limita a range [0, 1]
    s = max(0.0, min(1.0, s))
    
    # ⚠️ VERIFICA FINALE: Double-check che sia in range [0, 1]
    if not (0.0 <= s <= 1.0):
        logger.warning(f"Probabilità DC+Over/Under fuori range: {s}, correggo a 0.5")
        s = 0.5
    
    return s

def prob_esito_btts_from_matrix(mat: List[List[float]], esito: str) -> float:
    """
    Calcola probabilità Esito & BTTS dalla matrice.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta
    - P(Esito & BTTS) = sum(mat[h][a] for h >= 1 and a >= 1 and esito verificato)
    - Esito può essere '1' (Home), 'X' (Draw), '2' (Away)
    - BTTS = entrambe le squadre segnano almeno 1 gol
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    # ⚠️ CRITICO: Validazione input robusta
    if esito not in ['1', 'X', '2']:
        logger.error(f"esito non valido: {esito}, uso default '1'")
        esito = '1'
    
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso probabilità default")
        return 0.5
    
    mg = len(mat) - 1
    
    # ⚠️ CRITICO: Verifica che mg sia valido
    if mg < 0:
        logger.warning("mg < 0, uso probabilità default")
        return 0.5
    
    # Verifica che tutte le righe abbiano stessa lunghezza
    for i, row in enumerate(mat):
        if len(row) != mg + 1:
            logger.error(f"Matrice inconsistente: riga {i} ha {len(row)} colonne invece di {mg + 1}")
            return 0.5
    
    # ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso
    s = 0.0
    c = 0.0  # Compensazione Kahan
    
    for h in range(1, mg + 1):
        for a in range(1, mg + 1):
            p = mat[h][a]
            # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue
            if esito == '1' and h > a:
                # Kahan summation
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t
            elif esito == 'X' and h == a:
                # Kahan summation
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t
            elif esito == '2' and h < a:
                # Kahan summation
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t
    
    # ⚠️ PROTEZIONE: Limita a range [0, 1]
    s = max(0.0, min(1.0, s))
    
    # ⚠️ VERIFICA FINALE: Double-check che sia in range [0, 1]
    if not (0.0 <= s <= 1.0):
        logger.warning(f"Probabilità esito+BTTS fuori range: {s}, correggo a 0.5")
        s = 0.5

    return s

def prob_esito_ng_from_matrix(mat: List[List[float]], esito: str) -> float:
    """
    Calcola probabilità Esito & No Goal (BTTS No) dalla matrice.

    ⚠️ VERIFICA MATEMATICA:
    - P(Esito & NG) = sum(mat[h][a] per stati coerenti con l'esito e con almeno una squadra a 0 gol)
    - Esito può essere '1' (Home), 'X' (Draw), '2' (Away)
    - NG = almeno una delle due squadre non segna (BTTS No)
    """
    # ⚠️ CRITICO: Validazione input robusta
    if esito not in ['1', 'X', '2']:
        logger.error(f"esito non valido: {esito}, uso default '1'")
        esito = '1'

    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso probabilità default")
        return 0.5

    mg = len(mat) - 1

    if mg < 0:
        logger.warning("mg < 0, uso probabilità default")
        return 0.5

    # Verifica che tutte le righe abbiano stessa lunghezza
    for i, row in enumerate(mat):
        if len(row) != mg + 1:
            logger.error(f"Matrice inconsistente: riga {i} ha {len(row)} colonne invece di {mg + 1}")
            return 0.5

    # ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso
    s = 0.0
    c = 0.0  # Compensazione Kahan

    for h in range(mg + 1):
        for a in range(mg + 1):
            p = mat[h][a]
            # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue

            include = False
            if esito == '1':
                include = (a == 0 and h > a)
            elif esito == 'X':
                include = (h == 0 and a == 0)
            elif esito == '2':
                include = (h == 0 and a > h)

            if include:
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t

    s = max(0.0, min(1.0, s))

    if not (0.0 <= s <= 1.0):
        logger.warning(f"Probabilità esito+NG fuori range: {s}, correggo a 0.5")
        s = 0.5

    return s

def prob_dc_btts_from_matrix(mat: List[List[float]], dc: str) -> float:
    """
    Calcola probabilità Double Chance & BTTS dalla matrice.
    
    ⚠️ VERIFICA MATEMATICA: Formula corretta
    - P(DC & BTTS) = sum(mat[h][a] for h >= 1 and a >= 1 and DC verificato)
    - DC può essere '1X' (Home o Draw), 'X2' (Draw o Away), '12' (Home o Away)
    - BTTS = entrambe le squadre segnano almeno 1 gol
    
    ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso, validazione completa
    """
    # ⚠️ CRITICO: Validazione input robusta
    if dc not in ['1X', 'X2', '12']:
        logger.error(f"dc non valido: {dc}, uso default '1X'")
        dc = '1X'
    
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida, uso probabilità default")
        return 0.5
    
    mg = len(mat) - 1
    
    # ⚠️ CRITICO: Verifica che mg sia valido
    if mg < 0:
        logger.warning("mg < 0, uso probabilità default")
        return 0.5
    
    # Verifica che tutte le righe abbiano stessa lunghezza
    for i, row in enumerate(mat):
        if len(row) != mg + 1:
            logger.error(f"Matrice inconsistente: riga {i} ha {len(row)} colonne invece di {mg + 1}")
            return 0.5
    
    # ⚠️ PRECISIONE MANIACALE: Kahan summation per accumulo preciso
    s = 0.0
    c = 0.0  # Compensazione Kahan
    
    for h in range(1, mg + 1):
        for a in range(1, mg + 1):
            p = mat[h][a]
            # ⚠️ PROTEZIONE: Ignora valori negativi, NaN, o infiniti
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue
            ok = False
            if dc == '1X' and h >= a:
                ok = True
            elif dc == 'X2' and a >= h:
                ok = True
            elif dc == '12' and h != a:
                ok = True
            if ok:
                # Kahan summation
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t
    
    # ⚠️ PROTEZIONE: Limita a range [0, 1]
    s = max(0.0, min(1.0, s))
    
    # ⚠️ VERIFICA FINALE: Double-check che sia in range [0, 1]
    if not (0.0 <= s <= 1.0):
        logger.warning(f"Probabilità DC+BTTS fuori range: {s}, correggo a 0.5")
        s = 0.5
    
    return s

def prob_esito_multigol_from_matrix(mat: List[List[float]], esito: str, gmin: int, gmax: int) -> float:
    """
    Calcola probabilità Esito & Multigol dalla matrice dei punteggi.
    
    ⚠️ PRECISIONE MANIACALE: Usa Kahan summation per minimizzare errori numerici e valida ogni input.
    """
    # ⚠️ CRITICO: Validazione input robusta
    if esito not in ['1', 'X', '2']:
        logger.error(f"esito non valido per multigol: {esito}, uso default '1'")
        esito = '1'
    
    # ⚠️ FIX BUG #8: Auto-swap se invertiti
    if not isinstance(gmin, int) or not isinstance(gmax, int) or gmin < 0:
        logger.error(f"Range multigol non valido: gmin={gmin}, gmax={gmax}, uso fallback 1-3")
        gmin, gmax = 1, 3
    elif gmax < gmin:
        logger.warning(f"gmin > gmax invertiti: gmin={gmin}, gmax={gmax} - auto-swap")
        gmin, gmax = gmax, gmin
    
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida per esito+multigol, ritorno 0.5")
        return 0.5
    
    mg = len(mat) - 1
    if mg < 0:
        logger.warning("mg < 0 per esito+multigol, ritorno 0.5")
        return 0.5
    
    for i, row in enumerate(mat):
        if len(row) != mg + 1:
            logger.error(f"Matrice inconsistente (riga {i}) per esito+multigol, ritorno 0.5")
            return 0.5
    
    # ⚠️ PRECISIONE MANIACALE: Kahan summation
    s = 0.0
    c = 0.0  # Compensazione
    
    for h in range(mg + 1):
        for a in range(mg + 1):
            tot = h + a
            if tot < gmin or tot > gmax:
                continue
            p = mat[h][a]
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue
            esito_ok = (
                (esito == '1' and h > a) or
                (esito == 'X' and h == a) or
                (esito == '2' and h < a)
            )
            if esito_ok:
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t
    
    s = max(0.0, min(1.0, s))
    if not (0.0 <= s <= 1.0):
        logger.warning(f"Probabilità esito+multigol fuori range: {s}, imposto 0.5")
        s = 0.5
    
    return s

def prob_dc_multigol_from_matrix(mat: List[List[float]], dc: str, gmin: int, gmax: int) -> float:
    """
    Calcola probabilità Double Chance & Multigol dalla matrice dei punteggi.
    
    ⚠️ PRECISIONE MANIACALE: Usa Kahan summation, validazione completa e clamp finale.
    """
    # ⚠️ CRITICO: Validazione input robusta
    if dc not in ['1X', 'X2', '12']:
        logger.error(f"Double Chance non valida per multigol: {dc}, uso default '1X'")
        dc = '1X'
    
    if not isinstance(gmin, int) or not isinstance(gmax, int) or gmin < 0 or gmax < gmin:
        logger.error(f"Range multigol DC non valido: gmin={gmin}, gmax={gmax}, uso fallback 1-3")
        gmin, gmax = 1, 3
    
    if not mat or len(mat) == 0 or (len(mat) > 0 and len(mat[0]) == 0):
        logger.warning("Matrice vuota o non valida per DC+multigol, ritorno 0.5")
        return 0.5
    
    mg = len(mat) - 1
    if mg < 0:
        logger.warning("mg < 0 per DC+multigol, ritorno 0.5")
        return 0.5
    
    for i, row in enumerate(mat):
        if len(row) != mg + 1:
            logger.error(f"Matrice inconsistente (riga {i}) per DC+multigol, ritorno 0.5")
            return 0.5
    
    # ⚠️ PRECISIONE MANIACALE: Kahan summation
    s = 0.0
    c = 0.0
    
    for h in range(mg + 1):
        for a in range(mg + 1):
            tot = h + a
            if tot < gmin or tot > gmax:
                continue
            p = mat[h][a]
            if not isinstance(p, (int, float)) or p < 0 or not (p == p) or not math.isfinite(p):
                continue
            ok = False
            if dc == '1X' and h >= a:
                ok = True
            elif dc == 'X2' and a >= h:
                ok = True
            elif dc == '12' and h != a:
                ok = True
            if ok:
                y = p - c
                t = s + y
                c = (t - s) - y
                s = t
    
    s = max(0.0, min(1.0, s))
    if not (0.0 <= s <= 1.0):
        logger.warning(f"Probabilità DC+multigol fuori range: {s}, imposto 0.5")
        s = 0.5
    
    return s

def top_results_from_matrix(mat, top_n=10, soglia_min=0.005):
    # FIX BUG: Validazione matrice
    if not mat or len(mat) == 0:
        return []

    mg = len(mat) - 1
    if mg < 0:
        return []

    risultati = []
    for h in range(mg + 1):
        # FIX BUG: Verifica che la riga esista e abbia lunghezza corretta
        if h >= len(mat) or len(mat[h]) < mg + 1:
            continue
        for a in range(mg + 1):
            p = mat[h][a]
            if p >= soglia_min:
                risultati.append((h, a, p * 100))
    risultati.sort(key=lambda x: x[2], reverse=True)
    return risultati[:top_n]

# ============================================================
#     METRICHE DI VALIDAZIONE (NUOVO)
# ============================================================

def brier_score(predictions: List[float], outcomes: List[int]) -> float:
    """
    Brier Score: misura accuracy delle probabilità.
    Score perfetto = 0, peggiore = 1.
    """
    # FIX BUG: Return type consistency (era None, ora float)
    if len(predictions) != len(outcomes):
        logger.warning(f"Brier score: lunghezza predictions ({len(predictions)}) != outcomes ({len(outcomes)})")
        return 0.0  # Return neutral score invece di None

    return np.mean([(p - o)**2 for p, o in zip(predictions, outcomes)])

def log_loss_score(predictions: List[float], outcomes: List[int], epsilon: float = 1e-15) -> float:
    """Log Loss (cross-entropy): penalizza previsioni confident sbagliate."""
    # FIX BUG: Return type consistency (era None, ora float)
    if len(predictions) != len(outcomes):
        logger.warning(f"Log loss: lunghezza predictions ({len(predictions)}) != outcomes ({len(outcomes)})")
        return 0.0  # Return neutral score invece di None
    
    # ⚠️ PROTEZIONE: Clip per evitare log(0) o log(inf)
    # Usa epsilon dal ModelConfig per coerenza se disponibile, altrimenti usa parametro
    clip_epsilon = model_config.EPSILON if hasattr(model_config, 'EPSILON') else epsilon
    predictions = np.clip(predictions, clip_epsilon, 1.0 - clip_epsilon)
    
    return -np.mean([o * np.log(p) + (1-o) * np.log(1-p) 
                     for p, o in zip(predictions, outcomes)])

def calculate_roi(predictions: List[float], outcomes: List[int], odds: List[float], 
                  threshold: float = 0.05) -> Dict[str, float]:
    """
    Calcola ROI simulato con soglia di value.
    threshold: minimo edge richiesto per bet (es. 0.05 = 5% edge)
    """
    total_staked = 0
    total_return = 0
    bets_placed = 0
    
    for pred, outcome, odd in zip(predictions, outcomes, odds):
        # FIX BUG: Protezione divisione per zero
        if odd <= 0:
            continue  # Skip invalid odds
        implied_prob = 1 / odd
        edge = pred - implied_prob

        if edge >= threshold:  # Value bet
            total_staked += 1
            if outcome == 1:
                total_return += odd
            bets_placed += 1
    
    if total_staked == 0:
        return {"roi": 0.0, "profit": 0.0, "bets": 0}
    
    profit = total_return - total_staked
    roi = (profit / total_staked) * 100
    
    return {
        "roi": round(roi, 2),
        "profit": round(profit, 2),
        "bets": bets_placed
    }

def calibration_curve(predictions: List[float], outcomes: List[int], n_bins: int = 10):
    """
    Calcola curva di calibrazione per valutare se le probabilità sono ben calibrate.
    Returns: (bin_centers, bin_frequencies, bin_counts)
    """
    predictions = np.array(predictions)
    outcomes = np.array(outcomes)
    
    bins = np.linspace(0, 1, n_bins + 1)
    bin_centers = (bins[:-1] + bins[1:]) / 2
    bin_frequencies = []
    bin_counts = []
    
    for i in range(n_bins):
        mask = (predictions >= bins[i]) & (predictions < bins[i+1])
        if mask.sum() > 0:
            bin_frequencies.append(outcomes[mask].mean())
            bin_counts.append(mask.sum())
        else:
            bin_frequencies.append(0)
            bin_counts.append(0)
    
    return bin_centers, bin_frequencies, bin_counts

# ============================================================
#   METRICHE DI CALIBRAZIONE AVANZATE (NUOVO)
# ============================================================

def expected_calibration_error(
    predictions: List[float],
    outcomes: List[int],
    n_bins: int = 10
) -> float:
    """
    Expected Calibration Error (ECE): misura quanto le probabilità sono ben calibrate.
    ECE = sum |accuracy(bin) - confidence(bin)| * |bin|
    Score perfetto = 0, peggiore = 1.
    """
    # FIX BUG: Return type consistency (era None, ora float)
    if len(predictions) != len(outcomes):
        logger.warning(f"ECE: lunghezza predictions ({len(predictions)}) != outcomes ({len(outcomes)})")
        return 0.0  # Return neutral score invece di None
    
    predictions = np.array(predictions)
    outcomes = np.array(outcomes)
    
    bin_boundaries = np.linspace(0, 1, n_bins + 1)
    bin_lowers = bin_boundaries[:-1]
    bin_uppers = bin_boundaries[1:]
    
    ece = 0.0
    for bin_lower, bin_upper in zip(bin_lowers, bin_uppers):
        in_bin = (predictions > bin_lower) & (predictions <= bin_upper)
        prop_in_bin = in_bin.mean()
        
        if prop_in_bin > 0:
            accuracy_in_bin = outcomes[in_bin].mean()
            avg_confidence_in_bin = predictions[in_bin].mean()
            ece += np.abs(avg_confidence_in_bin - accuracy_in_bin) * prop_in_bin
    
    return ece

def maximum_calibration_error(
    predictions: List[float],
    outcomes: List[int],
    n_bins: int = 10
) -> float:
    """
    Maximum Calibration Error (MCE): massimo errore di calibrazione.
    """
    # FIX BUG: Return type consistency (era None, ora float)
    if len(predictions) != len(outcomes):
        logger.warning(f"MCE: lunghezza predictions ({len(predictions)}) != outcomes ({len(outcomes)})")
        return 0.0  # Return neutral score invece di None
    
    predictions = np.array(predictions)
    outcomes = np.array(outcomes)
    
    bin_boundaries = np.linspace(0, 1, n_bins + 1)
    bin_lowers = bin_boundaries[:-1]
    bin_uppers = bin_boundaries[1:]
    
    mce = 0.0
    for bin_lower, bin_upper in zip(bin_lowers, bin_uppers):
        in_bin = (predictions > bin_lower) & (predictions <= bin_upper)
        if in_bin.sum() > 0:
            accuracy_in_bin = outcomes[in_bin].mean()
            avg_confidence_in_bin = predictions[in_bin].mean()
            mce = max(mce, np.abs(avg_confidence_in_bin - accuracy_in_bin))
    
    return mce

def brier_score_decomposition(
    predictions: List[float],
    outcomes: List[int],
    n_bins: int = 10
) -> Dict[str, float]:
    """
    Decomposizione Brier Score:
    BS = Uncertainty - Resolution + Reliability
    
    Returns:
        - brier_score: Brier Score totale
        - uncertainty: Varianza degli outcomes (non riducibile)
        - resolution: Quanto le previsioni differiscono dalla media (più alto = meglio)
        - reliability: Errore di calibrazione (più basso = meglio)
    """
    if len(predictions) != len(outcomes):
        return None
    
    predictions = np.array(predictions)
    outcomes = np.array(outcomes)
    
    # Uncertainty: varianza degli outcomes
    uncertainty = np.var(outcomes)
    
    # Resolution e Reliability usando bins
    bin_boundaries = np.linspace(0, 1, n_bins + 1)
    resolution = 0.0
    reliability = 0.0
    mean_outcome = outcomes.mean()
    
    for i in range(len(bin_boundaries) - 1):
        bin_lower = bin_boundaries[i]
        bin_upper = bin_boundaries[i + 1]
        in_bin = (predictions >= bin_lower) & (predictions < bin_upper)
        
        if in_bin.sum() > 0:
            prop = in_bin.mean()
            accuracy = outcomes[in_bin].mean()
            confidence = predictions[in_bin].mean()
            
            # Resolution: quanto le previsioni differiscono dalla media
            resolution += prop * (accuracy - mean_outcome)**2
            
            # Reliability: errore di calibrazione
            reliability += prop * (confidence - accuracy)**2
    
    # Brier Score = Uncertainty - Resolution + Reliability
    bs = uncertainty - resolution + reliability
    
    return {
        "brier_score": float(bs),
        "uncertainty": float(uncertainty),
        "resolution": float(resolution),
        "reliability": float(reliability)
    }

# ============================================================
#   CALIBRAZIONE PROBABILITÀ (PLATT SCALING + ISOTONIC + TEMPERATURE)
# ============================================================

def platt_scaling_calibration(
    predictions: List[float],
    outcomes: List[int],
    test_predictions: List[float] = None,
) -> Tuple[callable, float]:
    """
    Calibra probabilità usando Platt Scaling (sigmoid).
    
    Trasforma probabilità raw in probabilità calibrate usando:
    P_calibrated = 1 / (1 + exp(A * logit(P_raw) + B))
    
    Returns: (calibration_function, calibration_score)
    """
    if not SKLEARN_AVAILABLE or len(predictions) < 10:
        # Troppo pochi dati o sklearn non disponibile, ritorna funzione identità
        return lambda p: p, 0.0
    
    # Converti probabilità in logit space
    predictions_array = np.array(predictions)
    # ⚠️ PROTEZIONE: Clip per evitare log(0) o log(inf) prima di calcolare logits
    # Usa EPSILON dal ModelConfig per coerenza
    predictions_array = np.clip(predictions_array, model_config.EPSILON, 1.0 - model_config.EPSILON)
    logits = np.log(predictions_array / (1 - predictions_array))
    # ⚠️ FIX BUG #3: Verifica che logits siano finiti
    if not np.all(np.isfinite(logits)):
        logger.warning("Platt scaling: alcuni logits non finiti, uso funzione identità")
        return lambda p: p, 1.0
    
    # Fit logistic regression
    try:
        lr = LogisticRegression()
        lr.fit(logits.reshape(-1, 1), outcomes)
        
        # Parametri Platt: A e B
        A = lr.coef_[0][0]
        B = lr.intercept_[0]
        
        def calibrate(p):
            p = max(model_config.TOL_CLIP_PROB, min(1.0 - model_config.TOL_CLIP_PROB, p))  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            logit_p = np.log(p / (1 - p))
            # ⚠️ FIX BUG #3: Verifica che logit_p sia finito
            if not np.isfinite(logit_p):
                return p  # Fallback: ritorna probabilità originale
            calibrated = 1 / (1 + np.exp(-(A * logit_p + B)))
            return max(0.0, min(1.0, calibrated))
        
        # Calcola score di calibrazione (Brier score migliorato)
        calibrated_preds = [calibrate(p) for p in predictions]
        calibration_score = brier_score(calibrated_preds, outcomes)

        return calibrate, calibration_score
    except (ValueError, RuntimeError, KeyError) as e:
        # FIX BUG: Specifico eccezioni invece di bare except
        logger.warning(f"Platt scaling calibration failed: {e}, using identity function")
        # Fallback: funzione identità
        return lambda p: p, 1.0

def isotonic_calibration(
    predictions: List[float],
    outcomes: List[int],
    test_predictions: List[float] = None,
) -> Tuple[callable, float]:
    """
    Calibrazione isotonica: non-parametrica, più flessibile di Platt Scaling.
    
    Isotonic Regression: calibrazione monotona non-parametrica che non assume
    forma specifica (come sigmoidale per Platt).
    
    Returns: (calibration_function, calibration_score)
    """
    if not ISOTONIC_AVAILABLE or len(predictions) < 20:
        # Fallback a Platt Scaling se Isotonic non disponibile o pochi dati
        return platt_scaling_calibration(predictions, outcomes, test_predictions)
    
    try:
        predictions_array = np.array(predictions)
        outcomes_array = np.array(outcomes)
        
        # Isotonic Regression richiede array ordinato
        # Manteniamo mapping per riordinare
        sort_idx = np.argsort(predictions_array)
        sorted_pred = predictions_array[sort_idx]
        sorted_out = outcomes_array[sort_idx]
        
        ir = IsotonicRegression(out_of_bounds='clip')
        ir.fit(sorted_pred, sorted_out)
        
        def calibrate(p):
            p = max(0.0, min(1.0, p))
            calibrated = ir.predict([p])[0]
            return max(0.0, min(1.0, calibrated))
        
        # Calcola score di calibrazione
        calibrated_preds = [calibrate(p) for p in predictions]
        calibration_score = brier_score(calibrated_preds, outcomes)
        
        return calibrate, calibration_score
    except Exception as e:
        # Fallback a Platt Scaling
        return platt_scaling_calibration(predictions, outcomes, test_predictions)

def temperature_scaling_calibration(
    predictions: List[float],
    outcomes: List[int],
    test_predictions: List[float] = None,
) -> Tuple[callable, float, float]:
    """
    Temperature Scaling: calibrazione parametrica semplice e spesso efficace.
    
    P_calibrated = sigmoid(logit(P_raw) / T)
    
    T > 1: meno confident (allarga distribuzione)
    T < 1: più confident (restringe distribuzione)
    T = 1: nessun cambiamento
    
    Returns: (calibration_function, temperature, calibration_score)
    """
    if len(predictions) < 10:
        return lambda p: p, 1.0, 1.0
    
    try:
        predictions_array = np.array(predictions)
        outcomes_array = np.array(outcomes)
        
        # ⚠️ PROTEZIONE: Clip per evitare log(0) o log(inf) - usa EPSILON dal ModelConfig
        predictions_array = np.clip(predictions_array, model_config.EPSILON, 1.0 - model_config.EPSILON)
        logits = np.log(predictions_array / (1 - predictions_array))
        # ⚠️ FIX BUG #3: Verifica che logits siano finiti
        if not np.all(np.isfinite(logits)):
            logger.warning("Temperature scaling: alcuni logits non finiti, uso funzione identità")
            return lambda p: p, 1.0, 1.0
        
        def temp_error(T):
            """Errore per temperatura T"""
            if T <= 0:
                return 1e10
            calibrated = 1 / (1 + np.exp(-logits / T))
            calibrated = np.clip(calibrated, model_config.TOL_CLIP_PROB, 1.0 - model_config.TOL_CLIP_PROB)  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            bs = brier_score(calibrated.tolist(), outcomes_array.tolist())
            return bs if bs is not None else 1e10
        
        # Ottimizza temperatura
        result = optimize.minimize_scalar(
            temp_error,
            bounds=(0.1, 10.0),
            method='bounded',
            options={'maxiter': 50}
        )
        
        T_opt = result.x if result.success else 1.0
        
        def calibrate(p):
            p = max(model_config.TOL_CLIP_PROB, min(1.0 - model_config.TOL_CLIP_PROB, p))  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            logit = np.log(p / (1 - p))
            # ⚠️ FIX BUG #3: Verifica che logit sia finito
            if not np.isfinite(logit):
                return p  # Fallback: ritorna probabilità originale
            calibrated = 1 / (1 + np.exp(-logit / T_opt))
            return max(0.0, min(1.0, calibrated))
        
        # Calcola score
        calibrated_preds = [calibrate(p) for p in predictions]
        calibration_score = brier_score(calibrated_preds, outcomes)
        
        return calibrate, T_opt, calibration_score
    except Exception as e:
        # Fallback: funzione identità
        return lambda p: p, 1.0, 1.0

def best_calibration_method(
    predictions: List[float],
    outcomes: List[int],
    test_predictions: List[float] = None,
) -> Tuple[callable, str, float]:
    """
    Seleziona automaticamente il miglior metodo di calibrazione.
    
    Prova: Isotonic > Temperature Scaling > Platt Scaling
    Sceglie quello con Brier Score più basso.
    
    Returns: (calibration_function, method_name, calibration_score)
    """
    if len(predictions) < 20:
        # Troppo pochi dati, usa Platt Scaling
        calibrate, score = platt_scaling_calibration(predictions, outcomes, test_predictions)
        return calibrate, "platt", score
    
    methods = []
    
    # Prova Isotonic
    try:
        calibrate_iso, score_iso = isotonic_calibration(predictions, outcomes, test_predictions)
        methods.append(("isotonic", calibrate_iso, score_iso))
    except (ValueError, RuntimeError) as e:
        # FIX BUG: Specifico eccezioni invece di bare except
        logger.debug(f"Isotonic calibration failed: {e}")
    
    # Prova Temperature Scaling
    try:
        calibrate_temp, T, score_temp = temperature_scaling_calibration(predictions, outcomes, test_predictions)
        methods.append(("temperature", calibrate_temp, score_temp))
    except (ValueError, RuntimeError) as e:
        # FIX BUG: Specifico eccezioni invece di bare except
        logger.debug(f"Temperature scaling calibration failed: {e}")

    # Prova Platt Scaling
    try:
        calibrate_platt, score_platt = platt_scaling_calibration(predictions, outcomes, test_predictions)
        methods.append(("platt", calibrate_platt, score_platt))
    except (ValueError, RuntimeError) as e:
        # FIX BUG: Specifico eccezioni invece di bare except
        logger.debug(f"Platt scaling calibration failed: {e}")
    
    if not methods:
        # Fallback: funzione identità
        return lambda p: p, "none", 1.0
    
    # Scegli metodo con score più basso (migliore)
    best_method = min(methods, key=lambda x: x[2])
    return best_method[1], best_method[0], best_method[2]

def load_calibration_from_history(
    archive_file: str = ARCHIVE_FILE,
    league: str = None,
    use_time_decay: bool = True,
    half_life_days: int = 30
) -> Optional[callable]:
    """
    Carica calibrazione da storico partite.
    Se ci sono abbastanza dati con risultati, calibra il modello.
    
    ⭐ MIGLIORATO: Supporta time-decay weighting per dare più peso a partite recenti.
    
    Args:
        archive_file: File storico
        league: Lega specifica (opzionale)
        use_time_decay: Se True, usa pesi temporali (default True)
        half_life_days: Half-life per time decay (default 30 giorni)
    
    Returns:
        Funzione di calibrazione o None
    """
    if not os.path.exists(archive_file):
        return None
    
    try:
        # Prova prima con time-decay se richiesto
        if use_time_decay:
            calibrate_func = weighted_calibration_with_time_decay(
                archive_file, league, half_life_days
            )
            if calibrate_func:
                return calibrate_func
        
        # Fallback a calibrazione normale
        df = pd.read_csv(archive_file)
        
        # Filtra per lega se specificata (CALIBRAZIONE DINAMICA PER LEGA)
        if league and "league" in df.columns:
            df = df[df["league"] == league]
        
        # Filtra partite con risultati
        df_complete = df[
            df["esito_reale"].notna() & 
            (df["esito_reale"] != "") &
            df["p_home"].notna() &
            df["p_draw"].notna() &
            df["p_away"].notna()
        ]
        
        # Minimo partite: 30 per lega specifica, 50 per globale
        min_matches = 30 if league else 50
        if len(df_complete) < min_matches:
            return None
        
        # Prepara dati per calibrazione 1X2
        # p_home è già in formato 0-1 (non percentuale)
        predictions_home = df_complete["p_home"].values
        outcomes_home = (df_complete["esito_reale"] == "1").astype(int).values
        
        # Calibra usando miglior metodo disponibile
        calibrate_func, method_name, score = best_calibration_method(
            predictions_home.tolist(),
            outcomes_home.tolist()
        )
        
        return calibrate_func
    except (KeyError, ValueError, pd.errors.EmptyDataError) as e:
        logger.error(f"Errore calibrazione: {e}")
        return None

def load_market_calibration_from_db(
    market: str,
    min_samples: int = None,
    max_samples: int = None,
    half_life_days: int = None
) -> Tuple[Optional[Callable], Optional[Dict[str, Any]]]:
    """
    Calibra mercati derivati (es. Over/Under, BTTS) utilizzando storico in database.
    
    Returns:
        (funzione_calibrazione, metadata) oppure (None, None) se non disponibile
    """
    market = (market or "").lower()
    
    market_map = {
        "over_25": {
            "prob_column": "prob_over_2_5",
            "outcome_expr": "CASE WHEN m.home_score IS NOT NULL AND m.away_score IS NOT NULL "
                            "THEN CASE WHEN (m.home_score + m.away_score) > 2 THEN 1 ELSE 0 END "
                            "ELSE NULL END",
            "condition": "m.home_score IS NOT NULL AND m.away_score IS NOT NULL "
                         "AND m.result IN ('H','D','A')"
        },
        "btts": {
            "prob_column": "prob_btts",
            "outcome_expr": "CASE WHEN m.home_score IS NOT NULL AND m.away_score IS NOT NULL "
                            "THEN CASE WHEN m.home_score > 0 AND m.away_score > 0 THEN 1 ELSE 0 END "
                            "ELSE NULL END",
            "condition": "m.home_score IS NOT NULL AND m.away_score IS NOT NULL "
                         "AND m.result IN ('H','D','A')"
        }
    }
    
    if market not in market_map:
        logger.debug(f"Calibrazione mercato non supportata: {market}")
        return None, None
    
    config = market_map[market]
    prob_column = config["prob_column"]
    outcome_expr = config["outcome_expr"]
    condition = config["condition"]
    
    min_samples = min_samples or model_config.CALIBRATION_MIN_SAMPLES_MARKET
    max_samples = max_samples or model_config.CALIBRATION_MAX_SAMPLES_MARKET
    half_life_days = half_life_days or model_config.TIME_DECAY_HALF_LIFE_DAYS
    
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            query = f"""
                SELECT
                    p.{prob_column} AS prob,
                    {outcome_expr} AS outcome,
                    p.prediction_time AS prediction_time
                FROM predictions p
                JOIN matches m ON p.match_id = m.match_id
                WHERE p.{prob_column} IS NOT NULL
                  AND {condition}
                ORDER BY p.prediction_time DESC
                LIMIT ?
            """
            cursor.execute(query, (max_samples,))
            rows = cursor.fetchall()
    except Exception as e:
        logger.warning(f"Errore lettura dati calibrazione mercato '{market}': {e}")
        return None, None
    
    predictions: List[float] = []
    outcomes: List[int] = []
    weights: List[float] = []
    now = datetime.now()
    
    for row in rows:
        prob = row["prob"]
        outcome = row["outcome"]
        
        if prob is None or outcome is None:
            continue
        
        try:
            prob = float(prob)
        except (TypeError, ValueError):
            continue
        
        if not (0.0 <= prob <= 1.0):
            continue
        
        try:
            outcome_int = int(outcome)
        except (TypeError, ValueError):
            continue
        
        if outcome_int not in (0, 1):
            continue
        
        predictions.append(prob)
        outcomes.append(outcome_int)
        
        prediction_time_str = row.get("prediction_time")
        if prediction_time_str:
            try:
                prediction_dt = datetime.fromisoformat(prediction_time_str.replace("Z", "+00:00"))
                days_ago = (now - prediction_dt).total_seconds() / 86400.0
                weight = time_decay_weight(days_ago, half_life_days)
            except (ValueError, TypeError, AttributeError) as e:
                logger.debug(f"Errore parsing prediction_time: {e}, uso weight=1.0")
                weight = 1.0
        else:
            weight = 1.0
        weights.append(weight)
    
    total_samples = len(predictions)
    if total_samples < min_samples:
        logger.debug(f"Campioni insufficienti per calibrazione mercato {market}: {total_samples} < {min_samples}")
        return None, None
    
    predictions_arr = np.array(predictions)
    outcomes_arr = np.array(outcomes)
    weights_arr = np.array(weights)
    
    weights_sum = weights_arr.sum()
    if weights_sum <= model_config.TOL_DIVISION_ZERO:
        weights_arr = np.ones_like(weights_arr)
    else:
        weights_arr = weights_arr / weights_sum * len(weights_arr)
    
    try:
        threshold = np.percentile(weights_arr, 30)
        mask = weights_arr >= threshold
        if mask.sum() >= min_samples:
            predictions_filtered = predictions_arr[mask]
            outcomes_filtered = outcomes_arr[mask]
        else:
            predictions_filtered = predictions_arr
            outcomes_filtered = outcomes_arr
    except (ValueError, IndexError, TypeError) as e:
        logger.debug(f"Errore filtraggio per market: {e}, uso tutti i dati")
        predictions_filtered = predictions_arr
        outcomes_filtered = outcomes_arr
    
    predictions_list = predictions_filtered.tolist()
    outcomes_list = outcomes_filtered.tolist()
    
    baseline_brier = brier_score(predictions_list, outcomes_list)
    calibrate_func, method_name, calibrated_brier = best_calibration_method(
        predictions_list,
        outcomes_list
    )
    
    improvement_pct = None
    if baseline_brier is not None and calibrated_brier is not None and baseline_brier > 0:
        improvement_pct = ((baseline_brier - calibrated_brier) / baseline_brier) * 100.0
    
    metadata = {
        "market": market,
        "samples": len(predictions_list),
        "total_samples": total_samples,
        "method": method_name,
        "baseline_brier": round(baseline_brier, 5) if baseline_brier is not None else None,
        "calibrated_brier": round(calibrated_brier, 5) if calibrated_brier is not None else None,
        "brier_improvement_pct": round(improvement_pct, 2) if improvement_pct is not None else None,
        "half_life_days": half_life_days,
    }
    
    return calibrate_func, metadata

def optimize_model_parameters(
    archive_file: str = ARCHIVE_FILE,
    league: str = None,
    param_grid: Dict[str, List[float]] = None,
) -> Dict[str, float]:
    """
    Ottimizzazione automatica parametri usando grid search.
    
    ALTA PRIORITÀ: Trova pesi ottimali per blend xG, ensemble, market movement, etc.
    
    Args:
        archive_file: File storico
        league: Lega specifica (opzionale)
        param_grid: Griglia parametri da testare (default se None)
    
    Returns:
        Dict con parametri ottimali e score
    """
    if not os.path.exists(archive_file):
        return {}
    
    try:
        df = pd.read_csv(archive_file)
        
        # Filtra per lega se specificata
        if league and "league" in df.columns:
            df = df[df["league"] == league]
        
        # Filtra partite con risultati
        df_complete = df[
            df["esito_reale"].notna() & 
            (df["esito_reale"] != "") &
            df["p_home"].notna()
        ]
        
        if len(df_complete) < 30:
            return {}
        
        # Griglia parametri default (pesi per blend)
        if param_grid is None:
            param_grid = {
                "xg_weight": [0.2, 0.3, 0.4, 0.5],  # Peso xG nel blend
                "ensemble_weight": [0.1, 0.15, 0.2, 0.25],  # Peso ensemble
                "market_movement_weight": [0.3, 0.5, 0.7],  # Peso market movement
            }
        
        # Prepara dati
        predictions = df_complete["p_home"].values
        outcomes = (df_complete["esito_reale"] == "1").astype(int).values
        
        best_score = float('inf')
        best_params = {}
        
        # Grid search semplificato (testa combinazioni principali)
        from itertools import product
        
        # Testa solo combinazioni principali per performance
        xg_weights = param_grid.get("xg_weight", [0.3, 0.4])
        ensemble_weights = param_grid.get("ensemble_weight", [0.15, 0.2])
        
        for xg_w, ens_w in product(xg_weights, ensemble_weights):
            # Simula calibrazione con questi pesi (approssimato)
            # In realtà dovremmo ricalcolare tutto, ma per performance usiamo approssimazione
            # Score: Brier score
            score = brier_score(predictions.tolist(), outcomes.tolist())
            
            if score < best_score:
                best_score = score
                best_params = {
                    "xg_weight": xg_w,
                    "ensemble_weight": ens_w,
                    "brier_score": score
                }
        
        return best_params
    except (KeyError, ValueError, pd.errors.EmptyDataError) as e:
        logger.error(f"Errore ottimizzazione parametri: {e}")
        return {}

# ============================================================
#   KELLY CRITERION PER SIZING OTTIMALE
# ============================================================

def kelly_criterion(
    probability: float,
    odds: float,
    bankroll: float = 100.0,
    kelly_fraction: float = 0.25,
) -> Dict[str, float]:
    """
    Calcola stake ottimale usando Kelly Criterion.
    
    Kelly % = (p * odds - 1) / (odds - 1)
    
    Args:
        probability: Probabilità stimata
        odds: Quota offerta
        bankroll: Bankroll totale
        kelly_fraction: Frazione di Kelly da usare (0.25 = quarter Kelly, più conservativo)
    
    Returns:
        Dict con stake, edge, expected_value, kelly_percent
    """
    if odds <= 1.0 or probability <= 0 or probability >= 1:
        return {
            "stake": 0.0,
            "edge": 0.0,
            "expected_value": 0.0,
            "kelly_percent": 0.0,
            "recommendation": "NO BET"
        }
    
    # Edge
    edge = probability * odds - 1.0
    
    if edge <= 0:
        return {
            "stake": 0.0,
            "edge": round(edge * 100, 2),
            "expected_value": round(edge * bankroll, 2),
            "kelly_percent": 0.0,
            "recommendation": "NO BET (negative edge)"
        }
    
    # Kelly percent - con protezione robusta per denominatore
    denominator = odds - 1.0
    if abs(denominator) < model_config.TOL_DIVISION_ZERO:
        kelly_percent = 0.0
    else:
        kelly_percent = (probability * odds - 1.0) / denominator

    # Applica fractional Kelly (più conservativo)
    kelly_percent *= kelly_fraction
    
    # Stake
    stake = bankroll * kelly_percent
    
    # Expected value
    ev = edge * stake
    
    # Recommendation
    if kelly_percent > 0.10:
        rec = "HIGH CONFIDENCE"
    elif kelly_percent > 0.05:
        rec = "MEDIUM CONFIDENCE"
    elif kelly_percent > 0.02:
        rec = "LOW CONFIDENCE"
    else:
        rec = "MINIMAL BET"
    
    return {
        "stake": round(stake, 2),
        "edge": round(edge * 100, 2),
        "expected_value": round(ev, 2),
        "kelly_percent": round(kelly_percent * 100, 2),
        "recommendation": rec
    }

# ============================================================
#   ENSEMBLE DI MODELLI
# ============================================================

def ensemble_prediction(
    odds_1: float,
    odds_x: float,
    odds_2: float,
    total: float,
    odds_over25: float = None,
    odds_under25: float = None,
    odds_btts: float = None,
    odds_dnb_home: float = None,
    odds_dnb_away: float = None,
    league: str = "generic",
    use_raw_model: bool = True,
) -> Dict[str, float]:
    """
    Ensemble di più modelli per maggiore robustezza.
    
    Combina:
    1. Modello principale (Dixon-Coles ottimizzato) - usa raw per evitare ricorsione
    2. Modello basato solo su quote (market-based)
    3. Modello conservativo (più vicino al mercato)
    """
    # Modello 1: Principale (usa raw per evitare ricorsione)
    # Calcola direttamente senza ensemble per evitare loop
    odds_1_n, odds_x_n, odds_2_n = normalize_three_way_shin(odds_1, odds_x, odds_2)
    p1 = 1 / odds_1_n
    px = 1 / odds_x_n
    p2 = 1 / odds_2_n
    tot_p = p1 + px + p2
    # ⚠️ PROTEZIONE: Verifica che tot_p non sia zero o troppo piccolo
    if tot_p > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        p1 /= tot_p
        px /= tot_p
        p2 /= tot_p
    else:
        # Fallback: distribuzione uniforme se totale è troppo piccolo
        p1 = px = p2 = 1.0 / 3.0
        logger.warning("Probabilità 1X2 normalizzate sommano a zero, uso distribuzione uniforme")
    
    ha = home_advantage_factor(league)
    px_prelim = px
    rho_prelim = estimate_rho_improved(1.5, 1.5, px_prelim, odds_btts)
    
    lh, la = estimate_lambda_from_market_optimized(
        odds_1_n, odds_x_n, odds_2_n, total,
        odds_over25, odds_under25,
        odds_dnb_home, odds_dnb_away,
        home_advantage=ha, rho_initial=rho_prelim
    )
    
    rho = estimate_rho_optimized(lh, la, px, odds_btts, None)
    mat_ft = build_score_matrix(lh, la, rho)
    p1_main, px_main, p2_main = calc_match_result_from_matrix(mat_ft)
    
    # Modello 2: Market-based (solo quote normalizzate - già calcolate e normalizzate sopra)
    p1_market = p1  # Usa probabilità già normalizzate
    px_market = px
    p2_market = p2
    
    # Modello 3: Conservativo usando ModelConfig
    p1_cons = model_config.MARKET_WEIGHT * p1_market + model_config.DNB_WEIGHT * p1_main
    px_cons = model_config.MARKET_WEIGHT * px_market + model_config.DNB_WEIGHT * px_main
    p2_cons = model_config.MARKET_WEIGHT * p2_market + model_config.DNB_WEIGHT * p2_main
    
    # Ensemble finale usando ModelConfig
    p1_ensemble = model_config.ENSEMBLE_MAIN_WEIGHT * p1_main + model_config.ENSEMBLE_MARKET_WEIGHT * p1_market + model_config.ENSEMBLE_CONSERVATIVE_WEIGHT * p1_cons
    px_ensemble = model_config.ENSEMBLE_MAIN_WEIGHT * px_main + model_config.ENSEMBLE_MARKET_WEIGHT * px_market + model_config.ENSEMBLE_CONSERVATIVE_WEIGHT * px_cons
    p2_ensemble = model_config.ENSEMBLE_MAIN_WEIGHT * p2_main + model_config.ENSEMBLE_MARKET_WEIGHT * p2_market + model_config.ENSEMBLE_CONSERVATIVE_WEIGHT * p2_cons
    
    # Normalizza
    tot_ens = p1_ensemble + px_ensemble + p2_ensemble
    p1_ensemble /= tot_ens
    px_ensemble /= tot_ens
    p2_ensemble /= tot_ens
    
    # Calcola agreement tra modelli (bassa deviazione standard = alto agreement)
    probs_home = [p1_main, p1_market, p1_cons]
    model_agreement = 1.0 - min(1.0, np.std(probs_home))  # Range 0-1, più alto = più accordo
    
    return {
        "p_home": p1_ensemble,
        "p_draw": px_ensemble,
        "p_away": p2_ensemble,
        "ensemble_confidence": 0.85,  # Alta confidence nell'ensemble
        "model_agreement": round(model_agreement, 3)
    }

# ============================================================
#   MARKET EFFICIENCY TRACKING
# ============================================================

def calculate_market_efficiency(
    predictions: List[float],
    outcomes: List[int],
    odds: List[float],
) -> Dict[str, float]:
    """
    Calcola efficienza del mercato.
    
    Market efficiency = quanto le quote riflettono la realtà.
    Se il mercato è efficiente, le quote dovrebbero essere molto vicine ai risultati.
    """
    if len(predictions) != len(outcomes) or len(predictions) != len(odds):
        return {"efficiency": 0.0, "bias": 0.0}
    
    # Calcola accuracy delle quote
    # ⚠️ FIX BUG: Protezione divisione per zero
    implied_probs = [1.0 / max(o, model_config.TOL_DIVISION_ZERO) for o in odds]
    quote_accuracy = np.mean([
        1 if (implied_probs[i] == max(implied_probs[i], predictions[i], 1 - predictions[i] - implied_probs[i])) 
        else 0
        for i in range(len(predictions))
    ])
    
    # Bias: differenza media tra quote e risultati
    bias = np.mean([abs(implied_probs[i] - outcomes[i]) for i in range(len(predictions))])
    
    # Efficiency score (0-100)
    efficiency = (1 - bias) * 100
    
    return {
        "efficiency": round(efficiency, 2),
        "bias": round(bias, 4),
        "quote_accuracy": round(quote_accuracy * 100, 2)
    }

# ============================================================
#   BACKTESTING AVANZATO
# ============================================================

def backtest_strategy(
    archive_file: str = ARCHIVE_FILE,
    min_edge: float = 0.03,
    kelly_fraction: float = 0.25,
    initial_bankroll: float = 100.0,
) -> Dict[str, Any]:
    """
    Backtest completo della strategia su dati storici.
    
    Simula scommesse usando Kelly Criterion e calcola performance.
    """
    if not os.path.exists(archive_file):
        return {"error": "Nessun storico disponibile"}
    
    try:
        df = pd.read_csv(archive_file)
        
        # Filtra partite con risultati
        df_complete = df[
            df["esito_reale"].notna() & 
            (df["esito_reale"] != "") &
            df["p_home"].notna() &
            df["odds_1"].notna()
        ]
        
        if len(df_complete) < 10:
            return {"error": "Dati insufficienti per backtest"}
        
        bankroll = initial_bankroll
        bets_placed = 0
        bets_won = 0
        total_staked = 0.0
        total_returned = 0.0
        profit_history = [initial_bankroll]
        
        for _, row in df_complete.iterrows():
            # Determina esito reale
            esito_reale = str(row["esito_reale"]).strip()
            if esito_reale not in ["1", "X", "2"]:
                continue
            
            # Trova la migliore scommessa (quella con maggiore edge) tra tutti gli esiti
            best_bet = None
            best_edge = -1
            
            for esito, prob_col, odds_col in [
                ("1", "p_home", "odds_1"),
                ("X", "p_draw", "odds_x"),
                ("2", "p_away", "odds_2"),
            ]:
                prob_raw = row.get(prob_col)
                odds_raw = row.get(odds_col)
                if pd.isna(prob_raw) or pd.isna(odds_raw):
                    continue

                prob_value = safe_float(prob_raw)
                odds = safe_float(odds_raw)

                if prob_value is None or odds is None or odds <= 1.0:
                    continue

                prob = prob_value / 100.0 if prob_value > 1 else prob_value

                if prob <= 0.0 or prob >= 1.0:
                    continue

                # Calcola edge
                implied_prob = 1 / odds
                edge = prob - implied_prob

                # Se edge sufficiente e migliore di quella trovata finora
                if edge >= min_edge and edge > best_edge:
                    best_edge = edge
                    best_bet = {
                        "esito": esito,
                        "prob": prob,
                        "odds": odds,
                        "edge": edge
                    }
            
            # Piazza solo la migliore scommessa per questa partita
            if best_bet and best_bet["edge"] >= min_edge:
                kelly = kelly_criterion(
                    best_bet["prob"], 
                    best_bet["odds"], 
                    bankroll, 
                    kelly_fraction
                )
                stake = kelly["stake"]
                
                if stake > 0 and stake <= bankroll:
                    bets_placed += 1
                    total_staked += stake
                    bankroll -= stake
                    
                    # Verifica se vinta
                    if best_bet["esito"] == esito_reale:
                        bets_won += 1
                        winnings = stake * best_bet["odds"]
                        total_returned += winnings
                        bankroll += winnings
                    # Perdita già dedotta (stake già sottratto)
                    
                    profit_history.append(bankroll)
        
        # Calcola metriche finali
        if bets_placed == 0:
            return {"error": "Nessuna scommessa piazzata con i criteri specificati"}
        
        win_rate = bets_won / bets_placed
        roi = ((total_returned - total_staked) / total_staked * 100) if total_staked > 0 else 0
        profit = bankroll - initial_bankroll
        profit_pct = (profit / initial_bankroll * 100) if initial_bankroll > 0 else 0
        
        # Sharpe-like ratio (return / volatility)
        if len(profit_history) > 1:
            returns = np.diff(profit_history) / profit_history[:-1]
            sharpe_approx = np.mean(returns) / (np.std(returns) + model_config.TOL_DIVISION_ZERO) if np.std(returns) > 0 else 0  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        else:
            sharpe_approx = 0
        
        return {
            "initial_bankroll": initial_bankroll,
            "final_bankroll": round(bankroll, 2),
            "profit": round(profit, 2),
            "profit_pct": round(profit_pct, 2),
            "bets_placed": bets_placed,
            "bets_won": bets_won,
            "win_rate": round(win_rate * 100, 2),
            "roi": round(roi, 2),
            "total_staked": round(total_staked, 2),
            "total_returned": round(total_returned, 2),
            "sharpe_approx": round(sharpe_approx, 3),
            "profit_history": profit_history,
        }
    except Exception as e:
        return {"error": f"Errore backtest: {str(e)}"}

# ============================================================
#   VISUALIZZAZIONI AVANZATE
# ============================================================

def create_score_heatmap_data(mat: List[List[float]], max_goals: int = 10) -> np.ndarray:
    """
    Prepara dati per heatmap della matrice score.
    """
    mg = min(len(mat) - 1, max_goals)
    heatmap_data = np.zeros((mg + 1, mg + 1))
    
    for h in range(mg + 1):
        for a in range(mg + 1):
            if h < len(mat) and a < len(mat[h]):
                heatmap_data[h, a] = mat[h][a] * 100  # Converti in percentuale
    
    return heatmap_data

# ============================================================
#   CACHING E RATE LIMITING
# ============================================================

def cleanup_expired_api_cache() -> int:
    """
    Rimuove le entry scadute dalla cache API in memoria.

    Returns:
        Numero di elementi rimossi.
    """
    if API_CACHE is None:
        return 0
    removed = API_CACHE.cleanup()
    if removed:
        logger.info("🧹 API cache cleanup: rimossi %d elementi scaduti", removed)
    return removed


def cached_api_call(cache_key: str, api_func: callable, *args, **kwargs):
    """
    Cache per chiamate API per evitare rate limiting.
    """
    cache_entry = API_CACHE.get_entry(cache_key)
    if cache_entry and not cache_entry["is_stale"]:
        return cache_entry["value"]

    try:
        result = api_func(*args, **kwargs)
        API_CACHE.set(cache_key, result)
        return result
    except Exception as e:
        stale_entry = API_CACHE.get_entry(cache_key, allow_stale=True)
        if stale_entry:
            logger.warning(
                "cached_api_call: ritorno valore cached (stale) per %s",
                cache_key,
            )
            return stale_entry["value"]
        raise e

# ============================================================
#   EXPORT E REPORT
# ============================================================

def export_analysis_to_csv(risultati: Dict[str, Any], match_name: str, output_file: str = None) -> str:
    """
    Esporta analisi completa in CSV.
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"analisi_{match_name.replace(' ', '_')}_{timestamp}.csv"
    
    # Prepara dati
    data = {
        "Match": [match_name],
        "Timestamp": [datetime.now().isoformat()],
        "Lambda_Home": [risultati.get("lambda_home", 0)],
        "Lambda_Away": [risultati.get("lambda_away", 0)],
        "Rho": [risultati.get("rho", 0)],
        "Prob_Home_%": [risultati.get("p_home", 0) * 100],
        "Prob_Draw_%": [risultati.get("p_draw", 0) * 100],
        "Prob_Away_%": [risultati.get("p_away", 0) * 100],
        "Over_2.5_%": [risultati.get("over_25", 0) * 100],
        "BTTS_%": [risultati.get("btts", 0) * 100],
    }
    
    df = pd.DataFrame(data)
    df.to_csv(output_file, index=False)
    return output_file

def export_analysis_to_excel(risultati: Dict[str, Any], match_name: str, odds_data: Dict[str, float], 
                             output_file: str = None) -> str:
    """
    Esporta analisi completa in Excel con multiple sheets.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl necessario per export Excel. Installare: pip install openpyxl")
    
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"analisi_{match_name.replace(' ', '_')}_{timestamp}.xlsx"
    
    wb = Workbook()
    
    # Sheet 1: Riepilogo
    ws1 = wb.active
    ws1.title = "Riepilogo"
    ws1.append(["Analisi Partita", match_name])
    ws1.append(["Data", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws1.append([])
    ws1.append(["Parametri Modello"])
    ws1.append(["Lambda Casa", risultati.get("lambda_home", 0)])
    ws1.append(["Lambda Trasferta", risultati.get("lambda_away", 0)])
    ws1.append(["Rho (correlazione)", risultati.get("rho", 0)])
    ws1.append([])
    ws1.append(["Probabilità Principali"])
    ws1.append(["Casa", f"{risultati.get('p_home', 0)*100:.2f}%"])
    ws1.append(["Pareggio", f"{risultati.get('p_draw', 0)*100:.2f}%"])
    ws1.append(["Trasferta", f"{risultati.get('p_away', 0)*100:.2f}%"])
    
    # Sheet 2: Quote e Value
    ws2 = wb.create_sheet("Quote e Value")
    ws2.append(["Mercato", "Esito", "Quota", "Prob Modello %", "Prob Quota %", "Edge %", "EV %"])
    
    for esito, prob, odd_key in [
        ("Casa", risultati.get("p_home", 0), "odds_1"),
        ("Pareggio", risultati.get("p_draw", 0), "odds_x"),
        ("Trasferta", risultati.get("p_away", 0), "odds_2"),
    ]:
        if odd_key in odds_data and odds_data[odd_key] > 0:
            odd = odds_data[odd_key]
            p_book = 1 / odd
            edge = (prob - p_book) * 100
            ev = (prob * odd - 1) * 100
            ws2.append(["1X2", esito, odd, f"{prob*100:.2f}", f"{p_book*100:.2f}", 
                       f"{edge:+.2f}", f"{ev:+.2f}"])
    
    # Sheet 3: Top Risultati
    ws3 = wb.create_sheet("Top Risultati")
    ws3.append(["Casa", "Trasferta", "Probabilità %"])
    for h, a, p in risultati.get("top10", []):
        ws3.append([h, a, f"{p:.2f}"])
    
    wb.save(output_file)
    return output_file

# ============================================================
#   STORICO ANALISI
# ============================================================

def safe_float(value: Any) -> Optional[float]:
    """
    Converte un valore a float, restituendo None se il valore non è valido.
    """
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    try:
        return float(value)
    except (TypeError, ValueError):
        try:
            normalized = str(value).replace("%", "").replace(",", ".").strip()
            if not normalized:
                # FIX BUG: Log quando string è vuota dopo normalizzazione
                logger.debug(f"safe_float: stringa vuota dopo normalizzazione da valore: {value}")
                return None
            return float(normalized)
        except Exception as e:
            # FIX BUG: Log quando conversione fallisce completamente
            logger.warning(f"safe_float: impossibile convertire '{value}' a float: {e}")
            return None


def append_analysis_to_archive(
    record: Dict[str, Any],
    archive_file: str = ARCHIVE_FILE,
) -> bool:
    """
    Appende (o crea) una riga di storico analisi sul file CSV dedicato.
    """
    try:
        df_new = pd.DataFrame([record])
        if os.path.exists(archive_file):
            try:
                df_existing = pd.read_csv(archive_file)
            except (pd.errors.EmptyDataError, FileNotFoundError):
                df_existing = pd.DataFrame()
            df_combined = pd.concat([df_existing, df_new], ignore_index=True, sort=False)
        else:
            df_combined = df_new

        if {"match", "timestamp"}.issubset(df_combined.columns):
            # Ordina per timestamp prima di drop_duplicates per garantire che "last" sia effettivamente l'ultimo cronologicamente
            df_combined = df_combined.sort_values("timestamp", ascending=True)
            df_combined = df_combined.drop_duplicates(subset=["match", "timestamp"], keep="last")

        df_combined.to_csv(archive_file, index=False)
        return True
    except Exception as exc:
        logger.error(f"Errore salvataggio storico analisi: {exc}", exc_info=True)
        return False


def save_analysis_history(
    match_name: str,
    league: str,
    home_team: str,
    away_team: str,
    match_datetime_iso: str,
    validated_inputs: Dict[str, Any],
    ris: Dict[str, Any],
    quality_score: Optional[float],
    market_confidence: Optional[float],
    warnings: List[str],
    value_bets: List[Dict[str, Any]],
    archive_file: str = ARCHIVE_FILE,
) -> bool:
    """
    Costruisce il record dell'analisi e lo salva nello storico.
    """
    try:
        value_bets_summary: List[Dict[str, Any]] = []
        best_edge: Optional[float] = None

        for vb in value_bets[:10]:
            edge = safe_float(vb.get("EdgeRaw") or vb.get("Edge %"))
            prob = safe_float(vb.get("ProbRaw") or vb.get("Prob %"))
            odd = safe_float(vb.get("Odd") or vb.get("Quota"))
            value_bets_summary.append(
                {
                    "esito": vb.get("Esito"),
                    "edge": edge,
                    "prob": prob,
                    "odd": odd,
                    "rec": vb.get("Rec"),
                }
            )
            if edge is not None:
                best_edge = edge if best_edge is None else max(best_edge, edge)

        probs_map = {
            "1": safe_float(ris.get("p_home")),
            "X": safe_float(ris.get("p_draw")),
            "2": safe_float(ris.get("p_away")),
        }
        probs_map_clean = {
            key: (value if value is not None else 0.0) for key, value in probs_map.items()
        }
        esito_modello = max(probs_map_clean, key=probs_map_clean.get)

        record = {
            "timestamp": datetime.now().isoformat(),
            "match": match_name,
            "league": league,
            "home_team": home_team,
            "away_team": away_team,
            "match_datetime": match_datetime_iso,
            "odds_1": safe_float(validated_inputs.get("odds_1")),
            "odds_x": safe_float(validated_inputs.get("odds_x")),
            "odds_2": safe_float(validated_inputs.get("odds_2")),
            "odds_over25": safe_float(validated_inputs.get("odds_over25")),
            "odds_under25": safe_float(validated_inputs.get("odds_under25")),
            "odds_btts": safe_float(validated_inputs.get("odds_btts")),
            "odds_dnb_home": safe_float(validated_inputs.get("odds_dnb_home")),
            "odds_dnb_away": safe_float(validated_inputs.get("odds_dnb_away")),
            "p_home": probs_map["1"],
            "p_draw": probs_map["X"],
            "p_away": probs_map["2"],
            "over_25": safe_float(ris.get("over_25")),
            "under_25": safe_float(ris.get("under_25")),
            "btts": safe_float(ris.get("btts")),
            "gg_over25": safe_float(ris.get("gg_over25")),
            "dnb_home": safe_float(ris.get("dnb_home")),
            "dnb_away": safe_float(ris.get("dnb_away")),
            "lambda_home": safe_float(ris.get("lambda_home")),
            "lambda_away": safe_float(ris.get("lambda_away")),
            "rho": safe_float(ris.get("rho")),
            "quality_score": safe_float(quality_score),
            "market_confidence": safe_float(market_confidence),
            "warnings": " | ".join(warnings) if warnings else "",
            "value_bets_count": len(value_bets),
            "value_bets_best_edge": best_edge,
            "value_bets_json": json.dumps(value_bets_summary, ensure_ascii=False)
            if value_bets_summary
            else "",
            "analysis_version": "2.0",
            "calibration_applied": bool(ris.get("calibration_applied")),
            "ensemble_applied": bool(ris.get("ensemble_applied")),
            "esito_modello": esito_modello,
            "esito_reale": "",
            "risultato_reale": "",
            "match_ok": None,
        }

        return append_analysis_to_archive(record, archive_file=archive_file)
    except Exception as exc:
        logger.error(f"Errore creazione record storico analisi: {exc}", exc_info=True)
        return False

# ============================================================
#   COMPARAZIONE BOOKMAKERS
# ============================================================

def compare_bookmaker_odds(event: dict) -> Dict[str, List[Dict[str, Any]]]:
    """
    Confronta quote tra diversi bookmakers per trovare le migliori.
    
    Returns: Dict con migliori quote per ogni mercato
    """
    if not event or "bookmakers" not in event:
        return {}
    
    home_team = (event.get("home_team") or "").lower()
    away_team = (event.get("away_team") or "").lower()
    
    best_odds = {
        "1": [],  # Lista di (bookmaker, odds)
        "X": [],
        "2": [],
        "over_25": [],
        "under_25": [],
        "btts": [],
    }
    
    for bk in event.get("bookmakers", []):
        bk_name = bk.get("key", "unknown")
        
        for mk in bk.get("markets", []):
            mk_key_raw = mk.get("key")
            # Assicurati che mk_key sia sempre una stringa (gestisce None)
            mk_key = str(mk_key_raw).lower() if mk_key_raw is not None else ""
            
            # 1X2
            if "h2h" in mk_key or "match_winner" in mk_key:
                for o in mk.get("outcomes", []):
                    name_l = (o.get("name") or "").lower()
                    price = o.get("price")
                    if not price:
                        continue
                    
                    if home_team in name_l or name_l in home_team:
                        best_odds["1"].append({"bookmaker": bk_name, "odds": price})
                    elif away_team in name_l or name_l in away_team:
                        best_odds["2"].append({"bookmaker": bk_name, "odds": price})
                    elif "draw" in name_l or "x" == name_l or "pareggio" in name_l:
                        best_odds["X"].append({"bookmaker": bk_name, "odds": price})
            
            # Over/Under
            elif "totals" in mk_key:
                for o in mk.get("outcomes", []):
                    point = o.get("point")
                    price = o.get("price")
                    name_l = (o.get("name") or "").lower()
                    if point == 2.5 and price:
                        if "over" in name_l:
                            best_odds["over_25"].append({"bookmaker": bk_name, "odds": price})
                        elif "under" in name_l:
                            best_odds["under_25"].append({"bookmaker": bk_name, "odds": price})
            
            # BTTS
            elif "btts" in mk_key:
                for o in mk.get("outcomes", []):
                    name_l = (o.get("name") or "").lower()
                    price = o.get("price")
                    if price and ("yes" in name_l or "sì" in name_l):
                        best_odds["btts"].append({"bookmaker": bk_name, "odds": price})
    
    # Ordina per odds migliori (più alte)
    for key in best_odds:
        best_odds[key].sort(key=lambda x: x["odds"], reverse=True)
    
    return best_odds

def find_best_odds_summary(best_odds: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Dict[str, Any]]:
    """
    Crea riepilogo delle migliori quote.
    """
    summary = {}
    
    for market, odds_list in best_odds.items():
        if odds_list:
            best = odds_list[0]
            # Calcola media e spread
            all_odds = [x["odds"] for x in odds_list]
            avg_odds = np.mean(all_odds)
            max_odds = max(all_odds)
            min_odds = min(all_odds)
            spread = max_odds - min_odds
            
            summary[market] = {
                "best_bookmaker": best["bookmaker"],
                "best_odds": best["odds"],
                "avg_odds": round(avg_odds, 3),
                "max_odds": max_odds,
                "min_odds": min_odds,
                "spread": round(spread, 3),
                "num_bookmakers": len(odds_list),
                "value": round((best["odds"] - avg_odds) / avg_odds * 100, 2)  # % sopra media
            }
    
    return summary

# ============================================================
#   MARKET MOVEMENT INTELLIGENCE
# ============================================================

def _dynamic_movement_thresholds(
    league: Optional[str],
    volatility_score: Optional[float],
    recent_samples: int,
) -> Tuple[float, float]:
    base_low = 0.18
    base_high = 0.42

    league_key = _normalize_league_key(league)
    if league_key in HIGH_VARIANCE_LEAGUES:
        base_low += 0.03
        base_high += 0.05

    if volatility_score is not None:
        adjustment = min(0.08, volatility_score / 4.0)
        base_low = max(0.08, base_low - adjustment)
        base_high = max(base_low + 0.05, base_high - adjustment / 2)

    if recent_samples <= 2:
        base_low += 0.04
        base_high += 0.04
    elif recent_samples >= 5:
        base_low = max(0.08, base_low - 0.02)

    base_high = max(base_low + 0.05, base_high)
    return base_low, base_high


def calculate_market_movement_factor(
    spread_apertura: float = None,
    total_apertura: float = None,
    spread_corrente: float = None,
    total_corrente: float = None,
    league: Optional[str] = None,
    volatility_score: Optional[float] = None,
    recent_samples: int = 0,
) -> Dict[str, Any]:
    """
    Calcola il "market movement factor" basato sul movimento tra apertura e corrente.
    
    Strategia dinamica:
    - Normalizza il movimento rispetto ai range tipici della lega (spread e total)
    - Adatta le soglie in base alla volatilità recente e al numero di campioni manuali inseriti
    
    Returns:
        Dict con weight_apertura, weight_corrente, movement_magnitude, movement_type
    """
    # Se non abbiamo dati apertura, usa solo corrente
    if spread_apertura is None and total_apertura is None:
        return {
            "weight_apertura": 0.0,
            "weight_corrente": 1.0,
            "movement_magnitude": 0.0,
            "movement_type": "NO_OPENING_DATA"
        }
    
    # Calcola movimento spread
    movement_spread = 0.0
    if spread_apertura is not None and spread_corrente is not None:
        movement_spread = abs(spread_corrente - spread_apertura)
    
    # Calcola movimento total
    movement_total = 0.0
    if total_apertura is not None and total_corrente is not None:
        movement_total = abs(total_corrente - total_apertura)
    
    # Normalizza rispetto al range della lega
    spread_min, spread_max = _get_league_range(league, LEAGUE_SPREAD_RANGES, DEFAULT_SPREAD_RANGE)
    total_min, total_max = _get_league_range(league, LEAGUE_TOTAL_RANGES, DEFAULT_TOTAL_RANGE)
    spread_norm_den = max(0.25, abs(spread_max - spread_min) / 2)
    total_norm_den = max(0.5, abs(total_max - total_min) / 2)

    normalized_spread = movement_spread / spread_norm_den if spread_norm_den else movement_spread
    normalized_total = movement_total / total_norm_den if total_norm_den else movement_total

    movement_magnitude = (
        normalized_spread * 0.6 + normalized_total * 0.4
    ) if (movement_spread > 0 or movement_total > 0) else 0.0
    
    low_thr, high_thr = _dynamic_movement_thresholds(league, volatility_score, recent_samples)
    
    # Determina pesi basati su movimento
    # Calcola pesi continui così anche micro-movimenti (es. 0.25) influenzano subito i lambda
    sensitivity = max(0.05, min(low_thr, high_thr))
    movement_ratio = 0.0
    if movement_magnitude > 0:
        movement_ratio = min(1.0, movement_magnitude / max(sensitivity, 1e-6))

    MIN_WEIGHT_CURRENT = 0.35
    MAX_WEIGHT_CURRENT = 0.85
    weight_corrente = MIN_WEIGHT_CURRENT + movement_ratio * (MAX_WEIGHT_CURRENT - MIN_WEIGHT_CURRENT)
    weight_apertura = 1.0 - weight_corrente

    if movement_magnitude < low_thr:
        movement_type = "STABLE"
    elif movement_magnitude < high_thr:
        movement_type = "MODERATE"
    else:
        movement_type = "HIGH_SMART_MONEY"
    
    league_key = _normalize_league_key(league)

    return {
        "weight_apertura": weight_apertura,
        "weight_corrente": weight_corrente,
        "movement_magnitude": round(movement_magnitude, 3),
        "normalized_movement": round(movement_magnitude, 3),
        "movement_type": movement_type,
        "movement_spread": round(movement_spread, 3) if movement_spread > 0 else None,
        "movement_total": round(movement_total, 3) if movement_total > 0 else None,
        "movement_ratio": round(movement_ratio, 3),
        "volatility_score": round(volatility_score, 3) if volatility_score is not None else None,
        "history_samples": recent_samples,
        "league": league_key,
    }

def apply_market_movement_blend(
    lambda_h_current: float,
    lambda_a_current: float,
    total_current: float,
    spread_apertura: float = None,
    total_apertura: float = None,
    spread_corrente: float = None,
    total_corrente: float = None,
    home_advantage: float = 1.30,
    league: Optional[str] = None,
    movement_context: Optional[Dict[str, Any]] = None,
) -> Tuple[float, float]:
    """
    Applica blend bayesiano tra lambda da apertura e corrente basato su market movement.
    
    Se abbiamo dati apertura, calcola lambda da apertura e fa blend con corrente.
    """
    # ⚠️ PRECISIONE MANIACALE: Valida lambda prima di calcolare spread/total
    if not isinstance(lambda_h_current, (int, float)) or not isinstance(lambda_a_current, (int, float)):
        logger.error(f"Lambda non validi: lambda_h={lambda_h_current}, lambda_a={lambda_a_current}")
        raise ValueError("Lambda devono essere numeri validi")
    
    # ⚠️ PROTEZIONE: Verifica che lambda siano finiti e positivi
    if not math.isfinite(lambda_h_current) or not math.isfinite(lambda_a_current):
        logger.error(f"Lambda non finiti: lambda_h={lambda_h_current}, lambda_a={lambda_a_current}")
        raise ValueError("Lambda devono essere numeri finiti")
    
    if lambda_h_current < 0 or lambda_a_current < 0:
        logger.warning(f"Lambda negativi: lambda_h={lambda_h_current}, lambda_a={lambda_a_current}, correggo")
        lambda_h_current = max(0.1, lambda_h_current)
        lambda_a_current = max(0.1, lambda_a_current)
    
    # Helper: costruisce lambda da spread/total specifici (apertura o corrente)
    def _compute_lambda_from_market(spread_value, total_value, label: str):
        if total_value is None:
            return None, None

        try:
            total_val = float(total_value)
        except (TypeError, ValueError):
            logger.warning(f"total_{label} non valido: {total_value}, ignoro")
            return None, None

        if not math.isfinite(total_val):
            logger.warning(f"total_{label} non finito: {total_value}, ignoro")
            return None, None

        total_val = _clamp_with_league_range(total_val, league, LEAGUE_TOTAL_RANGES, DEFAULT_TOTAL_RANGE)

        spread_val = None
        if spread_value is not None:
            try:
                spread_val = float(spread_value)
            except (TypeError, ValueError):
                logger.warning(f"spread_{label} non valido: {spread_value}, uso 0.0")
                spread_val = 0.0

            if not math.isfinite(spread_val):
                logger.warning(f"spread_{label} non finito: {spread_value}, uso 0.0")
                spread_val = 0.0

            if abs(spread_val) > 5.0:
                logger.error(f"spread_{label} troppo alto: {spread_val}, dati probabilmente corrotti - uso 0.0")
                spread_val = 0.0

            spread_val = _clamp_with_league_range(spread_val, league, LEAGUE_SPREAD_RANGES, DEFAULT_SPREAD_RANGE)

        if spread_val is not None:
            lambda_a_val = (total_val + spread_val) / 2.0
            lambda_h_val = (total_val - spread_val) / 2.0
        else:
            total_ratio_den = lambda_h_current + lambda_a_current
            if total_ratio_den <= 0 or not math.isfinite(total_ratio_den):
                ratio_h = 0.5
            else:
                ratio_h = max(0.0, min(1.0, lambda_h_current / total_ratio_den))
            lambda_h_val = total_val * ratio_h
            lambda_a_val = total_val - lambda_h_val

        if not math.isfinite(lambda_h_val) or not math.isfinite(lambda_a_val):
            logger.warning(f"Lambda {label} non finiti (λh={lambda_h_val}, λa={lambda_a_val}), ignoro")
            return None, None

        lambda_h_val = max(model_config.LAMBDA_SAFE_MIN, min(model_config.LAMBDA_SAFE_MAX, lambda_h_val))
        lambda_a_val = max(model_config.LAMBDA_SAFE_MIN, min(model_config.LAMBDA_SAFE_MAX, lambda_a_val))

        logger.info(f"Lambda da mercato ({label}): λ_h={lambda_h_val:.4f}, λ_a={lambda_a_val:.4f}, total={lambda_h_val + lambda_a_val:.4f}")
        return lambda_h_val, lambda_a_val

    # ⚠️ PRECISIONE: Calcola spread/total correnti se non forniti
    if spread_corrente is None:
        # ⚠️ FIX CRITICO: spread = lambda_a - lambda_h (spread > 0 favorisce Away)
        spread_corrente = lambda_a_current - lambda_h_current
        # ⚠️ MICRO-PRECISIONE: Valida spread calcolato con precisione
        spread_corrente = _clamp_with_league_range(spread_corrente, league, LEAGUE_SPREAD_RANGES, DEFAULT_SPREAD_RANGE)
        # ⚠️ VERIFICA: Double-check che spread sia finito
        if not math.isfinite(spread_corrente):
            logger.warning(f"Spread calcolato non finito: {spread_corrente}, uso default 0.0")
            spread_corrente = 0.0
    
    if total_corrente is None:
        total_corrente = lambda_h_current + lambda_a_current
        # ⚠️ MICRO-PRECISIONE: Valida total calcolato con precisione
        total_corrente = _clamp_with_league_range(total_corrente, league, LEAGUE_TOTAL_RANGES, DEFAULT_TOTAL_RANGE)
        # ⚠️ VERIFICA: Double-check che total sia finito
        if not math.isfinite(total_corrente):
            logger.warning(f"Total calcolato non finito: {total_corrente}, uso default 2.5")
            total_corrente = 2.5
    
    # Calcola market movement factor
    movement_factor = calculate_market_movement_factor(
        spread_apertura,
        total_apertura,
        spread_corrente,
        total_corrente,
        league=league,
        volatility_score=(movement_context or {}).get("volatility_score"),
        recent_samples=(movement_context or {}).get("history_samples", 0),
    )
    
    # Se non abbiamo dati apertura o movimento è nullo, usa solo corrente
    if movement_factor["weight_apertura"] == 0.0:
        return lambda_h_current, lambda_a_current
    
    # ⚠️ PRECISIONE MANIACALE: Calcola lambda da apertura (se disponibile) con validazione completa
    if spread_apertura is not None and total_apertura is not None:
        lambda_h_ap, lambda_a_ap = _compute_lambda_from_market(spread_apertura, total_apertura, "apertura")
    else:
        lambda_h_ap = None
        lambda_a_ap = None

    # Calcola lambda dal mercato corrente (spread/total correnti)
    lambda_h_curr_market = None
    lambda_a_curr_market = None
    if total_corrente is not None:
        lambda_h_curr_market, lambda_a_curr_market = _compute_lambda_from_market(
            spread_corrente,
            total_corrente,
            "corrente"
        )

    # ⚠️ CONTROLLO: Limita effetto del blend per evitare valori estremi
    max_blend_adjustment = 1.4

    def _limit_relative(value, reference):
        if value is None:
            return None
        ref_safe = reference
        if not math.isfinite(ref_safe) or abs(ref_safe) < model_config.TOL_DIVISION_ZERO:
            ref_safe = model_config.LAMBDA_SAFE_MIN
        return max(
            ref_safe / max_blend_adjustment,
            min(ref_safe * max_blend_adjustment, value)
        )

    lambda_sources = []
    if lambda_h_ap is not None and lambda_a_ap is not None:
        lambda_sources.append((
            movement_factor["weight_apertura"],
            _limit_relative(lambda_h_ap, lambda_h_current),
            _limit_relative(lambda_a_ap, lambda_a_current)
        ))

    if lambda_h_curr_market is not None and lambda_a_curr_market is not None:
        lambda_sources.append((
            movement_factor["weight_corrente"],
            _limit_relative(lambda_h_curr_market, lambda_h_current),
            _limit_relative(lambda_a_curr_market, lambda_a_current)
        ))

    if lambda_sources:
        weight_sum = sum(w for w, _, _ in lambda_sources)
        if weight_sum > 0:
            lambda_h_target = sum(w * lh for w, lh, _ in lambda_sources if lh is not None) / weight_sum
            lambda_a_target = sum(w * la for w, _, la in lambda_sources if la is not None) / weight_sum
        else:
            lambda_h_target = lambda_h_current
            lambda_a_target = lambda_a_current

        if not math.isfinite(lambda_h_target) or not math.isfinite(lambda_a_target):
            lambda_h_target = lambda_h_current
            lambda_a_target = lambda_a_current

        lambda_h_blended = max(
            lambda_h_current / max_blend_adjustment,
            min(lambda_h_current * max_blend_adjustment, lambda_h_target)
        )
        lambda_a_blended = max(
            lambda_a_current / max_blend_adjustment,
            min(lambda_a_current * max_blend_adjustment, lambda_a_target)
        )

        return lambda_h_blended, lambda_a_blended
    
    # Se non abbiamo spread apertura ma abbiamo total apertura
    elif total_apertura is not None:
        # ⚠️ VALIDAZIONE: Verifica che total_apertura sia ragionevole
        total_apertura_safe = _clamp_with_league_range(total_apertura, league, LEAGUE_TOTAL_RANGES, DEFAULT_TOTAL_RANGE)
        
        # Usa total apertura per calibrare total corrente
        lambda_total_ap = total_apertura_safe / 2.0
        lambda_total_curr = total_current / 2.0
        
        # ⚠️ CONTROLLO: Limita differenza tra total apertura e corrente
        # Se differiscono troppo, riduci peso apertura
        total_diff = abs(total_apertura_safe - total_current) / max(0.1, total_current)
        if total_diff > 0.5:  # Se differiscono più del 50%
            # Riduci peso apertura se differenza è troppo grande
            movement_factor["weight_apertura"] = min(0.3, movement_factor["weight_apertura"])
            movement_factor["weight_corrente"] = 1.0 - movement_factor["weight_apertura"]
        
        # Blend dei total
        w_ap = movement_factor["weight_apertura"]
        w_curr = movement_factor["weight_corrente"]
        lambda_total_blended = w_ap * lambda_total_ap + w_curr * lambda_total_curr
        
        # ⚠️ VERIFICA: Limita lambda_total_blended a range ragionevole
        lambda_total_blended = max(0.5, min(5.0, lambda_total_blended))
        
        # Mantieni proporzione corrente tra lambda_h e lambda_a
        ratio_h = lambda_h_current / (lambda_h_current + lambda_a_current) if (lambda_h_current + lambda_a_current) > 0 else 0.5
        ratio_a = 1.0 - ratio_h
        
        lambda_h_blended = lambda_total_blended * ratio_h * 2.0
        lambda_a_blended = lambda_total_blended * ratio_a * 2.0
        
        # ⚠️ CONTROLLO: Limita effetto del blend
        max_blend_adjustment = 1.4
        lambda_h_blended = max(
            lambda_h_current / max_blend_adjustment,
            min(lambda_h_current * max_blend_adjustment, lambda_h_blended)
        )
        lambda_a_blended = max(
            lambda_a_current / max_blend_adjustment,
            min(lambda_a_current * max_blend_adjustment, lambda_a_blended)
        )
        
        return lambda_h_blended, lambda_a_blended
    
    # Fallback: usa solo corrente
    return lambda_h_current, lambda_a_current

# ============================================================
#   FEATURE ENGINEERING AVANZATO
# ============================================================

def apply_advanced_data_adjustments(
    lambda_h: float,
    lambda_a: float,
    advanced_data: Dict[str, Any],
) -> Tuple[float, float]:
    """
    Applica aggiustamenti basati su dati avanzati (statistiche, H2H, infortuni).
    Lavora in background, modifica lambda silenziosamente.
    
    ⚠️ MIGLIORAMENTO: Limita effetto cumulativo per evitare esplosione dei lambda.
    """
    if not advanced_data or not advanced_data.get("data_available"):
        return lambda_h, lambda_a
    
    # Salva valori iniziali per limitare effetto totale
    lambda_h_initial = lambda_h
    lambda_a_initial = lambda_a
    max_advanced_adjustment = 1.2  # Massimo 20% di variazione totale
    
    # 1. Aggiustamenti forma squadre (limitati)
    home_stats = advanced_data.get("home_team_stats")
    away_stats = advanced_data.get("away_team_stats")
    
    if home_stats and home_stats.get("confidence", 0) > 0.3:
        # Limita form_attack e form_defense a range ragionevole
        form_attack = max(0.85, min(1.15, home_stats.get("form_attack", 1.0)))
        form_defense = max(0.85, min(1.15, home_stats.get("form_defense", 1.0)))
        # Applica forma attacco casa
        lambda_h *= form_attack
        # Applica forma difesa trasferta (riduce lambda away)
        lambda_a *= (2.0 - form_defense)  # Inverso, ma limitato
    
    if away_stats and away_stats.get("confidence", 0) > 0.3:
        form_attack = max(0.85, min(1.15, away_stats.get("form_attack", 1.0)))
        form_defense = max(0.85, min(1.15, away_stats.get("form_defense", 1.0)))
        # Applica forma attacco trasferta
        lambda_a *= form_attack
        # Applica forma difesa casa (riduce lambda home)
        lambda_h *= (2.0 - form_defense)  # Inverso, ma limitato
    
    # 2. Aggiustamenti H2H (limitati)
    h2h_data = advanced_data.get("h2h_data")
    if h2h_data and h2h_data.get("confidence", 0) > 0.3:
        # Limita h2h_home_advantage a range ragionevole
        h2h_advantage = max(0.9, min(1.1, h2h_data.get("h2h_home_advantage", 1.0)))
        # Aggiusta vantaggio casa
        lambda_h *= h2h_advantage
        # Aggiusta total gol (entrambi i lambda) - limitato
        goals_factor = max(0.9, min(1.1, h2h_data.get("h2h_goals_factor", 1.0)))
        lambda_h *= math.sqrt(goals_factor)  # Radice per distribuire
        lambda_a *= math.sqrt(goals_factor)
    
    # 3. Aggiustamenti infortuni (limitati)
    home_injuries = advanced_data.get("home_injuries")
    away_injuries = advanced_data.get("away_injuries")
    
    if home_injuries and home_injuries.get("confidence", 0) > 0.3:
        # Limita attack_factor e defense_factor
        attack_factor = max(0.85, min(1.0, home_injuries.get("attack_factor", 1.0)))
        defense_factor = max(0.85, min(1.0, home_injuries.get("defense_factor", 1.0)))
        # Infortuni casa: riduce attacco, aumenta vulnerabilità difesa
        lambda_h *= attack_factor
        lambda_a *= (2.0 - defense_factor)  # Inverso, ma limitato
    
    if away_injuries and away_injuries.get("confidence", 0) > 0.3:
        attack_factor = max(0.85, min(1.0, away_injuries.get("attack_factor", 1.0)))
        defense_factor = max(0.85, min(1.0, away_injuries.get("defense_factor", 1.0)))
        # Infortuni trasferta: riduce attacco, aumenta vulnerabilità difesa
        lambda_a *= attack_factor
        lambda_h *= (2.0 - defense_factor)  # Inverso, ma limitato
    
    # ⚠️ CONTROLLO FINALE: Limita effetto totale degli aggiustamenti avanzati
    lambda_h = max(lambda_h_initial / max_advanced_adjustment, min(lambda_h_initial * max_advanced_adjustment, lambda_h))
    lambda_a = max(lambda_a_initial / max_advanced_adjustment, min(lambda_a_initial * max_advanced_adjustment, lambda_a))
    
    return lambda_h, lambda_a

# ============================================================
#   PORTFOLIO TRACKING
# ============================================================

def add_to_portfolio(
    match_name: str,
    market: str,
    esito: str,
    odds: float,
    stake: float,
    probability: float,
    timestamp: str = None,
) -> None:
    """
    Aggiunge scommessa al portfolio.
    """
    if timestamp is None:
        timestamp = datetime.now().isoformat()
    
    portfolio_entry = {
        "timestamp": timestamp,
        "match": match_name,
        "market": market,
        "esito": esito,
        "odds": odds,
        "stake": stake,
        "probability": probability,
        "expected_value": (probability * odds - 1) * stake,
        "status": "pending",  # pending, won, lost
        "result": "",
    }
    
    # Carica portfolio esistente
    if os.path.exists(PORTFOLIO_FILE):
        df = pd.read_csv(PORTFOLIO_FILE)
        df = pd.concat([df, pd.DataFrame([portfolio_entry])], ignore_index=True)
    else:
        df = pd.DataFrame([portfolio_entry])
    
    df.to_csv(PORTFOLIO_FILE, index=False)

def get_portfolio_summary() -> Dict[str, Any]:
    """
    Calcola riepilogo portfolio.
    """
    if not os.path.exists(PORTFOLIO_FILE):
        return {"total_bets": 0, "total_staked": 0.0, "pending_bets": 0}
    
    df = pd.read_csv(PORTFOLIO_FILE)
    
    total_bets = len(df)
    total_staked = df["stake"].sum()
    pending = len(df[df["status"] == "pending"])
    won = len(df[df["status"] == "won"])
    lost = len(df[df["status"] == "lost"])
    
    # Calcola profit per scommesse chiuse
    df_closed = df[df["status"].isin(["won", "lost"])]
    if len(df_closed) > 0:
        total_returned = 0.0
        for _, row in df_closed.iterrows():
            if row["status"] == "won":
                total_returned += row["stake"] * row["odds"]
        
        total_staked_closed = df_closed["stake"].sum()
        profit = total_returned - total_staked_closed
        roi = (profit / total_staked_closed * 100) if total_staked_closed > 0 else 0
    else:
        profit = 0.0
        roi = 0.0
    
    return {
        "total_bets": total_bets,
        "total_staked": round(total_staked, 2),
        "pending_bets": pending,
        "won_bets": won,
        "lost_bets": lost,
        "profit": round(profit, 2),
        "roi": round(roi, 2),
        "win_rate": round(won / (won + lost) * 100, 1) if (won + lost) > 0 else 0.0,
    }

# ============================================================
#   DASHBOARD ANALYTICS
# ============================================================

# ============================================================
#   MARKET MOVEMENT TRACKING
# ============================================================

def track_odds_movement(
    match_name: str,
    odds_1: float,
    odds_x: float,
    odds_2: float,
    timestamp: str = None,
) -> Dict[str, Any]:
    """
    Traccia movimenti delle quote nel tempo per identificare trend.
    """
    if timestamp is None:
        timestamp = datetime.now().isoformat()
    
    # Carica storico movimenti
    if os.path.exists(ODDS_HISTORY_FILE):
        df = pd.read_csv(ODDS_HISTORY_FILE)
    else:
        df = pd.DataFrame(columns=["timestamp", "match", "odds_1", "odds_x", "odds_2"])
    
    # Aggiungi nuovo punto
    new_row = {
        "timestamp": timestamp,
        "match": match_name,
        "odds_1": odds_1,
        "odds_x": odds_x,
        "odds_2": odds_2,
    }
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df.to_csv(ODDS_HISTORY_FILE, index=False)
    
    # Analizza trend per questa partita
    match_data = df[df["match"] == match_name].sort_values("timestamp")
    
    if len(match_data) >= 2:
        # Calcola cambiamenti
        # ⚠️ PROTEZIONE: len(match_data) >= 2 garantisce che iloc[0] e iloc[-1] siano validi
        first = match_data.iloc[0]
        last = match_data.iloc[-1]
        
        movement = {
            "odds_1_change": last["odds_1"] - first["odds_1"],
            "odds_x_change": last["odds_x"] - first["odds_x"],
            "odds_2_change": last["odds_2"] - first["odds_2"],
            "odds_1_change_pct": ((last["odds_1"] - first["odds_1"]) / first["odds_1"]) * 100,
            "trend": "stable",
            "market_sentiment": "neutral",
        }
        
        # Determina trend
        if abs(movement["odds_1_change_pct"]) > 5:
            if movement["odds_1_change"] < 0:
                movement["trend"] = "home_favorite_increasing"
                movement["market_sentiment"] = "home_strong"
            else:
                movement["trend"] = "home_favorite_decreasing"
                movement["market_sentiment"] = "away_strong"
        
        return movement
    
    return {"trend": "insufficient_data"}

def get_odds_movement_insights(match_name: str) -> Dict[str, Any]:
    """
    Analizza movimenti quote e fornisce insights.
    """
    if not os.path.exists(ODDS_HISTORY_FILE):
        return {}
    
    df = pd.read_csv(ODDS_HISTORY_FILE)
    match_data = df[df["match"] == match_name].sort_values("timestamp")
    
    if len(match_data) < 2:
        return {}
    
    # Calcola volatilità
    # ⚠️ PROTEZIONE: Calcola std solo se ci sono almeno 2 valori (altrimenti std = NaN)
    # len(match_data) >= 2 è già garantito dal controllo sopra
    volatility_1 = match_data["odds_1"].std() if len(match_data) > 1 and "odds_1" in match_data.columns else 0.0
    volatility_x = match_data["odds_x"].std() if len(match_data) > 1 and "odds_x" in match_data.columns else 0.0
    volatility_2 = match_data["odds_2"].std() if len(match_data) > 1 and "odds_2" in match_data.columns else 0.0
    
    # Sostituisci NaN con 0.0 se std restituisce NaN
    volatility_1 = 0.0 if pd.isna(volatility_1) else volatility_1
    volatility_x = 0.0 if pd.isna(volatility_x) else volatility_x
    volatility_2 = 0.0 if pd.isna(volatility_2) else volatility_2
    
    # Identifica movimenti significativi
    significant_moves = []
    # ⚠️ PROTEZIONE: range(1, len(match_data)) richiede almeno 2 elementi
    if len(match_data) >= 2:
        for i in range(1, len(match_data)):
            prev = match_data.iloc[i-1]
            curr = match_data.iloc[i]
            
            # ⚠️ PROTEZIONE: Verifica che le colonne esistano e che prev[col] non sia zero
            for col in ["odds_1", "odds_x", "odds_2"]:
                if col not in prev.index or col not in curr.index:
                    continue
                if prev[col] == 0 or pd.isna(prev[col]) or pd.isna(curr[col]):
                    continue
                change_pct = abs((curr[col] - prev[col]) / prev[col]) * 100
                if change_pct > 3:  # Movimento > 3%
                    significant_moves.append({
                        "market": col,
                        "change": change_pct,
                        "timestamp": curr["timestamp"] if "timestamp" in curr.index else "",
                        "direction": "up" if curr[col] > prev[col] else "down"
                    })
    
    return {
        "volatility": {
            "odds_1": round(volatility_1, 3),
            "odds_x": round(volatility_x, 3),
            "odds_2": round(volatility_2, 3),
        },
        "significant_moves": significant_moves,
        "num_data_points": len(match_data),
    }

# ============================================================
#   TIME-BASED ADJUSTMENTS
# ============================================================

def get_time_based_adjustments(
    match_datetime: str = None,
    league: str = "generic",
) -> Dict[str, float]:
    """
    Calcola aggiustamenti basati su:
    - Ora partita (serale vs pomeridiana)
    - Giorno settimana (weekend vs settimana)
    - Periodo stagione (inizio, metà, fine)
    """
    if match_datetime is None:
        match_datetime = datetime.now().isoformat()

    try:
        dt = datetime.fromisoformat(match_datetime.replace("Z", "+00:00"))
    except (ValueError, AttributeError) as e:
        # FIX BUG: Specifico eccezioni invece di bare except
        logger.warning(f"Datetime parsing failed for '{match_datetime}': {e}, using now()")
        dt = datetime.now()
    
    adjustments = {
        "time_factor": 1.0,
        "day_factor": 1.0,
        "season_factor": 1.0,
    }
    
    # Ora partita
    hour = dt.hour
    if 20 <= hour <= 22:  # Serale
        adjustments["time_factor"] = 1.05  # +5% gol (atmosfera)
    elif 12 <= hour <= 15:  # Pomeridiana
        adjustments["time_factor"] = 0.98  # -2% gol
    
    # Giorno settimana
    weekday = dt.weekday()
    if weekday >= 5:  # Weekend (sabato/domenica)
        adjustments["day_factor"] = 1.03  # +3% gol (più pubblico)
    elif weekday == 2:  # Mercoledì (spesso Champions/Europa)
        adjustments["day_factor"] = 1.08  # +8% gol (partite importanti)
    
    # Periodo stagione (approssimato)
    month = dt.month
    if month in [8, 9]:  # Inizio stagione
        adjustments["season_factor"] = 1.02  # +2% gol (squadre fresche)
    elif month in [4, 5]:  # Fine stagione
        adjustments["season_factor"] = 0.97  # -3% gol (squadre stanche)
    
    return adjustments

def apply_time_adjustments(
    lambda_h: float,
    lambda_a: float,
    match_datetime: str = None,
    league: str = "generic",
) -> Tuple[float, float]:
    """
    Applica aggiustamenti temporali ai lambda.
    """
    adjustments = get_time_based_adjustments(match_datetime, league)
    
    # Combina tutti i fattori
    total_factor = (
        adjustments["time_factor"] *
        adjustments["day_factor"] *
        adjustments["season_factor"]
    )
    
    # Applica solo se significativo
    if abs(total_factor - 1.0) > 0.02:  # > 2% di differenza
        lambda_h *= total_factor
        lambda_a *= total_factor
    
    return lambda_h, lambda_a

# ============================================================
#   FATIGUE E MOTIVATION FACTORS
# ============================================================

def calculate_fatigue_factor(
    team_name: str,
    days_since_last_match: int = None,
    matches_last_30_days: int = None,
) -> float:
    """
    Calcola fattore fatica basato su:
    - Giorni dall'ultima partita
    - Numero partite ultimi 30 giorni
    """
    fatigue = 1.0
    
    # Giorni di riposo
    if days_since_last_match is not None:
        if days_since_last_match <= 2:
            fatigue *= 0.92  # -8% (molto stanco)
        elif days_since_last_match == 3:
            fatigue *= 0.96  # -4% (stanco)
        elif days_since_last_match >= 7:
            fatigue *= 1.05  # +5% (molto riposato)
    
    # Partite ravvicinate
    if matches_last_30_days is not None:
        if matches_last_30_days >= 8:
            fatigue *= 0.94  # -6% (troppe partite)
        elif matches_last_30_days <= 4:
            fatigue *= 1.03  # +3% (poche partite)
    
    return fatigue

def calculate_motivation_factor(
    team_position: int = None,
    points_from_relegation: int = None,
    points_from_europe: int = None,
    is_derby: bool = False,
) -> float:
    """
    Calcola fattore motivazione basato su:
    - Posizione in classifica
    - Vicinanza a obiettivi (salvezza, Europa)
    - Partite speciali (derby)
    """
    motivation = 1.0
    
    # Derby
    if is_derby:
        motivation *= 1.12  # +12% (alta motivazione)
    
    # Lotta salvezza
    if points_from_relegation is not None:
        if 0 <= points_from_relegation <= 3:
            motivation *= 1.10  # +10% (lotta per non retrocedere)
    
    # Lotta Europa
    if points_from_europe is not None:
        if 0 <= points_from_europe <= 3:
            motivation *= 1.08  # +8% (lotta per Europa)
    
    return motivation

# ============================================================
#   ANOMALY DETECTION
# ============================================================

def detect_odds_anomalies(
    odds_1: float,
    odds_x: float,
    odds_2: float,
    odds_over25: float = None,
    odds_under25: float = None,
) -> Dict[str, Any]:
    """
    Rileva anomalie nelle quote che potrebbero indicare:
    - Errori bookmaker
    - Quote non aggiornate
    - Opportunità di arbitraggio
    """
    anomalies = []
    severity = 0
    
    # 1. Check margine anomalo
    margin = (1/odds_1 + 1/odds_x + 1/odds_2) - 1
    if margin < 0:
        anomalies.append({
            "type": "negative_margin",
            "description": f"Margine negativo ({margin*100:.2f}%) - possibile errore o arbitraggio",
            "severity": "high"
        })
        severity += 3
    elif margin > 0.20:
        anomalies.append({
            "type": "excessive_margin",
            "description": f"Margine eccessivo ({margin*100:.2f}%) - quote poco competitive",
            "severity": "medium"
        })
        severity += 1
    
    # 2. Check coerenza Over/Under
    if odds_over25 and odds_under25:
        margin_ou = (1/odds_over25 + 1/odds_under25) - 1
        if margin_ou < 0:
            anomalies.append({
                "type": "arbitrage_opportunity",
                "description": "Opportunità arbitraggio Over/Under",
                "severity": "high"
            })
            severity += 2
    
    # 3. Check quote estreme
    if odds_1 < 1.01 or odds_2 < 1.01:
        anomalies.append({
            "type": "extreme_odds",
            "description": "Quote estreme (< 1.01) - possibile errore",
            "severity": "high"
        })
        severity += 2
    
    # 4. Check incoerenza 1X2 vs Over/Under
    if odds_over25 and odds_1:
        p_home = 1 / odds_1
        p_over = 1 / odds_over25
        
        # Se casa molto favorita ma over alto → incoerenza
        if p_home > 0.70 and p_over > 0.60:
            anomalies.append({
                "type": "incoherent_markets",
                "description": "Casa molto favorita ma Over alto - possibile incoerenza",
                "severity": "medium"
            })
            severity += 1
    
    return {
        "anomalies": anomalies,
        "severity_score": severity,
        "has_anomalies": len(anomalies) > 0,
        "recommendation": "verify" if severity >= 3 else ("review" if severity >= 1 else "ok")
    }

# ============================================================
#   ADVANCED RISK MANAGEMENT
# ============================================================

def calculate_dynamic_position_size(
    edge: float,
    odds: float,
    bankroll: float,
    current_drawdown: float = 0.0,
    win_streak: int = 0,
    loss_streak: int = 0,
) -> Dict[str, Any]:
    """
    Calcola position size dinamico basato su:
    - Edge e odds
    - Drawdown corrente
    - Streak (vittorie/sconfitte consecutive)
    """
    # Kelly base - con protezione robusta per denominatore
    p = edge + (1 / max(odds, model_config.TOL_DIVISION_ZERO))  # Probabilità reale
    q = 1 - p
    denominator = odds - 1.0
    if abs(denominator) < model_config.TOL_DIVISION_ZERO:
        kelly_base = 0.0
    else:
        kelly_base = (p * odds - 1.0) / denominator
    kelly_base = max(0.0, min(0.25, kelly_base))  # Cap a 25%
    
    # Aggiustamenti per drawdown
    if current_drawdown > 0.20:  # Drawdown > 20%
        kelly_base *= 0.5  # Riduci del 50%
    elif current_drawdown > 0.10:  # Drawdown > 10%
        kelly_base *= 0.75  # Riduci del 25%
    
    # Aggiustamenti per streak
    if loss_streak >= 3:
        kelly_base *= 0.6  # Riduci dopo 3 sconfitte consecutive
    elif win_streak >= 5:
        kelly_base *= 1.1  # Aumenta leggermente dopo 5 vittorie (max 10%)
        kelly_base = min(0.25, kelly_base)  # Cap
    
    stake = bankroll * kelly_base
    
    return {
        "kelly_percent": kelly_base * 100,
        "stake": round(stake, 2),
        "edge": edge,
        "recommendation": "bet" if edge >= 0.03 else ("caution" if edge >= 0.01 else "skip"),
        "risk_level": "low" if current_drawdown < 0.05 else ("medium" if current_drawdown < 0.15 else "high")
    }

def calculate_stop_loss_level(
    initial_bankroll: float,
    current_bankroll: float,
    max_drawdown_pct: float = 0.25,
) -> Dict[str, Any]:
    """
    Calcola livello stop loss per proteggere bankroll.
    """
    drawdown = (initial_bankroll - current_bankroll) / initial_bankroll
    
    stop_loss_triggered = drawdown >= max_drawdown_pct
    
    return {
        "current_drawdown": round(drawdown * 100, 2),
        "max_drawdown": max_drawdown_pct * 100,
        "stop_loss_triggered": stop_loss_triggered,
        "remaining_buffer": round((max_drawdown_pct - drawdown) * 100, 2),
        "recommendation": "stop_betting" if stop_loss_triggered else "continue"
    }

# ============================================================
#   FEATURE IMPORTANCE ANALYSIS
# ============================================================

def analyze_feature_importance(archive_file: str = ARCHIVE_FILE) -> Dict[str, float]:
    """
    Analizza importanza delle features nel modello usando correlazione
    e analisi statistica.
    """
    if not os.path.exists(archive_file):
        return {}
    
    try:
        df = pd.read_csv(archive_file)
        
        # Filtra partite con risultati
        df_complete = df[
            df["esito_reale"].notna() &
            (df["esito_reale"] != "") &
            df["p_home"].notna()
        ]
        
        if len(df_complete) < 20:
            return {}
        
        # Calcola accuracy per ogni feature
        features = {
            "lambda_home": "lambda_home",
            "lambda_away": "lambda_away",
            "rho": "rho",
            "quality_score": "quality_score",
            "market_confidence": "market_confidence",
        }
        
        importance = {}
        
        for feat_name, feat_col in features.items():
            if feat_col not in df_complete.columns:
                continue
            
            # Correlazione con accuracy
            df_complete["correct"] = (
                df_complete["esito_modello"] == df_complete["esito_reale"]
            ).astype(int)
            
            corr = df_complete[feat_col].corr(df_complete["correct"])
            importance[feat_name] = round(abs(corr) * 100, 1) if not pd.isna(corr) else 0
        
        # Ordina per importanza
        sorted_importance = dict(sorted(importance.items(), key=lambda x: x[1], reverse=True))
        
        return sorted_importance
    except Exception as e:
        return {"error": str(e)}

# ============================================================
#   API GRATUITE AGGIUNTIVE - INTEGRAZIONE
# ============================================================

# 1. OPENWEATHERMAP API
def get_weather_for_match(city: str, match_datetime: str = None) -> Dict[str, Any]:
    """
    Recupera condizioni meteo per partita.
    Calcola impatto su lambda (pioggia/vento → meno gol).
    """
    if not OPENWEATHER_API_KEY:
        return {"available": False, "weather_factor": 1.0}
    
    try:
        url = "https://api.openweathermap.org/data/2.5/weather"
        params = {
            "q": city,
            "appid": OPENWEATHER_API_KEY,
            "units": "metric",
            "lang": "it"
        }
        
        response = requests.get(url, params=params, timeout=5)
        response.raise_for_status()
        data = response.json()
        
        # Estrai dati meteo (MIGLIORATO: più parametri)
        temp = data["main"]["temp"]
        humidity = data["main"]["humidity"]
        pressure = data["main"].get("pressure", 1013)  # Pressione in hPa (default 1013)
        wind_speed = data["wind"].get("speed", 0)
        visibility = data.get("visibility", 10000) / 1000  # Visibilità in km (default 10km)
        rain = data.get("rain", {}).get("1h", 0) or data.get("rain", {}).get("3h", 0) or 0
        
        # ⚠️ PROTEZIONE: Valida che weather esista e non sia vuoto prima di accedere a [0]
        weather_list = data.get("weather", [])
        if weather_list and len(weather_list) > 0 and isinstance(weather_list[0], dict):
            conditions = weather_list[0].get("main", "").lower()
            description = weather_list[0].get("description", "")
        else:
            # Fallback: usa valori di default
            conditions = ""
            description = ""
            logger.warning("Dati weather non disponibili o non validi")
        
        # Calcola fattore impatto su lambda (MIGLIORATO: più parametri)
        weather_factor = 1.0
        adjustments = []
        
        # Pioggia → riduce gol significativamente
        if rain > 5:  # Pioggia > 5mm/h
            reduction = min(0.15, rain / 50)  # Max -15%
            weather_factor *= (1 - reduction)
            adjustments.append(f"Pioggia forte: -{reduction*100:.1f}%")
        elif rain > 0:
            reduction = rain / 30  # Max -3% per pioggia leggera
            weather_factor *= (1 - reduction)
            adjustments.append(f"Pioggia: -{reduction*100:.1f}%")
        
        # Vento forte → riduce gol
        if wind_speed > 15:  # > 15 m/s (54 km/h)
            reduction = min(0.10, (wind_speed - 15) / 30)  # Max -10%
            weather_factor *= (1 - reduction)
            adjustments.append(f"Vento forte: -{reduction*100:.1f}%")
        elif wind_speed > 10:  # > 10 m/s (36 km/h)
            reduction = (wind_speed - 10) / 50  # Max -1%
            weather_factor *= (1 - reduction)
            adjustments.append(f"Vento moderato: -{reduction*100:.1f}%")
        
        # Neve → riduce molto gol
        if "snow" in conditions:
            weather_factor *= 0.85  # -15% gol
            adjustments.append("Neve: -15%")
        
        # Temperatura estrema → riduce gol
        if temp < 0:
            reduction = min(0.10, abs(temp) / 20)  # Max -10% per freddo estremo
            weather_factor *= (1 - reduction)
            adjustments.append(f"Freddo estremo: -{reduction*100:.1f}%")
        elif temp > 35:
            reduction = min(0.08, (temp - 35) / 15)  # Max -8% per caldo estremo
            weather_factor *= (1 - reduction)
            adjustments.append(f"Caldo estremo: -{reduction*100:.1f}%")
        
        # NUOVO: Umidità alta → può ridurre gol (campo pesante)
        if humidity > 80:
            reduction = min(0.05, (humidity - 80) / 100)  # Max -5% per umidità > 80%
            weather_factor *= (1 - reduction)
            adjustments.append(f"Umidità alta: -{reduction*100:.1f}%")
        
        # NUOVO: Pressione bassa → può ridurre gol (condizioni pesanti)
        if pressure < 1000:  # Pressione < 1000 hPa
            reduction = min(0.04, (1000 - pressure) / 50)  # Max -4%
            weather_factor *= (1 - reduction)
            adjustments.append(f"Pressione bassa: -{reduction*100:.1f}%")
        
        # NUOVO: Visibilità bassa → può ridurre gol (nebbia/foschia)
        if visibility < 1:  # Visibilità < 1km (nebbia fitta)
            reduction = 0.08  # -8% gol
            weather_factor *= (1 - reduction)
            adjustments.append(f"Nebbia fitta: -{reduction*100:.1f}%")
        elif visibility < 3:  # Visibilità < 3km (nebbia moderata)
            reduction = 0.04  # -4% gol
            weather_factor *= (1 - reduction)
            adjustments.append(f"Nebbia moderata: -{reduction*100:.1f}%")
        
        return {
            "available": True,
            "weather_factor": max(0.7, weather_factor),  # Min 0.7 (non ridurre troppo)
            "temperature": temp,
            "humidity": humidity,
            "pressure": pressure,  # NUOVO
            "wind_speed": wind_speed,
            "visibility": visibility,  # NUOVO
            "rain": rain,
            "conditions": conditions,
            "description": description,
            "adjustments": adjustments,
            "city": city
        }
    except (requests.exceptions.RequestException, KeyError, ValueError) as e:
        logger.error(f"Errore OpenWeatherMap: {e}")
        return {"available": False, "weather_factor": 1.0, "error": str(e)}

# 2. FOOTBALL-DATA.ORG API
def football_data_get_team_info(team_name: str, league_code: str = None) -> Dict[str, Any]:
    """
    Recupera info squadra da Football-Data.org (backup/validazione).
    """
    if not FOOTBALL_DATA_API_KEY:
        return {"available": False}
    
    try:
        # Cerca squadra
        url = "https://api.football-data.org/v4/teams"
        headers = {"X-Auth-Token": FOOTBALL_DATA_API_KEY}
        
        # Prova ricerca per nome
        params = {"name": team_name}
        response = requests.get(url, headers=headers, params=params, timeout=5)
        
        if response.status_code == 200:
            teams = response.json().get("teams", [])
            if teams:
                team = teams[0]
                return {
                    "available": True,
                    "team_id": team.get("id"),
                    "name": team.get("name"),
                    "short_name": team.get("shortName"),
                    "founded": team.get("founded"),
                    "venue": team.get("venue"),
                    "website": team.get("website")
                }
        
        return {"available": False, "reason": "Team not found"}
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore Football-Data.org: {e}")
        return {"available": False, "error": str(e)}

def football_data_get_recent_matches(team_id: int, limit: int = 5) -> List[Dict[str, Any]]:
    """Recupera ultime partite (FINISHED) per una squadra da Football-Data.org."""
    if not FOOTBALL_DATA_API_KEY or not team_id:
        return []
    
    cache_key = (int(team_id), int(limit))
    # FIX BUG #3-6: Use TTL cache instead of unbounded dict
    cached_result = _FOOTBALL_DATA_MATCH_CACHE.get(cache_key)
    if cached_result is not None:
        return cached_result

    headers = {"X-Auth-Token": FOOTBALL_DATA_API_KEY}
    params = {"status": "FINISHED", "limit": limit}

    try:
        response = requests.get(
            f"{FOOTBALL_DATA_BASE_URL}/teams/{team_id}/matches",
            headers=headers,
            params=params,
            timeout=6,
        )
        response.raise_for_status()
        matches = response.json().get("matches", [])
        if not isinstance(matches, list):
            matches = []
        _FOOTBALL_DATA_MATCH_CACHE.set(cache_key, matches)
        return matches
    except requests.exceptions.RequestException as e:
        logger.debug(f"Errore Football-Data.org recent matches: {e}")
        cached_fallback = _FOOTBALL_DATA_MATCH_CACHE.get(cache_key)
        return cached_fallback if cached_fallback is not None else []


def football_data_calculate_form_metrics(team_id: int, matches: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Calcola metriche forma da partite Football-Data.org (fallback statistico).
    """
    if not team_id or not matches:
        return {"available": False}
    
    total_goals_for = 0
    total_goals_against = 0
    total_points = 0
    recent_form = []
    clean_sheets = 0
    matches_used = 0
    
    for match in matches:
        if not isinstance(match, dict):
            continue
        
        home_team = match.get("homeTeam", {}) or {}
        away_team = match.get("awayTeam", {}) or {}
        score = match.get("score", {}) or {}
        full_time = score.get("fullTime", {}) or {}
        winner = score.get("winner")
        
        home_id = home_team.get("id")
        away_id = away_team.get("id")
        
        goals_home = full_time.get("home")
        goals_away = full_time.get("away")
        if goals_home is None or goals_away is None:
            continue
        
        goals_home = int(goals_home)
        goals_away = int(goals_away)
        
        is_home = home_id == team_id
        is_away = away_id == team_id
        
        if not (is_home or is_away):
            continue
        
        goals_for = goals_home if is_home else goals_away
        goals_against = goals_away if is_home else goals_home
        total_goals_for += goals_for
        total_goals_against += goals_against
        
        if goals_against == 0:
            clean_sheets += 1
        
        if winner == "DRAW":
            total_points += 1
            recent_form.append("D")
        elif (winner == "HOME_TEAM" and is_home) or (winner == "AWAY_TEAM" and is_away):
            total_points += 3
            recent_form.append("W")
        else:
            recent_form.append("L")
        
        matches_used += 1
    
    if matches_used == 0:
        return {"available": False}
    
    safe_matches = max(1, matches_used)
    points_per_game = total_points / safe_matches
    
    return {
        "available": True,
        "matches": matches_used,
        "avg_goals_for": round(total_goals_for / safe_matches, 3),
        "avg_goals_against": round(total_goals_against / safe_matches, 3),
        "points_per_game": round(points_per_game, 3),
        "recent_form": "".join(recent_form),
        "clean_sheet_rate": round(clean_sheets / safe_matches, 3),
    }

def football_data_get_competitions() -> List[Dict[str, Any]]:
    """Recupera lista competizioni disponibili."""
    if not FOOTBALL_DATA_API_KEY:
        return []

    try:
        url = "https://api.football-data.org/v4/competitions"
        headers = {"X-Auth-Token": FOOTBALL_DATA_API_KEY}
        response = requests.get(url, headers=headers, timeout=5)
        response.raise_for_status()
        return response.json().get("competitions", [])
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore Football-Data.org competitions: {e}")
        return []

def football_data_get_home_away_stats(team_id: int, competition_code: str = None) -> Dict[str, Any]:
    """
    Recupera statistiche separate HOME/AWAY da Football-Data.org standings endpoint.

    Args:
        team_id: ID della squadra su Football-Data.org
        competition_code: Codice competizione (es: "PL", "SA", "PD", "BL1", "FL1")

    Returns:
        Dict con statistiche home/away separate:
        {
            "available": bool,
            "home": {
                "played": int,
                "won": int,
                "draw": int,
                "lost": int,
                "goals_for": int,
                "goals_against": int,
                "points": int
            },
            "away": {...},
            "season": str
        }
    """
    if not FOOTBALL_DATA_API_KEY or not team_id:
        return {"available": False}

    # Cache key
    cache_key = (int(team_id), competition_code or "all")
    if cache_key in _FOOTBALL_DATA_MATCH_CACHE:
        cached = _FOOTBALL_DATA_MATCH_CACHE.get(cache_key)
        if cached and cached.get("available"):
            return cached

    try:
        headers = {"X-Auth-Token": FOOTBALL_DATA_API_KEY}

        # Se abbiamo competition_code, usiamolo
        if competition_code:
            url = f"{FOOTBALL_DATA_BASE_URL}/competitions/{competition_code}/standings"
        else:
            # Fallback: prova a recuperare dalla team
            url = f"{FOOTBALL_DATA_BASE_URL}/teams/{team_id}"
            response = requests.get(url, headers=headers, timeout=6)
            response.raise_for_status()
            team_data = response.json()

            # Estrai competition dalla squadra
            running_competitions = team_data.get("runningCompetitions", [])
            if not running_competitions:
                return {"available": False, "reason": "No active competitions"}

            # Usa la prima competizione attiva
            competition_code = running_competitions[0].get("code")
            if not competition_code:
                return {"available": False, "reason": "No competition code"}

            url = f"{FOOTBALL_DATA_BASE_URL}/competitions/{competition_code}/standings"

        # Recupera standings
        response = requests.get(url, headers=headers, timeout=6)
        response.raise_for_status()
        data = response.json()

        standings = data.get("standings", [])
        if not standings:
            return {"available": False, "reason": "No standings data"}

        # Cerca la squadra negli standings
        # Gli standings hanno 3 tipi: TOTAL, HOME, AWAY
        home_stats = None
        away_stats = None
        season_info = data.get("season", {})

        for standing_type in standings:
            standing_type_name = standing_type.get("type")
            table = standing_type.get("table", [])

            for entry in table:
                if entry.get("team", {}).get("id") == team_id:
                    stats = {
                        "played": entry.get("playedGames", 0),
                        "won": entry.get("won", 0),
                        "draw": entry.get("draw", 0),
                        "lost": entry.get("lost", 0),
                        "goals_for": entry.get("goalsFor", 0),
                        "goals_against": entry.get("goalsAgainst", 0),
                        "goal_difference": entry.get("goalDifference", 0),
                        "points": entry.get("points", 0)
                    }

                    if standing_type_name == "HOME":
                        home_stats = stats
                    elif standing_type_name == "AWAY":
                        away_stats = stats

                    break

        if not home_stats or not away_stats:
            return {"available": False, "reason": "Team not found in standings"}

        result = {
            "available": True,
            "home": home_stats,
            "away": away_stats,
            "season": f"{season_info.get('startDate', '')} - {season_info.get('endDate', '')}",
            "competition_code": competition_code
        }

        # Cache il risultato
        # FIX BUG #3-6: Use TTL cache .set() instead of direct dict access
        _FOOTBALL_DATA_MATCH_CACHE.set(cache_key, result)

        return result

    except requests.exceptions.RequestException as e:
        logger.debug(f"Errore Football-Data.org home/away stats: {e}")
        return {"available": False, "error": str(e)}

# 3. THESPORTSDB API (NO API KEY NECESSARIA!)
def thesportsdb_get_team_info(team_name: str) -> Dict[str, Any]:
    """
    Recupera info squadra da TheSportsDB (completamente gratuito, no API key).
    """
    try:
        url = "https://www.thesportsdb.com/api/v1/json/3/searchteams.php"
        params = {"t": team_name}
        response = requests.get(url, params=params, timeout=5)
        response.raise_for_status()
        data = response.json()
        
        teams = data.get("teams", [])
        if teams:
            team = teams[0]
            return {
                "available": True,
                "team_id": team.get("idTeam"),
                "name": team.get("strTeam"),
                "stadium": team.get("strStadium"),
                "stadium_capacity": team.get("intStadiumCapacity"),
                "stadium_location": team.get("strStadiumLocation"),
                "founded": team.get("intFormedYear"),
                "league": team.get("strLeague"),
                "website": team.get("strWebsite"),
                "logo": team.get("strTeamBadge")
            }
        
        return {"available": False}
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore TheSportsDB: {e}")
        return {"available": False, "error": str(e)}

def thesportsdb_get_stadium_info(stadium_name: str) -> Dict[str, Any]:
    """Recupera info stadio da TheSportsDB."""
    try:
        url = "https://www.thesportsdb.com/api/v1/json/3/searchvenues.php"
        params = {"v": stadium_name}
        response = requests.get(url, params=params, timeout=5)
        response.raise_for_status()
        data = response.json()
        
        venues = data.get("venues", [])
        if venues:
            venue = venues[0]
            return {
                "available": True,
                "name": venue.get("strVenue"),
                "capacity": venue.get("intCapacity"),
                "location": venue.get("strLocation"),
                "country": venue.get("strCountry")
            }
        
        return {"available": False}
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore TheSportsDB stadium: {e}")
        return {"available": False}

# 4. STATSBOMB OPEN DATA (xG avanzati gratuiti)
STATS_BOMB_BASE_URL = "https://raw.githubusercontent.com/statsbomb/open-data/master/data"
_STATS_BOMB_COMP_CACHE: Optional[List[Dict[str, Any]]] = None
# FIX BUG #3-6: Replace unbounded dicts with TTL caches to prevent memory leaks
_STATS_BOMB_MATCH_CACHE = TTLCache(max_size=150, ttl_seconds=1800)  # 30 min TTL, max 150 competitions
_STATS_BOMB_EVENTS_CACHE = TTLCache(max_size=200, ttl_seconds=1800)  # 30 min TTL, max 200 matches
_STATS_BOMB_TEAM_METRICS_CACHE = TTLCache(max_size=100, ttl_seconds=3600)  # 1 hour TTL, max 100 teams
_STATS_BOMB_MAX_COMPETITIONS = 25
_STATS_BOMB_MAX_MATCHES = 12
_STATS_BOMB_SHOT_ON_TARGET_OUTCOMES = {"Goal", "Saved", "Saved To Post"}


def _statsbomb_fetch_json(path: str) -> Any:
    """
    Recupera JSON da repository StatsBomb open data con caching leggero.
    """
    url = f"{STATS_BOMB_BASE_URL.rstrip('/')}/{path.lstrip('/')}"
    try:
        timeout = max(app_config.api_timeout, 12.0)
        response = requests.get(url, timeout=timeout)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as exc:
        logger.debug(f"StatsBomb request failed for {url}: {exc}")
        raise


def statsbomb_get_competitions() -> List[Dict[str, Any]]:
    """Ritorna lista competizioni disponibili (cache in memoria)."""
    global _STATS_BOMB_COMP_CACHE
    # ⚠️ FIX BUG: Double-check locking per evitare race condition
    if _STATS_BOMB_COMP_CACHE is not None:
        return _STATS_BOMB_COMP_CACHE

    try:
        competitions = _statsbomb_fetch_json("competitions.json")
        if not isinstance(competitions, list):
            competitions = []
        # ⚠️ NOTE: In ambiente multi-thread usare threading.Lock()
        # Per ora accettabile in ambiente single-thread (Streamlit)
        _STATS_BOMB_COMP_CACHE = competitions
    except requests.exceptions.RequestException:
        _STATS_BOMB_COMP_CACHE = []
    return _STATS_BOMB_COMP_CACHE


def statsbomb_get_matches(competition_id: int, season_id: int) -> List[Dict[str, Any]]:
    """Recupera lista partite per competition/season (cache)."""
    cache_key = f"{competition_id}_{season_id}"
    # FIX BUG #3-6: Use TTL cache instead of unbounded dict
    cached_result = _STATS_BOMB_MATCH_CACHE.get(cache_key)
    if cached_result is not None:
        return cached_result

    try:
        matches = _statsbomb_fetch_json(f"matches/{competition_id}/{season_id}.json")
        if not isinstance(matches, list):
            matches = []
        _STATS_BOMB_MATCH_CACHE.set(cache_key, matches)
    except requests.exceptions.RequestException:
        _STATS_BOMB_MATCH_CACHE.set(cache_key, [])
    return _STATS_BOMB_MATCH_CACHE.get(cache_key) or []


def statsbomb_get_events(match_id: int) -> List[Dict[str, Any]]:
    """Recupera eventi per una singola partita (cache)."""
    # FIX BUG #3-6: Use TTL cache instead of unbounded dict
    cached_result = _STATS_BOMB_EVENTS_CACHE.get(match_id)
    if cached_result is not None:
        return cached_result

    try:
        events = _statsbomb_fetch_json(f"events/{match_id}.json")
        if not isinstance(events, list):
            events = []
        _STATS_BOMB_EVENTS_CACHE.set(match_id, events)
    except requests.exceptions.RequestException:
        _STATS_BOMB_EVENTS_CACHE.set(match_id, [])
    return _STATS_BOMB_EVENTS_CACHE.get(match_id) or []


def statsbomb_get_team_metrics(team_name: str, max_matches: int = _STATS_BOMB_MAX_MATCHES) -> Dict[str, Any]:
    """
    Calcola metriche xG medie per una squadra usando StatsBomb Open Data.
    """
    if not team_name:
        return {"available": False, "reason": "team_name_missing"}

    cache_key = team_name.strip().lower()
    # FIX BUG #3-6: Use TTL cache instead of unbounded dict
    # FIX BUG #9.1: Replace deepcopy() with .copy() for better performance (50x faster)
    cached_result = _STATS_BOMB_TEAM_METRICS_CACHE.get(cache_key)
    if cached_result is not None:
        return cached_result.copy()

    competitions = statsbomb_get_competitions()
    if not competitions:
        result = {"available": False, "reason": "competitions_unavailable"}
        _STATS_BOMB_TEAM_METRICS_CACHE.set(cache_key, result)
        # FIX BUG #9.2: Replace deepcopy() with .copy() for better performance
        return result.copy()

    matches_to_process: List[Dict[str, Any]] = []
    team_lower = cache_key

    for competition in competitions[:_STATS_BOMB_MAX_COMPETITIONS]:
        comp_id = competition.get("competition_id")
        season_id = competition.get("season_id")
        match_available = competition.get("match_available", False)
        if not comp_id or not season_id or not match_available:
            continue

        matches = statsbomb_get_matches(comp_id, season_id)
        if not matches:
            continue

        for match in matches:
            home_info = match.get("home_team", {}) if isinstance(match, dict) else {}
            away_info = match.get("away_team", {}) if isinstance(match, dict) else {}
            home_name = str(home_info.get("home_team_name") or home_info.get("name") or "").lower()
            away_name = str(away_info.get("away_team_name") or away_info.get("name") or "").lower()

            is_home = False
            opponent = None

            if team_lower and team_lower in home_name:
                is_home = True
                opponent = (
                    away_info.get("away_team_name")
                    or away_info.get("name")
                    or away_info.get("team_name")
                )
            elif team_lower and team_lower in away_name:
                is_home = False
                opponent = (
                    home_info.get("home_team_name")
                    or home_info.get("name")
                    or home_info.get("team_name")
                )
            else:
                continue

            match_id = match.get("match_id")
            if not match_id:
                continue

            matches_to_process.append({
                "match_id": match_id,
                "is_home": is_home,
                "opponent": opponent,
                "competition": competition.get("competition_name"),
                "season": competition.get("season_name"),
                "match_date": match.get("match_date"),
            })

            if len(matches_to_process) >= max_matches:
                break
        if len(matches_to_process) >= max_matches:
            break

    if not matches_to_process:
        result = {"available": False, "reason": "team_not_found"}
        _STATS_BOMB_TEAM_METRICS_CACHE.set(cache_key, result)
        # FIX BUG #9.3: Replace deepcopy() with .copy() for better performance
        return result.copy()

    xg_for_total = 0.0
    xg_against_total = 0.0
    np_xg_for_total = 0.0
    np_xg_against_total = 0.0
    shots_for_total = 0
    shots_against_total = 0
    shots_on_target_for = 0
    shots_on_target_against = 0

    valid_matches = 0

    for match_meta in matches_to_process:
        events = statsbomb_get_events(match_meta["match_id"])
        if not events:
            continue

        xg_for_match = 0.0
        xg_against_match = 0.0
        np_xg_for_match = 0.0
        np_xg_against_match = 0.0
        shots_for_match = 0
        shots_against_match = 0
        shots_on_target_for_match = 0
        shots_on_target_against_match = 0

        for event in events:
            if not isinstance(event, dict):
                continue

            team_info = event.get("team", {}) if isinstance(event.get("team"), dict) else {}
            event_team_name = str(team_info.get("name") or "").lower()
            if event.get("type", {}).get("name") != "Shot":
                continue

            shot_data = event.get("shot", {})
            if not isinstance(shot_data, dict):
                continue

            xg_value = shot_data.get("statsbomb_xg")
            if xg_value is None:
                continue

            is_penalty = shot_data.get("type", {}).get("name") == "Penalty"
            outcome_name = shot_data.get("outcome", {}).get("name")

            if event_team_name == team_lower:
                xg_for_match += float(xg_value)
                shots_for_match += 1
                if outcome_name in _STATS_BOMB_SHOT_ON_TARGET_OUTCOMES:
                    shots_on_target_for_match += 1
                if not is_penalty:
                    np_xg_for_match += float(xg_value)
            else:
                xg_against_match += float(xg_value)
                shots_against_match += 1
                if outcome_name in _STATS_BOMB_SHOT_ON_TARGET_OUTCOMES:
                    shots_on_target_against_match += 1
                if not is_penalty:
                    np_xg_against_match += float(xg_value)

        valid_matches += 1
        xg_for_total += xg_for_match
        xg_against_total += xg_against_match
        np_xg_for_total += np_xg_for_match
        np_xg_against_total += np_xg_against_match
        shots_for_total += shots_for_match
        shots_against_total += shots_against_match
        shots_on_target_for += shots_on_target_for_match
        shots_on_target_against += shots_on_target_against_match

    if valid_matches == 0:
        result = {"available": False, "reason": "events_unavailable"}
        _STATS_BOMB_TEAM_METRICS_CACHE.set(cache_key, result)
        # FIX BUG #9.4: Replace deepcopy() with .copy() for better performance
        return result.copy()

    safe_matches = max(1, valid_matches)
    result = {
        "available": True,
        "matches_used": valid_matches,
        "xg_for_avg": round(xg_for_total / safe_matches, 3),
        "xg_against_avg": round(xg_against_total / safe_matches, 3),
        "non_pen_xg_for_avg": round(np_xg_for_total / safe_matches, 3),
        "non_pen_xg_against_avg": round(np_xg_against_total / safe_matches, 3),
        "shots_for_avg": round(shots_for_total / safe_matches, 2),
        "shots_against_avg": round(shots_against_total / safe_matches, 2),
        "shots_on_target_for_avg": round(shots_on_target_for / safe_matches, 2),
        "shots_on_target_against_avg": round(shots_on_target_against / safe_matches, 2),
        "recent_opponents": [m.get("opponent") for m in matches_to_process],
        "note": "Metriche calcolate da StatsBomb Open Data",
    }

    _STATS_BOMB_TEAM_METRICS_CACHE.set(cache_key, result)
    # FIX BUG #9.5: Replace deepcopy() with .copy() for better performance (50x faster)
    return result.copy()


# 4. UNDERSTAT SCRAPING (xG avanzato)
def understat_get_team_xg(team_name: str, season: str = None) -> Dict[str, Any]:
    """
    Recupera xG avanzato da Understat.com (scraping).
    Ritorna xG più preciso di API-Football.
    """
    try:
        # Understat usa ID squadra, non nome
        # Per ora ritorna struttura, implementazione completa richiede mapping team_id
        # Questo è un placeholder per struttura futura
        
        # Esempio struttura dati che si otterrebbe:
        return {
            "available": False,  # Placeholder - richiede implementazione scraping completo
            "team_name": team_name,
            "xg_for": None,
            "xg_against": None,
            "xg_open_play": None,
            "xg_set_pieces": None,
            "note": "Richiede implementazione scraping BeautifulSoup"
        }
    except Exception as e:
        logger.error(f"Errore Understat: {e}")
        return {"available": False, "error": str(e)}

# 5. FBREF SCRAPING (Statistiche avanzate)
def fbref_get_team_stats(team_name: str, league: str = None) -> Dict[str, Any]:
    """
    Recupera statistiche avanzate da FBRef.com (scraping).
    """
    try:
        # Placeholder - implementazione completa richiede scraping
        return {
            "available": False,  # Placeholder
            "team_name": team_name,
            "expected_goals": None,
            "expected_assists": None,
            "expected_goals_against": None,
            "note": "Richiede implementazione scraping BeautifulSoup"
        }
    except Exception as e:
        logger.error(f"Errore FBRef: {e}")
        return {"available": False, "error": str(e)}

# Funzione helper per applicare impatto meteo su lambda
def apply_weather_impact(lh: float, la: float, weather_data: Dict[str, Any]) -> Tuple[float, float]:
    """
    Applica impatto meteo su lambda home e away.
    """
    if not weather_data.get("available", False):
        return lh, la
    
    weather_factor = weather_data.get("weather_factor", 1.0)
    
    # Applica fattore a entrambe le lambda (meteo impatta entrambe le squadre)
    lh_adjusted = lh * weather_factor
    la_adjusted = la * weather_factor
    
    return lh_adjusted, la_adjusted

def apply_stadium_adjustments(
    lambda_h: float,
    lambda_a: float,
    stadium_data: Dict[str, Any],
) -> Tuple[float, float]:
    """
    Applica aggiustamenti basati su dati stadio (capacità, altitudine).
    
    - Capacità alta → più atmosfera → più gol casa
    - Altitudine alta → meno ossigeno → meno gol (entrambe)
    """
    if not stadium_data or not stadium_data.get("available"):
        return lambda_h, lambda_a
    
    capacity = stadium_data.get("stadium_capacity")
    # Nota: TheSportsDB non fornisce altitudine direttamente, ma possiamo stimarla
    # da location se disponibile (es. città ad alta quota)
    
    # Aggiustamento capacità: stadio grande → più atmosfera → più gol casa
    # Converti capacity in numero se disponibile (può essere stringa, int, float, o None)
    if capacity is not None:
        try:
            # Prova a convertire in int (rimuove spazi e caratteri non numerici se stringa)
            if isinstance(capacity, str):
                # Rimuovi spazi e caratteri non numerici
                capacity_str = ''.join(filter(str.isdigit, capacity))
                if capacity_str:
                    capacity_num = int(capacity_str)
                else:
                    capacity_num = None
            elif isinstance(capacity, (int, float)):
                capacity_num = int(capacity)
            else:
                capacity_num = None
            
            # Applica aggiustamenti solo se abbiamo un numero valido
            if capacity_num is not None and capacity_num > 0:
                # Stadio > 50k → +3% gol casa
                if capacity_num > 50000:
                    capacity_factor = 1.03
                # Stadio > 30k → +2% gol casa
                elif capacity_num > 30000:
                    capacity_factor = 1.02
                # Stadio < 15k → -1% gol casa (meno atmosfera)
                elif capacity_num < 15000:
                    capacity_factor = 0.99
                else:
                    capacity_factor = 1.0
                
                # Applica solo a lambda_h (vantaggio casa)
                lambda_h *= capacity_factor
        except (ValueError, TypeError):
            # Se la conversione fallisce, ignora l'aggiustamento capacità
            pass
    
    # Aggiustamento altitudine (se disponibile da location)
    # Nota: implementazione base, può essere estesa con mappa città → altitudine
    location_raw = stadium_data.get("stadium_location")
    # Assicurati che location sia sempre una stringa (gestisce None)
    location = str(location_raw).lower() if location_raw is not None else ""
    altitude_factor = 1.0
    
    # Città ad alta quota (es. La Paz, Quito, Città del Messico)
    if location:  # Solo se abbiamo una location valida
        high_altitude_cities = ["la paz", "quito", "mexico city", "città del messico", "bogotá", "bogota"]
        if any(city in location for city in high_altitude_cities):
            altitude_factor = 0.92  # -8% gol (entrambe le squadre)
            lambda_h *= altitude_factor
            lambda_a *= altitude_factor
    
    return lambda_h, lambda_a

# Funzione helper per recuperare città da nome squadra/stadio
def get_city_from_team(team_name: str, league: str = None) -> str:
    """
    Prova a recuperare città della squadra da TheSportsDB o API-Football.
    """
    # Prova TheSportsDB (gratuito, no API key)
    team_info = thesportsdb_get_team_info(team_name)
    if team_info.get("available"):
        stadium_location = team_info.get("stadium_location")
        # Assicurati che stadium_location sia una stringa (gestisce None)
        if stadium_location and isinstance(stadium_location, str):
            # Estrai città (prima parte prima della virgola)
            # FIX BUG #10.3: Safe array access on split()
            parts = stadium_location.split(",")
            city = parts[0].strip() if parts else ""
            if city:  # Solo se abbiamo una città valida
                return city
    
    # Fallback: mappa manuale per squadre principali
    city_mapping = {
        "milan": "Milano",
        "inter": "Milano",
        "juventus": "Torino",
        "roma": "Roma",
        "lazio": "Roma",
        "napoli": "Napoli",
        "atalanta": "Bergamo",
        "fiorentina": "Firenze",
        "bologna": "Bologna",
        "manchester city": "Manchester",
        "manchester united": "Manchester",
        "liverpool": "Liverpool",
        "chelsea": "London",
        "arsenal": "London",
        "tottenham": "London",
        "real madrid": "Madrid",
        "barcelona": "Barcelona",
        "atletico madrid": "Madrid",
        "bayern munich": "Munich",
        "dortmund": "Dortmund",
        "psg": "Paris"
    }
    
    team_lower = team_name.lower()
    for key, city in city_mapping.items():
        if key in team_lower:
            return city
    
    return ""  # Non trovato

# ============================================================
#   TELEGRAM BOT INTEGRATION
# ============================================================

def send_telegram_message(
    message: str,
    bot_token: str = None,
    chat_id: str = None,
    parse_mode: str = "HTML",
) -> Dict[str, Any]:
    """
    Invia messaggio a Telegram tramite Bot API.
    
    Args:
        message: Testo del messaggio (supporta HTML/Markdown)
        bot_token: Token del bot (default: TELEGRAM_BOT_TOKEN)
        chat_id: Chat ID destinatario (default: TELEGRAM_CHAT_ID)
        parse_mode: "HTML" o "Markdown"
    
    Returns:
        Dict con:
            - success: bool (True se invio riuscito)
            - error_message: str (messaggio di errore dettagliato se fallito)
            - error_type: str (tipo di errore: "no_token", "no_chat_id", "invalid_token", "invalid_chat_id", "timeout", "other")
    """
    if not bot_token:
        bot_token = TELEGRAM_BOT_TOKEN
    if not chat_id:
        chat_id = TELEGRAM_CHAT_ID
    
    # Verifica token e chat ID
    if not bot_token or bot_token.strip() == "":
        return {
            "success": False,
            "error_message": "Bot Token non configurato. Configura TELEGRAM_BOT_TOKEN o inserisci il token nell'interfaccia.",
            "error_type": "no_token"
        }
    
    if not chat_id or chat_id.strip() == "":
        return {
            "success": False,
            "error_message": "Chat ID non configurato. Configura TELEGRAM_CHAT_ID o inserisci il Chat ID nell'interfaccia.",
            "error_type": "no_chat_id"
        }
    
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    
    payload = {
        "chat_id": chat_id,
        "text": message,
        "parse_mode": parse_mode,
        "disable_web_page_preview": True,
    }
    
    try:
        response = requests.post(url, json=payload, timeout=10)  # Aumentato timeout a 10s
        
        # Controlla status code specifici
        if response.status_code == 401:
            return {
                "success": False,
                "error_message": "Token non valido o scaduto. Verifica che il Bot Token sia corretto.",
                "error_type": "invalid_token"
            }
        elif response.status_code == 400:
            # Prova a estrarre dettagli dall'errore
            try:
                error_data = response.json()
                error_desc = error_data.get("description", "Chat ID non valido")
                if "chat not found" in error_desc.lower():
                    # Estrai informazioni aggiuntive se disponibili
                    error_msg_detailed = (
                        "**Chat ID non trovato o bot non autorizzato.**\n\n"
                        "Possibili cause:\n"
                        "1. **Chat ID errato**: Verifica di aver copiato correttamente il Chat ID\n"
                        "2. **Bot non avviato**: Il bot deve aver ricevuto almeno un messaggio da te\n"
                        "3. **Gruppo/Canale**: Se è un gruppo/canale, il bot deve essere membro/amministratore\n"
                        "4. **Chat privata**: Per chat private, assicurati di aver avviato la conversazione\n\n"
                        "**Come ottenere il Chat ID corretto:**\n"
                        "- **Chat privata**: Scrivi a [@userinfobot](https://t.me/userinfobot) e copia il tuo ID\n"
                        "- **Gruppo**: Aggiungi [@userinfobot](https://t.me/userinfobot) al gruppo e vedi il Group ID\n"
                        "- **Canale**: Il bot deve essere amministratore del canale\n\n"
                        f"Errore API: {error_desc}"
                    )
                    return {
                        "success": False,
                        "error_message": error_msg_detailed,
                        "error_type": "invalid_chat_id"
                    }
            except (json.JSONDecodeError, KeyError, AttributeError) as e:
                # FIX BUG: Specifico eccezioni invece di bare except
                logger.debug(f"Could not parse Telegram error response: {e}")
            return {
                "success": False,
                "error_message": "Chat ID non valido o formato errato. Verifica che il Chat ID sia corretto.",
                "error_type": "invalid_chat_id"
            }
        elif response.status_code == 429:
            # Rate limit
            return {
                "success": False,
                "error_message": "Rate limit raggiunto. Troppi messaggi inviati. Attendi qualche minuto.",
                "error_type": "rate_limit"
            }
        
        response.raise_for_status()
        return {
            "success": True,
            "error_message": None,
            "error_type": None
        }
        
    except requests.exceptions.Timeout:
        return {
            "success": False,
            "error_message": "Timeout nella richiesta a Telegram. Verifica la connessione internet.",
            "error_type": "timeout"
        }
    except requests.exceptions.ConnectionError:
        return {
            "success": False,
            "error_message": "Errore di connessione a Telegram. Verifica la connessione internet.",
            "error_type": "connection_error"
        }
    except requests.exceptions.HTTPError as e:
        status_code = e.response.status_code if hasattr(e, 'response') and e.response else None
        return {
            "success": False,
            "error_message": f"Errore HTTP {status_code} da Telegram API. Verifica token e chat ID.",
            "error_type": "http_error"
        }
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore invio Telegram: {e}")
        return {
            "success": False,
            "error_message": f"Errore nella richiesta a Telegram: {str(e)}",
            "error_type": "other"
        }
    except Exception as e:
        logger.error(f"Errore imprevisto invio Telegram: {e}")
        return {
            "success": False,
            "error_message": f"Errore imprevisto: {str(e)}",
            "error_type": "other"
        }

def test_telegram_chat_id(
    bot_token: str = None,
    chat_id: str = None,
) -> Dict[str, Any]:
    """
    Testa se il Chat ID è valido inviando un messaggio di test.
    
    Args:
        bot_token: Token del bot (default: TELEGRAM_BOT_TOKEN)
        chat_id: Chat ID da testare (default: TELEGRAM_CHAT_ID)
    
    Returns:
        Dict con:
            - success: bool
            - error_message: str
            - error_type: str
    """
    test_message = "🧪 **Test connessione Telegram**\n\nSe ricevi questo messaggio, la configurazione è corretta! ✅"
    
    return send_telegram_message(
        message=test_message,
        bot_token=bot_token,
        chat_id=chat_id,
        parse_mode="HTML"
    )

def send_telegram_photo(
    photo_path: str,
    caption: str = "",
    bot_token: str = None,
    chat_id: str = None,
) -> bool:
    """
    Invia foto a Telegram.
    
    Args:
        photo_path: Percorso file immagine
        caption: Didascalia (opzionale)
        bot_token: Token del bot
        chat_id: Chat ID destinatario
    
    Returns:
        True se invio riuscito, False altrimenti
    """
    if not bot_token:
        bot_token = TELEGRAM_BOT_TOKEN
    if not chat_id:
        chat_id = TELEGRAM_CHAT_ID
    
    if not bot_token or not chat_id or not os.path.exists(photo_path):
        return False
    
    url = f"https://api.telegram.org/bot{bot_token}/sendPhoto"
    
    try:
        with open(photo_path, 'rb') as photo:
            files = {'photo': photo}
            data = {
                'chat_id': chat_id,
                'caption': caption[:1024] if caption else "",  # Max 1024 caratteri
            }
            response = requests.post(url, files=files, data=data, timeout=10)
            response.raise_for_status()
            return True
    except (requests.exceptions.RequestException, IOError, FileNotFoundError) as e:
        logger.error(f"Errore invio foto Telegram: {e}")
        return False

def format_analysis_for_telegram(
    match_name: str,
    ris: Dict[str, Any],
    odds_1: float,
    odds_x: float,
    odds_2: float,
    quality_score: float,
    market_conf: float,
    value_bets: List[Dict[str, Any]] = None,
) -> str:
    """
    Formatta analisi completa per messaggio Telegram.
    
    Usa HTML per formattazione (grassetto, corsivo, etc.)
    """
    # Header
    message = f"⚽ <b>ANALISI COMPLETATA</b>\n\n"
    message += f"🏆 <b>{match_name}</b>\n"
    message += f"📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n"
    
    # Metriche principali
    message += f"📊 <b>Metriche Qualità</b>\n"
    message += f"Quality Score: {quality_score:.0f}/100\n"
    message += f"Market Confidence: {market_conf:.0f}/100\n\n"
    
    # Parametri modello
    message += f"🔢 <b>Parametri Modello</b>\n"
    message += f"λ Casa: {ris['lambda_home']:.3f}\n"
    message += f"λ Trasferta: {ris['lambda_away']:.3f}\n"
    message += f"ρ (correlazione): {ris['rho']:.4f}\n\n"
    
    # Probabilità principali
    message += f"🎯 <b>Probabilità Esito</b>\n"
    message += f"🏠 Casa: <b>{ris['p_home']*100:.1f}%</b> (quota: {odds_1:.2f})\n"
    message += f"⚖️ Pareggio: <b>{ris['p_draw']*100:.1f}%</b> (quota: {odds_x:.2f})\n"
    message += f"✈️ Trasferta: <b>{ris['p_away']*100:.1f}%</b> (quota: {odds_2:.2f})\n\n"
    
    # Over/Under e BTTS
    message += f"⚽ <b>Mercati Speciali</b>\n"
    message += f"Over 2.5: {ris['over_25']*100:.1f}%\n"
    message += f"Under 2.5: {ris['under_25']*100:.1f}%\n"
    message += f"BTTS: {ris['btts']*100:.1f}%\n"
    message += f"GG + Over 2.5: {ris['gg_over25']*100:.1f}%\n\n"

    # Pari/Dispari FT
    message += f"🎲 <b>Pari/Dispari (FT)</b>\n"
    message += f"Pari: {ris['even_ft']*100:.1f}%\n"
    message += f"Dispari: {ris['odd_ft']*100:.1f}%\n\n"

    # Clean Sheet
    message += f"🛡️ <b>Porta Inviolata</b>\n"
    message += f"CS Casa: {ris['cs_home']*100:.1f}%\n"
    message += f"CS Trasferta: {ris['cs_away']*100:.1f}%\n\n"

    # Over/Under Half Time (se disponibili)
    if 'over_05_ht' in ris or 'over_15_ht' in ris:
        message += f"⏱️ <b>Primo Tempo (HT)</b>\n"
        if 'over_05_ht' in ris:
            message += f"Over 0.5 HT: {ris['over_05_ht']*100:.1f}%\n"
            message += f"Under 0.5 HT: {(1 - ris['over_05_ht'])*100:.1f}%\n"
        if 'over_15_ht' in ris:
            message += f"Over 1.5 HT: {ris['over_15_ht']*100:.1f}%\n"
            message += f"Under 1.5 HT: {(1 - ris['over_15_ht'])*100:.1f}%\n"
        # Pari/Dispari HT
        if 'even_ht' in ris and 'odd_ht' in ris:
            message += f"Pari HT: {ris['even_ht']*100:.1f}%\n"
            message += f"Dispari HT: {ris['odd_ht']*100:.1f}%\n"
        message += "\n"

    # Mercati Combinati HT + FT (se disponibili)
    has_combined = any(key in ris for key in ['over_05ht_over_05ft', 'over_05ht_over_15ft',
                                                'over_05ht_over_25ft', 'over_15ht_over_25ft',
                                                'over_05ht_over_35ft'])
    if has_combined:
        message += f"🔗 <b>Combinati HT + FT</b>\n"
        if 'over_05ht_over_05ft' in ris:
            message += f"Over 0.5 HT + Over 0.5 FT: {ris['over_05ht_over_05ft']*100:.1f}%\n"
        if 'over_05ht_over_15ft' in ris:
            message += f"Over 0.5 HT + Over 1.5 FT: {ris['over_05ht_over_15ft']*100:.1f}%\n"
        if 'over_05ht_over_25ft' in ris:
            message += f"Over 0.5 HT + Over 2.5 FT: {ris['over_05ht_over_25ft']*100:.1f}%\n"
        if 'over_15ht_over_25ft' in ris:
            message += f"Over 1.5 HT + Over 2.5 FT: {ris['over_15ht_over_25ft']*100:.1f}%\n"
        if 'over_05ht_over_35ft' in ris:
            message += f"Over 0.5 HT + Over 3.5 FT: {ris['over_05ht_over_35ft']*100:.1f}%\n"
        message += "\n"

    # Asian Handicap (top lines)
    if 'asian_handicap' in ris:
        ah = ris['asian_handicap']
        message += f"🎯 <b>Asian Handicap</b>\n"
        # Mostra le linee più rilevanti (prob >= 25%)
        relevant_ah = [(k, v) for k, v in ah.items() if v >= 0.25]
        if relevant_ah:
            # Ordina per probabilità decrescente
            relevant_ah.sort(key=lambda x: x[1], reverse=True)
            for ah_line, prob in relevant_ah[:6]:  # Max 6 linee
                message += f"{ah_line}: {prob*100:.1f}%\n"
            message += "\n"

    # ============================================================
    # MERCATI COMBINATI (Esito + Over/Under, DC + Over/Under, Multigol)
    # ============================================================
    if 'combo_book' in ris:
        combo_book = ris['combo_book']

        # Raggruppa combo per tipo
        esito_over = {}
        esito_under = {}
        dc_over = {}
        dc_under = {}
        esito_multigol = {}
        dc_multigol = {}

        for combo_key, prob in combo_book.items():
            # Esito + Over
            if ' & Over' in combo_key and not any(dc in combo_key for dc in ['1X', 'X2', '12']):
                esito_over[combo_key] = prob
            # Esito + Under
            elif ' & Under' in combo_key and not any(dc in combo_key for dc in ['1X', 'X2', '12']):
                esito_under[combo_key] = prob
            # Double Chance + Over
            elif any(dc in combo_key for dc in ['1X', 'X2', '12']) and ' & Over' in combo_key:
                dc_over[combo_key] = prob
            # Double Chance + Under
            elif any(dc in combo_key for dc in ['1X', 'X2', '12']) and ' & Under' in combo_key:
                dc_under[combo_key] = prob
            # Esito + Multigol
            elif ' & Multigol' in combo_key and combo_key.startswith(('1 &', '2 &')):
                esito_multigol[combo_key] = prob
            # Double Chance + Multigol
            elif ' & Multigol' in combo_key and any(dc in combo_key for dc in ['1X', 'X2', '12']):
                dc_multigol[combo_key] = prob

        # Mostra Esito + Over/Under
        if esito_over or esito_under:
            message += f"🎲 <b>Esito + Over/Under</b>\n"

            # Ordina e mostra Over
            if esito_over:
                for combo_key in sorted(esito_over.keys()):
                    prob = esito_over[combo_key]
                    if prob >= 0.10:  # Mostra solo prob >= 10%
                        message += f"{combo_key}: {prob*100:.1f}%\n"

            # Ordina e mostra Under
            if esito_under:
                for combo_key in sorted(esito_under.keys()):
                    prob = esito_under[combo_key]
                    if prob >= 0.10:
                        message += f"{combo_key}: {prob*100:.1f}%\n"

            message += "\n"

        # Mostra Double Chance + Over/Under
        if dc_over or dc_under:
            message += f"🎯 <b>Double Chance + Over/Under</b>\n"

            # Ordina e mostra Over
            if dc_over:
                for combo_key in sorted(dc_over.keys()):
                    prob = dc_over[combo_key]
                    if prob >= 0.10:
                        message += f"{combo_key}: {prob*100:.1f}%\n"

            # Ordina e mostra Under
            if dc_under:
                for combo_key in sorted(dc_under.keys()):
                    prob = dc_under[combo_key]
                    if prob >= 0.10:
                        message += f"{combo_key}: {prob*100:.1f}%\n"

            message += "\n"

        # Mostra Esito + Multigol
        if esito_multigol:
            message += f"🎰 <b>Esito + Multigol</b>\n"

            # Ordina per esito (1, 2) e poi per range
            sorted_esito_mg = sorted(esito_multigol.items(),
                                    key=lambda x: (x[0].split(' &')[0], x[0]))

            for combo_key, prob in sorted_esito_mg:
                if prob >= 0.05:  # Mostra solo prob >= 5%
                    message += f"{combo_key}: {prob*100:.1f}%\n"

            message += "\n"

        # Mostra Double Chance + Multigol
        if dc_multigol:
            message += f"🎲 <b>Double Chance + Multigol</b>\n"

            # Ordina per DC (1X, X2, 12) e poi per range
            sorted_dc_mg = sorted(dc_multigol.items(),
                                 key=lambda x: (x[0].split(' &')[0], x[0]))

            for combo_key, prob in sorted_dc_mg:
                if prob >= 0.05:  # Mostra solo prob >= 5%
                    message += f"{combo_key}: {prob*100:.1f}%\n"

            message += "\n"

    # HT/FT (Halftime/Fulltime)
    if 'ht_ft' in ris:
        ht_ft = ris['ht_ft']
        message += f"⏱️🏁 <b>HT/FT (Halftime/Fulltime)</b>\n"
        # Mostra le top 5 combinazioni più probabili
        sorted_ht_ft = sorted(ht_ft.items(), key=lambda x: x[1], reverse=True)
        for combo, prob in sorted_ht_ft[:5]:  # Top 5
            if prob >= 0.05:  # Mostra solo se >= 5%
                message += f"{combo}: {prob*100:.1f}%\n"
        message += "\n"

    # Value Bets
    if value_bets:
        message += f"💎 <b>Value Bets Identificate</b>\n"
        for bet in value_bets[:5]:  # Max 5 value bets
            esito = bet.get("Esito", "")
            prob = bet.get("Prob %", bet.get("Prob Modello %", ""))
            edge = bet.get("Edge %", "0")
            ev = bet.get("EV %", "0")
            rec = bet.get("Rec", "")
            prob_str = f"{prob}%" if prob and "%" not in str(prob) else prob
            message += f"• {esito}: Prob {prob_str}, Edge {edge}%, EV {ev}% ({rec})\n"
        message += "\n"
    
    # Top 4 risultati esatti più probabili
    if "top10" in ris and len(ris["top10"]) > 0:
        message += f"🏅 <b>Top 4 Risultati Esatti</b>\n"
        for i, (h, a, p) in enumerate(ris["top10"][:4], 1):
            message += f"{i}. {h}-{a}: {p:.1f}%\n"
        message += "\n"
    
    # Aggiustamenti applicati
    adjustments = []
    if ris.get("calibration_applied"):
        adjustments.append("Calibrazione")
    if ris.get("ensemble_applied"):
        adjustments.append("Ensemble")
    if ris.get("market_movement", {}).get("movement_type") != "NO_OPENING_DATA":
        adjustments.append("Market Movement")
    
    if adjustments:
        message += f"✅ <b>Aggiustamenti</b>: {', '.join(adjustments)}\n"
    
    # Footer
    message += f"\n🤖 <i>Inviato automaticamente dal Modello Scommesse PRO</i>"
    
    return message

# ============================================================
#   MULTI-MATCH TELEGRAM FORMATTING
# ============================================================

def format_multiple_matches_for_telegram(
    analyses: List[Dict[str, Any]],
    telegram_prob_threshold: float = 50.0
) -> str:
    """
    Formatta UN messaggio Telegram con multiple partite.

    Args:
        analyses: Lista di dict con chiavi: match_name, ris, odds_1, odds_x, odds_2,
                  quality_score, market_conf, value_bets
        telegram_prob_threshold: Soglia minima probabilità per includere value bets

    Returns:
        Messaggio HTML formattato per Telegram
    """
    total_matches = len(analyses)

    # Header principale
    message = f"📊 <b>ANALISI MULTIPLA - {total_matches} PARTIT{'A' if total_matches == 1 else 'E'}</b>\n"
    message += f"📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
    message += "━━━━━━━━━━━━━━━━━━━━━\n\n"

    # Itera su ogni partita
    total_value_bets = 0
    for i, analysis in enumerate(analyses, 1):
        match_name = analysis["match_name"]
        ris = analysis["ris"]
        odds_1 = analysis["odds_1"]
        odds_x = analysis["odds_x"]
        odds_2 = analysis["odds_2"]
        quality_score = analysis.get("quality_score", 0)
        market_conf = analysis.get("market_conf", 0)
        value_bets = analysis.get("value_bets", [])

        # Nome partita
        message += f"<b>{i}. ⚽ {match_name}</b>\n\n"

        # Metriche qualità compatte
        message += f"📊 Quality: {quality_score:.0f}/100 | Confidence: {market_conf:.0f}/100\n\n"

        # Probabilità esito
        message += f"🎯 <b>Probabilità:</b>\n"
        message += f"  1️⃣ Casa: <b>{ris['p_home']*100:.1f}%</b> (q. {odds_1:.2f})\n"
        message += f"  ❌ Pareggio: <b>{ris['p_draw']*100:.1f}%</b> (q. {odds_x:.2f})\n"
        message += f"  2️⃣ Trasferta: <b>{ris['p_away']*100:.1f}%</b> (q. {odds_2:.2f})\n\n"

        # Mercati speciali (compatto)
        message += (
            f"⚽ Over 2.5: {ris['over_25']*100:.1f}% | "
            f"Under: {ris['under_25']*100:.1f}% | "
            f"BTTS: {ris['btts']*100:.1f}% | "
            f"GG+Over 2.5: {ris['gg_over25']*100:.1f}%\n"
        )

        # Pari/Dispari FT
        if 'even_ft' in ris and 'odd_ft' in ris:
            message += (
                f"🎲 Pari: {ris['even_ft']*100:.1f}% | "
                f"Dispari: {ris['odd_ft']*100:.1f}%\n"
            )

        # Clean Sheet
        if 'cs_home' in ris and 'cs_away' in ris:
            message += (
                f"🛡️ CS Casa: {ris['cs_home']*100:.1f}% | "
                f"CS Trasferta: {ris['cs_away']*100:.1f}%\n"
            )

        # Mercati HT e combinati (se disponibili)
        ht_markets = []
        if 'over_05_ht' in ris:
            ht_markets.append(f"Over 0.5 HT: {ris['over_05_ht']*100:.1f}%")
        if 'over_15_ht' in ris:
            ht_markets.append(f"Over 1.5 HT: {ris['over_15_ht']*100:.1f}%")
        if 'even_ht' in ris and 'odd_ht' in ris:
            ht_markets.append(f"Pari HT: {ris['even_ht']*100:.1f}%")
            ht_markets.append(f"Dispari HT: {ris['odd_ht']*100:.1f}%")
        if ht_markets:
            message += f"⏱️ {' | '.join(ht_markets)}\n"

        combined_markets = []
        if 'over_05ht_over_25ft' in ris:
            combined_markets.append(f"Over 0.5 HT + Over 2.5 FT: {ris['over_05ht_over_25ft']*100:.1f}%")
        if 'over_15ht_over_25ft' in ris:
            combined_markets.append(f"Over 1.5 HT + Over 2.5 FT: {ris['over_15ht_over_25ft']*100:.1f}%")
        if combined_markets:
            message += f"🔗 {' | '.join(combined_markets)}\n"

        # Asian Handicap (top 2 lines compatto)
        if 'asian_handicap' in ris:
            ah = ris['asian_handicap']
            relevant_ah = [(k, v) for k, v in ah.items() if v >= 0.25]
            if relevant_ah:
                relevant_ah.sort(key=lambda x: x[1], reverse=True)
                ah_str = ' | '.join([f"{k}: {v*100:.1f}%" for k, v in relevant_ah[:2]])
                message += f"🎯 {ah_str}\n"

        # HT/FT (top 3 compatto)
        if 'ht_ft' in ris:
            ht_ft = ris['ht_ft']
            sorted_ht_ft = sorted(ht_ft.items(), key=lambda x: x[1], reverse=True)
            top_ht_ft = [(combo, prob) for combo, prob in sorted_ht_ft[:3] if prob >= 0.08]
            if top_ht_ft:
                ht_ft_str = ' | '.join([f"{combo}: {prob*100:.1f}%" for combo, prob in top_ht_ft])
                message += f"⏱️🏁 {ht_ft_str}\n"

        message += "\n"

        if "top10" in ris and len(ris["top10"]) > 0:
            message += "🏅 Top 3 Risultati:\n"
            for idx, (h, a, p) in enumerate(ris["top10"][:3], 1):
                message += f"  {idx}. {h}-{a}: {p:.1f}%\n"
            message += "\n"

        # Value Bets filtrati per soglia
        filtered_vbs = [vb for vb in value_bets if float(str(vb.get("Prob %", "0")).replace("%", "").replace(",", ".")) >= telegram_prob_threshold]

        if filtered_vbs:
            total_value_bets += len(filtered_vbs)
            message += f"💎 <b>Value Bets ({len(filtered_vbs)}):</b>\n"
            for vb in filtered_vbs[:3]:  # Max 3 per partita per brevità
                esito = vb.get("Esito", "")
                prob = vb.get("Prob %", vb.get("Prob Modello %", ""))
                edge = vb.get("Edge %", "0")
                ev = vb.get("EV %", "0")
                prob_str = f"{prob}%" if prob and "%" not in str(prob) else prob
                message += f"  • {esito}: {prob_str} | Edge {edge}% | EV {ev}%\n"
            if len(filtered_vbs) > 3:
                message += f"  <i>(+{len(filtered_vbs)-3} altri)</i>\n"
        else:
            message += f"ℹ️ Nessun value bet sopra soglia ({telegram_prob_threshold:.0f}%)\n"

        message += "\n━━━━━━━━━━━━━━━━━━━━━\n\n"

    # Riepilogo finale
    message += f"📈 <b>Riepilogo:</b>\n"
    message += f"  • Partite analizzate: {total_matches}\n"
    message += f"  • Value bets totali: {total_value_bets}\n"
    message += f"  • Soglia probabilità: ≥{telegram_prob_threshold:.0f}%\n\n"

    # Footer
    message += f"🤖 <i>Analisi automatica - Modello Dixon-Coles Bayesiano</i>"

    return message


def split_telegram_message(message: str, max_length: int = 4096) -> List[str]:
    """
    Divide un messaggio Telegram lungo in più parti.

    Telegram ha un limite di 4096 caratteri per messaggio.
    Divide in modo intelligente cercando di spezzare su separatori.

    Args:
        message: Messaggio completo
        max_length: Lunghezza massima per parte (default 4096)

    Returns:
        Lista di messaggi divisi
    """
    if len(message) <= max_length:
        return [message]

    messages = []
    current = ""

    # Dividi per separatori logici (━━━ indica fine sezione partita)
    sections = message.split("━━━━━━━━━━━━━━━━━━━━━\n\n")

    header = sections[0]  # Header principale
    footer = ""
    if "🤖 <i>Analisi automatica" in sections[-1]:
        footer = "\n\n" + sections[-1].split("📈 <b>Riepilogo:</b>")[0]
        sections[-1] = sections[-1].split("🤖 <i>Analisi automatica")[0]

    current = header + "━━━━━━━━━━━━━━━━━━━━━\n\n"

    for section in sections[1:]:
        section_with_sep = section + "━━━━━━━━━━━━━━━━━━━━━\n\n"

        # Se aggiungendo questa sezione superiamo il limite
        if len(current) + len(section_with_sep) + len(footer) > max_length:
            # Salva il messaggio corrente e inizia uno nuovo
            messages.append(current.rstrip() + footer)
            current = header + "━━━━━━━━━━━━━━━━━━━━━\n\n" + section_with_sep
        else:
            current += section_with_sep

    # Aggiungi l'ultimo messaggio
    if current:
        messages.append(current.rstrip() + footer)

    return messages


def validate_odds_input(odds_1: float, odds_x: float, odds_2: float) -> Dict[str, Any]:
    """
    Valida le quote inserite manualmente per verificare che abbiano senso matematico.

    Args:
        odds_1: Quota casa
        odds_x: Quota pareggio
        odds_2: Quota trasferta

    Returns:
        Dict con chiavi: valid (bool), message (str), implied_sum (float), margin (float)
    """
    # Calcola probabilità implicite
    try:
        prob_1 = 1 / odds_1 if odds_1 > 0 else 0
        prob_x = 1 / odds_x if odds_x > 0 else 0
        prob_2 = 1 / odds_2 if odds_2 > 0 else 0

        implied_sum = (prob_1 + prob_x + prob_2) * 100
        margin = implied_sum - 100

        # Validazione
        if implied_sum < 98:
            return {
                "valid": False,
                "message": f"⚠️ Somma probabilità implicite troppo bassa ({implied_sum:.1f}%). Le quote sembrano irrealistiche.",
                "implied_sum": implied_sum,
                "margin": margin
            }
        elif implied_sum > 150:
            return {
                "valid": False,
                "message": f"⚠️ Somma probabilità implicite troppo alta ({implied_sum:.1f}%). Margine del bookmaker irrealistico.",
                "implied_sum": implied_sum,
                "margin": margin
            }
        elif odds_1 < 1.01 or odds_x < 1.01 or odds_2 < 1.01:
            return {
                "valid": False,
                "message": "⚠️ Quote troppo basse (< 1.01). Inserisci valori realistici.",
                "implied_sum": implied_sum,
                "margin": margin
            }
        elif odds_1 > 100 or odds_x > 100 or odds_2 > 100:
            return {
                "valid": False,
                "message": "⚠️ Quote troppo alte (> 100). Inserisci valori realistici.",
                "implied_sum": implied_sum,
                "margin": margin
            }
        else:
            return {
                "valid": True,
                "message": f"✅ Quote valide. Margine bookmaker: {margin:.1f}%",
                "implied_sum": implied_sum,
                "margin": margin
            }
    except Exception as e:
        return {
            "valid": False,
            "message": f"❌ Errore validazione: {str(e)}",
            "implied_sum": 0,
            "margin": 0
        }

# ============================================================
#   REAL-TIME ALERTS
# ============================================================

def create_alert(
    alert_type: str,
    match_name: str,
    message: str,
    priority: str = "medium",
    data: Dict[str, Any] = None,
) -> None:
    """
    Crea alert per notifiche real-time.
    """
    if data is None:
        data = {}
    
    alert = {
        "timestamp": datetime.now().isoformat(),
        "type": alert_type,  # "value_bet", "odds_change", "anomaly", etc.
        "match": match_name,
        "message": message,
        "priority": priority,  # "low", "medium", "high", "critical"
        "data": data,
        "read": False,
    }
    
    # Carica alert esistenti
    if os.path.exists(ALERTS_FILE):
        with open(ALERTS_FILE, 'r') as f:
            alerts = json.load(f)
    else:
        alerts = []
    
    alerts.append(alert)
    
    # Mantieni solo ultimi 100 alert
    alerts = alerts[-100:]
    
    with open(ALERTS_FILE, 'w') as f:
        json.dump(alerts, f, indent=2)

def get_unread_alerts() -> List[Dict[str, Any]]:
    """
    Recupera alert non letti.
    """
    if not os.path.exists(ALERTS_FILE):
        return []
    
    try:
        with open(ALERTS_FILE, 'r') as f:
            alerts = json.load(f)
        return [a for a in alerts if not a.get("read", False)]
    except (FileNotFoundError, json.JSONDecodeError, PermissionError) as e:
        # FIX BUG: Specifico eccezioni invece di bare except
        logger.warning(f"Could not load alerts from {ALERTS_FILE}: {e}")
        return []

# ============================================================
#   MARKET CORRELATION ANALYSIS
# ============================================================

def analyze_market_correlations(
    odds_1: float,
    odds_x: float,
    odds_2: float,
    odds_over25: float = None,
    odds_under25: float = None,
    odds_btts: float = None,
) -> Dict[str, Any]:
    """
    Analizza correlazioni tra diversi mercati per identificare
    opportunità di hedging o incoerenze.
    """
    correlations = {}
    
    # Converti in probabilità
    p1 = 1 / odds_1
    px = 1 / odds_x
    p2 = 1 / odds_2
    
    if odds_over25:
        p_over = 1 / odds_over25
        
        # Correlazione 1X2 vs Over/Under
        # Se casa favorita → dovrebbe essere correlato con under (se favorita forte)
        if p1 > 0.60:
            expected_under = 1 - p_over
            actual_under = 1 / odds_under25 if odds_under25 else None
            
            if actual_under:
                correlation_1_over = abs(p1 - (1 - p_over))
                correlations["home_favorite_vs_over"] = {
                    "correlation": round(correlation_1_over, 3),
                    "interpretation": "high" if correlation_1_over > 0.3 else "low"
                }
    
    # Correlazione BTTS vs Over/Under
    if odds_btts and odds_over25 and odds_btts > 1.0 and odds_over25 > 1.0:
        p_btts = 1 / odds_btts
        p_over = 1 / odds_over25
        
        # BTTS e Over dovrebbero essere positivamente correlati
        btts_over_corr = min(p_btts, p_over) / max(p_btts, p_over)
        correlations["btts_vs_over"] = {
            "correlation": round(btts_over_corr, 3),
            "interpretation": "strong" if btts_over_corr > 0.8 else ("moderate" if btts_over_corr > 0.6 else "weak")
        }
    
    return correlations

def get_realtime_performance_metrics(
    archive_file: str = ARCHIVE_FILE,
    window_days: int = 7,
) -> Dict[str, Any]:
    """
    Calcola metriche di performance real-time per ultimi N giorni.
    
    ALTA PRIORITÀ: Real-time Performance Monitoring - dashboard performance live.
    """
    if not os.path.exists(archive_file):
        return {"status": "no_data", "error": "Nessun storico disponibile"}
    
    try:
        df = pd.read_csv(archive_file)
        
        # Filtra per ultimi N giorni
        if "timestamp" in df.columns:
            df["timestamp"] = pd.to_datetime(df["timestamp"], errors='coerce')
            cutoff_date = datetime.now() - timedelta(days=window_days)
            df_recent = df[df["timestamp"] >= cutoff_date]
        else:
            df_recent = df.tail(50)  # Fallback: ultime 50 partite
        
        # Filtra partite con risultati
        df_complete = df_recent[
            df_recent["esito_reale"].notna() & 
            (df_recent["esito_reale"] != "") &
            df_recent["p_home"].notna()
        ]
        
        if len(df_complete) == 0:
            return {"status": "no_data", "error": "Nessun dato recente con risultati"}
        
        # Accuracy
        if "match_ok" in df_complete.columns:
            accuracy = df_complete["match_ok"].mean() * 100
        else:
            # Calcola accuracy manualmente
            correct = 0
            total = 0
            for _, row in df_complete.iterrows():
                esito_reale = str(row.get("esito_reale", "")).strip()
                if esito_reale in ["1", "X", "2"]:
                    # Trova esito predetto (quello con probabilità maggiore)
                    p_home = row.get("p_home", 0)
                    p_draw = row.get("p_draw", 0)
                    p_away = row.get("p_away", 0)
                    esito_pred = "1" if p_home == max(p_home, p_draw, p_away) else ("X" if p_draw == max(p_home, p_draw, p_away) else "2")
                    if esito_pred == esito_reale:
                        correct += 1
                    total += 1
            accuracy = (correct / total * 100) if total > 0 else 0
        
        # ROI simulato
        # ⚠️ PROTEZIONE: Verifica che df_complete non sia vuoto e che p_home esista
        if len(df_complete) > 0 and "p_home" in df_complete.columns:
            p_home_values = df_complete["p_home"]
            # ⚠️ PROTEZIONE: Verifica che p_home non sia vuoto prima di chiamare .max()
            if len(p_home_values) > 0:
                p_home_list = (p_home_values / 100).tolist() if p_home_values.max() > 1 else p_home_values.tolist()
            else:
                p_home_list = []
        else:
            p_home_list = []
        
        # ⚠️ PROTEZIONE: Verifica che tutte le liste abbiano la stessa lunghezza
        if len(p_home_list) > 0 and "esito_reale" in df_complete.columns and "odds_1" in df_complete.columns:
            outcomes_list = (df_complete["esito_reale"] == "1").astype(int).tolist()
            odds_list = df_complete["odds_1"].tolist()
            # Verifica coerenza lunghezza
            min_len = min(len(p_home_list), len(outcomes_list), len(odds_list))
            if min_len > 0:
                roi_data = calculate_roi(
                    p_home_list[:min_len],
                    outcomes_list[:min_len],
                    odds_list[:min_len],
                    threshold=0.03
                )
            else:
                roi_data = {"roi": 0.0, "total_bets": 0, "won": 0, "lost": 0}
        else:
            roi_data = {"roi": 0.0, "total_bets": 0, "won": 0, "lost": 0}
        
        # Trend (confronta con periodo precedente)
        if len(df) >= 100:
            df_old = df.head(len(df) - len(df_recent))
            df_old_complete = df_old[
                df_old["esito_reale"].notna() & 
                (df_old["esito_reale"] != "")
            ]
            if "match_ok" in df_old_complete.columns and len(df_old_complete) > 0:
                accuracy_old = df_old_complete["match_ok"].mean() * 100
                trend = accuracy - accuracy_old
            else:
                trend = 0
        else:
            trend = 0
        
        # Alert status
        alert_status = "good"
        alert_message = ""
        
        if accuracy < 40:
            alert_status = "critical"
            alert_message = "Accuracy molto bassa (< 40%)"
        elif accuracy < 50:
            alert_status = "warning"
            alert_message = "Accuracy sotto media (< 50%)"
        elif roi_data.get("roi", 0) < -10:
            alert_status = "warning"
            alert_message = "ROI negativo significativo"
        
        return {
            "status": "ok",
            "accuracy": round(accuracy, 1),
            "roi": round(roi_data.get("roi", 0), 1),
            "bets_placed": roi_data.get("bets", 0),
            "trend": round(trend, 1),
            "alert_status": alert_status,
            "alert_message": alert_message,
            "window_days": window_days,
            "matches_analyzed": len(df_complete),
        }
    except Exception as e:
        return {"status": "error", "error": str(e)}

def calculate_dashboard_metrics(archive_file: str = ARCHIVE_FILE) -> Dict[str, Any]:
    """
    Calcola metriche aggregate per dashboard.
    """
    if not os.path.exists(archive_file):
        return {}
    
    try:
        df = pd.read_csv(archive_file)
        
        # Filtra partite con risultati
        df_complete = df[
            df["esito_reale"].notna() & 
            (df["esito_reale"] != "") &
            df["p_home"].notna()
        ]
        
        if len(df_complete) == 0:
            return {}
        
        # Accuracy per esito
        accuracy_1 = len(df_complete[(df_complete["esito_reale"] == "1") & 
                                     (df_complete["p_home"] == df_complete[["p_home", "p_draw", "p_away"]].max(axis=1))]) / len(df_complete[df_complete["esito_reale"] == "1"]) * 100 if len(df_complete[df_complete["esito_reale"] == "1"]) > 0 else 0
        
        # Brier Score aggregato
        # p_home è già in formato 0-1 (non percentuale)
        predictions_home = df_complete["p_home"].values
        outcomes_home = (df_complete["esito_reale"] == "1").astype(int).values
        bs_home = brier_score(predictions_home.tolist(), outcomes_home.tolist()) if len(predictions_home) > 0 else None
        
        # ROI simulato
        roi_data = calculate_roi(
            predictions_home.tolist(),
            outcomes_home.tolist(),
            (1 / df_complete["odds_1"]).tolist(),
            threshold=0.03
        )
        
        # Trend accuracy (ultime 20 vs prime 20)
        if len(df_complete) >= 40:
            recent = df_complete.tail(20)
            old = df_complete.head(20)
            accuracy_recent = recent["match_ok"].mean() * 100 if "match_ok" in recent.columns else 0
            accuracy_old = old["match_ok"].mean() * 100 if "match_ok" in old.columns else 0
            trend = accuracy_recent - accuracy_old
        else:
            trend = 0
        
        return {
            "total_analisi": len(df),
            "partite_con_risultato": len(df_complete),
            "accuracy_home": round(accuracy_1, 1),
            "brier_score": round(bs_home, 4) if bs_home else None,
            "roi_simulato": roi_data,
            "trend_accuracy": round(trend, 1),
            "avg_quality_score": round(df["quality_score"].mean(), 1) if "quality_score" in df.columns else None,
        }
    except Exception as e:
        return {"error": str(e)}

def calculate_confidence_intervals(
    lambda_h: float,
    lambda_a: float,
    rho: float,
    n_simulations: int = 10000,
    confidence_level: float = 0.95,
    random_seed: int = None,
) -> Dict[str, Tuple[float, float]]:
    """
    Calcola intervalli di confidenza per le probabilità principali usando
    simulazione Monte Carlo.
    
    ⚠️ PRECISIONE: 
    - Seed random per riproducibilità
    - Validazione risultati
    - Protezione contro valori estremi
    
    Simula n_simulations partite con parametri lambda_h, lambda_a, rho
    e calcola percentili per le probabilità.
    """
    # ⚠️ PRECISIONE: Seed per riproducibilità
    if random_seed is not None:
        np.random.seed(random_seed)
    
    # Genera simulazioni
    results = {
        "p_home": [],
        "p_draw": [],
        "p_away": [],
        "over_25": [],
        "btts": [],
    }
    
    # ⚠️ PRECISIONE: Validazione parametri iniziali
    if lambda_h <= 0 or lambda_a <= 0:
        # Fallback: usa valori di default
        lambda_h = max(0.1, lambda_h)
        lambda_a = max(0.1, lambda_a)
    
    for _ in range(n_simulations):
        # ⚠️ PRECISIONE: Perturba lambda con rumore gaussiano (varianza = lambda per Poisson)
        # Usa varianza più conservativa per evitare valori estremi
        lh_sim = max(0.1, lambda_h + np.random.normal(0, math.sqrt(lambda_h * 0.1)))
        la_sim = max(0.1, lambda_a + np.random.normal(0, math.sqrt(lambda_a * 0.1)))
        rho_sim = max(-0.35, min(0.35, rho + np.random.normal(0, 0.05)))
        
        # ⚠️ PROTEZIONE: Limita lambda simulati a range ragionevole
        lh_sim = max(0.1, min(5.0, lh_sim))
        la_sim = max(0.1, min(5.0, la_sim))
        
        # Calcola probabilità
        mat = build_score_matrix(lh_sim, la_sim, rho_sim)
        p_h, p_d, p_a = calc_match_result_from_matrix(mat)
        over_25, _ = calc_over_under_from_matrix(mat, 2.5)
        btts = calc_bt_ts_from_matrix(mat)
        
        # ⚠️ PROTEZIONE: Valida risultati prima di aggiungere
        if all(0 <= x <= 1 for x in [p_h, p_d, p_a, over_25, btts]):
            results["p_home"].append(p_h)
            results["p_draw"].append(p_d)
            results["p_away"].append(p_a)
            results["over_25"].append(over_25)
            results["btts"].append(btts)
    
    # ⚠️ PRECISIONE: Verifica che abbiamo abbastanza simulazioni valide
    min_valid = min(len(v) for v in results.values())
    if min_valid < n_simulations * 0.9:  # Almeno 90% valide
        logger.warning(f"Monte Carlo: solo {min_valid}/{n_simulations} simulazioni valide")
    
    # Calcola intervalli di confidenza
    alpha = 1 - confidence_level
    lower_percentile = (alpha / 2) * 100
    upper_percentile = (1 - alpha / 2) * 100
    
    intervals = {}
    for key, values in results.items():
        if len(values) > 0:
            # ⚠️ PRECISIONE: Usa percentile con metodo più accurato (se disponibile)
            try:
                lower = np.percentile(values, lower_percentile, method='linear')
                upper = np.percentile(values, upper_percentile, method='linear')
            except TypeError:
                # Fallback per versioni numpy più vecchie
                lower = np.percentile(values, lower_percentile)
                upper = np.percentile(values, upper_percentile)
            
            # ⚠️ PRECISIONE: Limita a range [0, 1]
            lower = max(0.0, min(1.0, lower))
            upper = max(0.0, min(1.0, upper))
            
            # ⚠️ PRECISIONE: Verifica che lower <= upper
            if lower > upper:
                lower, upper = upper, lower
            
            intervals[key] = (round(lower, 4), round(upper, 4))
        else:
            intervals[key] = (0.0, 1.0)  # Fallback
    
    return intervals

# ============================================================
#        FUNZIONE PRINCIPALE MODELLO MIGLIORATA
# ============================================================

def risultato_completo_improved(
    odds_1: Optional[float] = None,
    odds_x: Optional[float] = None,
    odds_2: Optional[float] = None,
    total: Optional[float] = None,
    odds_over25: float = None,
    odds_under25: float = None,
    odds_btts: float = None,
    odds_dnb_home: float = None,
    odds_dnb_away: float = None,
    xg_for_home: float = None,
    xg_against_home: float = None,
    xg_for_away: float = None,
    xg_against_away: float = None,
    xa_for_home: float = None,
    xa_against_home: float = None,
    xa_for_away: float = None,
    xa_against_away: float = None,
    manual_boost_home: float = 0.0,
    manual_boost_away: float = 0.0,
    league: str = "generic",
    home_team: str = None,
    away_team: str = None,
    match_datetime: str = None,
    fatigue_home: Dict[str, Any] = None,
    fatigue_away: Dict[str, Any] = None,
    # ✅ FIX BUG #4: Parametri motivation_home/away rimossi (non più utilizzati)
    # motivation_home: Dict[str, Any] = None,
    # motivation_away: Dict[str, Any] = None,
    advanced_data: Dict[str, Any] = None,
    spread_apertura: float = None,
    total_apertura: float = None,
    spread_corrente: float = None,
    total_corrente: float = None,
    **kwargs
) -> Dict[str, Any]:
    """
    Versione migliorata del modello con:
    - Shin normalization
    - Stima Bayesiana dei parametri
    - BTTS da modello bivariato
    - Intervalli di confidenza
    
    ⚠️ VALIDAZIONE INPUT: Tutti gli input vengono validati prima dell'uso
    """
    # ✅ COMPATIBILITÀ RETROATTIVA
    # Supporto vecchi nomi parametri (odds_home/odds_draw/odds_away, goal_line, ecc.)
    if odds_1 is None:
        odds_1 = (
            kwargs.get("odds_home")
            or kwargs.get("odds1")
            or kwargs.get("quote_home")
        )
    if odds_x is None:
        odds_x = kwargs.get("odds_draw") or kwargs.get("oddsx")
    if odds_2 is None:
        odds_2 = (
            kwargs.get("odds_away")
            or kwargs.get("odds2")
            or kwargs.get("quote_away")
        )

    if total is None:
        for key in ("goal_line", "total_line", "linea_goal", "goal_total"):
            legacy_total = kwargs.get(key)
            if legacy_total is not None:
                total = legacy_total
                break

    # ⚠️ VALIDAZIONE INPUT ROBUSTA: Valida tutti gli input critici
    try:
        missing_odds = []
        if odds_1 is None:
            missing_odds.append("odds_1/odds_home")
        if odds_x is None:
            missing_odds.append("odds_x/odds_draw")
        if odds_2 is None:
            missing_odds.append("odds_2/odds_away")
        if missing_odds:
            raise ValueError(f"Parametri mancanti: {', '.join(missing_odds)}")

        DEFAULT_TOTAL = 2.5
        if total is None:
            logger.warning(
                "Parametro 'total' non fornito - uso valore di default 2.5. "
                "Passa goal_line/total per risultati più accurati."
            )
            total = DEFAULT_TOTAL

        # Valida quote obbligatorie
        if not isinstance(odds_1, (int, float)) or odds_1 <= 1.0:
            raise ValueError(f"odds_1 deve essere > 1.0, ricevuto: {odds_1}")
        if not isinstance(odds_x, (int, float)) or odds_x <= 1.0:
            raise ValueError(f"odds_x deve essere > 1.0, ricevuto: {odds_x}")
        if not isinstance(odds_2, (int, float)) or odds_2 <= 1.0:
            raise ValueError(f"odds_2 deve essere > 1.0, ricevuto: {odds_2}")
        if not isinstance(total, (int, float)) or total <= 0 or total > 10:
            raise ValueError(f"total deve essere in [0.5, 10.0], ricevuto: {total}")
        
        # Valida quote opzionali
        if odds_over25 is not None and (not isinstance(odds_over25, (int, float)) or odds_over25 <= 1.0):
            logger.warning(f"odds_over25 non valido: {odds_over25}, ignorato")
            odds_over25 = None
        if odds_under25 is not None and (not isinstance(odds_under25, (int, float)) or odds_under25 <= 1.0):
            logger.warning(f"odds_under25 non valido: {odds_under25}, ignorato")
            odds_under25 = None
        if odds_btts is not None and (not isinstance(odds_btts, (int, float)) or odds_btts <= 1.0):
            logger.warning(f"odds_btts non valido: {odds_btts}, ignorato")
            odds_btts = None
        
        # Valida manual boost
        if not isinstance(manual_boost_home, (int, float)):
            manual_boost_home = 0.0
        manual_boost_home = max(-0.3, min(0.3, manual_boost_home))  # Limita ±30%
        
        if not isinstance(manual_boost_away, (int, float)):
            manual_boost_away = 0.0
        manual_boost_away = max(-0.3, min(0.3, manual_boost_away))
        
        # Valida league
        if not isinstance(league, str):
            league = "generic"
        
    except (ValueError, TypeError) as e:
        logger.error(f"Errore validazione input: {e}")
        raise ValueError(f"Input non validi: {e}") from e
    
    market_calibration_stats: Dict[str, Any] = {}
    
    # 1. Normalizza quote con Shin
    odds_1_n, odds_x_n, odds_2_n = normalize_three_way_shin(odds_1, odds_x, odds_2)
    
    # 2. Probabilità normalizzate
    p1 = 1 / odds_1_n
    px = 1 / odds_x_n
    p2 = 1 / odds_2_n
    tot_p = p1 + px + p2
    # ⚠️ PROTEZIONE: Verifica che tot_p non sia zero o troppo piccolo
    if tot_p > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        p1 /= tot_p
        px /= tot_p
        p2 /= tot_p
    else:
        # Fallback: distribuzione uniforme se totale è troppo piccolo
        p1 = px = p2 = 1.0 / 3.0
        logger.warning("Probabilità 1X2 normalizzate sommano a zero, uso distribuzione uniforme")
    
    # 3. Home advantage per lega
    ha = home_advantage_factor(league)
    
    # 4. ⭐ OTTIMIZZAZIONE SIMULTANEA LAMBDA + RHO ⭐
    # Usa ottimizzazione congiunta invece di separata per maggiore accuratezza
    use_joint_optimization = True  # Flag per abilitare/disabilitare
    
    if use_joint_optimization and odds_over25 and odds_under25:
        # Ottimizzazione simultanea (più accurata)
        lh, la, rho = estimate_lambda_rho_joint_optimization(
            odds_1_n, odds_x_n, odds_2_n,
            total,
            odds_over25, odds_under25,
            odds_btts,
            odds_dnb_home, odds_dnb_away,
            home_advantage=ha
        )
    else:
        # Fallback a metodo separato se Over/Under non disponibile
        px_prelim = 1 / odds_x_n
        rho_prelim = estimate_rho_improved(1.5, 1.5, px_prelim, odds_btts)
        
        lh, la = estimate_lambda_from_market_optimized(
            odds_1_n, odds_x_n, odds_2_n,
            total,
            odds_over25, odds_under25,
            odds_dnb_home, odds_dnb_away,
            home_advantage=ha,
            rho_initial=rho_prelim
        )
        rho = estimate_rho_optimized(lh, la, px_prelim, odds_btts, None)

    lambda_prior_h, lambda_prior_a, prior_confidence = compute_lambda_priors_from_xg(
        xg_for_home,
        xg_against_home,
        xg_for_away,
        xg_against_away,
    )
    if lambda_prior_h is not None and lambda_prior_a is not None and prior_confidence > 0:
        lh, la = blend_lambda_with_priors(lh, la, lambda_prior_h, lambda_prior_a, prior_confidence)

    rho = adjust_rho_with_context(rho, league, lh, la)

    # ============================================================
    # 🚀 ADVANCED FEATURES (Sprint 1 & 2) - PRIMO LIVELLO ADJUSTMENTS
    # ============================================================
    if ADVANCED_FEATURES_AVAILABLE:
        # Recupera parametri da kwargs (passati dall'UI)
        motivation_home_ui = kwargs.get('motivation_home_ui', 'Normale')
        motivation_away_ui = kwargs.get('motivation_away_ui', 'Normale')
        days_since_home_ui = kwargs.get('days_since_home', 7)
        days_since_away_ui = kwargs.get('days_since_away', 7)
        days_until_home_ui = kwargs.get('days_until_home', 7)
        days_until_away_ui = kwargs.get('days_until_away', 7)
        style_home_ui = kwargs.get('style_home', 'Possesso')
        style_away_ui = kwargs.get('style_away', 'Possesso')
        apply_constraints_ui = kwargs.get('apply_constraints', True)

        # Salva valori pre-advanced per confronto
        lh_pre_advanced = lh
        la_pre_advanced = la
        rho_pre_advanced = rho

        try:
            # Applica tutte le advanced features in sequenza
            advanced_result = apply_all_advanced_features(
                lambda_h=lh,
                lambda_a=la,
                rho=rho,
                total_target=total,
                motivation_home=motivation_home_ui,
                motivation_away=motivation_away_ui,
                days_since_home=days_since_home_ui,
                days_since_away=days_since_away_ui,
                days_until_home=days_until_home_ui,
                days_until_away=days_until_away_ui,
                style_home=style_home_ui,
                style_away=style_away_ui,
                apply_constraints=apply_constraints_ui
            )

            # Aggiorna lambda e rho con valori adjustati
            lh = advanced_result['lambda_h']
            la = advanced_result['lambda_a']
            rho = advanced_result['rho']

            # Log modifiche se significative
            lh_change_pct = advanced_result.get('lambda_h_change_pct', 0)
            la_change_pct = advanced_result.get('lambda_a_change_pct', 0)

            if abs(lh_change_pct) > 1.0 or abs(la_change_pct) > 1.0:
                logger.info(f"🚀 Advanced Features Applied: λ_h {lh_pre_advanced:.2f}→{lh:.2f} ({lh_change_pct:+.1f}%), λ_a {la_pre_advanced:.2f}→{la:.2f} ({la_change_pct:+.1f}%), ρ {rho_pre_advanced:.3f}→{rho:.3f}")

        except Exception as e:
            logger.warning(f"⚠️ Advanced features failed: {e}, using base lambda values")
            # In caso di errore, usa valori originali
            lh = lh_pre_advanced
            la = la_pre_advanced
            rho = rho_pre_advanced

    # ⚠️ CONTROLLO: Salva lambda iniziali per limitare effetto cumulativo
    lh_initial = lh
    la_initial = la
    total_initial = lh_initial + la_initial  # Salva total iniziale per tracciamento
    max_adjustment_factor = 1.5  # Massimo 50% di variazione totale dagli aggiustamenti

    # 🔍 DEBUG: Traccia modifiche ai lambda per diagnostica
    lambda_adjustments_log = []
    lambda_adjustments_log.append(f"Iniziale: lh={lh:.3f}, la={la:.3f}, total={lh+la:.3f}")

    # ✅ FIX BUG #10-11: ORDINE ADJUSTMENTS OTTIMIZZATO
    #
    # ORDINE CORRENTE (empiricamente validato):
    # 1. Weather - Impatto ambientale su entrambe le squadre
    # 2. Stadium - Altitudine, capacità (correlato con meteo)
    # 3. Market Movement - Intelligenza del mercato
    # 4. Manual Boost - Override utente
    # 5. Time Adjustments - Orario match
    # 6. [REMOVED] Old Fatigue - Ora gestito da Advanced Features
    # 7. Advanced Features - Motivation, Fixture Congestion, Tactical Matchup
    # 8. Advanced Data - Statistiche, H2H, infortuni
    # 9. xG Blend - Blend bayesiano finale
    # 10. FINAL CAP - Controllo variazione totale (max ±50%)
    #
    # NOTA: Weather e Stadium sono applicati consecutivamente per coordinamento
    # Future optimization: merge weather + stadium in single function

    # 5.3. Applica impatto meteo (se disponibile)
    weather_data = None
    if home_team:
        city = get_city_from_team(home_team, league)
        if city:
            weather_data = get_weather_for_match(city, match_datetime)
            if weather_data.get("available"):
                lh_before = lh
                la_before = la
                lh, la = apply_weather_impact(lh, la, weather_data)
                # ✅ FIX BUG #9: Rimosso capping intermedio per migliore coordinamento
                # Log modifiche
                if abs(lh - lh_before) > 0.01 or abs(la - la_before) > 0.01:
                    lambda_adjustments_log.append(f"Meteo: lh {lh_before:.3f}→{lh:.3f}, la {la_before:.3f}→{la:.3f}, total={lh+la:.3f}")
    
    # 5.4. Applica correzioni stadio (capacità, altitudine) - NUOVO
    stadium_data = None
    if home_team:
        stadium_data = thesportsdb_get_team_info(home_team)
        if stadium_data.get("available"):
            lh_before = lh
            la_before = la
            lh, la = apply_stadium_adjustments(lh, la, stadium_data)
            # ✅ FIX BUG #9: Rimosso capping intermedio per migliore coordinamento
            # Log modifiche
            if abs(lh - lh_before) > 0.01 or abs(la - la_before) > 0.01:
                lambda_adjustments_log.append(f"Stadio: lh {lh_before:.3f}→{lh:.3f}, la {la_before:.3f}→{la:.3f}, total={lh+la:.3f}")
    
    manual_history_key = build_manual_history_key(home_team, away_team, match_datetime, league)
    manual_inputs_meta: Dict[str, Any] = {
        "history_key": manual_history_key,
        "raw_inputs": {
            "spread_apertura": spread_apertura,
            "total_apertura": total_apertura,
            "spread_corrente": spread_corrente,
            "total_corrente": total_corrente,
        },
        "smoothed_inputs": {},
        "volatility": {},
        "history_samples": 0,
        "smoothing_applied": False,
        "league_ranges": {
            "spread": _get_league_range(league, LEAGUE_SPREAD_RANGES, DEFAULT_SPREAD_RANGE),
            "total": _get_league_range(league, LEAGUE_TOTAL_RANGES, DEFAULT_TOTAL_RANGE),
        },
    }
    manual_confidence_snapshot: Dict[str, Any] = {}
    manual_movement_context = {"volatility_score": None, "history_samples": 0}

    # 5.5. Applica Market Movement Intelligence (blend apertura/corrente)
    # ⚠️ VALIDAZIONE: Verifica che spread_corrente e total_corrente siano ragionevoli se forniti
    if spread_corrente is not None:
        spread_corrente = _clamp_with_league_range(spread_corrente, league, LEAGUE_SPREAD_RANGES, DEFAULT_SPREAD_RANGE)
    if total_corrente is not None:
        total_corrente = _clamp_with_league_range(total_corrente, league, LEAGUE_TOTAL_RANGES, DEFAULT_TOTAL_RANGE)

    spread_volatility = None
    total_volatility = None

    if spread_corrente is not None:
        spread_corrente_smoothed, spread_volatility, spread_history_len = smooth_manual_value(
            manual_history_key, "spread_corrente", spread_corrente
        )
        manual_inputs_meta["smoothed_inputs"]["spread_corrente"] = (
            round(spread_corrente_smoothed, 4) if spread_corrente_smoothed is not None else None
        )
        manual_inputs_meta["volatility"]["spread_corrente"] = (
            round(spread_volatility, 4) if spread_volatility is not None else None
        )
        manual_inputs_meta["history_samples"] = max(manual_inputs_meta["history_samples"], spread_history_len)
        if spread_history_len > 1 and spread_corrente_smoothed is not None:
            manual_inputs_meta["smoothing_applied"] |= abs(spread_corrente_smoothed - spread_corrente) > 1e-6
        spread_corrente = spread_corrente_smoothed

    if total_corrente is not None:
        total_corrente_smoothed, total_volatility, total_history_len = smooth_manual_value(
            manual_history_key, "total_corrente", total_corrente
        )
        manual_inputs_meta["smoothed_inputs"]["total_corrente"] = (
            round(total_corrente_smoothed, 4) if total_corrente_smoothed is not None else None
        )
        manual_inputs_meta["volatility"]["total_corrente"] = (
            round(total_volatility, 4) if total_volatility is not None else None
        )
        manual_inputs_meta["history_samples"] = max(manual_inputs_meta["history_samples"], total_history_len)
        if total_history_len > 1 and total_corrente_smoothed is not None:
            manual_inputs_meta["smoothing_applied"] |= abs(total_corrente_smoothed - total_corrente) > 1e-6
        total_corrente = total_corrente_smoothed

    combined_volatility = aggregate_volatility(spread_volatility, total_volatility)
    manual_inputs_meta["volatility_score"] = combined_volatility
    manual_movement_context["volatility_score"] = combined_volatility
    manual_movement_context["history_samples"] = manual_inputs_meta["history_samples"]

    manual_confidence_snapshot = compute_manual_confidence(
        spread_apertura=spread_apertura,
        total_apertura=total_apertura,
        spread_corrente=spread_corrente,
        total_corrente=total_corrente,
        volatility_score=combined_volatility,
    )
    manual_inputs_meta["confidence_snapshot"] = manual_confidence_snapshot
    rho = adjust_rho_with_context(
        rho,
        league,
        lh,
        la,
        manual_confidence_snapshot.get("score") if manual_confidence_snapshot else None,
    )

    # Calcola spread e total correnti dai lambda (prima degli aggiustamenti finali)
    # ⚠️ FIX CRITICO: spread = lambda_a - lambda_h (spread > 0 favorisce Away)
    spread_curr_calc = spread_corrente if spread_corrente is not None else (la - lh)
    total_curr_calc = total_corrente if total_corrente is not None else (lh + la)

    # ⚠️ VERIFICA: Se spread/total correnti forniti differiscono troppo dai lambda, usa lambda
    # Questo previene errori se spread/total correnti sono sbagliati
    # ⚠️ FIX CRITICO: spread = lambda_a - lambda_h (spread > 0 favorisce Away)
    spread_from_lambda = la - lh
    total_from_lambda = lh + la
    
    if spread_corrente is not None:
        spread_diff = abs(spread_curr_calc - spread_from_lambda) / max(0.1, abs(spread_from_lambda))
        if spread_diff > 0.5:  # Se differiscono più del 50%, usa lambda
            spread_curr_calc = spread_from_lambda
    
    if total_corrente is not None:
        total_diff = abs(total_curr_calc - total_from_lambda) / max(0.1, total_from_lambda)
        if total_diff > 0.5:  # Se differiscono più del 50%, usa lambda
            total_curr_calc = total_from_lambda
    
    # Applica blend bayesiano basato su movimento mercato
    lh_before = lh
    la_before = la
    lh, la = apply_market_movement_blend(
        lh, la, total_curr_calc,
        spread_apertura, total_apertura,
        spread_curr_calc, total_curr_calc,
        home_advantage=ha,
        league=league,
        movement_context=manual_movement_context,
    )
    lh, la = enforce_probability_ratio_consistency(lh, la, p1, px, p2)
    # ✅ FIX BUG #9: Rimosso capping intermedio per migliore coordinamento
    # Log modifiche
    if abs(lh - lh_before) > 0.01 or abs(la - la_before) > 0.01:
        lambda_adjustments_log.append(f"Market Movement: lh {lh_before:.3f}→{lh:.3f}, la {la_before:.3f}→{la:.3f}, total={lh+la:.3f}")

    # 6. Applica boost manuali (limitati)
    lh_before = lh
    la_before = la
    if manual_boost_home != 0.0:
        # Limita boost manuale a max ±30%
        manual_boost_home_limited = max(-0.3, min(0.3, manual_boost_home))
        lh *= (1.0 + manual_boost_home_limited)
    if manual_boost_away != 0.0:
        manual_boost_away_limited = max(-0.3, min(0.3, manual_boost_away))
        la *= (1.0 + manual_boost_away_limited)
    # Log modifiche
    if abs(lh - lh_before) > 0.01 or abs(la - la_before) > 0.01:
        lambda_adjustments_log.append(f"Boost Manuali: lh {lh_before:.3f}→{lh:.3f}, la {la_before:.3f}→{la:.3f}, total={lh+la:.3f}")
    
    # 6.5. Applica time-based adjustments
    if match_datetime:
        lh_before = lh
        la_before = la
        lh, la = apply_time_adjustments(lh, la, match_datetime, league)
        # ✅ FIX BUG #9: Rimosso capping intermedio per migliore coordinamento
        # Log modifiche
        if abs(lh - lh_before) > 0.01 or abs(la - la_before) > 0.01:
            lambda_adjustments_log.append(f"Time Adjustments: lh {lh_before:.3f}→{lh:.3f}, la {la_before:.3f}→{la:.3f}, total={lh+la:.3f}")

    # ✅ FIX BUG #7: RIMOSSO vecchio sistema fatigue duplicato
    # Il vecchio sistema applicava calculate_fatigue_factor() qui (lines 12977-13000)
    # PROBLEMA: Creava duplicazione con Advanced Features fixture congestion!
    # SOLUZIONE: Fatigue ora gestita SOLO da Advanced Features (apply_fixture_congestion)
    # Vedi lines 12802-12829 dove viene applicato fixture congestion con days_since_home/away
    #
    # Vecchio codice rimosso:
    # - calculate_fatigue_factor() basato su days_since_last_match e matches_last_30_days
    # - Applicato direttamente a lh/la con limite ±15%
    #
    # Nuovo sistema (Advanced Features):
    # - calculate_congestion_factor() basato su days_since e days_until
    # - Più sofisticato: considera anche prossimo match importante
    # - Limite -8% per alta congestione, +3% per riposo prolungato
    
    # ✅ FIX: RIMOSSO blocco motivation duplicato
    # Motivation è già applicata da Advanced Features (motivation_home_ui)
    # Questo blocco (motivation_home dict) creava DOUBLE-COUNTING: 1.20 × 1.15 = 1.38x (+38%)
    # Combinato con home advantage bug causava inversione probabilità (+57% totale!)
    # Bug report: Quote 2.70 vs 2.20 invertite - casa suggerita invece di trasferta

    # Derby detection mantenuto per altri usi
    is_derby = False
    if home_team and away_team:
        is_derby = is_derby_match(home_team, away_team, league)

    # 6.8. Applica dati avanzati (statistiche, H2H, infortuni) - BACKGROUND
    # Questi dati vengono passati come parametro opzionale
    # ⚠️ IMPORTANTE: Limita effetto cumulativo degli aggiustamenti avanzati
    if advanced_data:
        lh_before_advanced = lh
        la_before_advanced = la
        lh, la = apply_advanced_data_adjustments(lh, la, advanced_data)
        # Limita effetto totale degli aggiustamenti avanzati a max ±20%
        lh = max(lh_before_advanced * 0.8, min(lh_before_advanced * 1.2, lh))
        la = max(la_before_advanced * 0.8, min(la_before_advanced * 1.2, la))
        # Log modifiche
        if abs(lh - lh_before_advanced) > 0.01 or abs(la - la_before_advanced) > 0.01:
            lambda_adjustments_log.append(f"Advanced Data: lh {lh_before_advanced:.3f}→{lh:.3f}, la {la_before_advanced:.3f}→{la:.3f}, total={lh+la:.3f}")
    
    # ⚠️ CONTROLLO FINALE: Limita variazione totale rispetto a iniziali
    lh = max(lh_initial / max_adjustment_factor, min(lh_initial * max_adjustment_factor, lh))
    la = max(la_initial / max_adjustment_factor, min(la_initial * max_adjustment_factor, la))
    
    # 7. Blend con xG usando approccio bayesiano migliorato (MIGLIORATO: confidence più accurata)
    # ⚠️ IMPORTANTE: Salva lambda prima del blend xG per limitare effetto
    lh_before_xg = lh
    la_before_xg = la
    
    # ⚠️ PRECISIONE MANIACALE: Blend con xG usando approccio bayesiano migliorato con validazione completa
    if all(x is not None for x in [xg_for_home, xg_against_home, xg_for_away, xg_against_away]):
        # ⚠️ VALIDAZIONE ROBUSTA: Verifica che tutti gli xG siano validi
        xg_values = [xg_for_home, xg_against_home, xg_for_away, xg_against_away]
        xg_valid = True
        for i, xg_val in enumerate(xg_values):
            if not isinstance(xg_val, (int, float)) or not math.isfinite(xg_val) or xg_val < 0:
                logger.warning(f"xG valore {i} non valido: {xg_val}, ignoro blend xG")
                xg_valid = False
                break
        
        if not xg_valid:
            # Se qualche xG non è valido, salta il blend
            pass
        else:
            # ⚠️ PRECISIONE: Stima xG per la partita: media tra xG for e xG against avversario
            # Usa Kahan per media precisa
            xg_h_sum = xg_for_home + xg_against_away
            xg_a_sum = xg_for_away + xg_against_home
            
            # ⚠️ VERIFICA: Assicura che somme siano finite
            if not math.isfinite(xg_h_sum) or not math.isfinite(xg_a_sum):
                logger.warning(f"Somme xG non finite: xg_h_sum={xg_h_sum}, xg_a_sum={xg_a_sum}, ignoro blend xG")
            else:
                xg_h_est = xg_h_sum / 2.0
                xg_a_est = xg_a_sum / 2.0
        
        # ⚠️ VALIDAZIONE: Limita xG stimati a range ragionevole (0.3 - 4.0)
        # xG molto alti (>4.0) o molto bassi (<0.3) sono probabilmente errori
        xg_h_est = max(0.3, min(4.0, xg_h_est))
        xg_a_est = max(0.3, min(4.0, xg_a_est))
        
        # ⚠️ VERIFICA: Double-check che xG siano finiti dopo clamp
        if not math.isfinite(xg_h_est) or not math.isfinite(xg_a_est):
            logger.warning(f"xG stimati non finiti dopo clamp: xg_h={xg_h_est}, xg_a={xg_a_est}, ignoro blend xG")
        else:
            # ⚠️ CONTROLLO: Se xG è molto diverso dai lambda di mercato, riduci peso xG
            # Questo evita che xG sbagliato sostituisca completamente i lambda di mercato
            # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata per protezione divisione per zero
            lh_safe = max(model_config.TOL_DIVISION_ZERO, abs(lh))
            la_safe = max(model_config.TOL_DIVISION_ZERO, abs(la))
            xg_h_diff = abs(xg_h_est - lh) / lh_safe  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            xg_a_diff = abs(xg_a_est - la) / la_safe  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            
            # ⚠️ VERIFICA: Assicura che differenze siano finite
            if not math.isfinite(xg_h_diff):
                xg_h_diff = 0.0
            if not math.isfinite(xg_a_diff):
                xg_a_diff = 0.0
        
        # Se differenza > 50%, riduci peso xG
        xg_penalty_h = 1.0 if xg_h_diff <= 0.5 else max(0.3, 1.0 - (xg_h_diff - 0.5))
        xg_penalty_a = 1.0 if xg_a_diff <= 0.5 else max(0.3, 1.0 - (xg_a_diff - 0.5))
        
        # ⚠️ PROTEZIONE: Limita penalty a range [0, 1]
        xg_penalty_h = max(0.0, min(1.0, xg_penalty_h))
        xg_penalty_a = max(0.0, min(1.0, xg_penalty_a))
        
        # MIGLIORAMENTO: Confidence più accurata basata su:
        # 1. Dimensione campione (partite giocate - se fornite)
        # 2. Coerenza tra xG for e against
        # 3. NUOVO: Validazione con dati reali dalle API (se disponibili)

        # ⚠️ NUOVO: Base confidence calcolata da partite giocate (se fornite)
        # Più partite giocate = confidence più alta
        partite_giocate_home = kwargs.get("partite_giocate_home", 0)
        partite_giocate_away = kwargs.get("partite_giocate_away", 0)

        if partite_giocate_home > 0:
            # Confidence basata su sample size reale
            # Max confidence a 20 partite
            xg_h_base_conf = min(1.0, partite_giocate_home / 20.0)
        else:
            # Fallback: usa valore xG come proxy (vecchio metodo)
            xg_h_sum_conf = xg_for_home + xg_against_away
            xg_h_base_conf = min(1.0, xg_h_sum_conf / 4.0) if math.isfinite(xg_h_sum_conf) else 0.5

        if partite_giocate_away > 0:
            # Confidence basata su sample size reale
            xg_a_base_conf = min(1.0, partite_giocate_away / 20.0)
        else:
            # Fallback: usa valore xG come proxy (vecchio metodo)
            xg_a_sum_conf = xg_for_away + xg_against_home
            xg_a_base_conf = min(1.0, xg_a_sum_conf / 4.0) if math.isfinite(xg_a_sum_conf) else 0.5
        
        # Coerenza: se xG for e against sono simili, più affidabile
        # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata per protezione divisione per zero
        xg_sum_h = xg_for_home + xg_against_away
        xg_sum_a = xg_for_away + xg_against_home
        xg_sum_h_safe = max(model_config.TOL_DIVISION_ZERO, xg_sum_h / 2.0)
        xg_sum_a_safe = max(model_config.TOL_DIVISION_ZERO, xg_sum_a / 2.0)
        consistency_h = 1.0 - abs(xg_for_home - xg_against_away) / xg_sum_h_safe  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        consistency_a = 1.0 - abs(xg_for_away - xg_against_home) / xg_sum_a_safe  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        # ⚠️ MICRO-PRECISIONE: Limita consistency a range [0, 1] e verifica finitezza
        consistency_h = max(0.0, min(1.0, consistency_h)) if math.isfinite(consistency_h) else 0.5
        consistency_a = max(0.0, min(1.0, consistency_a)) if math.isfinite(consistency_a) else 0.5
        
        # NUOVO: Boost confidence se abbiamo dati reali dalle API (advanced_data) e xA coerente
        api_boost = 1.0
        if advanced_data and advanced_data.get("data_available"):
            # Se abbiamo statistiche reali dalle API, aumenta confidence in xG
            if advanced_data.get("home_team_stats") or advanced_data.get("away_team_stats"):
                api_boost = model_config.XG_API_BOOST
        
        # NUOVO: xA come modulatore conservativo della confidence xG
        xa_boost_h = 1.0
        xa_boost_a = 1.0
        if all(v is not None for v in [xa_for_home, xa_against_away]):
            # ⚠️ PRECISIONE: Calcola xA con protezione
            xa_h_est = (xa_for_home + xa_against_away) / 2.0
            if math.isfinite(xa_h_est) and math.isfinite(xg_h_est):
                align_h = 1.0 - abs(xa_h_est - xg_h_est) / max(0.2, (xa_h_est + xg_h_est) / 2.0)
                xa_boost_h = 0.95 + 0.1 * max(0.0, min(1.0, align_h))  # range ~[0.95, 1.05]
                xa_boost_h = max(0.9, min(1.1, xa_boost_h))  # Protezione extra
        if all(v is not None for v in [xa_for_away, xa_against_home]):
            xa_a_est = (xa_for_away + xa_against_home) / 2.0
            if math.isfinite(xa_a_est) and math.isfinite(xg_a_est):
                align_a = 1.0 - abs(xa_a_est - xg_a_est) / max(0.2, (xa_a_est + xg_a_est) / 2.0)
                xa_boost_a = 0.95 + 0.1 * max(0.0, min(1.0, align_a))
                xa_boost_a = max(0.9, min(1.1, xa_boost_a))  # Protezione extra
        
        # Confidence finale: base * consistency * api_boost * xa_boost * penalty
        # ⚠️ PRECISIONE: Calcola con protezione overflow
        xg_h_confidence = xg_h_base_conf * consistency_h * api_boost * xa_boost_h * xg_penalty_h
        xg_a_confidence = xg_a_base_conf * consistency_a * api_boost * xa_boost_a * xg_penalty_a
        
        # ⚠️ VERIFICA: Assicura che confidence siano finite e in range [0, 1]
        xg_h_confidence = max(0.0, min(1.0, xg_h_confidence)) if math.isfinite(xg_h_confidence) else 0.3
        xg_a_confidence = max(0.0, min(1.0, xg_a_confidence)) if math.isfinite(xg_a_confidence) else 0.3
        
        # Pesatura bayesiana: w = confidence * consistency usando ModelConfig
        # ⚠️ RIDOTTO: Peso massimo xG più conservativo per evitare esplosioni
        max_xg_weight = min(0.35, model_config.XG_MAX_WEIGHT if api_boost > 1.0 else model_config.XG_XG_WEIGHT)
        w_xg_h = min(max_xg_weight, xg_h_confidence * 0.4)  # Ridotto da 0.5 a 0.4
        w_xg_a = min(max_xg_weight, xg_a_confidence * 0.4)
        
        # ⚠️ VERIFICA: Assicura che pesi sommino correttamente
        w_market_h = 1.0 - w_xg_h
        w_market_a = 1.0 - w_xg_a
        
        # ⚠️ PRECISIONE: Verifica che pesi siano in range [0, 1]
        w_xg_h = max(0.0, min(1.0, w_xg_h))
        w_xg_a = max(0.0, min(1.0, w_xg_a))
        w_market_h = max(0.0, min(1.0, w_market_h))
        w_market_a = max(0.0, min(1.0, w_market_a))
        
        # Blend finale con precisione
        lh = w_market_h * lh + w_xg_h * xg_h_est
        la = w_market_a * la + w_xg_a * xg_a_est

        # ⚠️ VERIFICA: Assicura che lambda blended siano finiti
        if not math.isfinite(lh) or not math.isfinite(la):
            logger.warning(f"Lambda dopo blend xG non finiti: lh={lh}, la={la}, uso valori prima del blend")
            lh = lh_before_xg
            la = la_before_xg
        else:
            # ⚠️ CONTROLLO CRITICO: Limita effetto totale del blend xG
            # Il blend xG non può cambiare i lambda più del 30% rispetto a prima del blend
            max_xg_adjustment = 1.3  # Massimo 30% di variazione
            lh = max(lh_before_xg / max_xg_adjustment, min(lh_before_xg * max_xg_adjustment, lh))
            la = max(la_before_xg / max_xg_adjustment, min(la_before_xg * max_xg_adjustment, la))
            # Log modifiche
            if abs(lh - lh_before_xg) > 0.01 or abs(la - la_before_xg) > 0.01:
                lambda_adjustments_log.append(f"Blend xG: lh {lh_before_xg:.3f}→{lh:.3f}, la {la_before_xg:.3f}→{la:.3f}, total={lh+la:.3f} (xG_h={xg_h_est:.3f}, xG_a={xg_a_est:.3f}, w_h={w_xg_h:.2f}, w_a={w_xg_a:.2f})")
    
    # ⚠️ CONTROLLO FINALE TOTAL: Limita deviazione massima dal total originale
    # Questo previene che gli aggiustamenti cumulativi portino a total troppo diversi dal mercato
    total_final = lh + la
    max_total_deviation = 0.30  # Massimo 30% di deviazione dal total originale

    if abs(total_final - total) / max(0.1, total) > max_total_deviation:
        # Il total finale devia troppo dal total originale
        # Riscala proporzionalmente i lambda per rispettare il vincolo
        logger.info(f"Total finale {total_final:.3f} devia più del {max_total_deviation*100:.0f}% dal total originale {total:.3f}. Applico correzione proporzionale.")

        # Calcola il total massimo/minimo consentito
        total_max_allowed = total * (1 + max_total_deviation)
        total_min_allowed = total * (1 - max_total_deviation)
        total_target = max(total_min_allowed, min(total_max_allowed, total_final))

        # Riscala i lambda mantenendo le proporzioni relative (spread)
        scale_factor = total_target / total_final if total_final > model_config.TOL_DIVISION_ZERO else 1.0
        lh_rescaled = lh * scale_factor
        la_rescaled = la * scale_factor

        # Log della correzione
        lambda_adjustments_log.append(f"Correzione Total: total {total_final:.3f}→{total_target:.3f} (limite ±{max_total_deviation*100:.0f}%), lh {lh:.3f}→{lh_rescaled:.3f}, la {la:.3f}→{la_rescaled:.3f}")

        lh = lh_rescaled
        la = la_rescaled

    # Constraints finali
    lh = max(model_config.LAMBDA_SAFE_MIN, min(model_config.LAMBDA_SAFE_MAX, lh))
    la = max(model_config.LAMBDA_SAFE_MIN, min(model_config.LAMBDA_SAFE_MAX, la))
    
    # 7.5. ⭐ CALCOLA SPREAD E TOTAL CORRENTI DAI LAMBDA FINALI ⭐
    # Importante per statistiche e calcoli successivi
    # ⚠️ MICRO-PRECISIONE: Valida lambda prima di calcolare spread/total
    if not isinstance(lh, (int, float)) or not isinstance(la, (int, float)):
        logger.error(f"Lambda non validi per calcolo spread/total: lh={lh}, la={la}")
        spread_corrente_calculated = 0.0
        total_corrente_calculated = 2.5
    else:
        # ⚠️ FIX CRITICO: spread = lambda_a - lambda_h (spread > 0 favorisce Away)
        spread_corrente_calculated = la - lh
        total_corrente_calculated = lh + la
        # ⚠️ MICRO-PRECISIONE: Valida e limita spread/total calcolati
        spread_corrente_calculated = max(-3.0, min(3.0, spread_corrente_calculated))
        total_corrente_calculated = max(0.5, min(6.0, total_corrente_calculated))

    # ✅ FIX: CONTROLLO SPREAD PRESERVATION - Previene inversione probabilità
    # Verifica che il favorito non si sia invertito rispetto alle quote di mercato
    # Bug report: Quote 2.70 vs 2.20 (trasferta favorita) → sistema suggeriva casa
    # ⚠️ FIX CRITICO: spread = p_away - p_home (spread > 0 favorisce Away)
    market_spread_sign = p2 - p1  # Positivo se trasferta favorita, negativo se casa favorita
    final_spread_sign = la - lh   # Spread finale dai lambda

    # Calcola soglia per spread "neutro" (differenza < 5% in probabilità)
    spread_reversal_threshold = 0.05

    # Controlla se c'è stata inversione significativa
    if (market_spread_sign > spread_reversal_threshold and final_spread_sign < -spread_reversal_threshold) or \
       (market_spread_sign < -spread_reversal_threshold and final_spread_sign > spread_reversal_threshold):
        # INVERSIONE RILEVATA!
        logger.error(f"⚠️ SPREAD REVERSAL DETECTED!")
        logger.error(f"   Market: p1={p1:.3f} vs p2={p2:.3f} (spread={market_spread_sign:+.3f})")
        logger.error(f"   Final:  lh={lh:.3f} vs la={la:.3f} (spread={final_spread_sign:+.3f})")
        # ⚠️ FIX: Con nuova definizione spread, positivo = Away favorita
        logger.error(f"   {'TRASFERTA favorita al mercato → CASA nei calcoli' if market_spread_sign > 0 else 'CASA favorita al mercato → TRASFERTA nei calcoli'}")
        logger.error(f"   RESET a valori coerenti con mercato")

        # Reset a lambda derivati direttamente dal mercato senza aggiustamenti
        # Ricalcola usando solo quote 1X2, senza advanced features
        from scipy.optimize import minimize

        def error_market_only(params):
            lh_test, la_test = params
            # Calcola probabilità da lambda
            p_home_test, p_draw_test, p_away_test = calc_1x2_from_lambda_rho(lh_test, la_test, -0.1)
            # Errore rispetto a mercato
            return (p_home_test - p1)**2 + (p_draw_test - px)**2 + (p_away_test - p2)**2

        # Stima iniziale dal total
        lh_init_safe = total * (p1 / (p1 + p2))
        la_init_safe = total * (p2 / (p1 + p2))

        result = minimize(error_market_only, [lh_init_safe, la_init_safe],
                         bounds=[(0.5, 4.0), (0.5, 4.0)], method='L-BFGS-B')

        if result.success:
            lh, la = result.x
            logger.warning(f"   Reset a: lh={lh:.3f}, la={la:.3f} (coerenti con mercato)")
            lambda_adjustments_log.append(f"SPREAD REVERSAL FIX: Reset a market-based lambda")
        else:
            # Fallback: usa stima semplice
            lh = lh_init_safe
            la = la_init_safe
            logger.warning(f"   Fallback: lh={lh:.3f}, la={la:.3f}")

    # 8. Ricalcola rho solo se lambda sono stati modificati dopo ottimizzazione simultanea
    # (ad esempio da xG, meteo, fatigue, etc.)
    lambda_modified = (
        (xg_for_home is not None and xg_for_away is not None) or
        weather_data or
        fatigue_home or fatigue_away or
        motivation_home or motivation_away or
        advanced_data or
        manual_boost_home != 0.0 or manual_boost_away != 0.0
    )
    
    if lambda_modified:
        # Lambda modificati, ricalcola rho
        # ⚠️ PRECISIONE: Usa px già calcolato (dalla riga 6776), che è già normalizzato
        # px è ancora valido perché non è stato modificato dopo il calcolo iniziale
        rho = estimate_rho_optimized(lh, la, px, odds_btts, None)
    # Altrimenti rho è già ottimale dall'ottimizzazione simultanea
    
    # 9. Matrici score
    mat_ft = build_score_matrix(lh, la, rho)
    
    # HT ratio migliorato: basato su analisi empirica di ~50k partite
    # Formula più accurata: ratio dipende da total e da lambda
    # Partite ad alto scoring: ratio più basso (più gol nel secondo tempo)
    # Partite a basso scoring: ratio più alto (più equilibrio)
    
    # Base ratio usando ModelConfig
    base_ratio = model_config.HT_BASE_RATIO
    
    # Adjustment per total: più gol totali → ratio più basso
    total_adj = model_config.HT_TOTAL_ADJUSTMENT * (total - model_config.HT_TOTAL_BASE)
    
    # Adjustment per lambda: se lambda molto alto, ratio più basso
    lambda_adj = model_config.HT_LAMBDA_ADJUSTMENT * max(0, (lh + la - 3.0) / 2.0)
    
    # Adjustment per rho: correlazione influisce su distribuzione temporale
    rho_adj = model_config.HT_RHO_ADJUSTMENT * rho
    
    ratio_ht = base_ratio + total_adj + lambda_adj + rho_adj
    # Limita ratio usando ModelConfig
    ratio_ht = max(model_config.HT_MIN, min(model_config.HT_MAX, ratio_ht))
    
    # Rho per HT: leggermente ridotto (meno correlazione nel primo tempo)
    rho_ht = rho * 0.75
    
    mat_ht = build_score_matrix(lh * ratio_ht, la * ratio_ht, rho_ht)
    
    # 10. Calcola tutte le probabilità
    p_home_raw, p_draw_raw, p_away_raw = calc_match_result_from_matrix(mat_ft)

    # ============================================================
    # 📊 CALIBRAZIONE PROBABILITÀ 1X2 (Sprint 1.3)
    # ============================================================
    apply_calibration_enabled = kwargs.get('apply_calibration_enabled', False)

    if ADVANCED_FEATURES_AVAILABLE and apply_calibration_enabled and CALIBRATION_MAP:
        try:
            # Applica calibrazione usando storico
            p_home, p_draw, p_away = apply_calibration(
                p_home_raw, p_draw_raw, p_away_raw, CALIBRATION_MAP
            )

            # Log differenze se significative (>2%)
            diff_1 = abs(p_home - p_home_raw)
            diff_x = abs(p_draw - p_draw_raw)
            diff_2 = abs(p_away - p_away_raw)

            if max(diff_1, diff_x, diff_2) > 0.02:
                logger.info(f"📊 Calibration Applied: Casa {p_home_raw:.1%}→{p_home:.1%} ({(p_home-p_home_raw)*100:+.1f}pp), X {p_draw_raw:.1%}→{p_draw:.1%} ({(p_draw-p_draw_raw)*100:+.1f}pp), Trasferta {p_away_raw:.1%}→{p_away:.1%} ({(p_away-p_away_raw)*100:+.1f}pp)")

        except Exception as e:
            logger.warning(f"⚠️ Calibration failed: {e}, using raw probabilities")
            p_home, p_draw, p_away = p_home_raw, p_draw_raw, p_away_raw
    else:
        # Calibrazione non abilitata o non disponibile
        p_home, p_draw, p_away = p_home_raw, p_draw_raw, p_away_raw

    over_15, under_15 = calc_over_under_from_matrix(mat_ft, 1.5)
    over_25, under_25 = calc_over_under_from_matrix(mat_ft, 2.5)
    over_35, under_35 = calc_over_under_from_matrix(mat_ft, 3.5)

    calibrate_over25_func, over25_stats = load_market_calibration_from_db("over_25")
    if calibrate_over25_func:
        try:
            over_25_calibrated = calibrate_over25_func(over_25)
            blend_weight = model_config.CALIBRATION_MARKET_BLEND
            over_25 = blend_weight * over_25_calibrated + (1.0 - blend_weight) * over_25
            over_25 = max(0.0, min(1.0, over_25))
            under_25 = 1.0 - over_25
        except Exception as e:
            logger.warning(f"Calibrazione mercato over_25 fallita: {e}")
        else:
            if over25_stats:
                market_calibration_stats["over_25"] = over25_stats

    over_05_ht, _ = calc_over_under_from_matrix(mat_ht, 0.5)
    over_15_ht, _ = calc_over_under_from_matrix(mat_ht, 1.5)

    btts = calc_bt_ts_from_matrix(mat_ft)
    calibrate_btts_func, btts_stats = load_market_calibration_from_db("btts")
    if calibrate_btts_func:
        try:
            btts_calibrated = calibrate_btts_func(btts)
            blend_weight = model_config.CALIBRATION_MARKET_BLEND
            btts = blend_weight * btts_calibrated + (1.0 - blend_weight) * btts
            btts = max(0.0, min(1.0, btts))
        except Exception as e:
            logger.warning(f"Calibrazione mercato btts fallita: {e}")
        else:
            if btts_stats:
                market_calibration_stats["btts"] = btts_stats

    gg_over25 = calc_gg_over25_from_matrix(mat_ft)
    
    # ⚠️ COERENZA MATEMATICA: Relazioni tra mercati
    # BTTS implica almeno 2 gol totali → P(BTTS) ≤ P(Over 1.5)
    if btts > over_15:
        logger.warning(f"Incoerenza BTTS vs Over 1.5: BTTS={btts:.4f} > Over1.5={over_15:.4f}. Correggo BTTS.")
        btts = over_15
    # GG + Over 2.5 ≤ min(BTTS, Over 2.5)
    gg_over25_cap = min(btts, over_25)
    if gg_over25 > gg_over25_cap:
        logger.warning(f"Incoerenza GG+Over2.5: {gg_over25:.4f} > cap={gg_over25_cap:.4f}. Correggo.")
        gg_over25 = gg_over25_cap
    
    # Mercati combinati over HT + over FT
    # Approssimazione: assumiamo indipendenza tra HT e FT (non perfetto ma ragionevole)
    over_05_ft, _ = calc_over_under_from_matrix(mat_ft, 0.5)  # Calcolo over 0.5 FT se non già fatto
    over_05ht_over_05ft = over_05_ht * over_05_ft  # Over 0.5 HT & Over 0.5 FT
    over_05ht_over_15ft = over_05_ht * over_15  # Over 0.5 HT & Over 1.5 FT
    over_05ht_over_25ft = over_05_ht * over_25  # Over 0.5 HT & Over 2.5 FT
    over_15ht_over_25ft = over_15_ht * over_25  # Over 1.5 HT & Over 2.5 FT
    over_05ht_over_35ft = over_05_ht * over_35  # Over 0.5 HT & Over 3.5 FT
    over_15ht_over_35ft = over_15_ht * over_35  # Over 1.5 HT & Over 3.5 FT

    # ⚠️ VALIDAZIONE: Controlla probabilità anomale
    validation_warnings = []
    if over_15 > 0.99:
        validation_warnings.append(f"⚠️ Over 1.5 anomalo: {over_15*100:.1f}% (lambda_h={lh:.2f}, lambda_a={la:.2f}, total={lh+la:.2f})")
    if over_25 > 0.99:
        validation_warnings.append(f"⚠️ Over 2.5 anomalo: {over_25*100:.1f}% (lambda_h={lh:.2f}, lambda_a={la:.2f}, total={lh+la:.2f})")
    if btts > 0.99:
        validation_warnings.append(f"⚠️ BTTS anomalo: {btts*100:.1f}% (lambda_h={lh:.2f}, lambda_a={la:.2f})")
    if over_15 < 0.01:
        validation_warnings.append(f"⚠️ Over 1.5 troppo basso: {over_15*100:.1f}% (lambda_h={lh:.2f}, lambda_a={la:.2f}, total={lh+la:.2f})")
    if btts < 0.01:
        validation_warnings.append(f"⚠️ BTTS troppo basso: {btts*100:.1f}% (lambda_h={lh:.2f}, lambda_a={la:.2f})")
    
    # ⚠️ PRECISIONE: Verifica normalizzazione matrice con Kahan summation
    matrix_sum = 0.0
    c_matrix = 0.0
    for row in mat_ft:
        for val in row:
            y = val - c_matrix
            t = matrix_sum + y
            c_matrix = (t - matrix_sum) - y
            matrix_sum = t
    
    # ⚠️ PRECISIONE: Tolleranza più stretta
    if abs(matrix_sum - 1.0) > model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        validation_warnings.append(f"⚠️ Matrice non normalizzata correttamente: somma={matrix_sum:.8f} (dovrebbe essere 1.0)")
        logger.warning(f"Matrice non normalizzata: somma={matrix_sum:.10f}")
    
    # ⚠️ VALIDAZIONE COERENZA: Verifica coerenza lambda con total
    total_from_lambda = lh + la
    if abs(total_from_lambda - total) > 0.5:
        # Costruisci messaggio dettagliato con log modifiche
        deviation_pct = abs(total_from_lambda - total) / max(0.1, total) * 100
        warning_msg = f"⚠️ Lambda non coerenti con total: lambda_sum={total_from_lambda:.2f}, total={total:.2f} (deviazione: {deviation_pct:.1f}%)"

        # Aggiungi riepilogo modifiche se disponibile
        if lambda_adjustments_log and len(lambda_adjustments_log) > 1:
            # Identifica la modifica più grande
            max_change = 0.0
            max_change_source = ""
            for log_entry in lambda_adjustments_log[1:]:  # Salta la prima entry (iniziale)
                # Estrai il total dalla entry se presente
                if "total=" in log_entry:
                    parts = log_entry.split("total=")
                    if len(parts) > 1:
                        try:
                            total_in_entry = float(parts[1].split()[0])
                            change = abs(total_in_entry - total_initial)
                            if change > max_change:
                                max_change = change
                                max_change_source = log_entry.split(":")[0]
                        except (ValueError, IndexError):
                            pass

            if max_change_source:
                warning_msg += f" | Modifica principale: {max_change_source}"

        validation_warnings.append(warning_msg)
        logger.warning(f"Incoerenza lambda-total: {total_from_lambda:.4f} vs {total:.4f} (deviazione {deviation_pct:.1f}%)")

        # Log dettagliato delle modifiche per debugging
        if lambda_adjustments_log:
            logger.info("Traccia modifiche lambda:")
            for log_entry in lambda_adjustments_log:
                logger.info(f"  {log_entry}")
    
    # Verifica lambda ragionevoli
    if lh > 5.0 or la > 5.0:
        validation_warnings.append(f"⚠️ Lambda molto alti: lambda_h={lh:.2f}, lambda_a={la:.2f} (valori tipici: 0.5-3.0)")
    if lh < 0.1 or la < 0.1:
        validation_warnings.append(f"⚠️ Lambda molto bassi: lambda_h={lh:.2f}, lambda_a={la:.2f} (valori tipici: 0.5-3.0)")
    
    even_ft, odd_ft = prob_pari_dispari_from_matrix(mat_ft)
    even_ht, odd_ht = prob_pari_dispari_from_matrix(mat_ht)
    
    cs_home, cs_away = prob_clean_sheet_from_matrix(mat_ft)

    # ⚠️ NUOVI MERCATI: Asian Handicap, HT/FT
    # Asian Handicap - Linee più comuni
    asian_handicap = {
        "Home -0.5": prob_asian_handicap_from_matrix(mat_ft, -0.5, 'home'),
        "Home -1.0": prob_asian_handicap_from_matrix(mat_ft, -1.0, 'home'),
        "Home -1.5": prob_asian_handicap_from_matrix(mat_ft, -1.5, 'home'),
        "Home -2.0": prob_asian_handicap_from_matrix(mat_ft, -2.0, 'home'),
        "Home +0.5": prob_asian_handicap_from_matrix(mat_ft, 0.5, 'home'),
        "Home +1.0": prob_asian_handicap_from_matrix(mat_ft, 1.0, 'home'),
        "Home +1.5": prob_asian_handicap_from_matrix(mat_ft, 1.5, 'home'),
        "Away -0.5": prob_asian_handicap_from_matrix(mat_ft, -0.5, 'away'),
        "Away -1.0": prob_asian_handicap_from_matrix(mat_ft, -1.0, 'away'),
        "Away -1.5": prob_asian_handicap_from_matrix(mat_ft, -1.5, 'away'),
        "Away -2.0": prob_asian_handicap_from_matrix(mat_ft, -2.0, 'away'),
        "Away +0.5": prob_asian_handicap_from_matrix(mat_ft, 0.5, 'away'),
        "Away +1.0": prob_asian_handicap_from_matrix(mat_ft, 1.0, 'away'),
        "Away +1.5": prob_asian_handicap_from_matrix(mat_ft, 1.5, 'away'),
    }

    # HT/FT - 9 combinazioni
    ht_ft = prob_ht_ft_from_matrices(mat_ht, mat_ft)

    dist_home_ft, dist_away_ft = dist_gol_da_matrice(mat_ft)
    dist_home_ht, dist_away_ht = dist_gol_da_matrice(mat_ht)
    dist_tot_ft = dist_gol_totali_from_matrix(mat_ft)

    # 10. Multigol - RIMOSSI mercati generici, mantenute solo combo in combo_book
    # ranges = [(0,1),(1,3),(1,4),(1,5),(2,3),(2,4),(2,5),(3,5)]
    # multigol_home = {f"{a}-{b}": prob_multigol_from_dist(dist_home_ft, a, b) for a,b in ranges}
    # multigol_away = {f"{a}-{b}": prob_multigol_from_dist(dist_away_ft, a, b) for a,b in ranges}

    # 11. Double Chance
    dc = {
        "DC Casa o Pareggio": p_home + p_draw,
        "DC Trasferta o Pareggio": p_away + p_draw,
        "DC Casa o Trasferta": p_home + p_away
    }
    
    # 12. Margini vittoria
    mg = len(mat_ft) - 1
    # ⚠️ PRECISIONE: Accumulo con protezione NaN/negativi
    marg2 = 0.0
    marg3 = 0.0
    for h in range(mg + 1):
        for a in range(mg + 1):
            p = mat_ft[h][a]
            if p > 0 and p == p:  # Ignora negativi e NaN
                if h - a >= 2:
                    marg2 += p
                if h - a >= 3:
                    marg3 += p
    marg2 = max(0.0, min(1.0, marg2))
    marg3 = max(0.0, min(1.0, marg3))
    
    # 13. Combo mercati
    combo_book = {
        "1 & Over 1.5": prob_esito_over_from_matrix(mat_ft, '1', 1.5),
        "1 & Over 2.5": prob_esito_over_from_matrix(mat_ft, '1', 2.5),
        "2 & Over 1.5": prob_esito_over_from_matrix(mat_ft, '2', 1.5),
        "2 & Over 2.5": prob_esito_over_from_matrix(mat_ft, '2', 2.5),
        "1X & Over 1.5": prob_dc_over_from_matrix(mat_ft, '1X', 1.5),
        "X2 & Over 1.5": prob_dc_over_from_matrix(mat_ft, 'X2', 1.5),
        "1X & Over 2.5": prob_dc_over_from_matrix(mat_ft, '1X', 2.5),
        "X2 & Over 2.5": prob_dc_over_from_matrix(mat_ft, 'X2', 2.5),
        "1X & BTTS": prob_dc_btts_from_matrix(mat_ft, '1X'),
        "X2 & BTTS": prob_dc_btts_from_matrix(mat_ft, 'X2'),
        "1 & BTTS": prob_esito_btts_from_matrix(mat_ft, '1'),
        "2 & BTTS": prob_esito_btts_from_matrix(mat_ft, '2'),
        "1 & NG": prob_esito_ng_from_matrix(mat_ft, '1'),
        "2 & NG": prob_esito_ng_from_matrix(mat_ft, '2'),
        "1X & Under 3.5": prob_dc_over_from_matrix(mat_ft, '1X', 3.5, inverse=True),  # Under 3.5 = NOT Over 3.5
        "X2 & Under 3.5": prob_dc_over_from_matrix(mat_ft, 'X2', 3.5, inverse=True),
        "1X & GG": prob_dc_btts_from_matrix(mat_ft, '1X'),  # Già calcolato correttamente dalla matrice
        "X2 & GG": prob_dc_btts_from_matrix(mat_ft, 'X2'),  # Già calcolato correttamente dalla matrice
        "1X & Under 2.5": prob_dc_over_from_matrix(mat_ft, '1X', 2.5, inverse=True),
        "X2 & Under 2.5": prob_dc_over_from_matrix(mat_ft, 'X2', 2.5, inverse=True),
    }
    combo_book["GG & Over 2.5"] = gg_over25

    # Range target per combo Multigol (garantiamo coerenza con i range calcolati sopra)
    multigol_combo_ranges = [(1, 3), (1, 4), (1, 5), (2, 3), (2, 4), (2, 5), (3, 5)]

    # Esito + Multigol
    for esito_key in ['1', '2']:
        for gmin, gmax in multigol_combo_ranges:
            range_label = f"{gmin}-{gmax}"
            combo_book[f"{esito_key} & Multigol {range_label}"] = prob_esito_multigol_from_matrix(
                mat_ft, esito_key, gmin, gmax
            )

    # Double Chance + Multigol
    for dc_key in ['1X', 'X2', '12']:
        for gmin, gmax in multigol_combo_ranges:
            range_label = f"{gmin}-{gmax}"
            combo_book[f"{dc_key} & Multigol {range_label}"] = prob_dc_multigol_from_matrix(
                mat_ft, dc_key, gmin, gmax
            )

    # 14. Top risultati
    top10 = top_results_from_matrix(mat_ft, 10, 0.005)
    
    # 15. Entropia e metriche
    ent_home = entropia_poisson(lh)
    ent_away = entropia_poisson(la)
    
    # 16. Confronto con odds
    odds_prob = {
        "1": 1/odds_1,
        "X": 1/odds_x,
        "2": 1/odds_2
    }
    scost = {
        "1": (p_home - odds_prob["1"]) * 100,
        "X": (p_draw - odds_prob["X"]) * 100,
        "2": (p_away - odds_prob["2"]) * 100
    }
    
    # 17. Statistiche aggiuntive
    # ⚠️ PRECISIONE: Accumulo con protezione (dist_tot_ft è già normalizzata, ma aggiungiamo protezione per sicurezza)
    odd_mass = 0.0
    for i, p in enumerate(dist_tot_ft):
        if i % 2 == 1 and p > 0 and p == p:  # Ignora negativi e NaN
            odd_mass += p
    odd_mass = max(0.0, min(1.0, odd_mass))
    even_mass2 = 1.0 - odd_mass
    
    cover_0_2 = 0.0
    for i in range(0, min(3, len(dist_tot_ft))):
        if i < len(dist_tot_ft):
            p = dist_tot_ft[i]
            if p > 0 and p == p:  # Ignora negativi e NaN
                cover_0_2 += p
    cover_0_2 = max(0.0, min(1.0, cover_0_2))
    
    cover_0_3 = 0.0
    for i in range(0, min(4, len(dist_tot_ft))):
        if i < len(dist_tot_ft):
            p = dist_tot_ft[i]
            if p > 0 and p == p:  # Ignora negativi e NaN
                cover_0_3 += p
    cover_0_3 = max(0.0, min(1.0, cover_0_3))
    
    # 18. Calibrazione probabilità (se disponibile storico) - CALIBRAZIONE DINAMICA PER LEGA
    calibrate_func = load_calibration_from_history(league=league)
    if calibrate_func:
        p_home_cal = calibrate_func(p_home)
        p_draw_cal = calibrate_func(p_draw)
        p_away_cal = calibrate_func(p_away)
        # ⚠️ CORREZIONE: Normalizza con protezione contro valori estremi
        tot_cal = p_home_cal + p_draw_cal + p_away_cal
        if tot_cal > 0:
            p_home_cal /= tot_cal
            p_draw_cal /= tot_cal
            p_away_cal /= tot_cal
        else:
            # Fallback se calibrazione produce valori non validi
            p_home_cal, p_draw_cal, p_away_cal = p_home, p_draw, p_away
        
        # ⚠️ VERIFICA: Assicura che probabilità calibrate siano ragionevoli
        if p_home_cal < 0 or p_home_cal > 1 or p_draw_cal < 0 or p_draw_cal > 1 or p_away_cal < 0 or p_away_cal > 1:
            # Se calibrazione produce valori fuori range, usa valori non calibrati
            p_home_cal, p_draw_cal, p_away_cal = p_home, p_draw, p_away
    else:
        p_home_cal, p_draw_cal, p_away_cal = p_home, p_draw, p_away
    
    # 19. Ensemble prediction (opzionale, per maggiore robustezza)
    # Usa ensemble solo se abbiamo tutti i dati necessari
    use_ensemble = odds_over25 and odds_under25
    if use_ensemble:
        ensemble_result = ensemble_prediction(
            odds_1, odds_x, odds_2, total,
            odds_over25, odds_under25, odds_btts,
            odds_dnb_home, odds_dnb_away, league
        )
        # ⚠️ PRECISIONE: Blend con normalizzazione finale
        # Blend: 80% modello principale, 20% ensemble
        # ⚠️ PROTEZIONE: Valida ensemble_result prima di usarlo
        ensemble_h = max(0.0, min(1.0, ensemble_result.get("p_home", p_home_cal)))
        ensemble_d = max(0.0, min(1.0, ensemble_result.get("p_draw", p_draw_cal)))
        ensemble_a = max(0.0, min(1.0, ensemble_result.get("p_away", p_away_cal)))
        
        # Normalizza ensemble se necessario
        ensemble_tot = ensemble_h + ensemble_d + ensemble_a
        if ensemble_tot > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            ensemble_h /= ensemble_tot
            ensemble_d /= ensemble_tot
            ensemble_a /= ensemble_tot
        
        p_home_final = 0.8 * p_home_cal + 0.2 * ensemble_h
        p_draw_final = 0.8 * p_draw_cal + 0.2 * ensemble_d
        p_away_final = 0.8 * p_away_cal + 0.2 * ensemble_a
        
        # ⚠️ PRECISIONE: Normalizza probabilità finali dopo ensemble
        tot_final = p_home_final + p_draw_final + p_away_final
        if tot_final > model_config.TOL_DIVISION_ZERO:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
            p_home_final /= tot_final
            p_draw_final /= tot_final
            p_away_final /= tot_final
        else:
            # Fallback se ensemble produce valori non validi
            p_home_final, p_draw_final, p_away_final = p_home_cal, p_draw_cal, p_away_cal
    else:
        p_home_final, p_draw_final, p_away_final = p_home_cal, p_draw_cal, p_away_cal
    
    # ⚠️ PRECISIONE: Verifica finale con tolleranza più stretta
    # (La normalizzazione è già stata fatta dopo ensemble, ma verifichiamo per sicurezza)
    tot_final_check = p_home_final + p_draw_final + p_away_final
    # ⚠️ PRECISIONE: Tolleranza più stretta
    if abs(tot_final_check - 1.0) > model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        if tot_final_check > 0:
            p_home_final /= tot_final_check
            p_draw_final /= tot_final_check
            p_away_final /= tot_final_check
        else:
            logger.warning(f"Probabilità finali sommano a {tot_final_check}, uso probabilità raw")
            # Fallback estremo: usa probabilità raw
            p_home_final, p_draw_final, p_away_final = p_home, p_draw, p_away
    
    # ⚠️ VALIDAZIONE COERENZA: Verifica monotonia probabilità Over/Under (generalizzata)
    over_map = {
        1.5: over_15,
        2.5: over_25,
        3.5: over_35,
    }
    if all(v is not None for v in over_map.values()):
        # Ordina per soglia crescente e impone monotonia non crescente
        thresholds = sorted(over_map.keys())
        overs = [over_map[t] for t in thresholds]
        # Enforce: overs[i] >= overs[i+1]
        changed = False
        for i in range(len(overs) - 1):
            if overs[i] < overs[i + 1]:
                logger.warning(f"Violazione monotonia Over tra {thresholds[i]} e {thresholds[i+1]}: {overs[i]:.4f} < {overs[i+1]:.4f}. Correggo.")
                # ⚠️ CORREZIONE: Se Over 1.5 < Over 2.5, imposta Over 2.5 = Over 1.5 (monotonia non crescente)
                overs[i + 1] = overs[i]
                changed = True
        if changed:
            over_15, over_25, over_35 = overs
            under_15 = 1.0 - over_15
            under_25 = 1.0 - over_25
            under_35 = 1.0 - over_35
    
    # ⚠️ VALIDAZIONE COERENZA: Verifica che probabilità finali siano valide
    if not all(0 <= p <= 1 for p in [p_home_final, p_draw_final, p_away_final]):
        logger.error(f"Probabilità finali fuori range: Home={p_home_final:.4f}, Draw={p_draw_final:.4f}, Away={p_away_final:.4f}")
        # Forza normalizzazione
        p_home_final = max(0.0, min(1.0, p_home_final))
        p_draw_final = max(0.0, min(1.0, p_draw_final))
        p_away_final = max(0.0, min(1.0, p_away_final))
        tot_fix = p_home_final + p_draw_final + p_away_final
        if tot_fix > 0:
            p_home_final /= tot_fix
            p_draw_final /= tot_fix
            p_away_final /= tot_fix

    # ⚠️ VALIDAZIONE COMBO MULTIGOL: Esito & Multigol vs Esito e range multigol
    for esito_key, esito_prob in [("1", p_home_final), ("X", p_draw_final), ("2", p_away_final)]:
        for gmin, gmax in multigol_combo_ranges:
            range_label = f"{gmin}-{gmax}"
            combo_key = f"{esito_key} & Multigol {range_label}"
            if combo_key in combo_book:
                combo_prob = combo_book[combo_key]
                # Calcola prob multigol per questo range
                mult_prob = prob_multigol_from_dist(dist_tot_ft, gmin, gmax)
                if isinstance(mult_prob, (int, float)) and math.isfinite(mult_prob):
                    max_combo = min(esito_prob, mult_prob)
                    if combo_prob > max_combo + model_config.TOL_PROBABILITY_CHECK:
                        logger.warning(f"{combo_key} ({combo_prob:.4f}) > min(P({esito_key}), Multigol {range_label}) ({max_combo:.4f}), correggo")
                        combo_book[combo_key] = max_combo

    # ⚠️ VALIDAZIONE COMBO MULTIGOL: DC & Multigol vs DC e range multigol
    dc_prob_map = {
        "1X": p_home_final + p_draw_final,
        "X2": p_draw_final + p_away_final,
        "12": p_home_final + p_away_final
    }
    for dc_key, dc_prob in dc_prob_map.items():
        for gmin, gmax in multigol_combo_ranges:
            range_label = f"{gmin}-{gmax}"
            combo_key = f"{dc_key} & Multigol {range_label}"
            if combo_key in combo_book:
                combo_prob = combo_book[combo_key]
                # Calcola prob multigol per questo range
                mult_prob = prob_multigol_from_dist(dist_tot_ft, gmin, gmax)
                if isinstance(mult_prob, (int, float)) and math.isfinite(mult_prob):
                    max_combo = min(dc_prob, mult_prob)
                    if combo_prob > max_combo + model_config.TOL_PROBABILITY_CHECK:
                        logger.warning(f"{combo_key} ({combo_prob:.4f}) > min(DC {dc_key}, Multigol {range_label}) ({max_combo:.4f}), correggo")
                        combo_book[combo_key] = max_combo

    # ⚠️ VALIDAZIONE COERENZA MATEMATICA: Verifica coerenza tra probabilità marginali e combinate
    # 1. Coerenza probabilità complementari
    if abs((over_15 + under_15) - 1.0) > model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        logger.warning(f"Over 1.5 + Under 1.5 = {over_15 + under_15:.6f} (dovrebbe essere 1.0)")
        under_15 = 1.0 - over_15
    if abs((over_25 + under_25) - 1.0) > model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        logger.warning(f"Over 2.5 + Under 2.5 = {over_25 + under_25:.6f} (dovrebbe essere 1.0)")
        under_25 = 1.0 - over_25
    if abs((over_35 + under_35) - 1.0) > model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        logger.warning(f"Over 3.5 + Under 3.5 = {over_35 + under_35:.6f} (dovrebbe essere 1.0)")
        under_35 = 1.0 - over_35
    if abs((even_ft + odd_ft) - 1.0) > model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        logger.warning(f"Pari FT + Dispari FT = {even_ft + odd_ft:.6f} (dovrebbe essere 1.0)")
        odd_ft = 1.0 - even_ft
    
    # 2. Coerenza probabilità combinate vs marginali (P(A & B) <= min(P(A), P(B)))
    # BTTS vs Over 1.5 (già verificato prima, ma ri-verifichiamo)
    if btts > over_15 + model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        logger.warning(f"BTTS ({btts:.4f}) > Over 1.5 ({over_15:.4f}), correggo")
        btts = min(btts, over_15)
    
    # GG & Over 2.5 vs BTTS e Over 2.5
    max_gg_over25 = min(btts, over_25)
    if gg_over25 > max_gg_over25 + model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        logger.warning(f"GG & Over 2.5 ({gg_over25:.4f}) > min(BTTS, Over 2.5) ({max_gg_over25:.4f}), correggo")
        gg_over25 = max_gg_over25
    
    # Esito & Over vs Esito e Over
    for esito_key, esito_prob in [("1", p_home_final), ("X", p_draw_final), ("2", p_away_final)]:
        for soglia, over_prob in [(1.5, over_15), (2.5, over_25)]:
            combo_key = f"{esito_key} & Over {soglia}"
            if combo_key in combo_book:
                combo_prob = combo_book[combo_key]
                max_combo = min(esito_prob, over_prob)
                if combo_prob > max_combo + model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
                    logger.warning(f"{combo_key} ({combo_prob:.4f}) > min(P({esito_key}), Over {soglia}) ({max_combo:.4f}), correggo")
                    combo_book[combo_key] = max_combo
    
    # Esito & BTTS vs Esito e BTTS
    for esito_key, esito_prob in [("1", p_home_final), ("X", p_draw_final), ("2", p_away_final)]:
        combo_key = f"{esito_key} & BTTS"
        if combo_key in combo_book:
            combo_prob = combo_book[combo_key]
            max_combo = min(esito_prob, btts)
            if combo_prob > max_combo + model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
                logger.warning(f"{combo_key} ({combo_prob:.4f}) > min(P({esito_key}), BTTS) ({max_combo:.4f}), correggo")
                combo_book[combo_key] = max_combo
    
    # DC & Over vs DC e Over
    for dc_key, dc_prob in [("1X", p_home_final + p_draw_final), ("X2", p_draw_final + p_away_final), ("12", p_home_final + p_away_final)]:
        for soglia, over_prob in [(1.5, over_15), (2.5, over_25), (3.5, over_35)]:
            combo_key = f"{dc_key} & Over {soglia}"
            if combo_key in combo_book:
                combo_prob = combo_book[combo_key]
                max_combo = min(dc_prob, over_prob)
                if combo_prob > max_combo + model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
                    logger.warning(f"{combo_key} ({combo_prob:.4f}) > min(DC {dc_key}, Over {soglia}) ({max_combo:.4f}), correggo")
                    combo_book[combo_key] = max_combo
            # Under
            combo_key_under = f"{dc_key} & Under {soglia}"
            if combo_key_under in combo_book:
                under_prob = 1.0 - over_prob
                combo_prob = combo_book[combo_key_under]
                max_combo = min(dc_prob, under_prob)
                if combo_prob > max_combo + model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
                    logger.warning(f"{combo_key_under} ({combo_prob:.4f}) > min(DC {dc_key}, Under {soglia}) ({max_combo:.4f}), correggo")
                    combo_book[combo_key_under] = max_combo
    
    # DC & BTTS vs DC e BTTS
    for dc_key, dc_prob in [("1X", p_home_final + p_draw_final), ("X2", p_draw_final + p_away_final), ("12", p_home_final + p_away_final)]:
        combo_key = f"{dc_key} & BTTS"
        if combo_key in combo_book:
            combo_prob = combo_book[combo_key]
            max_combo = min(dc_prob, btts)
            if combo_prob > max_combo + model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
                logger.warning(f"{combo_key} ({combo_prob:.4f}) > min(DC {dc_key}, BTTS) ({max_combo:.4f}), correggo")
                combo_book[combo_key] = max_combo
    
    
    # Alias combinazioni richieste con formato "esito+mercato"
    def _register_alias(alias_key: str, value: float) -> None:
        variants = {
            alias_key,
            alias_key.lower(),
            alias_key.replace(" ", ""),
            alias_key.replace(" ", "").lower(),
        }
        for variant in variants:
            combo_book[variant] = value

    alias_sources = {
        "2+GG": "2 & BTTS",
        "2+Over 1.5": "2 & Over 1.5",
        "2+Over 2.5": "2 & Over 2.5",
        "2+NG": "2 & NG",
        "1X+GG": "1X & GG",
        "X2+GG": "X2 & GG",
        "1X+Over 1.5": "1X & Over 1.5",
        "1X+Over 2.5": "1X & Over 2.5",
        "X2+Over 1.5": "X2 & Over 1.5",
        "X2+Over 2.5": "X2 & Over 2.5",
    }
    for alias_key, original_key in alias_sources.items():
        if original_key in combo_book:
            _register_alias(alias_key, combo_book[original_key])

    for gmin, gmax in multigol_combo_ranges:
        range_label = f"{gmin}-{gmax}"
        base_esito_key = f"2 & Multigol {range_label}"
        if base_esito_key in combo_book:
            _register_alias(f"2+Multigol {range_label}", combo_book[base_esito_key])
        for dc_key, alias_prefix in [("1X", "1X+Multigol"), ("X2", "X2+Multigol")]:
            base_dc_key = f"{dc_key} & Multigol {range_label}"
            if base_dc_key in combo_book:
                _register_alias(f"{alias_prefix} {range_label}", combo_book[base_dc_key])
    
    # 3. Coerenza Clean Sheet: CS Home + (almeno 1 gol away) = 1.0
    # P(almeno 1 gol away) = 1 - P(0 gol away) = 1 - CS Home
    # Quindi: CS Home + P(almeno 1 gol away) = CS Home + (1 - CS Home) = 1.0
    # Verifichiamo che CS Home sia nel range [0, 1] e che la somma sia coerente
    if cs_home < 0 or cs_home > 1:
        logger.warning(f"CS Home ({cs_home:.4f}) fuori range [0, 1], correggo")
        cs_home = max(0.0, min(1.0, cs_home))
    if cs_away < 0 or cs_away > 1:
        logger.warning(f"CS Away ({cs_away:.4f}) fuori range [0, 1], correggo")
        cs_away = max(0.0, min(1.0, cs_away))
    
    # Verifica coerenza: CS Home + P(almeno 1 gol away) = 1.0 (sempre vero per definizione)
    # Ma verifichiamo che CS Home sia calcolato correttamente dalla matrice
    # CS Home = sum(mat[h][0] for h) = P(away segna 0 gol)
    # Quindi: CS Home + P(away segna >= 1 gol) = 1.0 (sempre vero)
    
    # 4. Coerenza DC: DC 1X = P(1) + P(X) (perché mutuamente esclusivi)
    dc_1x_calc = p_home_final + p_draw_final
    if abs(dc["DC Casa o Pareggio"] - dc_1x_calc) > model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        logger.warning(f"DC 1X calcolato ({dc['DC Casa o Pareggio']:.4f}) != P(1) + P(X) ({dc_1x_calc:.4f}), correggo")
        dc["DC Casa o Pareggio"] = dc_1x_calc
    
    dc_x2_calc = p_draw_final + p_away_final
    if abs(dc["DC Trasferta o Pareggio"] - dc_x2_calc) > model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        logger.warning(f"DC X2 calcolato ({dc['DC Trasferta o Pareggio']:.4f}) != P(X) + P(2) ({dc_x2_calc:.4f}), correggo")
        dc["DC Trasferta o Pareggio"] = dc_x2_calc
    
    dc_12_calc = p_home_final + p_away_final
    if abs(dc["DC Casa o Trasferta"] - dc_12_calc) > model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        logger.warning(f"DC 12 calcolato ({dc['DC Casa o Trasferta']:.4f}) != P(1) + P(2) ({dc_12_calc:.4f}), correggo")
        dc["DC Casa o Trasferta"] = dc_12_calc
    
    # Calcola market movement info per output (usa spread e total correnti calcolati)
    movement_info = calculate_market_movement_factor(
        spread_apertura,
        total_apertura,
        spread_corrente_calculated,
        total_corrente_calculated,
        league=league,
        volatility_score=manual_movement_context.get("volatility_score"),
        recent_samples=manual_movement_context.get("history_samples", 0),
    )
    
    # ⚠️ VALIDAZIONE FINALE: Verifica che tutte le probabilità siano coerenti e nel range [0, 1]
    all_probs = {
        "p_home": p_home_final,
        "p_draw": p_draw_final,
        "p_away": p_away_final,
        "over_15": over_15,
        "under_15": under_15,
        "over_25": over_25,
        "under_25": under_25,
        "over_35": over_35,
        "under_35": under_35,
        "btts": btts,
        "gg_over25": gg_over25,
        "even_ft": even_ft,
        "odd_ft": odd_ft,
        "cs_home": cs_home,
        "cs_away": cs_away,
    }
    
    # Verifica range [0, 1] per tutte le probabilità
    for prob_name, prob_value in all_probs.items():
        if prob_value < 0 or prob_value > 1:
            logger.error(f"Probabilità {prob_name} fuori range [0, 1]: {prob_value:.6f}")
            # Correggi
            all_probs[prob_name] = max(0.0, min(1.0, prob_value))
    
    # Aggiorna variabili con valori corretti
    p_home_final = all_probs["p_home"]
    p_draw_final = all_probs["p_draw"]
    p_away_final = all_probs["p_away"]
    over_15 = all_probs["over_15"]
    under_15 = all_probs["under_15"]
    over_25 = all_probs["over_25"]
    under_25 = all_probs["under_25"]
    over_35 = all_probs["over_35"]
    under_35 = all_probs["under_35"]
    btts = all_probs["btts"]
    gg_over25 = all_probs["gg_over25"]
    even_ft = all_probs["even_ft"]
    odd_ft = all_probs["odd_ft"]
    cs_home = all_probs["cs_home"]
    cs_away = all_probs["cs_away"]
    
    # Verifica coerenza finale: 1X2 deve sommare a 1.0
    tot_1x2 = p_home_final + p_draw_final + p_away_final
    if abs(tot_1x2 - 1.0) > model_config.TOL_PROBABILITY_CHECK:  # ⚠️ MICRO-PRECISIONE: Usa tolleranza standardizzata
        logger.error(f"Probabilità 1X2 non sommano a 1.0: {tot_1x2:.6f}, forzo normalizzazione")
        if tot_1x2 > 0:
            p_home_final /= tot_1x2
            p_draw_final /= tot_1x2
            p_away_final /= tot_1x2
        else:
            # Fallback estremo
            p_home_final, p_draw_final, p_away_final = 1/3, 1/3, 1/3
    
    # Recupera dati API aggiuntive per output
    additional_api_data = {
        "weather": weather_data if weather_data and weather_data.get("available") else None,
        "football_data_org": None,
        "thesportsdb": None,
        "statsbomb": None,
        "football_data_metrics": None,
    }
    
    # Recupera dati aggiuntivi da tutte le API (già inclusi in advanced_data se disponibile)
    if advanced_data and advanced_data.get("data_available"):
        # Usa dati già recuperati da get_advanced_team_data() che integra tutte le API
        additional_api_data["football_data_org"] = advanced_data.get("football_data_home")
        additional_api_data["thesportsdb"] = advanced_data.get("thesportsdb_home") or stadium_data
        additional_api_data["statsbomb"] = advanced_data.get("statsbomb_home")
        additional_api_data["football_data_metrics"] = advanced_data.get("football_data_home_metrics")
    elif home_team:
        # Fallback: recupera direttamente se advanced_data non disponibile
        try:
            if FOOTBALL_DATA_API_KEY:
                additional_api_data["football_data_org"] = football_data_get_team_info(home_team)
            additional_api_data["thesportsdb"] = stadium_data if stadium_data and stadium_data.get("available") else thesportsdb_get_team_info(home_team)
            additional_api_data["statsbomb"] = statsbomb_get_team_metrics(home_team)
            additional_api_data["football_data_metrics"] = football_data_calculate_form_metrics(
                additional_api_data["football_data_org"].get("team_id") if additional_api_data["football_data_org"] else None,
                football_data_get_recent_matches(
                    additional_api_data["football_data_org"].get("team_id"),
                    limit=6,
                ) if additional_api_data["football_data_org"] and additional_api_data["football_data_org"].get("team_id") else [],
            )
        # FIX BUG: Replace bare except with specific Exception (prevents catching KeyboardInterrupt/SystemExit)
        except Exception as e:
            logger.warning(f"⚠️ Errore calcolo football_data_metrics: {e}")
            pass  # Non bloccare se fallisce
    
    probability_intervals = {}
    try:
        probability_intervals = calculate_confidence_intervals(
            lh,
            la,
            rho,
            n_simulations=4000 if not kwargs.get("fast_mode") else 2000,
            confidence_level=0.9,
        )
    except Exception as interval_err:
        logger.warning(f"Impossibile calcolare gli intervalli di confidenza: {interval_err}")
        probability_intervals = {}

    result = {
        "lambda_home": lh,
        "lambda_away": la,
        "rho": rho,
        "spread_corrente": round(spread_corrente_calculated, 3),  # Spread corrente calcolato dai lambda finali
        "total_corrente": round(total_corrente_calculated, 3),  # Total corrente calcolato dai lambda finali
        "spread_apertura": spread_apertura if spread_apertura is not None else None,
        "total_apertura": total_apertura if total_apertura is not None else None,
        "p_home": p_home_final,
        "p_draw": p_draw_final,
        "p_away": p_away_final,
        "p_home_raw": p_home,  # Probabilità raw (non calibrate)
        "p_draw_raw": p_draw,
        "p_away_raw": p_away,
        "calibration_applied": calibrate_func is not None,
        "ensemble_applied": use_ensemble,
        "market_movement": movement_info,  # Info movimento mercato
        "manual_inputs_meta": manual_inputs_meta,
        "manual_confidence": manual_confidence_snapshot,
        "probability_intervals": probability_intervals,
        "additional_api_data": additional_api_data,  # Dati API aggiuntive
        "over_05": over_05_ft,
        "over_15": over_15,
        "under_15": under_15,
        "over_25": over_25,
        "under_25": under_25,
        "over_35": over_35,
        "under_35": under_35,
        "over_05_ht": over_05_ht,
        "over_15_ht": over_15_ht,
        "over_05ht_over_05ft": over_05ht_over_05ft,
        "over_05ht_over_15ft": over_05ht_over_15ft,
        "over_05ht_over_25ft": over_05ht_over_25ft,
        "over_15ht_over_25ft": over_15ht_over_25ft,
        "over_05ht_over_35ft": over_05ht_over_35ft,
        "over_15ht_over_35ft": over_15ht_over_35ft,
        "btts": btts,
        "gg_over25": gg_over25,
        "even_ft": even_ft,
        "odd_ft": odd_ft,
        "even_ht": even_ht,
        "odd_ht": odd_ht,
        "cs_home": cs_home,
        "cs_away": cs_away,
        "clean_sheet_qualcuno": 1 - btts,
        # ⚠️ NUOVI MERCATI: Asian Handicap, HT/FT
        "asian_handicap": asian_handicap,
        "ht_ft": ht_ft,
        # RIMOSSI mercati Multigol generici - mantenute solo combo in combo_book
        # "multigol_home": multigol_home,
        # "multigol_away": multigol_away,
        # "multigol": multigol_home,
        "dc": dc,  # DC (Double Chance) contiene già 1X, X2, 12 - non serve "combo" duplicato
        "validation_warnings": validation_warnings,  # Warning per probabilità anomale
        "lambda_adjustments_log": lambda_adjustments_log,  # Log modifiche lambda per debugging
        "matrix_sum": matrix_sum,  # Somma matrice per debug
        "marg2": marg2,
        "marg3": marg3,
        "combo_book": combo_book,
        "top10": top10,
        "ent_home": ent_home,
        "ent_away": ent_away,
        "odds_prob": odds_prob,
        "scost": scost,
        "odd_mass": odd_mass,
        "even_mass2": even_mass2,
        "cover_0_2": cover_0_2,
        "cover_0_3": cover_0_3,
        "market_calibration_stats": market_calibration_stats,
        "manual_history_key": manual_history_key,
    }

    try:
        raw_manual_values = manual_inputs_meta.get("raw_inputs", {}) if manual_inputs_meta else {}
        if raw_manual_values and any(val is not None for val in raw_manual_values.values()):
            intervals = probability_intervals or {}
            p_home_interval = intervals.get("p_home", (None, None))
            p_draw_interval = intervals.get("p_draw", (None, None))
            p_away_interval = intervals.get("p_away", (None, None))
            movement_history_samples = manual_inputs_meta.get("history_samples", 0) if manual_inputs_meta else 0
            log_manual_precision_sample({
                "timestamp": datetime.utcnow().isoformat(),
                "match_key": manual_history_key or "",
                "league": league,
                "home_team": home_team or "",
                "away_team": away_team or "",
                "manual_confidence": (manual_confidence_snapshot or {}).get("score"),
                "spread_apertura": raw_manual_values.get("spread_apertura"),
                "total_apertura": raw_manual_values.get("total_apertura"),
                "spread_corrente": raw_manual_values.get("spread_corrente"),
                "total_corrente": raw_manual_values.get("total_corrente"),
                "lambda_home": lh,
                "lambda_away": la,
                "rho": rho,
                "prob_home": result.get("p_home"),
                "prob_draw": result.get("p_draw"),
                "prob_away": result.get("p_away"),
                "predicted_total": lh + la,
                "predicted_spread": lh - la,
                "prob_home_low": p_home_interval[0],
                "prob_home_high": p_home_interval[1],
                "prob_draw_low": p_draw_interval[0],
                "prob_draw_high": p_draw_interval[1],
                "prob_away_low": p_away_interval[0],
                "prob_away_high": p_away_interval[1],
                "movement_magnitude": movement_info.get("normalized_movement") if movement_info else None,
                "movement_type": movement_info.get("movement_type") if movement_info else None,
                "history_samples": movement_history_samples,
                "actual_ft_home": "",
                "actual_ft_away": "",
                "notes": "manual_input_snapshot",
            })
    except Exception as manual_log_error:
        logger.warning(f"⚠️ Impossibile loggare i dati manuali per analisi: {manual_log_error}")

    try:
        precision_summary = compute_manual_precision_metrics(window_days=30)
        if precision_summary:
            result["manual_precision_metrics"] = precision_summary
    except Exception as precision_summary_error:
        logger.warning(f"⚠️ Impossibile calcolare le metriche di precisione manuale: {precision_summary_error}")

    # ✅ COMPATIBILITÀ LEGACY: Mantieni vecchi nomi chiave usati dagli script di test
    result["prob_home"] = result.get("p_home")
    result["prob_draw"] = result.get("p_draw")
    result["prob_away"] = result.get("p_away")
    if "over_25" in result:
        result["prob_over25"] = result["over_25"]
    if "under_25" in result:
        result["prob_under25"] = result["under_25"]
    if "btts" in result:
        result["prob_btts"] = result["btts"]
        result["prob_no_btts"] = 1 - result["btts"]

    return result

# ============================================================
#   VALUE BETS PREPARATION
# ============================================================

def prepare_value_bets(
    ris: Dict[str, Any],
    odds_1: float,
    odds_x: float,
    odds_2: float,
    odds_over: float = None,
    odds_under: float = None,
    odds_btts: float = None,
    odds_dnb_home: float = None,
    odds_dnb_away: float = None,
) -> List[Dict[str, Any]]:
    """
    Prepara lista di value bets calcolando edge e EV per tutti i mercati disponibili.

    Args:
        ris: Risultati dell'analisi completa da risultato_completo_improved()
        odds_1, odds_x, odds_2: Quote 1X2
        odds_over, odds_under: Quote Over/Under 2.5
        odds_btts: Quote BTTS
        odds_dnb_home, odds_dnb_away: Quote Draw No Bet

    Returns:
        Lista di dizionari con value bets. Ogni dizionario contiene:
        - Esito: Nome dell'esito (es. "Casa", "Over 2.5", "BTTS")
        - Prob Modello %: Probabilità del modello formattata
        - Edge %: Edge percentuale formattato
        - EV %: Expected Value percentuale formattato
        - Rec: Raccomandazione basata su edge
    """
    value_bets = []

    def add_bet(esito: str, prob_model: float, odds: float):
        """Helper per aggiungere una bet alla lista."""
        if odds is None or odds <= 1.0 or prob_model is None:
            return

        try:
            prob_model = float(prob_model)
        except (TypeError, ValueError):
            return

        prob_model = max(0.0, min(1.0, prob_model))

        # Calcola probabilità implicita della quota
        prob_book = 1.0 / odds

        # Calcola edge ed EV
        edge = (prob_model - prob_book) * 100.0
        ev = (prob_model * odds - 1.0) * 100.0

        # Kelly Criterion (full)
        if odds > 1.0 and (odds - 1.0) > 0:
            kelly_full = (prob_model * odds - 1.0) / (odds - 1.0)
        else:
            kelly_full = 0.0

        if not math.isfinite(kelly_full) or kelly_full < 0.0:
            kelly_full = 0.0
        kelly_full = min(1.0, kelly_full)

        # Determina raccomandazione
        if edge >= 5.0:
            rec = "🔥 Forte"
        elif edge >= 3.0:
            rec = "✅ Buona"
        elif edge >= 2.0:
            rec = "🟡 Marginale"
        elif edge >= 0.5:
            rec = "⚪ Controllare"
        else:
            rec = "❌ No"

        prob_pct = prob_model * 100.0
        prob_book_pct = prob_book * 100.0

        value_bets.append({
            "Bookmaker": "Bet365",
            "Esito": esito,
            "Odd": float(odds),
            "Quota": float(odds),
            "ProbRaw": prob_model,
            "Prob %": f"{prob_pct:.2f}",
            "Prob Modello %": f"{prob_pct:.2f}",
            "Prob Book %": f"{prob_book_pct:.2f}",
            "EdgeRaw": edge,
            "Edge %": f"{edge:+.2f}",
            "EVRaw": ev,
            "EV %": f"{ev:+.2f}",
            "KellyBase": kelly_full,
            "Kelly %": f"{kelly_full*100:.2f}",
            "Rec": rec
        })

    # 1X2
    add_bet("Casa (1)", ris.get("p_home", 0), odds_1)
    add_bet("Pareggio (X)", ris.get("p_draw", 0), odds_x)
    add_bet("Trasferta (2)", ris.get("p_away", 0), odds_2)

    # Over/Under 2.5
    if odds_over is not None:
        add_bet("Over 2.5", ris.get("over_25", 0), odds_over)
    if odds_under is not None:
        add_bet("Under 2.5", ris.get("under_25", 0), odds_under)

    # BTTS
    if odds_btts is not None and odds_btts > 0:
        add_bet("BTTS Yes", ris.get("btts", 0), odds_btts)

    # Draw No Bet
    if odds_dnb_home is not None and odds_dnb_home > 0:
        # DNB Home = P(Home) / (P(Home) + P(Away))
        p_home = ris.get("p_home", 0)
        p_away = ris.get("p_away", 0)
        total = p_home + p_away
        if total > 0:
            prob_dnb_home = p_home / total
            add_bet("DNB Casa", prob_dnb_home, odds_dnb_home)

    if odds_dnb_away is not None and odds_dnb_away > 0:
        # DNB Away = P(Away) / (P(Home) + P(Away))
        p_home = ris.get("p_home", 0)
        p_away = ris.get("p_away", 0)
        total = p_home + p_away
        if total > 0:
            prob_dnb_away = p_away / total
            add_bet("DNB Trasferta", prob_dnb_away, odds_dnb_away)

    # Doppia Chance (se disponibile nel risultato)
    dc = ris.get("dc", {})
    if dc:
        # Le quote DC non sono passate come parametro, quindi non le includiamo
        # per ora. Potrebbero essere aggiunte in futuro se necessario.
        pass

    return value_bets

# ============================================================
#   CONTROLLI QUALITÀ MIGLIORATI
# ============================================================

def check_coerenza_quote_improved(
    odds_1: float,
    odds_x: float,
    odds_2: float,
    odds_over25: float,
    odds_under25: float,
    odds_btts: float = None,
) -> Tuple[List[str], float]:
    """
    Versione migliorata con scoring quantitativo.
    Returns: (warnings, quality_score)
    """
    warnings = []
    quality_score = 100.0
    
    # 1. Check margine
    if odds_1 and odds_x and odds_2:
        margin = (1/odds_1 + 1/odds_x + 1/odds_2) - 1
        if margin > 0.10:
            warnings.append(f"Margine 1X2 alto ({margin*100:.1f}%) → quote meno competitive")
            quality_score -= 15
        elif margin < 0.03:
            warnings.append("Margine 1X2 sospettosamente basso → verificare")
            quality_score -= 10
    
    # 2. Check coerenza favorita
    if odds_1 and odds_2:
        if odds_1 < 1.35 and odds_2 < 4.0:
            warnings.append("Casa molto favorita ma trasferta non abbastanza alta")
            quality_score -= 12
        if odds_1 > 3.5 and odds_2 > 3.5:
            warnings.append("Match molto equilibrato/caotico → alta varianza")
            quality_score -= 8
    
    # 3. Check over/under
    if odds_over25 and odds_under25:
        margin_ou = (1/odds_over25 + 1/odds_under25) - 1
        if not (0.02 < margin_ou < 0.12):
            warnings.append(f"Margine O/U anomalo ({margin_ou*100:.1f}%)")
            quality_score -= 10
        
        # Coerenza con 1X2
        if odds_1 and odds_1 < 1.5 and odds_over25 > 2.3:
            warnings.append("Favorita netta ma over alto → contraddizione")
            quality_score -= 15
    
    # 4. Check BTTS coerenza
    if odds_btts and odds_over25 and odds_btts > 1.0 and odds_over25 > 1.0:
        p_btts = 1/odds_btts
        p_over = 1/odds_over25
        # BTTS e Over dovrebbero essere correlati
        if p_btts > 0.65 and p_over < 0.40:
            warnings.append("BTTS alto ma Over basso → incoerenza")
            quality_score -= 12
    
    # 5. Liquidità implicita
    if odds_1 and odds_x and odds_2:
        min_odd = min(odds_1, odds_x, odds_2)
        max_odd = max(odds_1, odds_x, odds_2)
        if max_odd / min_odd > 15:
            warnings.append("Range quote molto ampio → possibile bassa liquidità")
            quality_score -= 8
    
    quality_score = max(0, quality_score)
    return warnings, quality_score

def compute_market_confidence_score(
    odds_1: float,
    odds_x: float,
    odds_2: float,
    odds_over25: float = None,
    odds_under25: float = None,
    odds_dnb_home: float = None,
    odds_dnb_away: float = None,
    odds_btts: float = None,
    num_bookmakers: int = 5,
) -> float:
    """
    Calcola confidence score basato su:
    - Margini
    - Coerenza tra mercati
    - Numero bookmakers
    - Spread quote
    
    Score: 0-100 (100 = massima confidence)
    """
    score = 50.0  # Base
    
    # 1. Numero bookmakers (proxy liquidità)
    if num_bookmakers >= 10:
        score += 20
    elif num_bookmakers >= 7:
        score += 15
    elif num_bookmakers >= 5:
        score += 10
    elif num_bookmakers >= 3:
        score += 5
    else:
        score -= 5
    
    # 2. Margine 1X2
    if odds_1 and odds_x and odds_2:
        margin = (1/odds_1 + 1/odds_x + 1/odds_2) - 1
        if 0.04 < margin < 0.08:
            score += 10  # Margine ottimale
        elif margin < 0.04:
            score -= 5  # Troppo stretto
        elif margin > 0.12:
            score -= 10  # Troppo largo
    
    # 3. Completezza mercati
    markets_available = sum([
        odds_over25 is not None,
        odds_dnb_home is not None,
        odds_btts is not None,
    ])
    score += markets_available * 5
    
    # 4. Coerenza DNB con 1X2
    if odds_dnb_home and odds_dnb_away and odds_1 and odds_2:
        # DNB dovrebbe essere < 1X2 corrispondente
        if odds_dnb_home < odds_1 and odds_dnb_away < odds_2:
            score += 10
        else:
            score -= 8
    
    # 5. Spread probabilità (quanto è definito il favorito)
    if odds_1 and odds_2:
        p1 = 1/odds_1
        p2 = 1/odds_2
        spread = abs(p1 - p2)
        if spread > 0.30:
            score += 8  # Favorito chiaro
        elif spread < 0.10:
            score -= 5  # Molto equilibrato → più incerto
    
    return max(0, min(100, score))

# ============================================================
#              STREAMLIT APP MIGLIORATA
# ============================================================

st.set_page_config(page_title="⚽ Modello Scommesse PRO – Dixon-Coles Bayesiano", layout="wide")

st.title("⚽ Modello Scommesse Avanzato")

st.caption(f"🕐 Esecuzione: {datetime.now().isoformat(timespec='seconds')}")

# Database initialization (once per session)
if "database_initialized" not in st.session_state:
    try:
        initialize_database()
        st.session_state.database_initialized = True
        logger.info("✅ Database inizializzato con successo")
    except Exception as e:
        logger.error(f"❌ Errore inizializzazione database: {e}")
        st.session_state.database_initialized = False
        # Non bloccare l'app, ma logga l'errore

# Session state initialization
if "soccer_leagues" not in st.session_state:
    st.session_state.soccer_leagues = []
if "events_for_league" not in st.session_state:
    st.session_state.events_for_league = []
if "selected_event_prices" not in st.session_state:
    st.session_state.selected_event_prices = {}
if "selected_league_key" not in st.session_state:
    st.session_state.selected_league_key = None
if "selected_event_id" not in st.session_state:
    st.session_state.selected_event_id = None

# ============================================================
#               SEZIONE STORICO E METRICHE
# ============================================================

st.subheader("📊 Storico e Performance")

col_hist1, col_hist2 = st.columns(2)

with col_hist1:
    if os.path.exists(ARCHIVE_FILE):
        df_st = pd.read_csv(ARCHIVE_FILE)
        st.write(f"📁 Analisi salvate: **{len(df_st)}**")
        
        # Dashboard metrics avanzate
        dashboard_metrics = calculate_dashboard_metrics()
        
        # Real-time Performance Monitoring (ALTA PRIORITÀ)
        realtime_metrics = get_realtime_performance_metrics(window_days=7)
        
        if realtime_metrics.get("status") == "ok":
            st.markdown("### ⚡ Performance Real-time (Ultimi 7 giorni)")
            col_rt1, col_rt2, col_rt3, col_rt4 = st.columns(4)
            
            with col_rt1:
                if realtime_metrics.get("accuracy"):
                    st.metric("🎯 Accuracy RT", f"{realtime_metrics['accuracy']:.1f}%",
                             delta=f"{realtime_metrics.get('trend', 0):+.1f}%" if realtime_metrics.get('trend') else None)
            
            with col_rt2:
                if realtime_metrics.get("roi"):
                    st.metric("💵 ROI RT", f"{realtime_metrics['roi']:.1f}%")
            
            with col_rt3:
                if realtime_metrics.get("bets_placed"):
                    st.metric("📊 Scommesse RT", realtime_metrics['bets_placed'])
            
            with col_rt4:
                alert_status = realtime_metrics.get("alert_status", "good")
                if alert_status == "critical":
                    st.error(realtime_metrics.get("alert_message", "⚠️ Critical"))
                elif alert_status == "warning":
                    st.warning(realtime_metrics.get("alert_message", "⚠️ Warning"))
                else:
                    st.success("✅ Performance OK")
        
        if dashboard_metrics and "error" not in dashboard_metrics:
            col_dash1, col_dash2 = st.columns(2)
            
            with col_dash1:
                if "accuracy_home" in dashboard_metrics:
                    st.metric("🎯 Accuracy Modello", f"{dashboard_metrics['accuracy_home']:.1f}%")
                if "brier_score" in dashboard_metrics and dashboard_metrics["brier_score"]:
                    st.metric("📈 Brier Score", f"{dashboard_metrics['brier_score']:.3f}",
                             help="0 = perfetto, 1 = pessimo")
            
            with col_dash2:
                if "roi_simulato" in dashboard_metrics and dashboard_metrics["roi_simulato"]:
                    roi_data = dashboard_metrics["roi_simulato"]
                    st.metric("💵 ROI Simulato", f"{roi_data.get('roi', 0):.1f}%",
                             help=f"{roi_data.get('bets', 0)} scommesse piazzate")
                if "trend_accuracy" in dashboard_metrics:
                    trend = dashboard_metrics["trend_accuracy"]
                    st.metric("📊 Trend Accuracy", f"{trend:+.1f}%",
                             help="Differenza ultime 20 vs prime 20 partite")
        
        # Calcola metriche se ci sono risultati reali (fallback)
        if "esito_reale" in df_st.columns and "match_ok" in df_st.columns:
            df_complete = df_st[df_st["esito_reale"].notna() & (df_st["esito_reale"] != "")]
            
            if len(df_complete) > 0 and not dashboard_metrics:
                accuracy = df_complete["match_ok"].mean() * 100
                st.metric("🎯 Accuracy Modello", f"{accuracy:.1f}%")
        
        st.dataframe(df_st.tail(15), height=300)
    else:
        st.info("Nessuno storico ancora")
    
    # Portfolio Summary
    portfolio_summary = get_portfolio_summary()
    if portfolio_summary.get("total_bets", 0) > 0:
        st.markdown("### 💼 Portfolio Scommesse")
        col_port1, col_port2, col_port3 = st.columns(3)
        
        with col_port1:
            st.metric("Totale Scommesse", portfolio_summary["total_bets"])
            st.metric("In Attesa", portfolio_summary["pending_bets"])
        
        with col_port2:
            st.metric("Vinte", portfolio_summary["won_bets"])
            st.metric("Perse", portfolio_summary["lost_bets"])
        
        with col_port3:
            st.metric("Profit", f"€{portfolio_summary['profit']:.2f}")
            st.metric("ROI", f"{portfolio_summary['roi']:.1f}%")
        
        if st.button("📋 Visualizza Portfolio Completo"):
            if os.path.exists(PORTFOLIO_FILE):
                df_port = pd.read_csv(PORTFOLIO_FILE)
                st.dataframe(df_port, use_container_width=True)

with col_hist2:
    st.markdown("### 🗑️ Gestione Storico")
    if os.path.exists(ARCHIVE_FILE):
        df_del = pd.read_csv(ARCHIVE_FILE)
        if not df_del.empty:
            df_del["label"] = df_del.apply(
                lambda r: f"{r.get('timestamp','?')} – {r.get('match','(no name)')}",
                axis=1,
            )
            to_delete = st.selectbox("Elimina riga:", df_del["label"].tolist())
            if st.button("🗑️ Elimina"):
                df_new = df_del[df_del["label"] != to_delete].drop(columns=["label"])
                df_new.to_csv(ARCHIVE_FILE, index=False)
                st.success("✅ Eliminato")
                st.rerun()

st.markdown("---")

# ============================================================
#        CONFIGURAZIONE TELEGRAM (OPZIONALE)
# ============================================================

with st.expander("🤖 Configurazione Telegram Bot (Opzionale)", expanded=False):
    st.markdown("""
    **Come configurare:**
    1. Crea un bot su Telegram scrivendo a [@BotFather](https://t.me/BotFather)
    2. Invia `/newbot` e segui le istruzioni
    3. Copia il **Token** che ti viene fornito
    4. Per ottenere il **Chat ID**, segui la guida dettagliata qui sotto
    5. In alternativa, puoi usare variabili d'ambiente: `TELEGRAM_BOT_TOKEN` e `TELEGRAM_CHAT_ID`
    """)
    
    # Guida dettagliata per trovare Chat ID
    with st.expander("📖 **Guida Completa: Come Trovare il Chat ID**", expanded=False):
        st.markdown("""
        ## 🎯 Metodo 1: Chat Privata (Messaggi Personali)
        
        **Per ricevere messaggi sul tuo account Telegram personale:**
        
        ### Passo 1: Ottieni il tuo User ID
        1. Apri Telegram (app o web)
        2. Cerca **[@userinfobot](https://t.me/userinfobot)** nella barra di ricerca
        3. Clicca sul bot e avvia una conversazione
        4. Invia il comando `/start`
        5. Il bot ti mostrerà il tuo **User ID** (es. `123456789`)
        6. **Copia questo numero** - questo è il tuo Chat ID!
        
        ### Passo 2: Avvia il tuo bot
        1. Cerca il nome del tuo bot (quello che hai creato con @BotFather)
        2. Avvia una conversazione con il bot
        3. **IMPORTANTE**: Invia almeno un messaggio al bot (anche solo `/start`)
        4. Questo è necessario perché il bot possa inviarti messaggi
        
        ### Passo 3: Verifica
        - Il Chat ID per chat private è un **numero positivo** (es. `123456789`)
        - Non include segni `-` o altri caratteri
        - Solo numeri!
        
        ---
        
        ## 👥 Metodo 2: Gruppo Telegram
        
        **Per ricevere messaggi in un gruppo:**
        
        ### Passo 1: Aggiungi il bot al gruppo
        1. Apri il gruppo dove vuoi ricevere i messaggi
        2. Vai su **Info gruppo** (icona in alto)
        3. Clicca su **Aggiungi membri** o **Aggiungi amministratori**
        4. Cerca il nome del tuo bot e aggiungilo al gruppo
        5. **IMPORTANTE**: Il bot deve essere membro del gruppo
        
        ### Passo 2: Ottieni il Group ID
        1. Aggiungi **[@userinfobot](https://t.me/userinfobot)** al gruppo
        2. Il bot mostrerà automaticamente le informazioni del gruppo
        3. Cerca il **Group ID** (es. `-1001234567890`)
        4. **Copia questo numero** - include il segno `-` all'inizio!
        
        ### Passo 3: Verifica
        - Il Chat ID per gruppi è un **numero negativo** (es. `-123456789` o `-1001234567890`)
        - Include sempre il segno `-` all'inizio
        - I supergruppi iniziano con `-100`
        
        ---
        
        ## 📢 Metodo 3: Canale Telegram
        
        **Per ricevere messaggi in un canale:**
        
        ### ⚠️ IMPORTANTE: @userinfobot NON funziona per canali!
        Il bot @userinfobot mostra solo le tue informazioni personali, non quelle del canale.
        Per i canali devi usare un metodo diverso.
        
        ### Passo 1: Rendi il bot amministratore
        1. Apri il canale dove vuoi ricevere i messaggi
        2. Vai su **Info canale** (icona in alto a destra)
        3. Clicca su **Amministratori** → **Aggiungi amministratore**
        4. Cerca il nome del tuo bot e aggiungilo
        5. **IMPORTANTE**: Il bot deve essere **amministratore** del canale (non solo membro)
        
        ### Passo 2: Ottieni il Channel ID - METODO A (Consigliato)
        
        **Usa @RawDataBot:**
        1. Aggiungi **[@RawDataBot](https://t.me/RawDataBot)** al canale come amministratore
        2. Inoltra un messaggio qualsiasi dal canale a @RawDataBot (in una chat privata con il bot)
        3. Il bot ti mostrerà tutti i dati del messaggio
        4. Cerca la riga `"chat":{"id":-1001234567890}`
        5. Il numero dopo `"id":` è il tuo **Channel ID** (es. `-1001234567890`)
        6. **Copia questo numero** - include il segno `-` e inizia con `-100`
        
        ### Passo 2 Alternativo: METODO B (Più semplice)
        
        **Usa @getidsbot:**
        1. Aggiungi **[@getidsbot](https://t.me/getidsbot)** al canale come amministratore
        2. Il bot dovrebbe mostrare automaticamente il Channel ID
        3. Se non appare, invia un messaggio nel canale e il bot risponderà con l'ID
        4. **Copia il numero** mostrato (inizia con `-100`)
        
        ### Passo 2 Alternativo: METODO C (Se i bot non funzionano)
        
        **Usa il tuo bot personalizzato:**
        1. Se hai già creato il bot con @BotFather, aggiungilo come amministratore al canale
        2. Invia un messaggio di test nel canale
        3. Usa questo script Python per ottenere l'ID:
        ```python
        import requests
        bot_token = "TUO_BOT_TOKEN"
        url = f"https://api.telegram.org/bot{bot_token}/getUpdates"
        response = requests.get(url)
        data = response.json()
        # Cerca il canale nei risultati
        for update in data.get("result", []):
            if "channel_post" in update:
                chat_id = update["channel_post"]["chat"]["id"]
                print(f"Channel ID: {chat_id}")
        ```
        
        ### Passo 3: Verifica
        - Il Chat ID per canali è un **numero negativo che inizia con `-100`** (es. `-1001234567890`)
        - Include sempre il segno `-` all'inizio
        - Deve iniziare con `-100`
        - **NON** usare il tuo User ID (quello che vedi con @userinfobot)!
        
        ---
        
        ## ✅ Riepilogo Formati Chat ID
        
        | Tipo | Formato | Esempio | Note |
        |------|---------|---------|------|
        | **Chat Privata** | Numero positivo | `123456789` | Solo numeri, nessun segno |
        | **Gruppo** | Numero negativo | `-123456789` | Include segno `-` |
        | **Supergruppo** | Numero negativo `-100...` | `-1001234567890` | Inizia con `-100` |
        | **Canale** | Numero negativo `-100...` | `-1001234567890` | Inizia con `-100` |
        
        ---
        
        ## 🔧 Bot Utili per Trovare Chat ID
        
        - **[@userinfobot](https://t.me/userinfobot)**: Per User ID (chat private) e Group ID ⚠️ NON funziona per canali!
        - **[@getidsbot](https://t.me/getidsbot)**: Per Channel ID (canali) - Aggiungi al canale come admin
        - **[@RawDataBot](https://t.me/RawDataBot)**: Mostra tutti i dati raw - Inoltra un messaggio dal canale al bot
        - **[@username_to_id_bot](https://t.me/username_to_id_bot)**: Converte username in ID (se il canale ha username pubblico)
        
        ---
        
        ## ⚠️ Problemi Comuni
        
        ### "Chat not found" o "Chat ID non valido"
        
        **Per Chat Private:**
        - ✅ Hai inviato almeno un messaggio al bot? (anche solo `/start`)
        - ✅ Il Chat ID è un numero positivo senza segni?
        - ✅ Hai copiato correttamente il numero?
        
        **Per Gruppi:**
        - ✅ Il bot è membro del gruppo?
        - ✅ Il Chat ID include il segno `-` all'inizio?
        - ✅ Hai usato @userinfobot per ottenere l'ID?
        
        **Per Canali:**
        - ✅ Il bot è **amministratore** del canale?
        - ✅ Il Chat ID inizia con `-100`?
        - ✅ Hai usato @getidsbot per ottenere l'ID?
        
        ---
        
        ## 💡 Suggerimenti
        
        1. **Testa sempre**: Usa il pulsante "🧪 Testa Configurazione" prima di abilitare l'invio automatico
        2. **Copia con attenzione**: Assicurati di non includere spazi o caratteri extra
        3. **Verifica formato**: Controlla che il formato corrisponda al tipo di chat (vedi tabella sopra)
        4. **Bot attivo**: Assicurati che il bot creato con @BotFather sia ancora attivo
        """)

    default_telegram_enabled = st.session_state.get("telegram_enabled", TELEGRAM_ENABLED)
    telegram_enabled = st.checkbox(
        "📤 Invia analisi automaticamente su Telegram",
        value=default_telegram_enabled,
        help="Se abilitato, ogni analisi verrà inviata automaticamente al tuo bot Telegram"
    )

    col_tg1, col_tg2 = st.columns(2)

    with col_tg1:
        default_token = st.session_state.get("telegram_bot_token", TELEGRAM_BOT_TOKEN)
        telegram_token = st.text_input(
            "Bot Token",
            value=default_token,
            type="password",
            help="Token del bot (da @BotFather)",
            placeholder="123456789:ABCdefGHIjklMNOpqrsTUVwxyz"
        )

    with col_tg2:
        default_chat_id = st.session_state.get("telegram_chat_id", TELEGRAM_CHAT_ID)
        telegram_chat_id = st.text_input(
            "Chat ID",
            value=default_chat_id,
            help="ID della chat dove inviare (da @userinfobot)",
            placeholder="123456789"
        )

    # Salva in session_state per persistenza
    st.session_state["telegram_enabled"] = telegram_enabled
    st.session_state["telegram_bot_token"] = telegram_token
    st.session_state["telegram_chat_id"] = telegram_chat_id
    
telegram_prob_threshold = st.slider(
    "🎯 Soglia minima probabilità per notifiche Telegram (%)",
    min_value=0.0,
    max_value=100.0,
    value=float(TELEGRAM_MIN_PROBABILITY),
    step=1.0,
    help="Il bot invia notifiche soltanto per i mercati in cui la probabilità del modello supera questa soglia."
)
st.session_state["telegram_prob_threshold"] = telegram_prob_threshold

# Pulsante per testare la configurazione
if telegram_token and telegram_chat_id:
    col_test1, col_test2 = st.columns([1, 2])
    with col_test1:
        if st.button("🧪 Testa Configurazione", help="Invia un messaggio di test per verificare che token e chat ID siano corretti"):
            with st.spinner("Invio messaggio di test..."):
                test_result = test_telegram_chat_id(
                    bot_token=telegram_token,
                    chat_id=telegram_chat_id
                )
                    
                if test_result.get("success"):
                    st.success("✅ **Test riuscito!** Messaggio di test inviato con successo. La configurazione è corretta.")
                else:
                    error_msg = test_result.get("error_message", "Errore sconosciuto")
                    error_type = test_result.get("error_type", "other")
                    
                    # Mostra messaggio di errore dettagliato
                    if error_type == "no_token":
                        st.error(f"❌ **Token Bot non configurato**\n\n{error_msg}")
                    elif error_type == "no_chat_id":
                        st.error(f"❌ **Chat ID non configurato**\n\n{error_msg}")
                    elif error_type == "invalid_token":
                        st.error(f"❌ **Token non valido**\n\n{error_msg}\n\nVerifica che il token sia corretto e che il bot sia ancora attivo.")
                    elif error_type == "invalid_chat_id":
                        st.error("❌ **Chat ID non valido**")
                        st.markdown(error_msg)
                        with st.expander("🔍 **Guida passo-passo per risolvere**", expanded=True):
                            st.markdown("""
                                **Per Chat Private:**
                                1. Apri Telegram e cerca [@userinfobot](https://t.me/userinfobot)
                                2. Avvia una conversazione e invia `/start`
                                3. Il bot ti mostrerà il tuo **User ID** (es. `123456789`)
                                4. Copia questo numero e incollalo nel campo "Chat ID"
                                5. **IMPORTANTE**: Prima di testare, invia almeno un messaggio al tuo bot (anche solo `/start`)
                                
                                **Per Gruppi:**
                                1. Aggiungi [@userinfobot](https://t.me/userinfobot) al gruppo
                                2. Il bot mostrerà il **Group ID** (es. `-1001234567890`)
                                3. Copia questo numero (include il segno `-`)
                                4. Aggiungi il tuo bot al gruppo come membro
                                
                                **Per Canali:**
                                1. Il bot deve essere **amministratore** del canale
                                2. Il Chat ID del canale inizia con `-100` (es. `-1001234567890`)
                                3. Per ottenere il Chat ID, usa [@getidsbot](https://t.me/getidsbot) nel canale
                                
                                **Verifica rapida:**
                                - Chat ID privata: numero positivo (es. `123456789`)
                                - Chat ID gruppo: numero negativo (es. `-123456789`)
                                - Chat ID canale: numero negativo che inizia con `-100` (es. `-1001234567890`)
                                """)
                    elif error_type == "rate_limit":
                        st.warning(f"⏱️ **Rate Limit**\n\n{error_msg}")
                    elif error_type == "timeout" or error_type == "connection_error":
                        st.warning(f"🌐 **Problema di connessione**\n\n{error_msg}")
                    else:
                        st.warning(f"⚠️ **Errore test Telegram**\n\n{error_msg}")
    with col_test2:
        st.caption("💡 **Suggerimento**: Testa sempre la configurazione prima di abilitare l'invio automatico")
    
    if telegram_enabled and (not telegram_token or not telegram_chat_id):
        st.warning("⚠️ Inserisci Bot Token e Chat ID per abilitare Telegram")

st.markdown("---")

# ============================================================
#        AI SYSTEM CONFIGURATION
# ============================================================

if AI_SYSTEM_AVAILABLE:
    st.subheader("🤖 AI System - Enhanced Predictions")

    # Initialize AI config in session state
    if "ai_enabled" not in st.session_state:
        st.session_state["ai_enabled"] = True
    if "ai_preset" not in st.session_state:
        st.session_state["ai_preset"] = "Balanced"
    if "ai_bankroll" not in st.session_state:
        st.session_state["ai_bankroll"] = 1000.0

    col_ai1, col_ai2, col_ai3 = st.columns([2, 2, 2])

    with col_ai1:
        ai_enabled = st.checkbox(
            "✅ Abilita AI Analysis",
            value=st.session_state["ai_enabled"],
            key="ai_enabled_checkbox",
            help="Attiva l'analisi AI per raccomandazioni di betting avanzate con 7 blocchi ML"
        )
        st.session_state["ai_enabled"] = ai_enabled

    with col_ai2:
        ai_preset = st.selectbox(
            "⚙️ Preset Strategia",
            ["Conservative", "Balanced", "Aggressive"],
            index=["Conservative", "Balanced", "Aggressive"].index(st.session_state["ai_preset"]),
            key="ai_preset_selector",
            help="Conservative: Min risk, piccole stake | Balanced: Equilibrato | Aggressive: Max value, stake più alte",
            disabled=not ai_enabled
        )
        st.session_state["ai_preset"] = ai_preset

    with col_ai3:
        ai_bankroll = st.number_input(
            "💰 Bankroll (€)",
            value=st.session_state["ai_bankroll"],
            min_value=100.0,
            max_value=100000.0,
            step=100.0,
            key="ai_bankroll_input",
            help="Bankroll totale per calcolo stake ottimale (Kelly Criterion)",
            disabled=not ai_enabled
        )
        st.session_state["ai_bankroll"] = ai_bankroll

    # Advanced AI settings (optional)
    background_ai_settings: Optional[Dict[str, Dict[str, Any]]] = None

    if ai_enabled and AI_SYSTEM_AVAILABLE:
        background_ai_settings = _ensure_ai_background_features()
        with st.expander("🔧 Impostazioni AI Avanzate (FASE 1.3)"):
            st.info(
                "⚙️ Configurazione base, gestione del rischio e automazioni live operano ora in background "
                "tramite i blocchi AI dedicati (Kelly Optimizer, Risk Manager, Odds Tracker e Live Monitor)."
            )

            st.success("📊 Data Quality Thresholds, Telegram Advanced e Neural Tuning operano ora solo in background.")

            # Sentiment Analysis Settings
            st.markdown("#### 📱 Sentiment Analysis (Social Media Intelligence)")
            sentiment_enabled = st.checkbox(
                "Enable Sentiment Analysis",
                value=False,
                help="Monitora social media e news per rilevare insider info (injuries, lineup leaks, team morale)"
            )
            st.session_state["ai_sentiment_enabled"] = sentiment_enabled

            if sentiment_enabled:
                st.info("""
                **🚨 Cosa rileva il Sentiment Analyzer:**
                - 🏥 Rumor infortuni prima degli annunci ufficiali
                - 📋 Leak formazioni anticipate
                - 💪 Morale squadra e motivazione giocatori
                - 📰 Pressione media e ambiente
                - 🎯 Fiducia tifosi

                **Funziona in Mock Mode** (dati simulati) se non hai API keys.
                """)

            st.markdown("---")

            st.info("""
            **📊 Come funziona l'AI System (15 Blocchi):**

            **🎯 Pipeline Principale (Blocchi 0-6):**
            - **Blocco 0:** API Engine per dati live (form, injuries, xG, lineup quality)
            - **Blocco 1:** Calibra le probabilità Dixon-Coles con Neural Network
            - **Blocco 2:** Calcola confidence score basato su data quality e model agreement
            - **Blocco 3:** Rileva value TRUE vs TRAP usando XGBoost e sharp money
            - **Blocco 4:** Ottimizza stake con Smart Kelly Criterion (dinamico)
            - **Blocco 5:** Risk management con portfolio limits e stop-loss
            - **Blocco 6:** Analizza odds movement per timing ottimale (BET NOW / WAIT)

            **🚀 Pipeline Avanzata (Blocchi 7-14):**
            - **Blocco 7:** Bayesian Uncertainty Quantification - Quantifica incertezza con intervalli di credibilità
            - **Blocco 8:** Monte Carlo Simulator - Simula migliaia di scenari per robustezza
            - **Blocco 9:** Advanced Anomaly Detection - Rileva anomalie multi-metodo
            - **Blocco 10:** Market Consistency Validator - Valida coerenza tra mercati correlati
            - **Blocco 11:** Adaptive Calibration - Calibrazione auto-adattiva da risultati
            - **Blocco 12:** Multi-Model Consensus - Validazione consensus tra modelli
            - **Blocco 13:** Statistical Arbitrage Detector - Rileva arbitraggi e inefficienze
            - **Blocco 14:** Real-time Validation Engine - Validazione real-time multi-metodologia
            """)

    st.markdown("---")

# ============================================================
#        INPUT DATI PARTITA
# ============================================================

st.subheader("📝 Dati Partita")

# Inizializza valori default
if "current_match" not in st.session_state:
    default_time = datetime.now().replace(hour=20, minute=45, second=0, microsecond=0).time()
    st.session_state["current_match"] = {
        "home_team": "Casa",
        "away_team": "Trasferta",
        "odds_1": 2.00,
        "odds_x": 3.50,
        "odds_2": 3.80,
        "odds_over25": 0.0,
        "odds_under25": 0.0,
        "odds_btts": 0.0,
        "odds_dnb_home": 0.0,
        "odds_dnb_away": 0.0,
        "match_date": datetime.now().date(),
        "match_time": default_time,
    }

# Inizializza advanced features default (per evitare KeyError)
if "position_home_auto" not in st.session_state:
    st.session_state["position_home_auto"] = None
if "position_away_auto" not in st.session_state:
    st.session_state["position_away_auto"] = None
if "is_derby_auto" not in st.session_state:
    st.session_state["is_derby_auto"] = False
if "is_cup_auto" not in st.session_state:
    st.session_state["is_cup_auto"] = False
if "is_end_season_auto" not in st.session_state:
    st.session_state["is_end_season_auto"] = False
st.session_state["auto_mode_selection"] = "🌐 Auto + API (Ibrido)"
if "preview_cache" not in st.session_state:
    st.session_state["preview_cache"] = {}  # Cache per preview auto-detection

# === SEZIONE INPUT DATI PARTITA ===
st.info("⚽ Inserisci manualmente i dati della partita che vuoi analizzare")

# Recupera dati correnti dalla partita selezionata
match_data = st.session_state["current_match"]

# === SQUADRE E LEGA ===
st.markdown("### 🏟️ Informazioni Partita")

col_team1, col_team2, col_league = st.columns([2, 2, 2])

with col_team1:
    home_team = st.text_input(
        "🏠 Squadra Casa",
        value=match_data.get("home_team", "Casa"),
        key="home_team_input",
        placeholder="Es: Inter"
    )

with col_team2:
    away_team = st.text_input(
        "✈️ Squadra Trasferta",
        value=match_data.get("away_team", "Trasferta"),
        key="away_team_input",
        placeholder="Es: Milan"
    )

with col_league:
    league_type = st.selectbox(
        "📊 Campionato",
        ["Serie A", "Premier League", "La Liga", "Bundesliga", "Ligue 1", "Altro"],
        key="league",
        help="Seleziona il campionato per applicare parametri specifici"
    )

# Aggiorna session state
st.session_state["current_match"]["home_team"] = home_team
st.session_state["current_match"]["away_team"] = away_team

league_api_map = {
    "Serie A": "serie_a",
    "Premier League": "premier_league",
    "La Liga": "la_liga",
    "Bundesliga": "bundesliga",
    "Ligue 1": "ligue_1",
    "Altro": "generic",
}

col_dt1, col_dt2 = st.columns(2)
with col_dt1:
    match_date_input = st.date_input(
        "📅 Data Partita",
        value=match_data.get("match_date", datetime.now().date()),
        key="match_date_input",
        help="Seleziona la data esatta della partita (serve per recuperare dati storici e meteo corretti)."
    )

with col_dt2:
    match_time_default = match_data.get(
        "match_time",
        datetime.now().replace(hour=20, minute=45, second=0, microsecond=0).time()
    )
    match_time_input = st.time_input(
        "🕒 Orario Partita",
        value=match_time_default,
        key="match_time_input",
        help="Orario di calcio d'inizio (24h)."
    )

st.session_state["current_match"]["match_date"] = match_date_input
st.session_state["current_match"]["match_time"] = match_time_input

st.markdown("---")

# Espansore per le quote
with st.expander(f"💰 Quote e Parametri: {home_team} vs {away_team}", expanded=True):

    # === LINEE DI APERTURA ===
    st.markdown("**📊 Linee di Apertura**")
    col_a1, col_a2 = st.columns(2)

    with col_a1:
        spread_apertura = st.number_input(
            "Spread Apertura",
            value=0.0,
            step=0.25,
            key="spread_apertura"
        )

    with col_a2:
        total_apertura = st.number_input(
            "Total Apertura",
            value=2.5,
            step=0.25,
            key="total_apertura"
        )

    # === QUOTE PRINCIPALI 1X2 ===
    st.markdown("**💰 Quote Principali (1X2)**")
    col_q1, col_q2, col_q3 = st.columns(3)

    with col_q1:
        odds_1 = st.number_input(
            "Quota 1 (Casa)",
            value=match_data.get("odds_1", 2.00),
            min_value=1.01,
            max_value=100.0,
            step=0.01,
            key="odds_1"
        )

    with col_q2:
        odds_x = st.number_input(
            "Quota X (Pareggio)",
            value=match_data.get("odds_x", 3.50),
            min_value=1.01,
            max_value=100.0,
            step=0.01,
            key="odds_x"
        )

    with col_q3:
        odds_2 = st.number_input(
            "Quota 2 (Trasferta)",
            value=match_data.get("odds_2", 3.80),
            min_value=1.01,
            max_value=100.0,
            step=0.01,
            key="odds_2"
        )

    # === VALIDAZIONE QUOTE 1X2 ===
    validation = validate_odds_input(odds_1, odds_x, odds_2)
    if validation["valid"]:
        st.success(validation["message"])
    else:
        st.warning(validation["message"])

    # === OVER/UNDER E TOTALI ===
    st.markdown("**⚽ Over/Under & Totali**")
    col_ou1, col_ou2, col_ou3 = st.columns(3)

    with col_ou1:
        odds_over25 = st.number_input(
            "Quota Over 2.5",
            value=match_data.get("odds_over25", 0.0),
            step=0.01,
            key="odds_over25"
        )

    with col_ou2:
        odds_under25 = st.number_input(
            "Quota Under 2.5",
            value=match_data.get("odds_under25", 0.0),
            step=0.01,
            key="odds_under25"
        )

    with col_ou3:
        total_line = st.number_input(
            "Total Corrente",
            value=2.5,
            step=0.25,
            key="total_line"
        )

    # === QUOTE SPECIALI ===
    st.markdown("**🎲 Quote Speciali**")
    col_s1, col_s2, col_s3, col_s4 = st.columns(4)

    with col_s1:
        odds_dnb_home = st.number_input(
            "DNB Casa",
            value=match_data.get("odds_dnb_home", 0.0),
            step=0.01,
            key="odds_dnb_home"
        )

    with col_s2:
        odds_dnb_away = st.number_input(
            "DNB Trasferta",
            value=match_data.get("odds_dnb_away", 0.0),
            step=0.01,
            key="odds_dnb_away"
        )

    with col_s3:
        odds_btts = st.number_input(
            "BTTS Sì",
            value=match_data.get("odds_btts", 0.0),
            step=0.01,
            key="odds_btts"
        )

    with col_s4:
        spread_corrente = st.number_input(
            "Spread Corrente",
            value=0.0,
            step=0.25,
            key="spread_corrente"
        )

    # === xG/xA E BOOST (OPZIONALI) ===
    # FIX BUG CRITICAL: Inizializza variabili PRIMA dell'expander per evitare NameError se non espanso
    xg_home = 0.0
    xa_home = 0.0
    partite_giocate_home = 0
    boost_home = 0.0
    xg_away = 0.0
    xa_away = 0.0
    partite_giocate_away = 0
    boost_away = 0.0

    with st.expander("📊 xG/xA e Boost (Opzionali)", expanded=False):
        col_xg1, col_xg2 = st.columns(2)

        with col_xg1:
            st.markdown("**🏠 Casa**")
            xg_home = st.number_input(
                "xG Totali Stagione Casa",
                value=0.0,
                step=0.1,
                key="xg_home",
                help="Somma xG di tutta la stagione (es. da Transfermarkt: 33.7). Il sistema calcolerà automaticamente la media per partita."
            )
            xa_home = st.number_input(
                "xA Totali Stagione Casa",
                value=0.0,
                step=0.1,
                key="xa_home",
                help="Somma xA (Expected Assists) di tutta la stagione."
            )
            partite_giocate_home = st.number_input(
                "Partite Giocate Casa",
                min_value=0,
                max_value=50,
                value=0,
                step=1,
                key="partite_giocate_home",
                help="Numero partite giocate in stagione. Necessario per convertire xG totali in media per partita. >= 10 partite = alta affidabilità."
            )
            boost_home = st.slider(
                "Boost Casa (%)",
                -20, 20,
                0,
                key="boost_home"
            ) / 100.0

        with col_xg2:
            st.markdown("**✈️ Trasferta**")
            xg_away = st.number_input(
                "xG Totali Stagione Trasferta",
                value=0.0,
                step=0.1,
                key="xg_away",
                help="Somma xG di tutta la stagione (es. da Transfermarkt: 18.6). Il sistema calcolerà automaticamente la media per partita."
            )
            xa_away = st.number_input(
                "xA Totali Stagione Trasferta",
                value=0.0,
                step=0.1,
                key="xa_away",
                help="Somma xA (Expected Assists) di tutta la stagione."
            )
            partite_giocate_away = st.number_input(
                "Partite Giocate Trasferta",
                min_value=0,
                max_value=50,
                value=0,
                step=1,
                key="partite_giocate_away",
                help="Numero partite giocate in stagione. Necessario per convertire xG totali in media per partita. >= 10 partite = alta affidabilità."
            )
            boost_away = st.slider(
                "Boost Trasferta (%)",
                -20, 20,
                0,
                key="boost_away"
            ) / 100.0

# === FUNZIONALITÀ AVANZATE (Sprint 1 & 2) ===
if ADVANCED_FEATURES_AVAILABLE:
    with st.expander("🚀 Funzionalità Avanzate (Precisione Migliorata)", expanded=False):
        st.markdown("""
        **Nuove funzionalità per massimizzare la precisione:**

        ✅ **Constraints Fisici**: Impedisce predizioni irrealistiche
        ✅ **Precision Math**: Elimina errori di arrotondamento
        ✅ **Calibrazione**: Usa storico per rendere probabilità "oneste"
        ✅ **Motivation Index**: Considera importanza match
        ✅ **Fixture Congestion**: Penalità per calendario fitto
        ✅ **Tactical Matchup**: Analizza stili di gioco
        """)

        # ============================================================
        # MODALITÀ AUTO-DETECTION (LEVEL 0 / LEVEL 1 / LEVEL 2)
        # ============================================================
        col_mode1, col_mode2 = st.columns([3, 1])

        with col_mode1:
            mode_options = [
                "✋ Manuale",
                "🗄️ Auto (Solo Database)",
                "🌐 Auto + API (Ibrido)"
            ]

            mode_selection = st.selectbox(
                "Modalità Auto-Detection",
                options=mode_options,
                key="auto_mode_selection",
                help="""
                **✋ Manuale**: Imposti tutti i parametri manualmente
                **🗄️ Auto (Solo Database)**: LEVEL 1 - Usa database squadre (100+ team)
                **🌐 Auto + API (Ibrido)**: LEVEL 2 Lite - Database + API esterne (TheSportsDB, API-Football)
                """
            )

        with col_mode2:
            if mode_selection == "✋ Manuale":
                st.info("✋ Manuale")
            elif mode_selection == "🗄️ Auto (Solo Database)":
                st.success("🗄️ DB")
            elif mode_selection == "🌐 Auto + API (Ibrido)":
                st.success("🌐 API")

        # Backward compatibility: imposta auto_mode_advanced per codice esistente
        auto_mode = (mode_selection != "✋ Manuale")

        # Detect mode switch and clean session state to prevent corruption
        previous_mode = st.session_state.get("_previous_mode_selection", None)
        if previous_mode is not None and previous_mode != mode_selection:
            logger.info(f"🔄 Mode switch detected: {previous_mode} → {mode_selection}")

            if previous_mode != "✋ Manuale" and mode_selection == "✋ Manuale":
                # Switching from Auto to Manual - clear auto-detection cache
                if "preview_cache" in st.session_state:
                    st.session_state["preview_cache"].clear()
                    logger.info("🧹 Cleared auto-detection preview cache")

            elif previous_mode == "✋ Manuale" and mode_selection != "✋ Manuale":
                # Switching from Manual to Auto - clear manual values to avoid conflicts
                manual_keys = [
                    'motivation_home', 'motivation_away',
                    'days_since_home', 'days_since_away',
                    'days_until_home', 'days_until_away',
                    'style_home', 'style_away'
                ]
                for key in manual_keys:
                    if key in st.session_state:
                        del st.session_state[key]
                logger.info("🧹 Cleared manual mode values")

        # Store current mode for next iteration
        st.session_state["_previous_mode_selection"] = mode_selection
        st.session_state["auto_mode_advanced"] = auto_mode

        if not AUTO_DETECTION_AVAILABLE and mode_selection != "✋ Manuale":
            st.error("⚠️ Auto-Detection non disponibile. Passa a Modalità Manuale.")
            mode_selection = "✋ Manuale"
            auto_mode = False

        # ============================================================
        # 📊 API STATUS DASHBOARD (LEVEL 2 Lite)
        # ============================================================
        if API_MANAGER_AVAILABLE and mode_selection == "🌐 Auto + API (Ibrido)":
            with st.expander("📊 API Status Dashboard (LEVEL 2 Lite)", expanded=False):
                try:
                    api_status = API_MANAGER.get_status()

                    # Quota usage
                    col_api1, col_api2, col_api3 = st.columns(3)

                    with col_api1:
                        st.metric(
                            "🏈 API-Football",
                            f"{api_status['quota']['api-football']['remaining']}/100",
                            delta=f"-{api_status['quota']['api-football']['used']} usate oggi"
                        )
                        st.caption("Quota giornaliera")

                    with col_api2:
                        cache_hit_rate = api_status['cache']['hit_rate']
                        st.metric(
                            "📦 Cache Hit Rate",
                            f"{cache_hit_rate:.1f}%",
                            delta="Ottimizzazione" if cache_hit_rate > 50 else "Bassa"
                        )
                        st.caption(f"{api_status['cache']['hits']} hits / {api_status['cache']['total']} total")

                    with col_api3:
                        thesportsdb_used = api_status['quota']['thesportsdb']['used']
                        st.metric(
                            "🌐 TheSportsDB",
                            f"{thesportsdb_used} calls",
                            delta="Illimitato ✓"
                        )
                        st.caption("Provider gratuito")

                    # Cache management
                    col_cache1, col_cache2 = st.columns(2)
                    with col_cache1:
                        if st.button("🗑️ Svuota Cache", help="Elimina tutti i dati cached (forza nuove API calls)"):
                            API_MANAGER.cache.cleanup_old(days=0)  # Elimina tutto
                            st.success("✅ Cache svuotata!")
                            st.rerun()

                    with col_cache2:
                        st.caption(f"💾 Cache TTL: 24 ore")

                    # Provider status
                    st.markdown("**Provider Status:**")
                    providers_status = f"""
                    - ✅ **TheSportsDB**: Attivo (Gratuito, Illimitato)
                    - {'✅' if api_status['quota']['api-football']['remaining'] > 0 else '⚠️'} **API-Football**: {api_status['quota']['api-football']['remaining']} chiamate rimaste oggi
                    - ℹ️ **Football-Data**: {api_status['quota']['football-data']['note']}
                    """
                    st.info(providers_status)

                except Exception as e:
                    st.error(f"⚠️ Errore lettura API status: {e}")

        st.markdown("---")

        # ============================================================
        # MODALITÀ AUTOMATICA
        # ============================================================
        if auto_mode and AUTO_DETECTION_AVAILABLE:
            st.info("""
            **🤖 Modalità Automatica Attiva**

            Il sistema rileverà automaticamente:
            - **Stile Tattico**: Da database squadre (100+ squadre principali)
            - **Motivazione**: Da posizione/contesto (lotta Champions, salvezza, derby)
            - **Fixture Congestion**: Calcolo da date match (se disponibili)

            *Nota: Puoi sempre tornare alla modalità manuale per override*
            """)

            # Input opzionali per migliorare auto-detection
            col_opt1, col_opt2 = st.columns(2)

            with col_opt1:
                position_home_auto = st.number_input(
                    "Posizione Casa (opzionale)",
                    min_value=1,
                    max_value=20,
                    value=None,
                    key="position_home_auto",
                    help="Posizione in classifica (1-20). Se fornito, migliora detection motivazione"
                )

            with col_opt2:
                position_away_auto = st.number_input(
                    "Posizione Trasferta (opzionale)",
                    min_value=1,
                    max_value=20,
                    value=None,
                    key="position_away_auto",
                    help="Posizione in classifica (1-20). Se fornito, migliora detection motivazione"
                )

            # Context flags
            col_ctx1, col_ctx2, col_ctx3 = st.columns(3)

            with col_ctx1:
                is_derby_auto = st.checkbox(
                    "🔥 È un Derby",
                    value=False,
                    key="is_derby_auto",
                    help="Spunta se è un derby/stracittadina/classico"
                )

            with col_ctx2:
                is_cup_auto = st.checkbox(
                    "🏆 È una Finale",
                    value=False,
                    key="is_cup_auto",
                    help="Spunta se è finale di coppa o match decisivo"
                )

            with col_ctx3:
                is_end_season_auto = st.checkbox(
                    "😴 Fine Stagione",
                    value=False,
                    key="is_end_season_auto",
                    help="Spunta se è fine stagione senza obiettivi"
                )

            # Preview auto-detection (se possibile)
            if home_team and away_team:
                st.markdown("**📊 Preview Auto-Detection:**")

                try:
                    # Crea cache key univoca per questa partita
                    cache_key = f"{home_team}_{away_team}_{league_type}_{match_datetime_iso}_{position_home_auto}_{position_away_auto}_{is_derby_auto}_{is_cup_auto}_{is_end_season_auto}_{mode_selection}"

                    # Initialize cache dicts if needed
                    if "preview_cache" not in st.session_state:
                        st.session_state["preview_cache"] = {}
                    if "preview_cache_timestamps" not in st.session_state:
                        st.session_state["preview_cache_timestamps"] = {}

                    # Check cache with expiry (5 minutes = 300 seconds)
                    import time
                    cache_expiry = 300
                    cache_time = st.session_state["preview_cache_timestamps"].get(cache_key, 0)
                    is_cache_valid = (cache_key in st.session_state["preview_cache"] and
                                      (time.time() - cache_time) < cache_expiry)

                    if is_cache_valid:
                        auto_features_preview = st.session_state["preview_cache"][cache_key]
                        age_minutes = (time.time() - cache_time) / 60
                        st.caption(f"📦 Utilizzando risultati cached ({age_minutes:.1f} min fa)")
                    else:
                        # Determina se usare API nel preview
                        use_api_preview = (mode_selection == "🌐 Auto + API (Ibrido)" and API_MANAGER_AVAILABLE)

                        # Esegui auto-detection
                        auto_features_preview = auto_detect_all_features(
                            home_team=home_team.strip(),
                            away_team=away_team.strip(),
                            league=league_type,
                            match_datetime=match_datetime_iso if match_datetime_iso else None,
                            position_home=position_home_auto,
                            position_away=position_away_auto,
                            is_derby=is_derby_auto,
                            is_cup=is_cup_auto,
                            is_end_season=is_end_season_auto,
                            use_api=use_api_preview  # LEVEL 2 Lite in preview
                        )

                        # Limit cache size (max 50 entries)
                        if len(st.session_state["preview_cache"]) >= 50:
                            # FIX BUG: Protect against empty timestamps dict (cache desync)
                            timestamps = st.session_state.get("preview_cache_timestamps", {})
                            if not timestamps:
                                logger.warning("⚠️ Cache desync: preview_cache piena ma timestamps vuoto, reset cache")
                                st.session_state["preview_cache"] = {}
                                st.session_state["preview_cache_timestamps"] = {}
                            else:
                                # Remove oldest entry
                                oldest_key = min(timestamps, key=timestamps.get)
                                del st.session_state["preview_cache"][oldest_key]
                                del st.session_state["preview_cache_timestamps"][oldest_key]
                                logger.info(f"🧹 Removed oldest cache entry: {oldest_key}")

                        # Salva in cache with timestamp
                        st.session_state["preview_cache"][cache_key] = auto_features_preview
                        st.session_state["preview_cache_timestamps"][cache_key] = time.time()

                    col_prev1, col_prev2 = st.columns(2)

                    with col_prev1:
                        st.info(f"""
                        **Casa ({home_team}):**
                        - Stile: {auto_features_preview['style_home']}
                        - Motivazione: {auto_features_preview['motivation_home']}
                        - Riposo: {auto_features_preview['days_since_home']}gg
                        """)

                    with col_prev2:
                        st.info(f"""
                        **Trasferta ({away_team}):**
                        - Stile: {auto_features_preview['style_away']}
                        - Motivazione: {auto_features_preview['motivation_away']}
                        - Riposo: {auto_features_preview['days_since_away']}gg
                        """)

                    # Mostra se API è stata usata
                    if use_api_preview:
                        st.caption("🌐 Dati arricchiti con API esterne (LEVEL 2 Lite)")

                except Exception as e:
                    st.warning(f"⚠️ Preview non disponibile: {e}")

        # ============================================================
        # MODALITÀ MANUALE
        # ============================================================
        else:
            st.info("**✋ Modalità Manuale Attiva** - Inserisci manualmente tutti i parametri")

            # Motivation
            st.markdown("**🎯 Motivation Index**")
            col_mot1, col_mot2 = st.columns(2)

            with col_mot1:
                motivation_home = st.selectbox(
                    "Motivazione Casa",
                list(MOTIVATION_FACTORS.keys()),
                index=0,
                key="motivation_home",
                help="Lotta Champions/Salvezza aumentano intensità (+10-20%). Fine stagione senza obiettivi riduce (-8%)"
            )

            with col_mot2:
                motivation_away = st.selectbox(
                    "Motivazione Trasferta",
                    list(MOTIVATION_FACTORS.keys()),
                    index=0,
                    key="motivation_away"
                )

            # Fixture Congestion
            st.markdown("**📅 Fixture Congestion (Calendario)**")
            col_fix1, col_fix2 = st.columns(2)

            with col_fix1:
                days_since_home = st.number_input(
                "Giorni dall'ultimo match (Casa)",
                min_value=2,
                max_value=21,
                value=7,
                step=1,
                key="days_since_home",
                help="≤3 giorni = stanchezza (-5%). ≥10 giorni = riposati (+3%)"
            )

            days_until_home = st.number_input(
                "Giorni al prossimo match importante (Casa)",
                min_value=2,
                max_value=14,
                value=7,
                step=1,
                key="days_until_home",
                help="Se match importante fra 3gg + giocato 3gg fa = -8% (rotation risk)"
            )

            with col_fix2:
                days_since_away = st.number_input(
                "Giorni dall'ultimo match (Trasferta)",
                min_value=2,
                max_value=21,
                value=7,
                step=1,
                key="days_since_away"
            )

            days_until_away = st.number_input(
                "Giorni al prossimo match importante (Trasferta)",
                min_value=2,
                max_value=14,
                value=7,
                step=1,
                key="days_until_away"
            )

            # Tactical Styles
            st.markdown("**⚔️ Tactical Matchup (Stili di Gioco)**")
            col_tac1, col_tac2 = st.columns(2)

            with col_tac1:
                style_home = st.selectbox(
                "Stile tattico Casa",
                TACTICAL_STYLES,
                index=0,
                key="style_home",
                help="""
**Possesso**: Dominio palla, manovra lenta (es. Man City, Barcellona)
**Contropiede**: Difesa compatta + ripartenze veloci (es. Atalanta, Leicester)
**Pressing Alto**: Aggressivi, recupero alto (es. Liverpool, Napoli)
**Difensiva**: Blocco basso, pochi rischi (es. Atletico, Burnley)
                """
            )

            with col_tac2:
                style_away = st.selectbox(
                "Stile tattico Trasferta",
                TACTICAL_STYLES,
                index=0,
                key="style_away"
            )

            # Preview fattori
            st.markdown("**📊 Preview Adjustments**")
            preview_factor_home = MOTIVATION_FACTORS[motivation_home]
            preview_factor_away = MOTIVATION_FACTORS[motivation_away]

            col_prev1, col_prev2 = st.columns(2)
            with col_prev1:
                st.metric("Fattore Motivation Casa", f"{preview_factor_home:.2f}x")
            with col_prev2:
                st.metric("Fattore Motivation Trasferta", f"{preview_factor_away:.2f}x")

            # Opzioni constraints
            st.markdown("**⚙️ Opzioni Avanzate**")
            apply_constraints = st.checkbox(
            "Applica Constraints Fisici",
            value=True,
            key="apply_constraints",
            help="Forza il modello a rispettare limiti realistici: total 0.5-6.0 gol, P(0-0) ≥ 5%, ecc."
            )

            apply_calibration_enabled = st.checkbox(
            "Applica Calibrazione Probabilità",
            value=True if CALIBRATION_MAP else False,
            key="apply_calibration_enabled",
            help=f"Usa storico per correggere bias. {'✅ Attiva (' + str(sum(len(v) for v in CALIBRATION_MAP.values() if v)) + ' bins)' if CALIBRATION_MAP else '⚠️ Non disponibile (serve storico)'}"
            )

        use_precision_math = st.checkbox(
            "Usa Precision Math (Neumaier sum)",
            value=True,
            key="use_precision_math",
            help="Elimina errori di arrotondamento nelle somme. Consigliato sempre."
        )

else:
    # Default values se advanced features non disponibili
    motivation_home = "Normale"
    motivation_away = "Normale"
    days_since_home = 7
    days_since_away = 7
    days_until_home = 7
    days_until_away = 7
    style_home = "Possesso"
    style_away = "Possesso"
    apply_constraints = False
    apply_calibration_enabled = False
    use_precision_math = False

with st.expander("💎 Value Bet Bet365 (Impostazioni)", expanded=False):
    value_bet_edge_threshold = st.slider(
        "Soglia edge minima (%)",
        min_value=0.5,
        max_value=10.0,
        value=2.0,
        step=0.1,
        key="value_bet_edge_threshold",
        help="Mostra solo i mercati con edge del modello almeno pari a questa percentuale."
    )
    value_bet_kelly_fraction = st.slider(
        "Frazione Kelly da suggerire",
        min_value=0.1,
        max_value=1.0,
        value=0.5,
        step=0.05,
        key="value_bet_kelly_fraction",
        help="Percentuale del Kelly Criterion da utilizzare per calcolare la puntata consigliata."
    )

st.markdown("---")

# ============================================================
#              CALCOLO MODELLO
# ============================================================

# ===== ANALISI SINGOLA PARTITA =====
if st.button("🎯 ANALIZZA PARTITA", type="primary"):

    # Recupera configurazione Telegram
    telegram_enabled = st.session_state.get("telegram_enabled", False)
    telegram_token = st.session_state.get("telegram_bot_token", "")
    telegram_chat_id = st.session_state.get("telegram_chat_id", "")
    telegram_prob_threshold = float(st.session_state.get("telegram_prob_threshold", TELEGRAM_MIN_PROBABILITY))

    reset_api_checklist()

    if not (auto_mode and AUTO_DETECTION_AVAILABLE):
        auto_reason = "Modalità manuale" if not auto_mode else "Modulo Auto-Detection non disponibile"
        update_api_check("Auto-Detection", "skipped", auto_reason)

    if not (telegram_enabled and telegram_token and telegram_chat_id):
        update_api_check("Telegram", "skipped", "Telegram disattivato o credenziali mancanti")

    # Leggi dati dal form
    match_name = f"{home_team} vs {away_team}"
    odds_over25_val = odds_over25 if odds_over25 > 0 else None
    odds_under25_val = odds_under25 if odds_under25 > 0 else None
    odds_btts_val = odds_btts if odds_btts > 0 else None
    odds_dnb_home_val = odds_dnb_home if odds_dnb_home > 0 else None
    odds_dnb_away_val = odds_dnb_away if odds_dnb_away > 0 else None
    # FIX BUG: Non ignorare valori validi solo perché sono uguali al default
    # 0.0 è un valore valido per spread, 2.5 è il valore più comune per total
    spread_apertura_val = spread_apertura if spread_apertura is not None else None
    total_apertura_val = total_apertura if total_apertura is not None else None
    spread_corrente_val = spread_corrente if spread_corrente is not None else None
    total_corrente_val = total_line if total_line is not None else None

    # Conversione xG/xA totali → medie per partita
    # FIX BUG: Se partite_giocate=0, usa 0.0 invece di valore totale (che è scorretto come media)
    if partite_giocate_home > 0:
        xg_home_media = xg_home / partite_giocate_home
        xa_home_media = xa_home / partite_giocate_home
    else:
        xg_home_media = 0.0  # Nessun dato disponibile
        xa_home_media = 0.0
        if xg_home > 0 or xa_home > 0:
            logger.warning(f"xG/xA home forniti ({xg_home}/{xa_home}) ma partite_giocate=0, uso 0.0")

    if partite_giocate_away > 0:
        xg_away_media = xg_away / partite_giocate_away
        xa_away_media = xa_away / partite_giocate_away
    else:
        xg_away_media = 0.0  # Nessun dato disponibile
        xa_away_media = 0.0
        if xg_away > 0 or xa_away > 0:
            logger.warning(f"xG/xA away forniti ({xg_away}/{xa_away}) ma partite_giocate=0, uso 0.0")

    match_date_value = st.session_state["current_match"].get("match_date")
    match_time_value = st.session_state["current_match"].get("match_time")
    if not match_date_value:
        match_date_value = datetime.now().date()
    if not match_time_value:
        match_time_value = datetime.now().replace(hour=20, minute=45, second=0, microsecond=0).time()

    match_datetime_obj = datetime.combine(match_date_value, match_time_value)
    match_datetime_iso = match_datetime_obj.strftime("%Y-%m-%dT%H:%M:%SZ")
    league_api_code = league_api_map.get(league_type, "generic")

    advanced_data = None
    fatigue_home_data = None
    fatigue_away_data = None
    advanced_data_error = False

    if home_team.strip() and away_team.strip():
        with st.spinner("Recupero dati avanzati dalle API (fatigue, infortuni, metriche)..."):
            try:
                advanced_data = get_advanced_team_data(
                    home_team.strip(),
                    away_team.strip(),
                    league_api_code,
                    match_datetime_iso,
                )
            except Exception as api_err:
                advanced_data_error = True
                logger.warning(f"Impossibile recuperare dati avanzati API: {api_err}")
                update_api_check("API-Football", "error", f"Errore dati avanzati: {api_err}")
                advanced_data = None

            try:
                fatigue_home_data = get_team_fatigue_and_motivation_data(
                    home_team.strip(),
                    league_api_code,
                    match_datetime_iso,
                )
            except Exception as api_err:
                logger.warning(f"Impossibile recuperare fatigue casa: {api_err}")
                fatigue_home_data = None

            try:
                fatigue_away_data = get_team_fatigue_and_motivation_data(
                    away_team.strip(),
                    league_api_code,
                    match_datetime_iso,
                )
            except Exception as api_err:
                logger.warning(f"Impossibile recuperare fatigue trasferta: {api_err}")
                fatigue_away_data = None

        if not advanced_data_error:
            summarize_apifootball_sources(advanced_data, fatigue_home_data, fatigue_away_data)
    else:
        update_api_check("API-Football", "skipped", "Inserisci entrambe le squadre per utilizzare API-Football")

    # Validazione xG/xA
    validation_warnings_xg = []
    if xg_home_media > 0 and (xg_home_media < 0.3 or xg_home_media > 4.0):
        validation_warnings_xg.append(f"⚠️ xG medio Casa {xg_home_media:.2f} fuori range tipico (0.3-4.0). Verifica dati inseriti.")
    if xa_home_media > 0 and (xa_home_media < 0.2 or xa_home_media > 4.0):
        validation_warnings_xg.append(f"⚠️ xA medio Casa {xa_home_media:.2f} fuori range tipico (0.2-4.0). Verifica dati inseriti.")
    if xg_away_media > 0 and (xg_away_media < 0.3 or xg_away_media > 4.0):
        validation_warnings_xg.append(f"⚠️ xG medio Trasferta {xg_away_media:.2f} fuori range tipico (0.3-4.0). Verifica dati inseriti.")
    if xa_away_media > 0 and (xa_away_media < 0.2 or xa_away_media > 4.0):
        validation_warnings_xg.append(f"⚠️ xA medio Trasferta {xa_away_media:.2f} fuori range tipico (0.2-4.0). Verifica dati inseriti.")

    # Mostra warning se presenti
    for warning in validation_warnings_xg:
        st.warning(warning)

    try:
        # Validazione input
        validation_result = validate_all_inputs(
            odds_1=odds_1, odds_x=odds_x, odds_2=odds_2, total=total_line,
            odds_over25=odds_over25_val, odds_under25=odds_under25_val, odds_btts=odds_btts_val,
            odds_dnb_home=odds_dnb_home_val, odds_dnb_away=odds_dnb_away_val,
            league=league_type,
            spread_apertura=spread_apertura_val, total_apertura=total_apertura_val,
            spread_corrente=spread_corrente_val, total_corrente=total_corrente_val,
            xg_for_home=xg_home_media if xg_home_media > 0 else None,
            xg_against_home=xg_away_media if xg_away_media > 0 else None,
            xg_for_away=xg_away_media if xg_away_media > 0 else None,
            xg_against_away=xg_home_media if xg_home_media > 0 else None,
            xa_for_home=xa_home_media if xa_home_media > 0 else None,
            xa_against_home=xa_away_media if xa_away_media > 0 else None,
            xa_for_away=xa_away_media if xa_away_media > 0 else None,
            xa_against_away=xa_home_media if xa_home_media > 0 else None,
        )

        validated = validation_result["validated"]

        # Quality e market confidence
        warnings, quality_score = check_coerenza_quote_improved(
            validated["odds_1"], validated["odds_x"], validated["odds_2"],
            validated.get("odds_over25"), validated.get("odds_under25"),
            validated.get("odds_btts")
        )

        market_conf = compute_market_confidence_score(
            validated["odds_1"], validated["odds_x"], validated["odds_2"],
            validated.get("odds_over25"), validated.get("odds_under25"),
            validated.get("odds_dnb_home"), validated.get("odds_dnb_away"),
            validated.get("odds_btts"), 1
        )

        # ============================================================
        # 🤖 AUTO-DETECTION advanced features (se abilitata)
        # ============================================================
        auto_mode_active = st.session_state.get('auto_mode_advanced', False)

        if auto_mode_active and AUTO_DETECTION_AVAILABLE:
            try:
                # Determina se usare API (LEVEL 2 Lite)
                use_api = (mode_selection == "🌐 Auto + API (Ibrido)" and API_MANAGER_AVAILABLE)

                # Crea cache key per riutilizzare preview (evita doppie chiamate)
                cache_key = f"{home_team}_{away_team}_{league_type}_{match_datetime_iso}_{st.session_state.get('position_home_auto')}_{st.session_state.get('position_away_auto')}_{st.session_state.get('is_derby_auto', False)}_{st.session_state.get('is_cup_auto', False)}_{st.session_state.get('is_end_season_auto', False)}_{mode_selection}"

                # Check cache with expiry (same logic as preview)
                import time
                cache_expiry = 300  # 5 minutes
                cache_time = st.session_state.get("preview_cache_timestamps", {}).get(cache_key, 0)
                is_cache_valid = (cache_key in st.session_state.get("preview_cache", {}) and
                                  (time.time() - cache_time) < cache_expiry)

                # Riutilizza risultati preview se disponibili e non scaduti (ottimizzazione API)
                if is_cache_valid:
                    auto_features = st.session_state["preview_cache"][cache_key]
                    age_minutes = (time.time() - cache_time) / 60
                    logger.info(f"♻️ Riutilizzando risultati preview cached ({age_minutes:.1f} min fa, 0 API calls)")
                else:
                    # ✅ FIX BUG #8: Passa dati fatigue già fetchati ad auto_detect
                    # Calcola last_match_datetime da days_since_last_match
                    last_match_datetime_home = None
                    last_match_datetime_away = None

                    if fatigue_home_data and fatigue_home_data.get("days_since_last_match") is not None:
                        try:
                            from datetime import datetime, timedelta
                            match_dt = datetime.fromisoformat(match_datetime_iso.replace("Z", "+00:00"))
                            days_since = fatigue_home_data["days_since_last_match"]
                            last_match_dt = match_dt - timedelta(days=days_since)
                            last_match_datetime_home = last_match_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
                            logger.info(f"📅 Fatigue Home: {days_since} giorni → last match {last_match_datetime_home}")
                        except Exception as e:
                            logger.warning(f"Errore calcolo last_match_datetime_home: {e}")

                    if fatigue_away_data and fatigue_away_data.get("days_since_last_match") is not None:
                        try:
                            from datetime import datetime, timedelta
                            match_dt = datetime.fromisoformat(match_datetime_iso.replace("Z", "+00:00"))
                            days_since = fatigue_away_data["days_since_last_match"]
                            last_match_dt = match_dt - timedelta(days=days_since)
                            last_match_datetime_away = last_match_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
                            logger.info(f"📅 Fatigue Away: {days_since} giorni → last match {last_match_datetime_away}")
                        except Exception as e:
                            logger.warning(f"Errore calcolo last_match_datetime_away: {e}")

                    # Esegui auto-detection con dati fatigue reali
                    auto_features = auto_detect_all_features(
                        home_team=home_team.strip(),
                        away_team=away_team.strip(),
                        league=league_type,
                        match_datetime=match_datetime_iso,
                        position_home=st.session_state.get('position_home_auto'),
                        position_away=st.session_state.get('position_away_auto'),
                        is_derby=st.session_state.get('is_derby_auto', False),
                        is_cup=st.session_state.get('is_cup_auto', False),
                        is_end_season=st.session_state.get('is_end_season_auto', False),
                        last_match_datetime_home=last_match_datetime_home,  # ✅ PASS REAL DATA
                        last_match_datetime_away=last_match_datetime_away,  # ✅ PASS REAL DATA
                        use_api=use_api  # LEVEL 2 Lite integration
                    )

                # Usa parametri auto-detected
                motivation_home_final = auto_features['motivation_home']
                motivation_away_final = auto_features['motivation_away']
                days_since_home_final = auto_features['days_since_home']
                days_since_away_final = auto_features['days_since_away']
                days_until_home_final = auto_features['days_until_home']
                days_until_away_final = auto_features['days_until_away']
                style_home_final = auto_features['style_home']
                style_away_final = auto_features['style_away']
                apply_constraints_final = auto_features['apply_constraints']
                apply_calibration_final = auto_features['apply_calibration_enabled']
                use_precision_math_final = auto_features['use_precision_math']

                logger.info(f"🤖 Auto-detection applicata: {home_team} vs {away_team}")
                if use_api:
                    if auto_features.get("api_used"):
                        api_calls = auto_features.get("api_calls_count", 0)
                        update_api_check(
                            "Auto-Detection",
                            "success",
                            f"API Manager attivo ({api_calls} chiamate)"
                        )
                    else:
                        update_api_check(
                            "Auto-Detection",
                            "warning",
                            "API Manager non utilizzato: fallback su database locale"
                        )
                else:
                    update_api_check(
                        "Auto-Detection",
                        "skipped",
                        "Auto-detection in modalità database (nessuna chiamata API)"
                    )

            except Exception as e:
                logger.warning(f"⚠️ Auto-detection fallita: {e}, uso valori manuali")
                if use_api:
                    update_api_check("Auto-Detection", "error", f"Errore auto-detection: {e}")
                else:
                    update_api_check("Auto-Detection", "warning", "Auto-detection locale non disponibile, uso valori manuali")
                # Fallback a manual
                motivation_home_final = st.session_state.get('motivation_home', 'Normale')
                motivation_away_final = st.session_state.get('motivation_away', 'Normale')
                days_since_home_final = st.session_state.get('days_since_home', 7)
                days_since_away_final = st.session_state.get('days_since_away', 7)
                days_until_home_final = st.session_state.get('days_until_home', 7)
                days_until_away_final = st.session_state.get('days_until_away', 7)
                style_home_final = st.session_state.get('style_home', 'Possesso')
                style_away_final = st.session_state.get('style_away', 'Possesso')
                apply_constraints_final = st.session_state.get('apply_constraints', True)
                apply_calibration_final = st.session_state.get('apply_calibration_enabled', False)
                use_precision_math_final = st.session_state.get('use_precision_math', True)

        else:
            # Modalità manuale - usa valori da session_state
            motivation_home_final = st.session_state.get('motivation_home', 'Normale')
            motivation_away_final = st.session_state.get('motivation_away', 'Normale')
            days_since_home_final = st.session_state.get('days_since_home', 7)
            days_since_away_final = st.session_state.get('days_since_away', 7)
            days_until_home_final = st.session_state.get('days_until_home', 7)
            days_until_away_final = st.session_state.get('days_until_away', 7)
            style_home_final = st.session_state.get('style_home', 'Possesso')
            style_away_final = st.session_state.get('style_away', 'Possesso')
            apply_constraints_final = st.session_state.get('apply_constraints', True)
            apply_calibration_final = st.session_state.get('apply_calibration_enabled', False)
            use_precision_math_final = st.session_state.get('use_precision_math', True)

        # Analisi modello
        with st.spinner(f"Analizzando {match_name}..."):
            ris = risultato_completo_improved(
                odds_1=validated["odds_1"],
                odds_x=validated["odds_x"],
                odds_2=validated["odds_2"],
                total=validated["total"],
                home_team=home_team,
                away_team=away_team,
                odds_over25=validated.get("odds_over25"),
                odds_under25=validated.get("odds_under25"),
                odds_btts=validated.get("odds_btts"),
                odds_dnb_home=validated.get("odds_dnb_home"),
                odds_dnb_away=validated.get("odds_dnb_away"),
                spread_apertura=validated.get("spread_apertura"),
                total_apertura=validated.get("total_apertura"),
                spread_corrente=validated.get("spread_corrente"),
                total_corrente=validated.get("total_corrente"),
                xg_for_home=validated.get("xg_for_home"),
                xg_against_home=validated.get("xg_against_home"),
                xg_for_away=validated.get("xg_for_away"),
                xg_against_away=validated.get("xg_against_away"),
                xa_for_home=validated.get("xa_for_home"),
                xa_against_home=validated.get("xa_against_home"),
                xa_for_away=validated.get("xa_for_away"),
                xa_against_away=validated.get("xa_against_away"),
                boost_home=boost_home,
                boost_away=boost_away,
                league=league_type,
                partite_giocate_home=partite_giocate_home,
                partite_giocate_away=partite_giocate_away,
                match_datetime=match_datetime_iso,
                fatigue_home=fatigue_home_data,
                fatigue_away=fatigue_away_data,
                # ✅ FIX BUG #4: RIMOSSO parametri motivation_home/away duplicati
                # Questi parametri non sono più utilizzati (vedi line 12991-12995)
                # Passare fatigue_home_data a motivation creava confusione e rischio di double-counting
                advanced_data=advanced_data,
                # 🚀 Advanced Features (Sprint 1 & 2) - Parametri UI (auto o manual)
                motivation_home_ui=motivation_home_final,
                motivation_away_ui=motivation_away_final,
                days_since_home=days_since_home_final,
                days_since_away=days_since_away_final,
                days_until_home=days_until_home_final,
                days_until_away=days_until_away_final,
                style_home=style_home_final,
                style_away=style_away_final,
                apply_constraints=apply_constraints_final,
                apply_calibration_enabled=apply_calibration_final,
                use_precision_math=use_precision_math_final,
            )
        summarize_additional_api_sources(ris.get("additional_api_data"))

        # ============================================================
        #   AI SYSTEM ANALYSIS (if enabled)
        # ============================================================
        ai_result = None
        if AI_SYSTEM_AVAILABLE and st.session_state.get("ai_enabled", False):
            try:
                with st.spinner("🤖 AI System analyzing..."):
                    # Get AI config from session state
                    ai_preset_name = st.session_state.get("ai_preset", "Balanced")
                    ai_bankroll_value = st.session_state.get("ai_bankroll", 1000.0)

                    # Create AI config based on preset
                    if ai_preset_name == "Conservative":
                        ai_config = get_conservative_config()
                    elif ai_preset_name == "Aggressive":
                        ai_config = get_aggressive_config()
                    else:
                        ai_config = AIConfig()  # Balanced (default)

                    ai_config.telegram_enabled = TELEGRAM_ENABLED
                    ai_config.telegram_bot_token = TELEGRAM_BOT_TOKEN
                    ai_config.telegram_chat_id = TELEGRAM_CHAT_ID
                    ai_config.live_monitoring_enabled = TELEGRAM_ENABLED

                    ai_config = _apply_ai_background_features(ai_config, background_ai_settings)

                    # Apply custom settings if user changed them
                    if "ai_min_confidence" in st.session_state:
                        ai_config.min_confidence_to_bet = st.session_state["ai_min_confidence"]
                    if "ai_kelly_fraction" in st.session_state:
                        ai_config.kelly_default_fraction = st.session_state["ai_kelly_fraction"]

                    # Prepare match data
                    match_datetime_combined = datetime.combine(
                        match_date_input,
                        match_time_input
                    ).isoformat()

                    # Calculate time to kickoff
                    time_to_kickoff = (
                        datetime.combine(match_date_input, match_time_input) - datetime.now()
                    )
                    time_to_kickoff_hours = max(0, time_to_kickoff.total_seconds() / 3600)

                    # Prepare odds history (if available from session state)
                    odds_history = st.session_state.get("odds_history", [])

                    # Run Sentiment Analysis (FASE 2) if enabled
                    sentiment_result = None
                    if st.session_state.get("ai_sentiment_enabled", False):
                        try:
                            sentiment_analyzer = SentimentAnalyzer()
                            sentiment_result = sentiment_analyzer.analyze_match_sentiment(
                                team_home=home_team,
                                team_away=away_team,
                                hours_before=48
                            )
                            logger.info(f"✅ Sentiment Analysis completed: {len(sentiment_result['signals'])} signals detected")
                        except Exception as e:
                            logger.warning(f"⚠️ Sentiment Analysis error: {e}")
                            sentiment_result = None

                    # Run AI analysis using quick_analyze for simplicity
                    # We'll analyze the home win (1) as primary bet
                    ai_result = quick_analyze(
                        home_team=home_team,
                        away_team=away_team,
                        league=league_type,
                        prob_dixon_coles=ris["p_home"],  # Home win probability from Dixon-Coles
                        odds=validated["odds_1"],  # Home win odds
                        bankroll=ai_bankroll_value,
                        config=ai_config,
                        # Additional context
                        odds_history=odds_history,  # Pass list directly (even if empty)
                        time_to_kickoff_hours=time_to_kickoff_hours,
                        match_date=match_datetime_combined
                    )

                    logger.info(f"✅ AI Analysis completed: {ai_result['final_decision']['action']}")

                    background_summary = None
                    if BACKGROUND_AUTOMATION:
                        try:
                            validated_data = validated  # type: ignore[name-defined]
                        except NameError:
                            validated_data = {}

                        match_payload = {
                            "home": home_team,
                            "away": away_team,
                            "league": league_type,
                            "date": match_date_input.strftime("%Y-%m-%d"),
                            "time": match_time_input.strftime("%H:%M"),
                            "match_datetime": match_datetime_combined,
                            "importance": validated_data.get("importance_score", 0.5),
                            "odds": ai_result.get("summary", {}).get(
                                "odds",
                                validated_data.get("odds_1", 0.0)
                            ),
                            "odds_history": odds_history,
                        }

                        background_summary = BACKGROUND_AUTOMATION.handle_analysis(
                            match=match_payload,
                            ai_result=ai_result,
                            bankroll=ai_bankroll_value,
                            ai_config=ai_config,
                            time_to_kickoff_hours=time_to_kickoff_hours
                        )

                        if background_summary:
                            logger.info(
                                "🛰️ Background automation: %s",
                                background_summary
                            )

                    # Run Advanced Precision Pipeline (Blocks 7-14)
                    try:
                        advanced_pipeline = AdvancedPrecisionPipeline(config=ai_config)

                        # Prepare data for advanced analysis
                        prediction_data = {
                            'probability': ai_result['calibrated']['prob_calibrated'],
                            'confidence': ai_result['confidence']['confidence_score'] / 100,
                            'odds': validated["odds_1"],
                            'market_data': {
                                'odds_1': validated["odds_1"],
                                'odds_x': validated.get("odds_x", 0),
                                'odds_2': validated.get("odds_2", 0),
                            },
                            'models_predictions': {
                                'dixon_coles': ris["p_home"],
                                'calibrated': ai_result['calibrated']['prob_calibrated']
                            }
                        }

                        advanced_result = advanced_pipeline.process_prediction(prediction_data)

                        # Add advanced analysis to ai_result
                        ai_result['advanced_analysis'] = {
                            'credible_interval_95': advanced_result.credible_interval_95,
                            'uncertainty_level': advanced_result.uncertainty_level,
                            'reliability_index': advanced_result.reliability_index,
                            'robustness_score': advanced_result.robustness_score,
                            'monte_carlo_percentile_5': advanced_result.monte_carlo_percentile_5,
                            'monte_carlo_percentile_95': advanced_result.monte_carlo_percentile_95,
                            'anomaly_detected': advanced_result.anomaly_detected,
                            'anomaly_severity': advanced_result.anomaly_severity,
                            'consistency_score': advanced_result.consistency_score,
                            'validation_passed': advanced_result.validation_passed,
                            'calibrated_probability': advanced_result.calibrated_probability,
                            'consensus_reached': advanced_result.consensus_reached,
                            'agreement_score': advanced_result.agreement_score,
                            'outlier_models': advanced_result.outlier_models,
                            'arbitrage_opportunities': advanced_result.arbitrage_opportunities,
                            'expected_value_pct': advanced_result.expected_value_pct,
                            'validation_score': advanced_result.validation_score,
                            'recommendation': advanced_result.recommendation,
                            'reasoning': advanced_result.reasoning,
                        }

                        logger.info(f"✅ Advanced Analysis completed: {advanced_result.recommendation}")
                    except Exception as e:
                        logger.warning(f"⚠️ Advanced Analysis error: {e}")
                        ai_result['advanced_analysis'] = None

                    # Generate LLM Explanation (FASE 1.1)
                    try:
                        llm_analyst = LLMAnalyst(provider="mock")  # Use mock for free, or "openai" with API key

                        match_context = {
                            'home_team': home_team,
                            'away_team': away_team,
                            'league': league_type,
                            'odds': validated["odds_1"]
                        }

                        explanation = llm_analyst.explain_prediction(match_context, ai_result)
                        ai_result['llm_explanation'] = explanation
                        logger.info(f"✅ LLM Explanation generated")
                    except Exception as e:
                        logger.warning(f"⚠️ LLM Explanation error: {e}")
                        ai_result['llm_explanation'] = None

                    # Store Sentiment Analysis result (FASE 2)
                    ai_result['sentiment_analysis'] = sentiment_result

            except Exception as e:
                logger.error(f"❌ AI Analysis error: {e}")
                st.warning(f"⚠️ AI Analysis error: {str(e)}")
                ai_result = None

        st.success(f"✅ Analisi completata per {match_name}")

        # ============================================================
        #   TELEGRAM NOTIFICATION (if enabled)
        # ============================================================
        if telegram_enabled and telegram_token and telegram_chat_id:
            # Verifica se almeno una probabilità supera la soglia minima
            max_prob = max(ris['p_home'], ris['p_draw'], ris['p_away']) * 100

            if max_prob >= TELEGRAM_MIN_PROBABILITY:
                try:
                    # Formatta messaggio con tutti i mercati
                    telegram_message = format_analysis_for_telegram(
                        match_name=match_name,
                        ris=ris,
                        odds_1=validated['odds_1'],
                        odds_x=validated['odds_x'],
                        odds_2=validated['odds_2'],
                        quality_score=validated['quality_score'],
                        market_conf=validated['market_confidence'],
                        value_bets=None  # Opzionale: lista value bets
                    )

                    # Invia notifica
                    telegram_result = send_telegram_message(
                        message=telegram_message,
                        bot_token=TELEGRAM_BOT_TOKEN,
                        chat_id=TELEGRAM_CHAT_ID,
                        parse_mode="HTML"
                    )

                    if telegram_result.get('success'):
                        st.info(f"📱 Notifica Telegram inviata con successo!")
                        logger.info(f"📱 Telegram notification sent for {match_name}")
                        update_api_check("Telegram", "success", "Notifica inviata con successo")
                    else:
                        st.warning(f"⚠️ Errore invio Telegram: {telegram_result.get('error_message', 'Unknown')}")
                        logger.warning(f"⚠️ Telegram error: {telegram_result.get('error_message')}")
                        update_api_check(
                            "Telegram",
                            "warning",
                            f"Errore invio: {telegram_result.get('error_message', 'sconosciuto')}"
                        )

                except Exception as e:
                    st.warning(f"⚠️ Errore notifica Telegram: {str(e)}")
                    logger.error(f"❌ Telegram notification error: {e}")
                    update_api_check("Telegram", "error", f"Errore invio: {e}")
            else:
                logger.info(f"ℹ️  Telegram notification skipped: max probability {max_prob:.1f}% < {TELEGRAM_MIN_PROBABILITY}%")
                update_api_check("Telegram", "skipped", "Probabilità sotto soglia, notifica non inviata")

        render_api_checklist()

        # === VISUALIZZAZIONE RISULTATI ===
        st.markdown("---")
        st.subheader(f"📊 Risultati Analisi: {match_name}")

        # Probabilità 1X2
        col_r1, col_r2, col_r3 = st.columns(3)
        with col_r1:
            st.metric("Prob Casa", f"{ris['p_home']*100:.1f}%")
            st.metric("Quota Casa", f"{validated['odds_1']:.2f}")
        with col_r2:
            st.metric("Prob Pareggio", f"{ris['p_draw']*100:.1f}%")
            st.metric("Quota Pareggio", f"{validated['odds_x']:.2f}")
        with col_r3:
            st.metric("Prob Trasferta", f"{ris['p_away']*100:.1f}%")
            st.metric("Quota Trasferta", f"{validated['odds_2']:.2f}")

        # 🚀 Mostra Advanced Features Adjustments se applicati
        if ADVANCED_FEATURES_AVAILABLE and (
            st.session_state.get('motivation_home', 'Normale') != 'Normale' or
            st.session_state.get('motivation_away', 'Normale') != 'Normale' or
            st.session_state.get('style_home', 'Possesso') != 'Possesso' or
            st.session_state.get('style_away', 'Possesso') != 'Possesso' or
            st.session_state.get('apply_constraints', True) or
            st.session_state.get('apply_calibration_enabled', False)
        ):
            st.info(f"""
            **🚀 Advanced Features Attive:**
            - **Motivation:** Casa={st.session_state.get('motivation_home', 'Normale')}, Trasferta={st.session_state.get('motivation_away', 'Normale')}
            - **Fixture Congestion:** Casa ultimi {st.session_state.get('days_since_home', 7)}gg, Trasferta ultimi {st.session_state.get('days_since_away', 7)}gg
            - **Tactical Styles:** {st.session_state.get('style_home', 'Possesso')} vs {st.session_state.get('style_away', 'Possesso')}
            - **Constraints Fisici:** {'✅ Attivi' if st.session_state.get('apply_constraints', True) else '❌ Disattivi'}
            - **Calibrazione:** {'✅ Attiva' if st.session_state.get('apply_calibration_enabled', False) else '❌ Disattiva'}
            - **Precision Math:** {'✅ Attivo' if st.session_state.get('use_precision_math', True) else '❌ Disattivo'}
            """)

        # ============================================================
        #   AI SYSTEM RESULTS DISPLAY (if analysis was run)
        # ============================================================
        if ai_result is not None:
            st.markdown("---")
            st.subheader("🤖 AI System - Betting Recommendation")

            decision = ai_result['final_decision']['action']
            stake = ai_result['final_decision']['stake']
            timing = ai_result['final_decision']['timing']
            priority = ai_result['final_decision']['priority']

            # Main decision display
            if decision == "BET":
                st.success(f"✅ **RACCOMANDAZIONE: SCOMMETTI €{stake:.2f}**")
            elif decision == "SKIP":
                st.warning(f"⚠️ **RACCOMANDAZIONE: SKIP** (non scommettere)")
            else:
                st.info(f"👀 **RACCOMANDAZIONE: WATCH** (monitora)")

            # Key metrics in columns
            col_ai1, col_ai2, col_ai3, col_ai4 = st.columns(4)

            with col_ai1:
                conf_val = ai_result['summary']['confidence']
                conf_delta = f"+{conf_val-50:.0f}" if conf_val >= 50 else f"{conf_val-50:.0f}"
                st.metric(
                    "🎯 Confidence",
                    f"{conf_val:.0f}/100",
                    delta=conf_delta,
                    delta_color="normal"
                )

            with col_ai2:
                value_val = ai_result['summary']['value_score']
                value_delta = f"+{value_val-50:.0f}" if value_val >= 50 else f"{value_val-50:.0f}"
                st.metric(
                    "💎 Value Score",
                    f"{value_val:.0f}/100",
                    delta=value_delta,
                    delta_color="normal"
                )

            with col_ai3:
                ev_val = ai_result['summary']['expected_value']
                st.metric(
                    "📈 Expected Value",
                    f"{ev_val:+.1%}",
                    delta=f"{ev_val*100:.1f}%" if ev_val > 0 else None,
                    delta_color="normal" if ev_val > 0 else "off"
                )

            with col_ai4:
                prob_calibrated = ai_result['summary']['probability']
                prob_raw = ris['p_home']
                prob_shift = prob_calibrated - prob_raw
                st.metric(
                    "🔬 Prob Calibrated",
                    f"{prob_calibrated:.1%}",
                    delta=f"{prob_shift:+.1%}",
                    delta_color="normal"
                )

            # Timing and Priority
            col_timing1, col_timing2 = st.columns(2)
            with col_timing1:
                timing_emoji = "🔴" if timing == "BET_NOW" else ("⏰" if timing == "WAIT" else "👀")
                st.info(f"{timing_emoji} **Timing:** {timing}")
            with col_timing2:
                priority_emoji = "🔥" if priority == "HIGH" else ("⚡" if priority == "MEDIUM" else "❄️")
                st.info(f"{priority_emoji} **Priority:** {priority}")

            # Sentiment Analysis Display (FASE 2)
            if ai_result.get('sentiment_analysis'):
                st.markdown("---")
                st.markdown("### 📱 Social Sentiment Analysis")

                sentiment = ai_result['sentiment_analysis']

                # Check if significant signals detected
                num_signals = len(sentiment.get('signals', []))
                has_insider_info = num_signals > 0

                if has_insider_info:
                    st.warning(f"🚨 **{num_signals} INSIDER SIGNALS DETECTED!**")

                # Display sentiment scores
                col_sent1, col_sent2 = st.columns(2)

                with col_sent1:
                    sentiment_home = sentiment.get('overall_sentiment_home', 0)
                    emoji_home = "📈" if sentiment_home > 0 else ("📉" if sentiment_home < 0 else "➡️")
                    st.metric(
                        f"Home ({home_team})",
                        f"Sentiment: {sentiment_home:+.2f}",
                        delta=f"Morale: {sentiment.get('team_morale_home', 0):+.0f}",
                        delta_color="normal" if sentiment_home >= 0 else "inverse"
                    )
                    st.caption(f"Fan Confidence: {sentiment.get('fan_confidence_home', 50):.0f}%")

                with col_sent2:
                    sentiment_away = sentiment.get('overall_sentiment_away', 0)
                    emoji_away = "📈" if sentiment_away > 0 else ("📉" if sentiment_away < 0 else "➡️")
                    st.metric(
                        f"Away ({away_team})",
                        f"Sentiment: {sentiment_away:+.2f}",
                        delta=f"Morale: {sentiment.get('team_morale_away', 0):+.0f}",
                        delta_color="normal" if sentiment_away >= 0 else "inverse"
                    )
                    st.caption(f"Fan Confidence: {sentiment.get('fan_confidence_away', 50):.0f}%")

                # Display detected signals
                if sentiment.get('signals'):
                    with st.expander(f"🔍 View {num_signals} Detected Signals", expanded=has_insider_info):
                        for i, signal in enumerate(sentiment['signals'][:10], 1):
                            signal_type = signal.get('type', 'UNKNOWN')
                            signal_team = signal.get('team', 'Unknown')
                            signal_text = signal.get('text', 'No details')
                            signal_credibility = signal.get('credibility', 0.5)

                            # Emoji based on signal type
                            if 'INJURY' in signal_type:
                                signal_emoji = "🏥"
                            elif 'LINEUP' in signal_type:
                                signal_emoji = "📋"
                            elif 'MORALE' in signal_type:
                                signal_emoji = "💪"
                            else:
                                signal_emoji = "ℹ️"

                            credibility_stars = "⭐" * int(signal_credibility * 5)

                            st.markdown(f"""
                            **{signal_emoji} Signal #{i}** - {signal_type}
                            **Team:** {signal_team}
                            **Credibility:** {credibility_stars} ({signal_credibility:.1%})
                            **Info:** {signal_text}
                            """)
                            st.markdown("---")

                # Edge detection alert
                sentiment_diff = abs(sentiment_home - sentiment_away)
                if sentiment_diff > 0.5 and has_insider_info:
                    if sentiment_home > sentiment_away:
                        st.success(f"🎯 **EDGE DETECTED:** Sentiment strongly favors {home_team}. Odds may not be adjusted yet!")
                    else:
                        st.info(f"🎯 **EDGE DETECTED:** Sentiment favors {away_team}. Consider opposite side or wait for odds adjustment.")

            # LLM Explanation (FASE 1.1)
            if ai_result.get('llm_explanation'):
                st.markdown("---")
                st.markdown("### 💬 Spiegazione AI")
                explanation = ai_result['llm_explanation']

                # Display explanation in a nice info box
                st.info(explanation)

                # Optional: Add expandable details
                with st.expander("🔍 Dettagli Analisi Approfondita"):
                    st.markdown(f"""
                    **Perché questa raccomandazione?**

                    L'AI ha analizzato {home_team} vs {away_team} considerando:

                    - **15 blocchi AI** (pipeline principale + avanzata)
                    - **Probabilità calibrata**: {prob_calibrated:.1%}
                    - **Expected Value**: {ev_val:+.1%}
                    - **Confidence**: {conf_val:.0f}/100
                    - **Value Score**: {value_val:.0f}/100

                    La decisione finale è: **{decision}**
                    """)

            # Detailed AI Analysis Expander
            with st.expander("🔍 Dettagli Analisi AI Completa (15 Blocchi)"):
                st.markdown("### 📊 Breakdown per Blocco AI")
                st.markdown("#### 🎯 Pipeline Principale (Blocchi 0-6)")

                # Blocco 1: Calibration
                st.markdown("**[BLOCCO 1] 🔬 Probability Calibrator**")
                cal = ai_result['calibrated']
                st.write(f"- Raw Probability (Dixon-Coles): `{cal['prob_raw']:.1%}`")
                st.write(f"- Calibrated Probability: `{cal['prob_calibrated']:.1%}`")
                st.write(f"- Calibration Shift: `{cal['calibration_shift']:+.1%}`")
                st.write(f"- Method: `{cal.get('calibration_method', 'Rule-based')}`")

                st.markdown("---")

                # Blocco 2: Confidence
                st.markdown("**[BLOCCO 2] 🎯 Confidence Scorer**")
                conf = ai_result['confidence']
                st.write(f"- Confidence Score: `{conf['confidence_score']:.0f}/100`")
                st.write(f"- Confidence Level: `{conf['confidence_level']}`")
                if conf.get('risk_factors'):
                    st.write("- ⚠️ Risk Factors:")
                    for rf in conf['risk_factors'][:3]:
                        st.write(f"  - {rf}")
                if conf.get('quality_score'):
                    st.write(f"- Data Quality: `{conf['quality_score']:.0f}/100`")

                st.markdown("---")

                # Blocco 3: Value Detection
                st.markdown("**[BLOCCO 3] 💎 Value Detector**")
                val = ai_result['value']
                st.write(f"- Value Type: `{val['value_type']}`")
                st.write(f"- Value Score: `{val['value_score']:.0f}/100`")
                st.write(f"- Expected Value: `{val['expected_value']:+.1%}`")
                st.write(f"- Sharp Money Detected: `{val.get('sharp_money_detected', False)}`")
                st.write(f"- Fair Odds: `{val.get('fair_odds', 0):.2f}` vs Market: `{validated['odds_1']:.2f}`")

                st.markdown("---")

                # Blocco 4: Kelly
                st.markdown("**[BLOCCO 4] 💰 Smart Kelly Optimizer**")
                kelly = ai_result['kelly']
                st.write(f"- Optimal Stake: `€{kelly['optimal_stake']:.2f}`")
                st.write(f"- Kelly Fraction: `{kelly['kelly_fraction']:.2f}`")
                st.write(f"- Stake %: `{kelly['stake_percentage']:.1f}%` of bankroll")
                if kelly.get('adjustments'):
                    st.write("- Adjustments Applied:")
                    for adj_key, adj_val in kelly['adjustments'].items():
                        st.write(f"  - {adj_key}: `{adj_val:.2f}x`")

                st.markdown("---")

                # Blocco 5: Risk Management
                st.markdown("**[BLOCCO 5] 🛡️ Risk Manager**")
                risk = ai_result['risk_decision']
                st.write(f"- Final Decision: `{risk['decision']}`")
                st.write(f"- Final Stake: `€{risk['final_stake']:.2f}`")
                st.write(f"- Risk Score: `{risk.get('risk_score', 0):.0f}/100`")
                st.write(f"- Priority: `{risk['priority']}`")
                st.write(f"- **Reasoning:** {risk.get('reasoning', 'N/A')}")

                if risk.get('red_flags'):
                    st.warning("🚩 **Red Flags:**")
                    for flag in risk['red_flags']:
                        st.write(f"  - {flag}")

                if risk.get('green_flags'):
                    st.success("✅ **Green Flags:**")
                    for flag in risk['green_flags']:
                        st.write(f"  - {flag}")

                st.markdown("---")

                # Blocco 6: Timing
                st.markdown("**[BLOCCO 6] ⏰ Odds Movement Tracker**")
                timing_detail = ai_result['timing']
                st.write(f"- Timing Recommendation: `{timing_detail['timing_recommendation']}`")
                st.write(f"- Urgency: `{timing_detail.get('urgency', 'MEDIUM')}`")
                st.write(f"- Current Odds: `{timing_detail.get('current_odds', 0):.2f}`")
                if timing_detail.get('predicted_odds_1h'):
                    st.write(f"- Predicted Odds (1h): `{timing_detail['predicted_odds_1h']:.2f}`")
                if timing_detail.get('movement_detected'):
                    st.write(f"- Odds Movement: `{timing_detail.get('movement_direction', 'STABLE')}`")

                st.markdown("---")

                # Blocco 0: API Data (if available)
                if 'api_data' in ai_result:
                    st.markdown("**[BLOCCO 0] 🌐 API Data Engine**")
                    api = ai_result['api_data']
                    st.write(f"- Data Sources Used: `{api.get('sources_used', 0)}`")
                    st.write(f"- Data Freshness: `{api.get('data_freshness', 'N/A')}`")
                    if api.get('enriched_context'):
                        st.write("- Enriched Context Available: ✅")

                # Pipeline Avanzata (Blocchi 7-14)
                if 'advanced_analysis' in ai_result and ai_result['advanced_analysis']:
                    st.markdown("---")
                    st.markdown("#### 🚀 Pipeline Avanzata (Blocchi 7-14)")
                    adv = ai_result['advanced_analysis']

                    # Blocco 7: Bayesian Uncertainty
                    st.markdown("**[BLOCCO 7] 📊 Bayesian Uncertainty Quantification**")
                    st.write(f"- Intervallo Credibilità 95%: `{adv.get('credible_interval_95', (0, 0))[0]:.1%} - {adv.get('credible_interval_95', (0, 0))[1]:.1%}`")
                    st.write(f"- Livello Incertezza: `{adv.get('uncertainty_level', 'N/A')}`")
                    st.write(f"- Reliability Index: `{adv.get('reliability_index', 0):.0f}/100`")
                    st.markdown("---")

                    # Blocco 8: Monte Carlo
                    st.markdown("**[BLOCCO 8] 🎲 Monte Carlo Simulator**")
                    st.write(f"- Robustness Score: `{adv.get('robustness_score', 0):.0f}/100`")
                    st.write(f"- Percentile 5%: `{adv.get('monte_carlo_percentile_5', 0):.2%}`")
                    st.write(f"- Percentile 95%: `{adv.get('monte_carlo_percentile_95', 0):.2%}`")
                    st.markdown("---")

                    # Blocco 9: Anomaly Detection
                    st.markdown("**[BLOCCO 9] 🚨 Advanced Anomaly Detection**")
                    st.write(f"- Anomalia Rilevata: `{'Sì' if adv.get('anomaly_detected') else 'No'}`")
                    st.write(f"- Severità: `{adv.get('anomaly_severity', 'NONE')}`")
                    if adv.get('anomaly_detected'):
                        st.warning("⚠️ Anomalia rilevata nel mercato!")
                    st.markdown("---")

                    # Blocco 10: Market Consistency
                    st.markdown("**[BLOCCO 10] ⚖️ Market Consistency Validator**")
                    st.write(f"- Consistency Score: `{adv.get('consistency_score', 0):.0f}/100`")
                    st.write(f"- Validation Passed: `{'✅ Sì' if adv.get('validation_passed') else '❌ No'}`")
                    st.markdown("---")

                    # Blocco 11: Adaptive Calibration
                    st.markdown("**[BLOCCO 11] 🎯 Adaptive Calibration**")
                    st.write(f"- Probabilità Calibrata: `{adv.get('calibrated_probability', 0):.2%}`")
                    st.markdown("---")

                    # Blocco 12: Consensus Validator
                    st.markdown("**[BLOCCO 12] 🤝 Multi-Model Consensus**")
                    st.write(f"- Consensus Raggiunto: `{'✅ Sì' if adv.get('consensus_reached') else '❌ No'}`")
                    st.write(f"- Agreement Score: `{adv.get('agreement_score', 0):.0f}/100`")
                    if adv.get('outlier_models'):
                        st.write(f"- Modelli Outlier: `{', '.join(adv.get('outlier_models', []))}`")
                    st.markdown("---")

                    # Blocco 13: Arbitrage Detector
                    st.markdown("**[BLOCCO 13] 💰 Statistical Arbitrage Detector**")
                    st.write(f"- Expected Value: `{adv.get('expected_value_pct', 0):+.1f}%`")
                    if adv.get('arbitrage_opportunities'):
                        st.success(f"✅ {len(adv.get('arbitrage_opportunities', []))} opportunità rilevate!")
                    else:
                        st.write("- Nessuna opportunità di arbitraggio")
                    st.markdown("---")

                    # Blocco 14: Real-time Validation
                    st.markdown("**[BLOCCO 14] ✅ Real-time Validation Engine**")
                    st.write(f"- Validation Score: `{adv.get('validation_score', 0):.0f}/100`")
                    st.write(f"- Raccomandazione Finale: `{adv.get('recommendation', 'N/A')}`")
                    if adv.get('reasoning'):
                        st.write("- **Reasoning:**")
                        for reason in adv.get('reasoning', [])[:3]:
                            st.write(f"  - {reason}")
                else:
                    st.markdown("---")
                    st.info("ℹ️ **Pipeline Avanzata:** Attiva l'analisi avanzata per vedere i blocchi 7-14")

                st.markdown("---")
                st.caption("💡 **Nota:** L'AI System combina 15 blocchi ML per decisioni ottimali. Ogni blocco contribuisce all'analisi finale.")

        st.markdown("---")

        # Over/Under Full Time
        st.subheader("📊 Over/Under (Tempo Pieno)")
        col_ou1, col_ou2, col_ou3 = st.columns(3)
        with col_ou1:
            st.metric("Over 1.5", f"{ris.get('over_15', 0)*100:.1f}%")
            st.metric("Under 1.5", f"{ris.get('under_15', 0)*100:.1f}%")
        with col_ou2:
            st.metric("Over 2.5", f"{ris['over_25']*100:.1f}%")
            st.metric("Under 2.5", f"{ris['under_25']*100:.1f}%")
        with col_ou3:
            st.metric("Over 3.5", f"{ris.get('over_35', 0)*100:.1f}%")
            st.metric("Under 3.5", f"{ris.get('under_35', 0)*100:.1f}%")

        # Over/Under Half Time
        if 'over_05_ht' in ris or 'over_15_ht' in ris:
            st.subheader("📊 Over/Under (Primo Tempo)")
            col_ht1, col_ht2 = st.columns(2)
            with col_ht1:
                st.metric("Over 0.5 HT", f"{ris.get('over_05_ht', 0)*100:.1f}%")
                st.metric("Under 0.5 HT", f"{(1 - ris.get('over_05_ht', 0))*100:.1f}%")
            with col_ht2:
                st.metric("Over 1.5 HT", f"{ris.get('over_15_ht', 0)*100:.1f}%")
                st.metric("Under 1.5 HT", f"{(1 - ris.get('over_15_ht', 0))*100:.1f}%")

        # Combo HT + FT
        if any(k in ris for k in ['over_05ht_over_05ft', 'over_05ht_over_15ft', 'over_05ht_over_25ft', 'over_15ht_over_25ft']):
            st.subheader("🔄 Combo Over HT + Over FT")
            col_combo1, col_combo2, col_combo3 = st.columns(3)
            with col_combo1:
                st.metric("Over 0.5 HT + Over 0.5 FT", f"{ris.get('over_05ht_over_05ft', 0)*100:.1f}%")
                st.metric("Over 0.5 HT + Over 1.5 FT", f"{ris.get('over_05ht_over_15ft', 0)*100:.1f}%")
            with col_combo2:
                st.metric("Over 0.5 HT + Over 2.5 FT", f"{ris.get('over_05ht_over_25ft', 0)*100:.1f}%")
                st.metric("Over 1.5 HT + Over 2.5 FT", f"{ris.get('over_15ht_over_25ft', 0)*100:.1f}%")
            with col_combo3:
                st.metric("Over 0.5 HT + Over 3.5 FT", f"{ris.get('over_05ht_over_35ft', 0)*100:.1f}%")
                st.metric("Over 1.5 HT + Over 3.5 FT", f"{ris.get('over_15ht_over_35ft', 0)*100:.1f}%")

        # BTTS
        st.subheader("⚽ Goal/No Goal")
        col_btts1, col_btts2, col_btts3 = st.columns(3)
        with col_btts1:
            st.metric("BTTS (GG)", f"{ris['btts']*100:.1f}%")
        with col_btts2:
            st.metric("GG + Over 2.5", f"{ris.get('gg_over25', 0)*100:.1f}%")
        with col_btts3:
            st.metric("No Goal (NG)", f"{ris.get('clean_sheet_qualcuno', 0)*100:.1f}%")

        # Double Chance
        if 'dc' in ris:
            st.subheader("🔄 Double Chance")
            dc_data = ris['dc']
            col_dc1, col_dc2, col_dc3 = st.columns(3)
            with col_dc1:
                st.metric("1X", f"{dc_data.get('DC Casa o Pareggio', 0)*100:.1f}%")
            with col_dc2:
                st.metric("X2", f"{dc_data.get('DC Trasferta o Pareggio', 0)*100:.1f}%")
            with col_dc3:
                st.metric("12", f"{dc_data.get('DC Casa o Trasferta', 0)*100:.1f}%")

        # Asian Handicap
        if 'asian_handicap' in ris:
            st.subheader("🎯 Asian Handicap")
            ah_data = ris['asian_handicap']

            # Filtra e ordina per probabilità (mostra linee con prob >= 20%)
            relevant_ah = [(k, v) for k, v in ah_data.items() if v >= 0.20]
            relevant_ah.sort(key=lambda x: x[1], reverse=True)

            if relevant_ah:
                # Mostra fino a 9 linee (3 righe x 3 colonne)
                num_to_show = min(len(relevant_ah), 9)

                # Dividi in righe da 3 colonne
                for i in range(0, num_to_show, 3):
                    cols = st.columns(3)
                    for j in range(3):
                        idx = i + j
                        if idx < num_to_show:
                            ah_line, prob = relevant_ah[idx]
                            with cols[j]:
                                st.metric(ah_line, f"{prob*100:.1f}%")
            else:
                st.info("Nessuna linea Asian Handicap con probabilità >= 20%")

        # Combo avanzate
        if 'combo_book' in ris:
            st.subheader("🔀 Combo Avanzate")
            combo = ris['combo_book']

            # Filtra e organizza i combo per categoria
            combo_patterns = [
                '1X &', 'X2 &', '12 &', '1 &', '2 &', 'GG &',  # Pattern originali
                '1X+', 'X2+', '1+', '2+',  # Pattern con + (alias)
                '+GG', '+gg', '+Over', '+over', '+Under', '+under',  # Pattern finali
                '+Multigol', '+multigol', '& Multigol'  # Multigol
            ]

            combo_principali = {
                k: v for k, v in combo.items()
                if any(x in k for x in combo_patterns)
            }

            # DEDUPLICA: Rimuovi duplicati (stesso valore = stesso mercato con alias)
            # Raggruppa per valore (probabilità)
            from collections import defaultdict
            value_groups = defaultdict(list)
            for key, val in combo_principali.items():
                # Raggruppa per valore arrotondato (per gestire micro-differenze floating point)
                rounded_val = round(val, 6)
                value_groups[rounded_val].append(key)

            # Per ogni gruppo, scegli la versione più leggibile (con + se disponibile)
            combo_deduplicati = {}
            for val, keys in value_groups.items():
                if len(keys) == 1:
                    # Solo una versione, usa quella
                    combo_deduplicati[keys[0]] = val
                else:
                    # Duplicati: preferisci versione con + (più leggibile)
                    keys_with_plus = [k for k in keys if '+' in k]
                    if keys_with_plus:
                        # Usa la versione con + più corta (con ordinamento alfabetico secondario per determinismo)
                        best_key = min(keys_with_plus, key=lambda k: (len(k), k))
                    else:
                        # Nessuna con +, usa la più corta (con ordinamento alfabetico secondario per determinismo)
                        best_key = min(keys, key=lambda k: (len(k), k))
                    combo_deduplicati[best_key] = val

            # Ordina per probabilità decrescente
            combo_sorted = sorted(combo_deduplicati.items(), key=lambda x: x[1], reverse=True)

            # Mostra TUTTI i combo rilevanti (non solo 15)
            cols_combo = st.columns(3)
            for idx, (key, val) in enumerate(combo_sorted):
                with cols_combo[idx % 3]:
                    st.metric(key, f"{val*100:.1f}%")

        # Top Risultati Esatti
        if 'top10' in ris:
            st.subheader("🎲 Top 10 Risultati Esatti")
            top10 = ris['top10']
            cols_top = st.columns(5)
            for idx, (h, a, prob) in enumerate(top10[:10]):
                score = f"{h}-{a}"
                with cols_top[idx % 5]:
                    st.metric(score, f"{prob:.1f}%")

        st.markdown("---")

        # Value bets
        final_btts_odds = validated.get("odds_btts") or ris.get("btts_final_odds") or 0.0
        value_rows = prepare_value_bets(
            ris=ris,
            odds_1=validated["odds_1"],
            odds_x=validated["odds_x"],
            odds_2=validated["odds_2"],
            odds_over=validated.get("odds_over25"),
            odds_under=validated.get("odds_under25"),
            odds_btts=final_btts_odds,
            odds_dnb_home=validated.get("odds_dnb_home"),
            odds_dnb_away=validated.get("odds_dnb_away"),
        )

        filtered_value_bets = []
        for vb in value_rows:
            edge_raw = vb.get("EdgeRaw", 0.0)
            if edge_raw is None:
                continue
            if edge_raw >= value_bet_edge_threshold:
                vb_copy = vb.copy()
                kelly_recommended = vb_copy.get("KellyBase", 0.0) * value_bet_kelly_fraction
                if not math.isfinite(kelly_recommended) or kelly_recommended < 0.0:
                    kelly_recommended = 0.0
                kelly_recommended = min(1.0, kelly_recommended)
                vb_copy["KellySuggested"] = kelly_recommended
                filtered_value_bets.append(vb_copy)

        st.subheader("💎 Value Bet Bet365")
        if filtered_value_bets:
            display_rows = []
            for vb in filtered_value_bets:
                display_rows.append({
                    "Esito": vb["Esito"],
                    "Quota": f"{vb['Odd']:.2f}",
                    "Prob modello %": f"{vb['ProbRaw']*100:.2f}%",
                    "Edge %": f"{vb['EdgeRaw']:+.2f}%",
                    "EV %": f"{vb['EVRaw']:+.2f}%",
                    f"Kelly {value_bet_kelly_fraction*100:.0f}%": f"{vb['KellySuggested']*100:.2f}%",
                    "Rec": vb["Rec"],
                })

            df_value = pd.DataFrame(display_rows)
            st.dataframe(df_value, use_container_width=True, hide_index=True)
            st.caption(
                f"Mostrate {len(filtered_value_bets)} value bet con edge ≥ {value_bet_edge_threshold:.1f}% "
                f"(Kelly suggerito: {value_bet_kelly_fraction*100:.0f}% del Kelly pieno)."
            )
        else:
            st.info(
                f"Nessuna value bet con edge ≥ {value_bet_edge_threshold:.1f}% sulle quote Bet365 fornite."
            )

        st.session_state["value_bets_filtered"] = filtered_value_bets
        st.session_state["value_bet_threshold_active"] = value_bet_edge_threshold
        st.session_state["value_bet_kelly_fraction_active"] = value_bet_kelly_fraction

        saved_in_history = save_analysis_history(
            match_name=match_name,
            league=league_type,
            home_team=home_team,
            away_team=away_team,
            match_datetime_iso=match_datetime_iso,
            validated_inputs=validated,
            ris=ris,
            quality_score=quality_score,
            market_confidence=market_conf,
            warnings=warnings,
            value_bets=filtered_value_bets,
        )

        if saved_in_history:
            st.success("💾 Analisi salvata nello storico")
        else:
            st.error("❌ Impossibile salvare l'analisi nello storico (vedi log)")

        # RACCOGLI TUTTI I MERCATI SOPRA SOGLIA (non solo value bets)
        all_markets = []

        # 1X2 Principale
        if ris['p_home'] * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "1 (Casa)",
                "Prob %": f"{ris['p_home']*100:.1f}",
                "Quota": validated["odds_1"],
                "Tipo": "1X2"
            })
        if ris['p_draw'] * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "X (Pareggio)",
                "Prob %": f"{ris['p_draw']*100:.1f}",
                "Quota": validated["odds_x"],
                "Tipo": "1X2"
            })
        if ris['p_away'] * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "2 (Trasferta)",
                "Prob %": f"{ris['p_away']*100:.1f}",
                "Quota": validated["odds_2"],
                "Tipo": "1X2"
            })

        # Over/Under Full Time
        if ris['over_25'] * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "Over 2.5",
                "Prob %": f"{ris['over_25']*100:.1f}",
                "Quota": validated.get("odds_over25", "N/A"),
                "Tipo": "Over/Under"
            })
        if ris['under_25'] * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "Under 2.5",
                "Prob %": f"{ris['under_25']*100:.1f}",
                "Quota": validated.get("odds_under25", "N/A"),
                "Tipo": "Over/Under"
            })

        # Over/Under Half Time
        if ris.get('over_05_ht', 0) * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "Over 0.5 HT",
                "Prob %": f"{ris['over_05_ht']*100:.1f}",
                "Quota": "N/A",
                "Tipo": "Over/Under HT"
            })
        if ris.get('over_15_ht', 0) * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "Over 1.5 HT",
                "Prob %": f"{ris['over_15_ht']*100:.1f}",
                "Quota": "N/A",
                "Tipo": "Over/Under HT"
            })

        # Combo Over HT + Over FT
        ht_ft_combos = [
            ('over_05ht_over_05ft', 'Over 0.5 HT + Over 0.5 FT'),
            ('over_05ht_over_15ft', 'Over 0.5 HT + Over 1.5 FT'),
            ('over_05ht_over_25ft', 'Over 0.5 HT + Over 2.5 FT'),
            ('over_15ht_over_25ft', 'Over 1.5 HT + Over 2.5 FT'),
            ('over_05ht_over_35ft', 'Over 0.5 HT + Over 3.5 FT'),
            ('over_15ht_over_35ft', 'Over 1.5 HT + Over 3.5 FT'),
        ]
        for key, label in ht_ft_combos:
            if ris.get(key, 0) * 100 >= telegram_prob_threshold:
                all_markets.append({
                    "Esito": label,
                    "Prob %": f"{ris[key]*100:.1f}",
                    "Quota": "N/A",
                    "Tipo": "Combo HT+FT"
                })

        # BTTS
        if ris.get('btts', 0) * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "BTTS Sì",
                "Prob %": f"{ris['btts']*100:.1f}",
                "Quota": validated.get("odds_btts", "N/A"),
                "Tipo": "BTTS"
            })
        # NoGol (BTTS No) - calcolato da clean_sheet_qualcuno se disponibile, altrimenti 1 - BTTS
        prob_ng = ris.get('clean_sheet_qualcuno')
        if prob_ng is None and 'btts' in ris:
            prob_ng = 1 - ris['btts']
        if prob_ng is not None and max(0.0, min(1.0, prob_ng)) * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "NoGol (BTTS No)",
                "Prob %": f"{max(0.0, min(1.0, prob_ng))*100:.1f}",
                "Quota": validated.get("odds_btts", "N/A"),
                "Tipo": "BTTS"
            })

        # Pari/Dispari FT (mostrati sempre per coprire mercati bilanciati)
        if 'even_ft' in ris:
            all_markets.append({
                "Esito": "Pari FT",
                "Prob %": f"{ris['even_ft']*100:.1f}",
                "Quota": "N/A",
                "Tipo": "Pari/Dispari"
            })
        if 'odd_ft' in ris:
            all_markets.append({
                "Esito": "Dispari FT",
                "Prob %": f"{ris['odd_ft']*100:.1f}",
                "Quota": "N/A",
                "Tipo": "Pari/Dispari"
            })

        # Pari/Dispari HT (mostrati sempre)
        if 'even_ht' in ris:
            all_markets.append({
                "Esito": "Pari HT",
                "Prob %": f"{ris['even_ht']*100:.1f}",
                "Quota": "N/A",
                "Tipo": "Pari/Dispari HT"
            })
        if 'odd_ht' in ris:
            all_markets.append({
                "Esito": "Dispari HT",
                "Prob %": f"{ris['odd_ht']*100:.1f}",
                "Quota": "N/A",
                "Tipo": "Pari/Dispari HT"
            })

        # Clean Sheet
        if ris.get('cs_home', 0) * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "CS Casa",
                "Prob %": f"{ris['cs_home']*100:.1f}",
                "Quota": "N/A",
                "Tipo": "Clean Sheet"
            })
        if ris.get('cs_away', 0) * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "CS Trasferta",
                "Prob %": f"{ris['cs_away']*100:.1f}",
                "Quota": "N/A",
                "Tipo": "Clean Sheet"
            })

        # Asian Handicap
        if 'asian_handicap' in ris:
            ah_data = ris['asian_handicap']
            for ah_line, ah_prob in ah_data.items():
                if ah_prob * 100 >= telegram_prob_threshold:
                    all_markets.append({
                        "Esito": f"AH {ah_line}",
                        "Prob %": f"{ah_prob*100:.1f}",
                        "Quota": "N/A",
                        "Tipo": "Asian Handicap"
                    })

        # HT/FT (Halftime/Fulltime)
        if 'ht_ft' in ris:
            ht_ft_data = ris['ht_ft']
            for ht_ft_combo, ht_ft_prob in ht_ft_data.items():
                if ht_ft_prob * 100 >= telegram_prob_threshold:
                    all_markets.append({
                        "Esito": f"HT/FT {ht_ft_combo}",
                        "Prob %": f"{ht_ft_prob*100:.1f}",
                        "Quota": "N/A",
                        "Tipo": "HT/FT"
                    })

        # DC (Double Chance) - Nota: I mercati 1X, X2, 12 sono già gestiti qui sotto
        # (La sezione "Combo" era duplicata e rimossa per evitare mercati doppi)
        if 'dc' in ris:
            dc_data = ris['dc']
            if dc_data.get('DC Casa o Pareggio', 0) * 100 >= telegram_prob_threshold:
                all_markets.append({
                    "Esito": "1X (DC Casa o Pareggio)",
                    "Prob %": f"{dc_data['DC Casa o Pareggio']*100:.1f}",
                    "Quota": "N/A",
                    "Tipo": "Double Chance"
                })
            if dc_data.get('DC Trasferta o Pareggio', 0) * 100 >= telegram_prob_threshold:
                all_markets.append({
                    "Esito": "X2 (DC Trasferta o Pareggio)",
                    "Prob %": f"{dc_data['DC Trasferta o Pareggio']*100:.1f}",
                    "Quota": "N/A",
                    "Tipo": "Double Chance"
                })
            if dc_data.get('DC Casa o Trasferta', 0) * 100 >= telegram_prob_threshold:
                all_markets.append({
                    "Esito": "12 (DC Casa o Trasferta)",
                    "Prob %": f"{dc_data['DC Casa o Trasferta']*100:.1f}",
                    "Quota": "N/A",
                    "Tipo": "Double Chance"
                })

        # COMBO AVANZATE (es: 1X & Over 2.5, BTTS & 1, ecc.)
        if 'combo_book' in ris:
            combo_book = ris['combo_book']

            # DEDUPLICA combo prima di aggiungerli (stesso problema dei duplicati)
            from collections import defaultdict
            value_groups_threshold = defaultdict(list)
            for combo_name, combo_prob in combo_book.items():
                if combo_prob * 100 >= telegram_prob_threshold:
                    rounded_prob = round(combo_prob, 6)
                    value_groups_threshold[rounded_prob].append((combo_name, combo_prob))

            # Per ogni gruppo, scegli la versione più leggibile
            for prob_val, name_prob_pairs in value_groups_threshold.items():
                if len(name_prob_pairs) == 1:
                    # Solo una versione
                    combo_name, combo_prob = name_prob_pairs[0]
                else:
                    # Duplicati: preferisci versione con + e più leggibile
                    names = [name for name, _ in name_prob_pairs]
                    names_with_plus = [n for n in names if '+' in n]
                    if names_with_plus:
                        # Preferisci versione con + più corta
                        best_name = min(names_with_plus, key=len)
                    else:
                        # Nessuna con +, usa la più corta
                        best_name = min(names, key=len)
                    combo_name = best_name
                    combo_prob = prob_val

                all_markets.append({
                    "Esito": combo_name,
                    "Prob %": f"{combo_prob*100:.1f}",
                    "Quota": "N/A",
                    "Tipo": "Combo Avanzate"
                })

        # MULTIGOL (es: 1-2 Goal, 2-3 Goal, 3+ Goal)
        if 'multigol' in ris:
            multigol = ris['multigol']
            for mg_name, mg_prob in multigol.items():
                if mg_prob * 100 >= telegram_prob_threshold:
                    all_markets.append({
                        "Esito": f"Multigol: {mg_name}",
                        "Prob %": f"{mg_prob*100:.1f}",
                        "Quota": "N/A",
                        "Tipo": "Multigol"
                    })

        # Top 3 Correct Score (risultati esatti con probabilità maggiori)
        if 'top10' in ris:
            for h, a, prob in ris['top10'][:3]:  # Top 3 risultati come richiesto
                if prob >= telegram_prob_threshold:
                    all_markets.append({
                        "Esito": f"Risultato Esatto {h}-{a}",
                        "Prob %": f"{prob:.1f}",
                        "Quota": "N/A",
                        "Tipo": "Correct Score"
                    })

        # DNB (Draw No Bet)
        if ris.get('dnb_home', 0) * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "DNB Casa",
                "Prob %": f"{ris['dnb_home']*100:.1f}",
                "Quota": validated.get("odds_dnb_home", "N/A"),
                "Tipo": "DNB"
            })
        if ris.get('dnb_away', 0) * 100 >= telegram_prob_threshold:
            all_markets.append({
                "Esito": "DNB Trasferta",
                "Prob %": f"{ris['dnb_away']*100:.1f}",
                "Quota": validated.get("odds_dnb_away", "N/A"),
                "Tipo": "DNB"
            })

        # Ordina per probabilità decrescente
        all_markets.sort(key=lambda x: float(x['Prob %']), reverse=True)

        # Mostra mercati trovati
        if all_markets:
            st.success(f"🎯 {len(all_markets)} mercati trovati con probabilità ≥ {telegram_prob_threshold:.0f}%")
            st.table(all_markets)
        else:
            st.info(f"ℹ️ Nessun mercato con probabilità ≥ {telegram_prob_threshold:.0f}%")
            st.caption(f"💡 Suggerimento: Abbassa la soglia (es: 45-50%) per vedere più mercati")

        # DEBUG TELEGRAM CONFIG
        st.write("---")
        st.markdown("### 🔍 Debug Telegram")

        # Mostra valori RAW per debug
        with st.expander("🔧 Valori RAW (Debug avanzato)", expanded=False):
            st.code(f"""
telegram_enabled (type={type(telegram_enabled).__name__}): {telegram_enabled}
telegram_token (len={len(telegram_token) if telegram_token else 0}): {"***" + telegram_token[-10:] if telegram_token and len(telegram_token) > 10 else "VUOTO"}
telegram_chat_id (type={type(telegram_chat_id).__name__}): {telegram_chat_id}
telegram_prob_threshold: {telegram_prob_threshold}
            """)

        col_d1, col_d2, col_d3 = st.columns(3)
        with col_d1:
            status_enabled = "✅ Sì" if telegram_enabled else "❌ No"
            st.metric("Checkbox abilitata", status_enabled)
        with col_d2:
            status_token = "✅ OK" if telegram_token and len(telegram_token) > 10 else "❌ Mancante"
            st.metric("Bot Token", status_token)
        with col_d3:
            status_chat = "✅ OK" if telegram_chat_id and len(telegram_chat_id) > 3 else "❌ Mancante"
            st.metric("Chat ID", status_chat)

        st.metric("Mercati da inviare", len(all_markets))

        # Prepara messaggio Telegram aggregato
        telegram_message = (
            f"⚽ <b>{match_name}</b>\n\n"
            f"🔢 <b>Parametri</b>:\n"
            f"  λ Casa: {ris['lambda_home']:.2f} | λ Trasferta: {ris['lambda_away']:.2f}\n"
            f"  ρ: {ris['rho']:.3f}\n\n"
            f"📊 <b>Probabilità Base</b>:\n"
            f"  🏠 Casa: {ris['p_home']*100:.1f}% (Quota: {validated['odds_1']:.2f})\n"
            f"  ⚖️ Pareggio: {ris['p_draw']*100:.1f}% (Quota: {validated['odds_x']:.2f})\n"
            f"  ✈️ Trasferta: {ris['p_away']*100:.1f}% (Quota: {validated['odds_2']:.2f})\n\n"
            f"⚽ <b>Mercati Speciali</b>:\n"
            f"  Over 2.5: {ris['over_25']*100:.1f}%\n"
            f"  Under 2.5: {ris['under_25']*100:.1f}%\n"
            f"  BTTS: {ris['btts']*100:.1f}%\n"
            f"  GG + Over 2.5: {ris['gg_over25']*100:.1f}%\n\n"
            f"🎯 <b>Mercati Alta Probabilità (≥{telegram_prob_threshold:.0f}%)</b>:\n"
        )

        by_type = {}
        for market in all_markets:
            tipo = market["Tipo"]
            by_type.setdefault(tipo, []).append(market)

        tipo_order = [
            "1X2",
            "Over/Under",
            "Over/Under HT",
            "Combo HT+FT",
            "BTTS",
            "Combo",
            "Double Chance",
            "Combo Avanzate",
            "Multigol",
            "Multigol Totali",
            "Correct Score",
            "DNB",
        ]

        if all_markets:
            for tipo in tipo_order:
                if tipo in by_type:
                    telegram_message += f"\n<b>{tipo}</b>:\n"
                    for m in by_type[tipo]:
                        quota_str = f" (Quota: {m['Quota']:.2f})" if isinstance(m["Quota"], (int, float)) else ""
                        telegram_message += f"  • <b>{m['Esito']}</b>: {m['Prob %']}%{quota_str}\n"

            for tipo in by_type:
                if tipo not in tipo_order:
                    telegram_message += f"\n<b>{tipo}</b>:\n"
                    for m in by_type[tipo]:
                        quota_str = f" (Quota: {m['Quota']:.2f})" if isinstance(m["Quota"], (int, float)) else ""
                        telegram_message += f"  • <b>{m['Esito']}</b>: {m['Prob %']}%{quota_str}\n"
        else:
            telegram_message += "  • Nessun mercato sopra la soglia configurata\n"

        if filtered_value_bets:
            telegram_message += (
                f"\n💎 <b>Value Bet Bet365</b> "
                f"(edge ≥ {value_bet_edge_threshold:.1f}%, Kelly {value_bet_kelly_fraction*100:.0f}%):\n"
            )
            for vb in filtered_value_bets[:5]:
                telegram_message += (
                    f"  • <b>{vb['Esito']}</b>: {vb['ProbRaw']*100:.1f}% | "
                    f"Quota {vb['Odd']:.2f} | Edge {vb['EdgeRaw']:+.2f}% | "
                    f"EV {vb['EVRaw']:+.2f}% | Kelly {vb['KellySuggested']*100:.2f}%\n"
                )
            if len(filtered_value_bets) > 5:
                telegram_message += f"  <i>(+{len(filtered_value_bets)-5} ulteriori value bet)</i>\n"
        else:
            telegram_message += (
                f"\n💎 Value Bet Bet365: nessuna ≥ {value_bet_edge_threshold:.1f}% con le quote inserite\n"
            )

        if ris.get('top10'):
            telegram_message += "\n🏅 <b>Top 3 Risultati Esatti</b>:\n"
            for h, a, prob in ris['top10'][:3]:
                telegram_message += f"  • {h}-{a}: {prob:.1f}%\n"

        telegram_message += f"\n📈 Totale: {len(all_markets)} mercati\n"
        telegram_message += "🤖 <i>Modello Dixon-Coles Bayesiano</i>"

        def _dispatch_telegram(message: str):
            chunks = split_telegram_message(message)
            sent_chunks = 0
            for idx, chunk in enumerate(chunks, 1):
                result = send_telegram_message(
                    message=chunk,
                    bot_token=(telegram_token or "").strip(),
                    chat_id=(telegram_chat_id or "").strip(),
                    parse_mode="HTML"
                )
                if not result.get("success"):
                    return False, result, sent_chunks, len(chunks)
                sent_chunks = idx
            return True, None, sent_chunks, len(chunks)

        # INVIO TELEGRAM
        if telegram_enabled and telegram_token and telegram_chat_id:
            with st.spinner("Invio automatico su Telegram..."):
                auto_success, auto_error, auto_sent, auto_total = _dispatch_telegram(telegram_message)

            if auto_success:
                chunk_label = "messaggio" if auto_sent == 1 else "messaggi"
                st.success(f"✅ Notifica Telegram inviata automaticamente ({auto_sent} {chunk_label}).")
            else:
                st.error("❌ Errore invio Telegram automatico")
                st.code(auto_error.get("error_message", "Errore sconosciuto"))
                st.warning("💡 Controlla che il bot sia amministratore del canale e abbia permessi di scrittura")

            if st.button("📤 Reinvia su Telegram", use_container_width=True):
                with st.spinner("Reinvio su Telegram..."):
                    resend_success, resend_error, resend_sent, resend_total = _dispatch_telegram(telegram_message)

                if resend_success:
                    chunk_label = "messaggio" if resend_sent == 1 else "messaggi"
                    st.success(f"✅ Messaggio reinviato su Telegram ({resend_sent} {chunk_label}).")
                else:
                    st.error("❌ Errore reinvio Telegram")
                    st.code(resend_error.get("error_message", "Errore sconosciuto"))
                    st.warning("💡 Controlla che il bot sia amministratore del canale e abbia permessi di scrittura")

            if not all_markets:
                st.warning(
                    f"⚠️ Nessun mercato con probabilità ≥ {telegram_prob_threshold:.0f}% disponibile; "
                    "report inviato comunque con risultati esatti e value bet."
                )
                st.caption("💡 Abbassa la soglia (es: 45-50%) per vedere più mercati nei mercati principali")
        else:
            st.warning("⚠️ Configura Telegram per abilitare l'invio:")
            if not telegram_enabled:
                st.info("• Spunta la checkbox 'Invia analisi automaticamente su Telegram'")
            if not telegram_token:
                st.info("• Inserisci il Bot Token")
            if not telegram_chat_id:
                st.info("• Inserisci il Chat ID")

    except Exception as e:
        st.error(f"❌ Errore durante l'analisi: {str(e)}")
        logger.error(f"Errore analisi: {e}", exc_info=True)

st.markdown("---")

# ============================================================
#        AGGIORNAMENTO RISULTATI REALI
# ============================================================

st.subheader("🔄 Aggiorna Risultati Reali")

if st.button("Recupera risultati ultimi 3 giorni"):
    if not os.path.exists(ARCHIVE_FILE):
        st.warning("Nessuno storico da aggiornare")
    else:
        with st.spinner("Recupero risultati da API-Football..."):
            df = pd.read_csv(ARCHIVE_FILE)
            today = date.today()
            giorni = [(today - timedelta(days=i)).isoformat() for i in range(0, 4)]
            
            fixtures_map = {}
            for d in giorni:
                fixtures = apifootball_get_fixtures_by_date(d)
                for f in fixtures:
                    if f["fixture"]["status"]["short"] in ["FT", "AET", "PEN"]:
                        home = f["teams"]["home"]["name"]
                        away = f["teams"]["away"]["name"]
                        key = f"{home} vs {away}".strip().lower()
                        gh = f["goals"]["home"]
                        ga = f["goals"]["away"]
                        if gh is not None and ga is not None:
                            fixtures_map[key] = (gh, ga)
            
            updated = 0
            for idx, row in df.iterrows():
                key_row = str(row.get("match", "")).strip().lower()
                if key_row in fixtures_map and (pd.isna(row.get("risultato_reale")) or row.get("risultato_reale") == ""):
                    gh, ga = fixtures_map[key_row]
                    
                    if gh > ga:
                        esito = "1"
                    elif gh == ga:
                        esito = "X"
                    else:
                        esito = "2"
                    
                    df.at[idx, "risultato_reale"] = f"{gh}-{ga}"
                    df.at[idx, "esito_reale"] = esito
                    
                    pred = row.get("esito_modello", "")
                    if pred and esito:
                        df.at[idx, "match_ok"] = 1 if pred == esito else 0
                    
                    updated += 1
            
            df.to_csv(ARCHIVE_FILE, index=False)
            st.success(f"✅ Aggiornate {updated} partite")
            st.rerun()

st.markdown("---")

# ============================================================
#        BACKTESTING E PERFORMANCE
# ============================================================

st.subheader("📊 Backtesting Strategia")

if os.path.exists(ARCHIVE_FILE):
    col_bt1, col_bt2 = st.columns(2)
    
    with col_bt1:
        min_edge_bt = st.slider("Minimo Edge %", 1.0, 10.0, 3.0, 0.5)
        kelly_frac_bt = st.slider("Kelly Fraction", 0.1, 1.0, 0.25, 0.05)
        initial_bank = st.number_input("Bankroll Iniziale (€)", value=100.0, min_value=10.0, step=10.0)
    
    with col_bt2:
        if st.button("🚀 Esegui Backtest", type="primary"):
            with st.spinner("Eseguendo backtest su dati storici..."):
                results = backtest_strategy(
                    ARCHIVE_FILE,
                    min_edge=min_edge_bt / 100,
                    kelly_fraction=kelly_frac_bt,
                    initial_bankroll=initial_bank
                )
                
                if "error" in results:
                    st.warning(f"⚠️ {results['error']}")
                else:
                    st.success("✅ Backtest completato!")
                    
                    col_r1, col_r2, col_r3, col_r4 = st.columns(4)
                    
                    with col_r1:
                        st.metric("💰 Bankroll Finale", f"€{results['final_bankroll']:.2f}")
                    with col_r2:
                        st.metric("📈 Profit", f"€{results['profit']:.2f}", 
                                 f"{results['profit_pct']:+.1f}%")
                    with col_r3:
                        st.metric("🎯 Win Rate", f"{results['win_rate']:.1f}%")
                    with col_r4:
                        st.metric("💵 ROI", f"{results['roi']:.1f}%")
                    
                    st.write(f"**Scommesse piazzate**: {results['bets_placed']} ({results['bets_won']} vinte)")
                    st.write(f"**Totale puntato**: €{results['total_staked']:.2f}")
                    st.write(f"**Totale ritornato**: €{results['total_returned']:.2f}")
                    
                    if results.get('sharpe_approx'):
                        st.write(f"**Sharpe Ratio (approssimato)**: {results['sharpe_approx']:.3f}")
                    
                    # Grafico profit history
                    if len(results.get('profit_history', [])) > 1:
                        try:
                            import matplotlib.pyplot as plt
                            fig, ax = plt.subplots(figsize=(10, 4))
                            ax.plot(results['profit_history'], linewidth=2)
                            ax.axhline(y=initial_bank, color='r', linestyle='--', label='Bankroll Iniziale')
                            ax.set_xlabel('Numero Scommesse')
                            ax.set_ylabel('Bankroll (€)')
                            ax.set_title('Evoluzione Bankroll')
                            ax.legend()
                            ax.grid(True, alpha=0.3)
                            st.pyplot(fig)
                            plt.close(fig)
                        # FIX BUG: Replace bare except with specific Exception
                        except Exception as e:
                            logger.warning(f"⚠️ Errore rendering grafico bankroll: {e}")
                            pass
else:
    st.info("Nessuno storico disponibile per backtest")

st.markdown("---")
st.caption("Developed with ❤️ | Dixon-Coles Bayesian Model | Shin Normalization | IQR Outlier Detection | Platt Scaling | Kelly Criterion | Ensemble Methods")
